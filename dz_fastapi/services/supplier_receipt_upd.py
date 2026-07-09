import asyncio
import re
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload, selectinload

from dz_fastapi.core.time import now_moscow
from dz_fastapi.crud.email_account import crud_email_account
from dz_fastapi.crud.settings import crud_customer_order_inbox_settings
from dz_fastapi.models.autopart import preprocess_oem_number
from dz_fastapi.models.partner import SupplierReceipt, SupplierReceiptItem
from dz_fastapi.models.settings import DiadocIntegrationSettings
from dz_fastapi.services.email import build_email_delivery_kwargs, send_email_with_attachment


def _money(value: Any) -> Decimal:
    if value is None:
        return Decimal("0.00")
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except Exception:
        return Decimal("0.00")


def _qty(value: Any) -> Decimal:
    if value is None:
        return Decimal("0")
    try:
        return Decimal(str(value))
    except Exception:
        return Decimal("0")


def _safe_filename_part(value: str) -> str:
    value = re.sub(r"[^\wа-яА-ЯёЁ.-]+", "_", str(value or ""), flags=re.U)
    value = value.strip("._")
    return value[:80] or "supplier"


def _date_text(value: date | None) -> str:
    actual = value or now_moscow().date()
    return actual.strftime("%d.%m.%Y")


def _document_date(receipt: SupplierReceipt) -> date:
    return receipt.document_date or now_moscow().date()


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def _receipt_item_history_key(item: SupplierReceiptItem) -> tuple[str, str] | None:
    oem = _clean_text(getattr(item, "oem_number", None))
    brand = _clean_text(getattr(item, "brand_name", None)).upper()
    if not oem or not brand:
        return None
    return preprocess_oem_number(oem), brand


async def _load_receipt(
    session: AsyncSession,
    receipt_id: int,
) -> SupplierReceipt:
    stmt = (
        select(SupplierReceipt)
        .where(SupplierReceipt.id == receipt_id)
        .options(
            joinedload(SupplierReceipt.provider),
            joinedload(SupplierReceipt.warehouse),
            selectinload(SupplierReceipt.items),
        )
    )
    receipt = (await session.execute(stmt)).unique().scalar_one_or_none()
    if receipt is None:
        raise LookupError(f"Документ поступления #{receipt_id} не найден")
    return receipt


async def _load_buyer_profile(session: AsyncSession) -> dict[str, str]:
    result = await session.execute(select(DiadocIntegrationSettings).limit(1))
    settings = result.scalar_one_or_none()
    if settings is None:
        return {}
    return {
        "name": settings.organization_name or "",
        "inn": settings.organization_inn or "",
        "kpp": settings.organization_kpp or "",
        "address": (
            settings.seller_legal_address
            or settings.seller_postal_address
            or ""
        ),
    }


async def _load_historical_gtd_overrides(
    session: AsyncSession,
    receipt: SupplierReceipt,
) -> dict[int, dict[str, str]]:
    missing_items = [
        item
        for item in receipt.items or []
        if not _clean_text(getattr(item, "gtd_code", None))
    ]
    keys = {
        key
        for item in missing_items
        if (key := _receipt_item_history_key(item)) is not None
    }
    if not keys:
        return {}

    raw_oems = {
        _clean_text(getattr(item, "oem_number", None))
        for item in missing_items
        if _clean_text(getattr(item, "oem_number", None))
    }
    raw_oems.update(preprocess_oem_number(oem) for oem in list(raw_oems))
    if not raw_oems:
        return {}

    stmt = (
        select(SupplierReceiptItem)
        .where(SupplierReceiptItem.receipt_id != receipt.id)
        .where(SupplierReceiptItem.oem_number.in_(raw_oems))
        .where(SupplierReceiptItem.gtd_code.is_not(None))
        .where(SupplierReceiptItem.gtd_code != "")
        .order_by(SupplierReceiptItem.id.desc())
        .limit(2000)
    )
    historical_items = (await session.execute(stmt)).scalars().all()
    candidates_by_key: dict[tuple[str, str], list[dict[str, str]]] = {}
    seen_by_key: dict[tuple[str, str], set[tuple[str, str, str]]] = {}

    for item in historical_items:
        key = _receipt_item_history_key(item)
        if key not in keys:
            continue
        gtd_code = _clean_text(item.gtd_code)
        if not gtd_code:
            continue
        country_code = _clean_text(item.country_code)
        country_name = _clean_text(item.country_name)
        candidate_key = (gtd_code, country_code, country_name)
        seen = seen_by_key.setdefault(key, set())
        if candidate_key in seen:
            continue
        seen.add(candidate_key)
        candidates_by_key.setdefault(key, []).append(
            {
                "gtd_code": gtd_code,
                "country_code": country_code,
                "country_name": country_name,
            }
        )

    overrides: dict[int, dict[str, str]] = {}
    used_offsets: dict[tuple[str, str], int] = {}
    for item in missing_items:
        item_id = getattr(item, "id", None)
        key = _receipt_item_history_key(item)
        if item_id is None or key is None:
            continue
        candidates = candidates_by_key.get(key) or []
        if not candidates:
            continue
        offset = used_offsets.get(key, 0)
        overrides[int(item_id)] = candidates[offset % len(candidates)]
        used_offsets[key] = offset + 1
    return overrides


def build_supplier_receipt_upd_attachment(
    *,
    receipt: SupplierReceipt,
    buyer_profile: dict[str, str] | None = None,
    historical_gtd_by_item_id: dict[int, dict[str, str]] | None = None,
) -> bytes:
    provider = receipt.provider
    buyer_profile = buyer_profile or {}
    historical_gtd_by_item_id = historical_gtd_by_item_id or {}
    doc_date = _document_date(receipt)
    is_vat_payer = bool(getattr(provider, "is_vat_payer", False))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "УПД"
    sheet.freeze_panes = "A17"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="EAF2F8")
    title_font = Font(bold=True, size=14)
    bold = Font(bold=True)
    small = Font(size=9)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    right = Alignment(horizontal="right", vertical="top", wrap_text=True)

    sheet.merge_cells("A1:L1")
    sheet["A1"] = "Универсальный передаточный документ"
    sheet["A1"].font = title_font
    sheet["A1"].alignment = center

    sheet.merge_cells("A2:L2")
    sheet["A2"] = (
        "Печатная форма сформирована покупателем на основании документа "
        "поступления. Требуется подтверждение поставщика."
    )
    sheet["A2"].font = small
    sheet["A2"].alignment = center

    info_rows = [
        ("Статус", "2"),
        ("Счет-фактура N", "б/н"),
        ("Дата", _date_text(doc_date)),
        ("Продавец", getattr(provider, "name", None) or "Поставщик"),
        ("ИНН/КПП продавца", " / ".join(
            part for part in [
                getattr(provider, "inn", None),
                getattr(provider, "kpp", None),
            ] if part
        ) or "—"),
        ("Покупатель", buyer_profile.get("name") or "Dragonzap"),
        ("ИНН/КПП покупателя", " / ".join(
            part for part in [
                buyer_profile.get("inn"),
                buyer_profile.get("kpp"),
            ] if part
        ) or "—"),
        ("Адрес покупателя", buyer_profile.get("address") or "—"),
        ("Склад", receipt.warehouse_name or "—"),
        ("Дата поступления", _date_text(doc_date)),
    ]
    row_num = 4
    for label, value in info_rows:
        sheet.cell(row=row_num, column=1, value=label).font = bold
        sheet.cell(row=row_num, column=2, value=value)
        sheet.merge_cells(
            start_row=row_num,
            start_column=2,
            end_row=row_num,
            end_column=12,
        )
        for col in range(1, 13):
            cell = sheet.cell(row=row_num, column=col)
            cell.alignment = left
            cell.border = border
        row_num += 1

    headers = [
        "N",
        "Код товара/артикул",
        "Бренд",
        "Наименование",
        "Ед.",
        "Кол-во",
        "Цена",
        "Стоимость без НДС",
        "НДС",
        "Стоимость с НДС",
        "Страна",
        "ГТД",
    ]
    header_row = row_num + 1
    for col, title in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=col, value=title)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    total_qty = Decimal("0")
    total_without_vat = Decimal("0.00")
    total_vat = Decimal("0.00")
    total_with_vat = Decimal("0.00")
    vat_rate = Decimal("0.20")
    current_row = header_row + 1
    for index, item in enumerate(receipt.items or [], start=1):
        historical_gtd = historical_gtd_by_item_id.get(int(item.id or 0), {})
        quantity = _qty(item.received_quantity)
        unit_price_with_vat = _money(item.price)
        line_with_vat = _money(item.total_price_with_vat)
        if line_with_vat <= 0:
            line_with_vat = (quantity * unit_price_with_vat).quantize(
                Decimal("0.01")
            )
        if is_vat_payer and line_with_vat > 0:
            line_without_vat = (
                line_with_vat / (Decimal("1.00") + vat_rate)
            ).quantize(Decimal("0.01"))
            vat_amount = line_with_vat - line_without_vat
            vat_label = f"{vat_amount:.2f}"
        else:
            line_without_vat = line_with_vat
            vat_amount = Decimal("0.00")
            vat_label = "без НДС"

        values = [
            index,
            item.oem_number or "",
            item.brand_name or "",
            item.autopart_name or "",
            "шт",
            float(quantity),
            float(unit_price_with_vat),
            float(line_without_vat),
            vat_label,
            float(line_with_vat),
            (
                item.country_name
                or item.country_code
                or historical_gtd.get("country_name")
                or historical_gtd.get("country_code")
                or ""
            ),
            item.gtd_code or historical_gtd.get("gtd_code") or "",
        ]
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row=current_row, column=col, value=value)
            cell.border = border
            cell.alignment = right if col in {6, 7, 8, 10} else left
            if col in {7, 8, 10}:
                cell.number_format = "#,##0.00"
        total_qty += quantity
        total_without_vat += line_without_vat
        total_vat += vat_amount
        total_with_vat += line_with_vat
        current_row += 1

    sheet.cell(row=current_row, column=1, value="Итого").font = bold
    sheet.merge_cells(
        start_row=current_row,
        start_column=1,
        end_row=current_row,
        end_column=5,
    )
    totals = {
        6: float(total_qty),
        8: float(total_without_vat),
        9: float(total_vat) if is_vat_payer else "без НДС",
        10: float(total_with_vat),
    }
    for col in range(1, 13):
        cell = sheet.cell(row=current_row, column=col)
        if col in totals:
            cell.value = totals[col]
        cell.font = bold
        cell.border = border
        cell.alignment = right if col in {6, 8, 9, 10} else left
        if col in {8, 10} or (is_vat_payer and col == 9):
            cell.number_format = "#,##0.00"

    sign_row = current_row + 3
    sheet.merge_cells(
        start_row=sign_row,
        start_column=1,
        end_row=sign_row,
        end_column=6,
    )
    sheet.cell(
        row=sign_row,
        column=1,
        value="Ответственный за оформление: __________________ / __________",
    )
    sheet.merge_cells(
        start_row=sign_row,
        start_column=7,
        end_row=sign_row,
        end_column=12,
    )
    sheet.cell(
        row=sign_row,
        column=7,
        value="Поставщик подтвердил: __________________ / __________",
    )

    widths = [6, 20, 16, 42, 8, 10, 12, 16, 14, 16, 18, 22]
    for col, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(col)].width = width

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def supplier_receipt_upd_filename(receipt: SupplierReceipt) -> str:
    provider_name = getattr(receipt.provider, "name", None) or "supplier"
    date_part = _document_date(receipt).isoformat()
    return f"UPD_{_safe_filename_part(provider_name)}_{date_part}.xlsx"


async def send_supplier_receipt_upd_email(
    session: AsyncSession,
    *,
    receipt_id: int,
    force: bool = False,
) -> dict[str, Any]:
    settings = await crud_customer_order_inbox_settings.get_or_create(session)
    enabled = bool(
        getattr(settings, "supplier_receipt_upd_email_enabled", False)
    )
    if not enabled and not force:
        return {"sent": False, "skipped": True, "reason": "disabled"}

    to_email = str(
        getattr(settings, "supplier_receipt_upd_email", "") or ""
    ).strip()
    if not to_email:
        raise ValueError("Не указан email для отправки УПД поставщика")

    receipt = await _load_receipt(session, receipt_id)
    buyer_profile = await _load_buyer_profile(session)
    historical_gtd_by_item_id = await _load_historical_gtd_overrides(
        session,
        receipt,
    )
    attachment = build_supplier_receipt_upd_attachment(
        receipt=receipt,
        buyer_profile=buyer_profile,
        historical_gtd_by_item_id=historical_gtd_by_item_id,
    )

    account = None
    account_id = getattr(settings, "supplier_receipt_upd_email_account_id", None)
    if account_id:
        account = await crud_email_account.get(session, int(account_id))
        if account is None or not bool(getattr(account, "is_active", False)):
            raise ValueError("Выбранный исходящий почтовый ящик неактивен")
    if account is None:
        accounts = await crud_email_account.get_active_by_purpose(
            session, "orders_out"
        )
        account = accounts[0] if accounts else None

    smtp_kwargs = build_email_delivery_kwargs(account) if account else {}
    provider_name = (
        str(getattr(receipt.provider, "name", "") or "").strip()
        or "Поставщик"
    )
    body = (
        "<p>Во вложении сформированная печатная форма УПД по документу "
        f"поступления #{receipt.id} от {_date_text(_document_date(receipt))}."
        "</p>"
        "<p>Документ сформирован покупателем на основании фактического "
        "поступления и требует подтверждения поставщика.</p>"
    )
    sent = await asyncio.to_thread(
        send_email_with_attachment,
        to_email,
        provider_name,
        body,
        attachment,
        supplier_receipt_upd_filename(receipt),
        True,
        **smtp_kwargs,
    )
    if not sent:
        raise RuntimeError("Не удалось отправить письмо с УПД")
    return {
        "sent": True,
        "skipped": False,
        "to_email": to_email,
        "subject": provider_name,
        "filename": supplier_receipt_upd_filename(receipt),
    }
