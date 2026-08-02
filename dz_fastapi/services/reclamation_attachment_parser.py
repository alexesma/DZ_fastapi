"""Structured extraction from customer reclamation attachments."""

from __future__ import annotations

import logging
import os
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Optional

import xlrd
from openpyxl import load_workbook

logger = logging.getLogger("dz_fastapi")

MAX_RECLAMATION_ATTACHMENT_BYTES = 15 * 1024 * 1024
MAX_SHEET_ROWS = 500
MAX_SHEET_COLS = 200

_SOURCE_DOCUMENT_RE = re.compile(
    r"(?:УПД|сч[её]т-фактура|товарн\w*\s+накладн\w*)\s*№?\s*"
    r"([A-Za-zА-Яа-я0-9][A-Za-zА-Яа-я0-9_./-]{0,79})\s+от\s+"
    r"(\d{1,2}[./-]\d{1,2}[./-]\d{2,4}|"
    r"\d{1,2}\s+[А-Яа-яЁё]+\s+\d{4}(?:\s*г\.?)?)",
    re.IGNORECASE,
)
_ITEM_REASON_RE = re.compile(
    r"^\s*([A-Za-zА-Яа-я0-9][A-Za-zА-Яа-я0-9./-]{3,39})\s*" r"[-–—]\s*(.+?)\s*$"
)
_RETURN_BASIS_RE = re.compile(
    r"возврат\s+по\s+(?:сч[её]ту-фактуре|упд)\s*"
    r"(?:no|№)?\s*([A-Za-zА-Яа-я0-9][A-Za-zА-Яа-я0-9_./-]{0,79})\s+от\s+"
    r"(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})",
    re.IGNORECASE,
)


def _normalize_oem(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9]", "", str(value or "")).upper()


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return " ".join(str(value).replace("\xa0", " ").split()).strip()


def _parse_date(value: str) -> Optional[date]:
    normalized = str(value or "").replace("/", ".").replace("-", ".")
    for fmt in ("%d.%m.%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(normalized, fmt).date()
        except ValueError:
            continue
    return None


def _number(value: Any) -> Optional[float]:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = _clean_text(value).replace(" ", "").replace(",", ".")
    if not re.fullmatch(r"[-+]?\d+(?:\.\d+)?", text):
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _decimal_text(value: Any) -> Optional[str]:
    number = _number(value)
    if number is None:
        return None
    try:
        return format(Decimal(str(number)).quantize(Decimal("0.01")), "f")
    except (InvalidOperation, ValueError):
        return None


def _xlsx_value_after_label(ws: Any, label: str) -> Any:
    normalized_label = label.casefold()
    for row in ws.iter_rows():
        for cell in row:
            if _clean_text(cell.value).casefold() != normalized_label:
                continue
            for candidate in row[cell.column:]:
                if _clean_text(candidate.value):
                    return candidate.value
    return None


def _xlsx_full_text(ws: Any) -> str:
    return "\n".join(
        " ".join(_clean_text(cell.value) for cell in row if _clean_text(cell.value))
        for row in ws.iter_rows()
    )


def _parse_russian_long_date(value: Any) -> Optional[date]:
    text = _clean_text(value).lower().replace(" г.", "").replace(" г", "")
    months = {
        "января": 1,
        "февраля": 2,
        "марта": 3,
        "апреля": 4,
        "мая": 5,
        "июня": 6,
        "июля": 7,
        "августа": 8,
        "сентября": 9,
        "октября": 10,
        "ноября": 11,
        "декабря": 12,
    }
    match = re.fullmatch(r"(\d{1,2})\s+([а-яё]+)\s+(\d{4})", text)
    if not match or match.group(2) not in months:
        return _parse_date(text)
    try:
        return date(int(match.group(3)), months[match.group(2)], int(match.group(1)))
    except ValueError:
        return None


def _split_inn_kpp(value: Any) -> tuple[Optional[str], Optional[str]]:
    numbers = re.findall(r"\d{9,12}", _clean_text(value))
    return (
        numbers[0] if numbers else None,
        numbers[1] if len(numbers) > 1 else None,
    )


def parse_customer_return_upd_xlsx(payload: bytes) -> Optional[dict[str, Any]]:
    """Parse a buyer-issued return UPD without treating it as our UKD."""
    workbook = load_workbook(BytesIO(payload), read_only=False, data_only=True)
    for ws in workbook.worksheets:
        full_text = _xlsx_full_text(ws)
        lowered = full_text.casefold()
        if (
            "универсальный передаточный" not in lowered
            or "основание передачи" not in lowered
            or "возврат по" not in lowered
        ):
            continue

        basis_match = _RETURN_BASIS_RE.search(full_text)
        if not basis_match:
            continue
        original_date = _parse_date(basis_match.group(2))

        document_number = _clean_text(_xlsx_value_after_label(ws, "Счет-фактура №"))
        document_date = _parse_russian_long_date(
            _xlsx_value_after_label(ws, "от")
        )
        status_value = _clean_text(_xlsx_value_after_label(ws, "Статус:"))
        if status_value != "2":
            continue

        seller_name = _clean_text(_xlsx_value_after_label(ws, "Продавец"))
        buyer_name = _clean_text(_xlsx_value_after_label(ws, "Покупатель"))
        seller_inn, seller_kpp = _split_inn_kpp(
            _xlsx_value_after_label(ws, "ИНН / КПП продавца")
        )
        buyer_inn, buyer_kpp = _split_inn_kpp(
            _xlsx_value_after_label(ws, "ИНН / КПП покупателя")
        )

        header_row = None
        for row_index in range(1, min(ws.max_row, MAX_SHEET_ROWS) + 1):
            if _clean_text(ws.cell(row_index, 1).value) == "№ п/п":
                header_row = row_index
                break
        if header_row is None:
            continue

        items: list[dict[str, Any]] = []
        for row_index in range(header_row + 3, min(ws.max_row, MAX_SHEET_ROWS) + 1):
            sequence = _number(ws.cell(row_index, 1).value)
            if sequence is None:
                if "всего к оплате" in _clean_text(ws.cell(row_index, 9).value).casefold():
                    break
                continue
            oem = _normalize_oem(_clean_text(ws.cell(row_index, 3).value))
            quantity = _number(ws.cell(row_index, 39).value)  # AM
            if not oem or quantity is None or quantity <= 0:
                continue
            product = _clean_text(ws.cell(row_index, 9).value)  # I
            product_parts = [part.strip() for part in product.split("|")]
            brand_name = product_parts[1] if len(product_parts) >= 3 else None
            item_name = product_parts[0] if product_parts else product
            vat_text = _clean_text(ws.cell(row_index, 53).value)  # BA
            vat_match = re.search(r"\d+(?:[.,]\d+)?", vat_text)
            vat_rate = vat_match.group(0).replace(",", ".") if vat_match else "22.00"
            items.append(
                {
                    "oem_number": oem,
                    "brand_name": brand_name,
                    "autopart_name": item_name or None,
                    "quantity": int(quantity),
                    "unit_price_without_vat": _decimal_text(ws.cell(row_index, 42).value),  # AP
                    "subtotal_without_vat": _decimal_text(ws.cell(row_index, 46).value),  # AT
                    "vat_rate": vat_rate,
                    "vat_amount": _decimal_text(ws.cell(row_index, 57).value),  # BE
                    "total_with_vat": _decimal_text(ws.cell(row_index, 60).value),  # BH
                    "country_code": _clean_text(ws.cell(row_index, 64).value) or None,  # BL
                    "country_name": _clean_text(ws.cell(row_index, 67).value) or None,  # BO
                    "gtd_number": _clean_text(ws.cell(row_index, 69).value) or None,  # BQ
                    "reason": "Возврат товара по документу клиента",
                }
            )
        if not items:
            continue
        return {
            "parser": "customer_return_upd_xlsx",
            "document_number": document_number or None,
            "document_date": document_date.isoformat() if document_date else None,
            "original_document_number": basis_match.group(1).strip(),
            "original_document_date": original_date.isoformat() if original_date else None,
            "document_status": 2,
            "seller_name": seller_name or None,
            "seller_inn": seller_inn,
            "seller_kpp": seller_kpp,
            "buyer_name": buyer_name or None,
            "buyer_inn": buyer_inn,
            "buyer_kpp": buyer_kpp,
            "reason": "Возврат товара по документу клиента",
            "items": items,
        }
    return None


def _sheet_rows(sheet: Any) -> list[list[tuple[int, Any, str]]]:
    rows: list[list[tuple[int, Any, str]]] = []
    for row_index in range(min(int(sheet.nrows), MAX_SHEET_ROWS)):
        row: list[tuple[int, Any, str]] = []
        for col_index in range(min(int(sheet.ncols), MAX_SHEET_COLS)):
            raw = sheet.cell_value(row_index, col_index)
            text = _clean_text(raw)
            if text:
                row.append((col_index, raw, text))
        rows.append(row)
    return rows


def _row_text(row: list[tuple[int, Any, str]]) -> str:
    return " ".join(cell[2] for cell in row)


def _looks_like_torg2(rows: list[list[tuple[int, Any, str]]]) -> bool:
    text = "\n".join(_row_text(row) for row in rows).lower()
    return "унифицированная форма № торг-2" in text or (
        "об установленном расхождении" in text and "фактически оказалось" in text
    )


def _extract_document(rows: list[list[tuple[int, Any, str]]]) -> dict[str, Any]:
    for row in rows:
        match = _SOURCE_DOCUMENT_RE.search(_row_text(row))
        if not match:
            continue
        document_date = _parse_russian_long_date(match.group(2))
        return {
            "document_number": match.group(1).strip(),
            "document_date": (document_date.isoformat() if document_date else None),
        }
    return {"document_number": None, "document_date": None}


def _row_cell_value(
    row: list[tuple[int, Any, str]],
    column: int,
) -> tuple[Any, str] | None:
    for col_index, raw, text in row:
        if col_index == column:
            return raw, text
    return None


def _extract_torg2_act(
    rows: list[list[tuple[int, Any, str]]],
) -> dict[str, Any]:
    """Извлекает номер и дату самого акта, не путая их с основанием."""
    for row_index, row in enumerate(rows[:40]):
        if not any(
            re.sub(r"[^а-яё]", "", cell[2].casefold()).startswith("акт")
            for cell in row
        ):
            continue
        nearby = rows[row_index:min(row_index + 4, len(rows))]
        dates: list[date] = []
        numbers: list[str] = []
        for candidate_row in nearby:
            for _, _, text in candidate_row:
                parsed_date = _parse_russian_long_date(text)
                if parsed_date:
                    dates.append(parsed_date)
                    continue
                if re.fullmatch(r"[A-Za-zА-Яа-я0-9][A-Za-zА-Яа-я0-9/-]{2,39}", text):
                    lowered = text.casefold()
                    if "торг" not in lowered and lowered != "акт":
                        numbers.append(text)
        if numbers or dates:
            return {
                "external_document_number": numbers[0] if numbers else None,
                "external_document_date": dates[0].isoformat() if dates else None,
            }
    return {
        "external_document_number": None,
        "external_document_date": None,
    }


def _extract_torg2_reason(
    rows: list[list[tuple[int, Any, str]]],
) -> Optional[str]:
    for row in rows:
        row_text = _row_text(row)
        match = re.search(
            r"(?i)причина\s+возврата\s*[:\-–—]?\s*(.+)$",
            row_text,
        )
        if match:
            return match.group(1).strip(" .,:;-–—") or None
    return None


def _extract_torg2_table_items(
    rows: list[list[tuple[int, Any, str]]],
) -> list[dict[str, Any]]:
    """Читает табличный вариант ТОРГ-2 с отдельной колонкой «Возврат»."""
    header_start = None
    for index, row in enumerate(rows):
        lowered = _row_text(row).casefold()
        if "товар" in lowered and "наимен" in lowered:
            header_start = index
            break
    if header_start is None:
        return []

    header_rows = rows[header_start:min(header_start + 10, len(rows))]
    product_col = None
    oem_col = None
    return_col = None
    for row in header_rows:
        for col_index, _, text in row:
            lowered = text.casefold().strip()
            if product_col is None and "товар" in lowered and "наимен" in lowered:
                product_col = col_index
            if oem_col is None and "артикул" in lowered:
                oem_col = col_index
            if return_col is None and lowered == "возврат":
                return_col = col_index
    if oem_col is None:
        return []

    global_reason = _extract_torg2_reason(rows)
    items: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in rows[header_start + 1:]:
        lowered = _row_text(row).casefold()
        if "причина возврата" in lowered or "заключение комиссии" in lowered:
            break
        oem_cell = _row_cell_value(row, oem_col)
        if oem_cell is None:
            continue
        oem = _normalize_oem(oem_cell[1])
        if len(oem) < 4 or oem in seen or not re.search(r"\d", oem):
            continue

        product = ""
        if product_col is not None:
            product_cell = _row_cell_value(row, product_col)
            product = product_cell[1] if product_cell else ""
        product_tokens = product.split()
        normalized_tokens = [_normalize_oem(token) for token in product_tokens]
        if oem in normalized_tokens:
            product_tokens = product_tokens[normalized_tokens.index(oem) + 1:]
        brand_name = product_tokens[0].strip(" ,.;") if product_tokens else None
        autopart_name = " ".join(product_tokens[1:]).strip() or None

        numeric_after_oem = [
            (col_index, number)
            for col_index, raw, _ in row
            if col_index > oem_col and (number := _number(raw)) is not None
        ]
        return_values = [
            number
            for col_index, number in numeric_after_oem
            if return_col is not None and col_index >= return_col
        ]
        values = return_values or [number for _, number in numeric_after_oem]
        quantity = max(1, int(values[0])) if values else 1
        line_sum = values[1] if len(values) > 1 else None
        unit_price = line_sum / quantity if line_sum is not None else None
        seen.add(oem)
        items.append(
            {
                "oem_number": oem,
                "reason": global_reason,
                "quantity": quantity,
                "unit_price": unit_price,
                "line_sum": line_sum,
                "total_with_vat": _decimal_text(line_sum),
                "brand_name": brand_name,
                "autopart_name": autopart_name,
            }
        )
    return items


def _extract_reason_items(
    rows: list[list[tuple[int, Any, str]]],
) -> list[dict[str, Any]]:
    start = None
    end = len(rows)
    for index, row in enumerate(rows):
        lowered = _row_text(row).lower()
        if "подробное описание дефектов" in lowered:
            start = index + 1
        elif start is not None and "заключение комиссии" in lowered:
            end = index
            break
    if start is None:
        return []

    result: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in rows[start:end]:
        for _, _, text in row:
            match = _ITEM_REASON_RE.match(text)
            if not match:
                continue
            oem = _normalize_oem(match.group(1))
            reason = match.group(2).strip(" .,:;-")
            if not oem or not reason or oem in seen:
                continue
            seen.add(oem)
            result.append({"oem_number": oem, "reason": reason})
    return result


def _find_actual_values(rows: list[list[tuple[int, Any, str]]], oem: str) -> list[float]:
    normalized = _normalize_oem(oem)
    for row in rows:
        oem_column = None
        for col_index, _, text in row:
            if _normalize_oem(text) == normalized:
                oem_column = col_index
                break
        if oem_column is None:
            continue
        values = [
            number
            for col_index, raw, _ in row
            if col_index > oem_column and (number := _number(raw)) is not None
        ]
        if values:
            return values
    return []


def _find_description(
    rows: list[list[tuple[int, Any, str]]], oem: str
) -> tuple[Optional[str], Optional[str]]:
    normalized = _normalize_oem(oem)
    for row in rows:
        for _, _, text in row:
            tokens = text.split()
            positions = [
                index for index, token in enumerate(tokens) if _normalize_oem(token) == normalized
            ]
            if not positions or len(tokens) < 2:
                continue
            oem_index = positions[-1]
            if oem_index + 1 >= len(tokens):
                continue
            brand = tokens[oem_index + 1].strip(" ,.;") or None
            name_tokens = tokens[slice(oem_index + 2, None)]
            name = " ".join(name_tokens).strip() or None
            return brand, name
    return None, None


def parse_torg2_sheet(sheet: Any) -> Optional[dict[str, Any]]:
    """Extracts stable fields from the customer's TORG-2 workbook sheet."""
    rows = _sheet_rows(sheet)
    if not _looks_like_torg2(rows):
        return None

    document = _extract_document(rows)
    act_document = _extract_torg2_act(rows)
    items = _extract_torg2_table_items(rows)
    if not items:
        items = _extract_reason_items(rows)
        for item in items:
            values = _find_actual_values(rows, item["oem_number"])
            brand, name = _find_description(rows, item["oem_number"])
            item["quantity"] = max(1, int(values[0])) if values else 1
            item["unit_price"] = values[1] if len(values) > 1 else None
            item["line_sum"] = values[2] if len(values) > 2 else None
            item["brand_name"] = brand
            item["autopart_name"] = name

    reasons = list(
        dict.fromkeys(item["reason"] for item in items if item.get("reason"))
    )
    return {
        "parser": "torg2_xls",
        **document,
        "original_document_number": document["document_number"],
        "original_document_date": document["document_date"],
        **act_document,
        "reason": "; ".join(reasons) or None,
        "items": items,
    }


def parse_reclamation_attachment(
    filename: Optional[str], payload: bytes
) -> Optional[dict[str, Any]]:
    """Returns structured data only for known, confidently detected files."""
    extension = os.path.splitext(str(filename or ""))[1].lower()
    if extension not in {".xls", ".xlsx"} or not payload:
        return None
    if len(payload) > MAX_RECLAMATION_ATTACHMENT_BYTES:
        logger.warning(
            "Reclamation attachment %s is too large for parsing: %s bytes",
            filename,
            len(payload),
        )
        return None
    if extension == ".xlsx":
        try:
            parsed = parse_customer_return_upd_xlsx(payload)
            if parsed:
                parsed["filename"] = filename
                return parsed
        except Exception as exc:  # noqa: BLE001
            logger.warning(
                "Failed to parse XLSX reclamation attachment %s: %s",
                filename,
                exc,
            )
        return None
    try:
        workbook = xlrd.open_workbook(
            file_contents=payload,
            on_demand=True,
            formatting_info=False,
        )
        for sheet in workbook.sheets():
            parsed = parse_torg2_sheet(sheet)
            if parsed:
                parsed["filename"] = filename
                return parsed
    except Exception as exc:  # noqa: BLE001
        logger.warning(
            "Failed to parse reclamation attachment %s: %s",
            filename,
            exc,
        )
    return None
