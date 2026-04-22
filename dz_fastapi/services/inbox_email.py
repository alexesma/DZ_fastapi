"""
Сервис для работы с входящими письмами (Inbox).

Логика:
1. fetch_inbox_for_account  — скачать письма с IMAP/Resend для одного ящика
2. fetch_and_store_emails   — скачать и сохранить в InboxEmail,
                              запустить авто-разметку
3. auto_detect_and_process  — попробовать найти паттерн и
                              обработать автоматически
4. assign_rule              — менеджер вручную назначает правило +
                              обрабатываем письмо
5. _process_email_by_rule   — диспетчер, вызывает нужный обработчик
6. cleanup_old_emails       — удаление писем старше max_days

Обработчики по типу правила:
  price_list       → process_provider_pricelist (существующий сервис)
  order_reply      → process_supplier_response_messages (существующий сервис)
  customer_order   → process_customer_orders (существующий сервис)
  document         → supplier_workflow: create_supplier_receipt
  shipment_notice  → уведомление менеджера + сохранение трекинг-данных
  claim            → уведомление менеджера + создание задачи
                     (заглушка → notify)
  error_report     → уведомление менеджера (заглушка)
  inquiry          → уведомление менеджера (заглушка)
  proposal         → уведомление менеджера (заглушка)
  spam             → пометить и скрыть
  ignore           → пометить и не трогать
"""

import asyncio
import csv
import logging
import os
import re
from datetime import date, datetime, timedelta
from email.header import decode_header
from typing import Dict, List, Optional, Tuple

import aiofiles
from imap_tools import AND, MailboxLoginError

try:
    from imap_tools.errors import MailboxFolderSelectError
except ImportError:
    MailboxFolderSelectError = Exception
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from dz_fastapi.core.constants import IMAP_SERVER
from dz_fastapi.core.email_folders import (DEFAULT_IMAP_FOLDER,
                                           parse_imap_additional_folders)
from dz_fastapi.core.time import now_moscow
from dz_fastapi.crud.email_account import crud_email_account
from dz_fastapi.crud.inbox_email import (cleanup_old_inbox_emails,
                                         create_inbox_email,
                                         create_rule_pattern,
                                         find_matching_pattern,
                                         get_inbox_email,
                                         increment_pattern_applied,
                                         increment_pattern_confirmed,
                                         mark_processed,
                                         update_inbox_email_rule)
from dz_fastapi.models.inbox_email import InboxEmail, InboxForceProcessAudit
from dz_fastapi.schemas.inbox_email import RULE_META, FetchInboxResponse
from dz_fastapi.services.email import (_create_mailbox,
                                       _dedupe_fetched_messages,
                                       _extract_email,
                                       _fetch_resend_price_messages,
                                       _FetchedAttachment,
                                       _FetchedInboxMessage)

logger = logging.getLogger('dz_fastapi')
PROJECT_ROOT = os.path.abspath(
    os.path.join(os.path.dirname(__file__), '..', '..')
)
PREVIEWABLE_ATTACHMENT_EXTENSIONS = {'.xls', '.xlsx', '.csv'}
MAX_PREVIEWABLE_ATTACHMENT_SIZE_BYTES = 5 * 1024 * 1024
FORCE_PROCESSABLE_RULES = {'order_reply', 'customer_order', 'document'}


# ---------------------------------------------------------------------------
# Вспомогательные функции
# ---------------------------------------------------------------------------

def _decode_subject(raw: str) -> str:
    """Декодирует тему письма из MIME-encoded-words."""
    parts = decode_header(raw or '')
    decoded_parts = []
    for part, enc in parts:
        if isinstance(part, bytes):
            decoded_parts.append(part.decode(enc or 'utf-8', errors='replace'))
        else:
            decoded_parts.append(str(part))
    return ' '.join(decoded_parts).strip()


def _get_attachment_extensions(attachment_info: list) -> List[str]:
    """Возвращает список расширений вложений в нижнем регистре."""
    exts = []
    for att in attachment_info:
        name = att.get('name', '') or ''
        _, ext = os.path.splitext(name)
        if ext:
            exts.append(ext.lower())
    return exts


def _rule_label(rule_type: str) -> str:
    return RULE_META.get(rule_type, {}).get('label', rule_type)


# ---------------------------------------------------------------------------
# Предпросмотр вложений
# ---------------------------------------------------------------------------

def _read_attachment_preview_sync(
    file_path: str,
    max_rows: int = 25,
) -> Dict:
    """Синхронное чтение файла вложения для предпросмотра (XLS/XLSX/CSV)."""
    def _stringify_cell(value: object) -> str:
        if value is None:
            return ''
        if isinstance(value, float):
            # NaN
            if value != value:
                return ''
            if value.is_integer():
                return str(int(value))
        text = str(value)
        return '' if text.lower() == 'nan' else text

    def _pad_rows(rows: list[list[str]], columns: int) -> list[list[str]]:
        if columns <= 0:
            return rows
        return [row + [''] * (columns - len(row)) for row in rows]

    _, ext = os.path.splitext(file_path)
    ext = ext.lower()

    if ext == '.xls':
        import xlrd

        workbook = xlrd.open_workbook(file_path, on_demand=True)
        try:
            if workbook.nsheets == 0:
                return {'rows': [], 'total_rows': 0, 'columns': 0}
            sheet = workbook.sheet_by_index(0)
            total_rows = int(sheet.nrows or 0)
            rows: list[list[str]] = []
            columns = 0
            for idx in range(min(max_rows, total_rows)):
                row_values = [
                    _stringify_cell(cell) for cell in sheet.row_values(idx)
                ]
                rows.append(row_values)
                columns = max(columns, len(row_values))
            return {
                'rows': _pad_rows(rows, columns),
                'total_rows': total_rows,
                'columns': columns,
            }
        finally:
            workbook.release_resources()
    elif ext == '.xlsx':
        from openpyxl import load_workbook

        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            if not workbook.worksheets:
                return {'rows': [], 'total_rows': 0, 'columns': 0}
            sheet = workbook.worksheets[0]
            total_rows = int(sheet.max_row or 0)
            rows: list[list[str]] = []
            columns = 0
            for row in sheet.iter_rows(
                min_row=1, max_row=max_rows, values_only=True
            ):
                row_values = [_stringify_cell(cell) for cell in row]
                rows.append(row_values)
                columns = max(columns, len(row_values))
            return {
                'rows': _pad_rows(rows, columns),
                'total_rows': total_rows,
                'columns': columns,
            }
        finally:
            workbook.close()
    elif ext == '.csv':
        rows: list[list[str]] = []
        total_rows = 0
        columns = 0
        with open(
            file_path,
            mode='r',
            encoding='utf-8-sig',
            errors='replace',
            newline='',
        ) as fh:
            reader = csv.reader(fh)
            for row in reader:
                total_rows += 1
                if len(rows) < max_rows:
                    row_values = [_stringify_cell(cell) for cell in row]
                    rows.append(row_values)
                    columns = max(columns, len(row_values))
        return {
            'rows': _pad_rows(rows, columns),
            'total_rows': total_rows,
            'columns': columns,
        }
    else:
        raise ValueError(f'Unsupported file extension: {ext}')


async def read_attachment_preview(
    file_path: str,
    max_rows: int = 25,
) -> Dict:
    """Асинхронная обёртка для чтения предпросмотра файла вложения."""
    return await asyncio.to_thread(
        _read_attachment_preview_sync, file_path, max_rows
    )


def resolve_inbox_attachment_fs_path(file_path: Optional[str]) -> str:
    """
    Преобразует путь вложения из БД в абсолютный путь ФС.
    В БД обычно хранится относительный путь вида uploads/inbox_attachments/...
    """
    if not file_path:
        return ''
    if os.path.isabs(file_path):
        return file_path
    return os.path.join(PROJECT_ROOT, file_path)


def inbox_attachment_exists(file_path: Optional[str]) -> bool:
    resolved = resolve_inbox_attachment_fs_path(file_path)
    return bool(resolved and os.path.exists(resolved))


def _build_attachment_relative_path(
    *,
    account_id: int,
    msg_uid: Optional[str],
    msg_date: object,
    filename: str,
) -> str:
    if isinstance(msg_date, datetime):
        date_str = msg_date.strftime('%Y%m%d')
    elif isinstance(msg_date, date):
        date_str = msg_date.strftime('%Y%m%d')
    else:
        date_str = datetime.today().strftime('%Y%m%d')
    safe_filename = re.sub(r'[^A-Za-z0-9._-]', '_', filename or '')
    if not safe_filename:
        safe_filename = 'attachment.bin'
    uid_str = msg_uid or 'nouid'
    return os.path.join(
        'uploads',
        'inbox_attachments',
        str(account_id),
        date_str,
        f'{uid_str}_{safe_filename}',
    )


async def _build_attachment_info_for_message(
    msg: _FetchedInboxMessage,
    *,
    account_id: int,
) -> list[dict]:
    att_info: list[dict] = []
    for att in (msg.attachments or []):
        att_entry = {
            'name': att.filename or '',
            'size': len(att.payload) if att.payload else 0,
            'path': None,
        }
        try:
            filename = att.filename or ''
            _, ext = os.path.splitext(filename)
            if (
                ext.lower() in PREVIEWABLE_ATTACHMENT_EXTENSIONS
                and att.payload
                and len(att.payload) <= MAX_PREVIEWABLE_ATTACHMENT_SIZE_BYTES
            ):
                rel_path = _build_attachment_relative_path(
                    account_id=account_id,
                    msg_uid=msg.uid,
                    msg_date=msg.date,
                    filename=filename,
                )
                abs_path = resolve_inbox_attachment_fs_path(rel_path)
                os.makedirs(os.path.dirname(abs_path), exist_ok=True)
                async with aiofiles.open(abs_path, 'wb') as fh:
                    await fh.write(att.payload)
                att_entry['path'] = rel_path
        except Exception as save_err:
            logger.warning(
                'Failed to save attachment %s: %s',
                att.filename,
                save_err,
            )
        att_info.append(att_entry)
    return att_info


def _fetch_inbox_message_by_uid_imap_sync(
    *,
    host: str,
    email: str,
    password: str,
    folder: str,
    uid: str,
    port: int,
) -> Optional[_FetchedInboxMessage]:
    """
    Загружает одно письмо из IMAP по UID в указанной папке.
    Возвращает None если письмо не найдено или возникла ошибка.
    """
    try:
        with _create_mailbox(
                host, port, True
        ).login(email, password) as mailbox:
            mailbox.folder.set(folder)
            raw_messages = list(
                mailbox.fetch(
                    f'UID {uid}:{uid}',
                    mark_seen=False,
                )
            )
            if not raw_messages:
                return None
            msg = raw_messages[-1]
            attachments = [
                _FetchedAttachment(
                    filename=att.filename,
                    payload=att.payload,
                )
                for att in (msg.attachments or [])
            ]
            return _FetchedInboxMessage(
                uid=str(msg.uid) if msg.uid else str(uid),
                from_=msg.from_ or '',
                subject=_decode_subject(msg.subject or ''),
                attachments=attachments,
                date=getattr(msg, 'date', None),
                folder_name=folder,
            )
    except Exception as exc:
        logger.warning(
            'Failed to fetch IMAP message by UID: '
            'host=%s folder=%s uid=%s error=%s',
            host,
            folder,
            uid,
            exc,
        )
        return None


async def restore_inbox_email_attachments_from_source(
    session: AsyncSession,
    *,
    inbox_email: InboxEmail,
) -> bool:
    """
    Пытается восстановить вложения inbox-письма из источника (IMAP) по UID.
    Возвращает True, если attachment_info обновлён и сохранён в БД.
    """
    if not inbox_email.uid or not inbox_email.email_account_id:
        return False

    account = await crud_email_account.get(
        session,
        inbox_email.email_account_id
    )
    if account is None:
        return False

    transport = (getattr(account, 'transport', '') or '').strip().lower()
    if transport == 'resend_api':
        # Для Resend в текущей реализации нет точечного fetch по UID.
        return False

    host = (getattr(account, 'imap_host', '') or '').strip()
    email = (getattr(account, 'email', '') or '').strip()
    password = str(getattr(account, 'password', '') or '')
    if not host or not email or not password:
        return False

    folders: list[str] = []
    seen_folders: set[str] = set()
    for raw_folder in (
        inbox_email.folder,
        getattr(account, 'imap_folder', None),
        DEFAULT_IMAP_FOLDER,
    ):
        folder = str(raw_folder or '').strip() or DEFAULT_IMAP_FOLDER
        key = folder.lower()
        if key in seen_folders:
            continue
        seen_folders.add(key)
        folders.append(folder)

    for folder in folders:
        fetched_msg = await asyncio.to_thread(
            _fetch_inbox_message_by_uid_imap_sync,
            host=host,
            email=email,
            password=password,
            folder=folder,
            uid=str(inbox_email.uid),
            port=int(getattr(account, 'imap_port', None) or IMAP_SERVER),
        )
        if fetched_msg is None:
            continue
        if not fetched_msg.attachments:
            continue

        rebuilt_att_info = await _build_attachment_info_for_message(
            fetched_msg, account_id=account.id
        )
        if not rebuilt_att_info:
            continue

        inbox_email.has_attachments = bool(rebuilt_att_info)
        inbox_email.attachment_info = rebuilt_att_info
        inbox_email.fetched_at = now_moscow()
        session.add(inbox_email)
        await session.commit()
        logger.info(
            'Restored inbox attachments from source: email_id=%s '
            'account_id=%s uid=%s folder=%s files=%s',
            inbox_email.id,
            account.id,
            inbox_email.uid,
            folder,
            len(rebuilt_att_info),
        )
        return True

    return False


# ---------------------------------------------------------------------------
# Fetch писем с IMAP/Resend для одного EmailAccount
# ---------------------------------------------------------------------------

def _fetch_inbox_imap_sync(
    host: str,
    email: str,
    password: str,
    folder: str,
    port: int = 993,
    since_date: Optional[date] = None,
) -> List[_FetchedInboxMessage]:
    """
    Синхронная загрузка писем из IMAP-папки за последние N дней.
    Запускается через asyncio.to_thread.
    """
    if since_date is None:
        since_date = date.today()

    result: List[_FetchedInboxMessage] = []
    try:
        mb = _create_mailbox(host, port, True).login(email, password)
        with mb as mailbox:
            mailbox.folder.set(folder)
            messages = list(
                mailbox.fetch(
                    AND(date_gte=since_date, all=True),
                    charset='utf-8',
                    mark_seen=False,
                )
            )
            for msg in messages:
                attachments = []
                for att in msg.attachments:
                    attachments.append(
                        _FetchedAttachment(
                            filename=att.filename,
                            payload=att.payload,
                        )
                    )
                result.append(
                    _FetchedInboxMessage(
                        uid=str(msg.uid) if msg.uid else None,
                        from_=msg.from_ or '',
                        subject=_decode_subject(msg.subject or ''),
                        attachments=attachments,
                        date=msg.date,
                        folder_name=folder,
                    )
                )
    except MailboxLoginError as e:
        logger.error('IMAP login failed for %s: %s', email, e)
    except Exception as e:
        logger.error('IMAP fetch error for %s folder=%s: %s', email, folder, e)
    return result


# Имена папок IMAP, которые точно содержат ИСХОДЯЩИЕ письма — исключаем
_SENT_FOLDER_NAMES: set[str] = {
    'sent', 'sent items', 'sent messages', 'sent mail',
    'отправленные', 'отправленные письма',
    'inbox.sent', '[gmail]/sent mail', '[gmail]/отправленные',
}


def _is_sent_folder(folder_name: str) -> bool:
    """Возвращает True если папка — папка отправленных."""
    return folder_name.strip().lower() in _SENT_FOLDER_NAMES


async def fetch_inbox_for_account(
    account,
    days: int = 3,
) -> List[_FetchedInboxMessage]:
    """
    Загружает ТОЛЬКО ВХОДЯЩИЕ письма для одного EmailAccount (IMAP или Resend).

    Фильтрация исходящих писем:
      1. Пропускаем папки типа Sent / Отправленные
      2. Пропускаем письма, где отправитель совпадает с адресом самого ящика
    """
    since_date = (now_moscow() - timedelta(days=days)).date()
    account_email = (account.email or '').lower().strip()
    messages: List[_FetchedInboxMessage] = []

    transport = (account.transport or '').strip().lower()

    if transport == 'resend_api':
        if not account.resend_api_key:
            logger.warning(
                'Resend API key missing for account id=%s', account.id
            )
            return []
        try:
            fetched, _ = await _fetch_resend_price_messages(account)
            for msg in fetched:
                # Фильтр: пропускаем исходящие (отправитель == сам ящик)
                msg_from = _extract_email(msg.from_ or '').lower()
                if msg_from == account_email:
                    continue
                msg_date = msg.date
                if isinstance(msg_date, datetime):
                    if msg_date.date() >= since_date:
                        messages.append(msg)
                else:
                    messages.append(msg)
        except Exception as e:
            logger.error(
                'Resend fetch error for account id=%s: %s', account.id, e
            )
        return messages

    host = account.imap_host
    if not host:
        logger.warning('No IMAP host for account id=%s', account.id)
        return []

    # Собираем список папок для опроса:
    # 1) основная папка (imap_folder или INBOX)
    # 2) дополнительные папки из imap_additional_folders
    primary_folder = (account.imap_folder or '').strip() or DEFAULT_IMAP_FOLDER
    additional_folders = parse_imap_additional_folders(
        getattr(account, 'imap_additional_folders', None)
    )

    # Дедупликация: не читаем одну папку дважды
    folders_to_fetch: list[str] = []
    seen_folders: set[str] = set()

    for folder in [primary_folder] + additional_folders:
        folder_norm = folder.strip().lower()
        if not folder_norm:
            continue
        if folder_norm in seen_folders:
            continue
        seen_folders.add(folder_norm)
        # Пропускаем папки отправленных
        if _is_sent_folder(folder):
            logger.debug(
                'Папка "%s" для account id=%s похожа '
                'на папку отправленных — пропускаем',
                folder, account.id,
            )
            continue
        folders_to_fetch.append(folder)

    if not folders_to_fetch:
        logger.warning(
            'Нет подходящих IMAP-папок для account id=%s', account.id
        )
        return []

    for folder in folders_to_fetch:
        try:
            raw_messages = await asyncio.to_thread(
                _fetch_inbox_imap_sync,
                host,
                account.email,
                account.password,
                folder,
                account.imap_port or IMAP_SERVER,
                since_date,
            )
        except MailboxFolderSelectError as e:
            logger.warning(
                'IMAP папка "%s" не найдена для account id=%s — пропускаем. '
                'Проверьте название папки в настройках. Ошибка: %s',
                folder, account.id, e,
            )
            continue
        except Exception as e:
            logger.error(
                'Error fetching inbox for account id=%s folder=%s: %s',
                account.id, folder, e,
            )
            continue

        # Фильтруем: убираем письма, где from == адрес самого ящика
        # (исходящие копии)
        for msg in raw_messages:
            msg_from = _extract_email(msg.from_ or '').lower()
            if msg_from == account_email:
                logger.debug(
                    'Пропускаем исходящее письмо (from=%s == account=%s)',
                    msg_from, account_email,
                )
                continue
            messages.append(msg)

    return _dedupe_fetched_messages(messages)


# ---------------------------------------------------------------------------
# Главная функция: загрузить, сохранить, авто-обработать
# ---------------------------------------------------------------------------

async def _get_existing_inbox_email_by_uid(
    session: AsyncSession,
    *,
    email_account_id: int,
    uid: str,
    folder: Optional[str] = None,
) -> Optional[InboxEmail]:
    from sqlalchemy import and_, select

    base_filters = [
        InboxEmail.email_account_id == email_account_id,
        InboxEmail.uid == uid,
    ]

    filters = list(base_filters)
    if folder is not None:
        filters.append(InboxEmail.folder == folder)

    result = await session.execute(
        select(InboxEmail)
        .where(and_(*filters))
        .order_by(InboxEmail.id.desc())
        .limit(1)
    )
    found = result.scalars().first()
    if found is not None:
        return found

    # Fallback: старые записи могли быть без folder.
    # Ищем по uid без учета folder, чтобы можно было восстановить вложения.
    if folder is not None:
        fallback_result = await session.execute(
            select(InboxEmail)
            .where(and_(*base_filters))
            .order_by(InboxEmail.id.desc())
            .limit(1)
        )
        return fallback_result.scalars().first()

    return None


async def _restore_existing_email_attachments_if_missing(
    session: AsyncSession,
    *,
    inbox_email: InboxEmail,
    msg: _FetchedInboxMessage,
    account_id: int,
) -> bool:
    existing_info = inbox_email.attachment_info or []
    has_missing_files = (
        not existing_info
        or any(
            not att.get('path') or not inbox_attachment_exists(att.get('path'))
            for att in existing_info
        )
    )
    if not has_missing_files or not msg.attachments:
        return False

    rebuilt_att_info = await _build_attachment_info_for_message(
        msg, account_id=account_id
    )
    if not rebuilt_att_info:
        return False

    inbox_email.has_attachments = bool(rebuilt_att_info)
    inbox_email.attachment_info = rebuilt_att_info
    inbox_email.fetched_at = now_moscow()
    session.add(inbox_email)
    await session.flush()
    return True


async def fetch_and_store_emails(
    session: AsyncSession,
    *,
    email_account_id: Optional[int] = None,
    days: int = 3,
) -> FetchInboxResponse:
    """
    Загружает письма из указанного ящика (или всех активных),
    сохраняет новые в InboxEmail, запускает авто-разметку.
    """
    days = max(1, min(days, 7))

    if email_account_id is not None:
        account = await crud_email_account.get(session, email_account_id)
        accounts = [account]
    else:
        accounts = await crud_email_account.get_multi(session)
        accounts = [a for a in accounts if a.is_active]

    total_fetched = 0
    total_stored = 0
    total_auto_processed = 0

    for account in accounts:
        try:
            messages = await fetch_inbox_for_account(account, days=days)
            total_fetched += len(messages)
        except Exception as e:
            logger.error(
                'Failed to fetch inbox for account id=%s: %s',
                account.id, e
            )
            continue

        for msg in messages:
            uid = msg.uid
            folder = getattr(msg, 'folder_name', None)

            if uid:
                existing_email = await _get_existing_inbox_email_by_uid(
                    session,
                    email_account_id=account.id,
                    uid=uid,
                    folder=folder,
                )
                if existing_email is not None:
                    restored = await (
                        _restore_existing_email_attachments_if_missing(
                            session,
                            inbox_email=existing_email,
                            msg=msg,
                            account_id=account.id,
                        )
                    )
                    if restored:
                        logger.info(
                            'Restored attachment files for inbox email '
                            'id=%s account_id=%s uid=%s',
                            existing_email.id,
                            account.id,
                            uid,
                        )
                        await session.commit()
                    continue

            from_email = _extract_email(msg.from_)
            from_name = (
                msg.from_.replace(f'<{from_email}>', '').strip().strip('"')
                if from_email in msg.from_ else None
            )

            # File I/O is done before opening the DB
            # transaction to keep it short
            att_info = await _build_attachment_info_for_message(
                msg, account_id=account.id
            )

            try:
                inbox_email = await create_inbox_email(
                    session,
                    email_account_id=account.id,
                    uid=uid,
                    folder=folder,
                    from_email=from_email,
                    from_name=from_name or None,
                    subject=msg.subject,
                    body_preview=None,
                    body_full=None,
                    has_attachments=bool(att_info),
                    attachment_info=att_info,
                    received_at=msg.date,
                )
                total_stored += 1

                processed = await auto_detect_and_process(
                    session,
                    inbox_email=inbox_email,
                    fetched_msg=msg,
                    account=account,
                )
                if processed:
                    total_auto_processed += 1

                await session.commit()
            except Exception as msg_err:
                logger.error(
                    'Failed to store/process inbox '
                    'email uid=%s account_id=%s: %s',
                    uid, account.id, msg_err,
                )
                await session.rollback()

    return FetchInboxResponse(
        fetched=total_fetched,
        stored=total_stored,
        auto_processed=total_auto_processed,
    )


# ---------------------------------------------------------------------------
# Определение правила по уже существующим настройкам системы
# ---------------------------------------------------------------------------

async def _detect_rule_from_existing_configs(
    session: AsyncSession,
    *,
    inbox_email: InboxEmail,
    fetched_msg: Optional[_FetchedInboxMessage] = None,
) -> Optional[str]:
    """
    Пытается определить тип правила на основе УЖЕ существующих настроек:
      1. ProviderPricelistConfig — поставщики с настроенными прайс-листами
      2. CustomerOrderConfig     — клиенты с настроенными заказами

    Возвращает строку rule_type или None если совпадений нет.
    Не создаёт никаких новых записей — только читает существующие конфиги.
    """
    from dz_fastapi.crud.partner import (crud_provider,
                                         crud_provider_pricelist_config)
    from dz_fastapi.services.email import _message_matches_provider_config

    from_email = (inbox_email.from_email or '').lower().strip()
    if not from_email:
        return None

    # ------------------------------------------------------------------
    # 1. Проверяем поставщиков: Provider.email_incoming_price
    # ------------------------------------------------------------------
    try:
        provider = await crud_provider.get_by_email_incoming_price(
            session=session, email=from_email
        )
        if provider is not None:
            # Если есть активные конфиги прайсов — точно price_list
            configs = await crud_provider_pricelist_config.get_configs(
                provider_id=provider.id, session=session, only_active=True
            )
            if configs:
                # Если есть fetched_msg — проверяем точное совпадение конфига
                if fetched_msg is not None:
                    for config in configs:
                        if _message_matches_provider_config(
                            fetched_msg, config
                        ):
                            logger.info(
                                'Письмо id=%s → price_list '
                                '(provider_id=%s, config_id=%s)',
                                inbox_email.id, provider.id, config.id,
                            )
                            return 'price_list'
                else:
                    # Нет fetched_msg — достаточно того, что email совпадает
                    logger.info(
                        'Письмо id=%s → price_list (provider_id=%s, по email)',
                        inbox_email.id, provider.id,
                    )
                    return 'price_list'
    except Exception as e:
        logger.warning(
            'Ошибка при поиске provider по email=%s: %s', from_email, e
        )

    # ------------------------------------------------------------------
    # 2. Проверяем конфиги заказов клиентов: CustomerOrderConfig
    # ------------------------------------------------------------------
    try:
        matched_config_ids = await _find_matching_customer_order_configs(
            session,
            from_email=inbox_email.from_email,
            subject=inbox_email.subject,
            email_account_id=inbox_email.email_account_id,
        )
        if matched_config_ids:
            logger.info(
                'Письмо id=%s → customer_order (config_ids=%s)',
                inbox_email.id,
                matched_config_ids,
            )
            return 'customer_order'
    except Exception as e:
        logger.warning('Ошибка при поиске CustomerOrderConfig: %s', e)

    return None


async def _find_matching_customer_order_configs(
    session: AsyncSession,
    *,
    from_email: Optional[str],
    subject: Optional[str],
    email_account_id: Optional[int] = None,
) -> list[int]:
    """
    Возвращает id активных CustomerOrderConfig,
    подходящих по отправителю и теме письма.
    """
    from sqlalchemy import select as _select

    from dz_fastapi.models.partner import CustomerOrderConfig as _COC

    normalized_email = (from_email or '').lower().strip()
    if not normalized_email:
        return []

    result = await session.execute(
        _select(_COC).where(_COC.is_active.is_(True))
    )
    configs = result.scalars().all()
    matched_ids: list[int] = []
    subject_text = subject or ''

    for config in configs:
        if (
            email_account_id is not None
            and config.email_account_id not in (None, email_account_id)
        ):
            continue

        emails_in_config: list[str] = []
        if config.order_email:
            emails_in_config.append(config.order_email.lower().strip())
        if config.order_emails:
            extra = (
                config.order_emails
                if isinstance(config.order_emails, list) else []
            )
            emails_in_config.extend(e.lower().strip() for e in extra if e)

        if normalized_email not in emails_in_config:
            continue

        if config.order_subject_pattern:
            try:
                if not re.search(
                    config.order_subject_pattern,
                    subject_text,
                    flags=re.IGNORECASE,
                ):
                    continue
            except re.error:
                if (
                    config.order_subject_pattern.lower()
                    not in subject_text.lower()
                ):
                    continue

        matched_ids.append(int(config.id))

    return matched_ids


# ---------------------------------------------------------------------------
# Авто-определение правила по паттернам
# ---------------------------------------------------------------------------

async def auto_detect_and_process(
    session: AsyncSession,
    *,
    inbox_email: InboxEmail,
    fetched_msg: Optional[_FetchedInboxMessage] = None,
    account=None,
) -> bool:
    """
    Определяет правило письма и обрабатывает его автоматически.

    Порядок проверки:
      1. EmailRulePattern  — паттерны, подтверждённые менеджером вручную
      2. Существующие конфиги системы
         (ProviderPricelistConfig, CustomerOrderConfig)
         — чтобы не настраивать одно и то же дважды

    Возвращает True если письмо было обработано.
    """
    att_extensions = _get_attachment_extensions(
        inbox_email.attachment_info or []
    )

    # --- Шаг 1: ищем подтверждённый паттерн ---
    pattern = await find_matching_pattern(
        session,
        email_account_id=inbox_email.email_account_id,
        from_email=inbox_email.from_email,
        subject=inbox_email.subject or '',
        has_attachments=inbox_email.has_attachments,
        attachment_extensions=att_extensions,
    )

    if pattern is not None:
        rule_type = pattern.rule_type
        source = f'паттерн id={pattern.id}'
        await increment_pattern_applied(session, pattern)
    else:
        # --- Шаг 2: ищем совпадение в существующих конфигах системы ---
        rule_type = await _detect_rule_from_existing_configs(
            session,
            inbox_email=inbox_email,
            fetched_msg=fetched_msg,
        )
        source = 'существующий конфиг'

    if rule_type is None:
        return False

    logger.info(
        'Авто-разметка письма id=%s rule=%s (источник: %s)',
        inbox_email.id, rule_type, source,
    )

    await update_inbox_email_rule(
        session,
        email=inbox_email,
        rule_type=rule_type,
        auto_detected=True,
    )

    result, error = await _process_email_by_rule(
        session,
        inbox_email=inbox_email,
        rule_type=rule_type,
        fetched_msg=fetched_msg,
        account=account,
    )
    await mark_processed(
        session, email=inbox_email, result=result, error=error
    )
    return True


# ---------------------------------------------------------------------------
# Назначение правила вручную менеджером
# ---------------------------------------------------------------------------

async def assign_rule(
    session: AsyncSession,
    *,
    email_id: int,
    rule_type: str,
    user_id: Optional[int],
    save_pattern: bool = True,
    process_now: bool = True,
    queued_note: Optional[str] = None,
) -> InboxEmail:
    """
    Менеджер вручную назначает правило письму.
    Если save_pattern=True — создаёт/обновляет паттерн
    для будущей авто-разметки.
    Если process_now=False — письмо не обрабатывается сразу, а помечается
    как queued (удобно для мастера, чтобы не держать UI в ожидании).
    """
    inbox_email = await get_inbox_email(session, email_id)
    if inbox_email is None:
        raise ValueError(f'InboxEmail id={email_id} не найдено')

    att_extensions = _get_attachment_extensions(
        inbox_email.attachment_info or []
    )
    existing_pattern = await find_matching_pattern(
        session,
        email_account_id=inbox_email.email_account_id,
        from_email=inbox_email.from_email,
        subject=inbox_email.subject or '',
        has_attachments=inbox_email.has_attachments,
        attachment_extensions=att_extensions,
    )

    if existing_pattern and existing_pattern.rule_type == rule_type:
        await increment_pattern_confirmed(session, existing_pattern)
    elif save_pattern:
        from_domain = (
            inbox_email.from_email.split('@')[-1]
            if '@' in inbox_email.from_email else None
        )
        words = [w for w in (inbox_email.subject or '').split() if len(w) > 3]
        subject_keywords = words[:5]

        await create_rule_pattern(
            session,
            email_account_id=inbox_email.email_account_id,
            from_email_pattern=inbox_email.from_email,
            from_domain_pattern=from_domain,
            subject_keywords=subject_keywords,
            requires_attachments=inbox_email.has_attachments or None,
            attachment_extensions=att_extensions,
            rule_type=rule_type,
            created_by_id=user_id,
        )
        logger.info(
            'Создан паттерн rule=%s для from=%s',
            rule_type, inbox_email.from_email
        )

    await update_inbox_email_rule(
        session,
        email=inbox_email,
        rule_type=rule_type,
        rule_set_by_id=user_id,
        auto_detected=False,
    )

    if process_now:
        result, error = await _process_email_by_rule(
            session,
            inbox_email=inbox_email,
            rule_type=rule_type,
        )
    else:
        result, error = (
            {
                'action': rule_type,
                'status': 'queued',
                'note': (
                    queued_note
                    or 'Настройки сохранены. Обработка будет запущена позже.'
                ),
            },
            None,
        )
    await mark_processed(
        session, email=inbox_email, result=result, error=error
    )
    await session.commit()
    return inbox_email


def _normalize_sender_emails(value: object) -> set[str]:
    if value is None:
        return set()
    if isinstance(value, str):
        chunks = value.split(',')
    elif isinstance(value, (list, tuple, set)):
        chunks = [str(item or '') for item in value]
    else:
        chunks = [str(value)]
    normalized: set[str] = set()
    for chunk in chunks:
        cleaned = chunk.strip().lower()
        if cleaned:
            normalized.add(cleaned)
    return normalized


def _subject_matches_pattern(
    pattern: Optional[str],
    subject: Optional[str],
) -> bool:
    if not pattern:
        return True
    subject_text = subject or ''
    try:
        return bool(re.search(pattern, subject_text, flags=re.IGNORECASE))
    except re.error:
        return pattern.lower() in subject_text.lower()


def _rule_matches_payload_type(
    *,
    rule_type: str,
    payload_type: Optional[str],
) -> bool:
    normalized_payload = (payload_type or 'response').strip().lower()
    if rule_type == 'order_reply':
        return normalized_payload == 'response'
    if rule_type == 'document':
        return normalized_payload == 'document'
    return False


async def _find_matching_supplier_response_configs(
    session: AsyncSession,
    *,
    inbox_email: InboxEmail,
    rule_type: str,
) -> list:
    from sqlalchemy import select as _select

    from dz_fastapi.models.partner import SupplierResponseConfig as _SRC

    sender_email = (inbox_email.from_email or '').strip().lower()
    if not sender_email:
        return []

    result = await session.execute(
        _select(_SRC).where(_SRC.is_active.is_(True))
    )
    configs = result.scalars().all()
    matched: list = []

    for config in configs:
        if not _rule_matches_payload_type(
            rule_type=rule_type,
            payload_type=getattr(config, 'file_payload_type', None),
        ):
            continue
        if (
            inbox_email.email_account_id is not None
            and config.inbox_email_account_id
            not in (None, inbox_email.email_account_id)
        ):
            continue
        senders = _normalize_sender_emails(config.sender_emails)
        if senders and sender_email not in senders:
            continue
        if not _subject_matches_pattern(
            config.subject_pattern,
            inbox_email.subject,
        ):
            continue
        matched.append(config)

    return matched


async def _create_supplier_response_registry_stub(
    session: AsyncSession,
    *,
    inbox_email: InboxEmail,
    rule_type: str,
    provider_id: Optional[int] = None,
    response_config_ids: Optional[list[int]] = None,
    reason_text: Optional[str] = None,
) -> int:
    """
    Создаёт служебную запись в SupplierOrderMessage, чтобы менеджер сразу видел
    письмо в "Реестре обработанных писем" конкретной конфигурации.

    Важно: source_uid не заполняем, чтобы не блокировать последующую
    полноценную обработку тем же письмом в шедулере.
    """
    if rule_type not in {'order_reply', 'document'}:
        return 0

    from dz_fastapi.models.partner import SupplierOrderAttachment as _SOA
    from dz_fastapi.models.partner import SupplierOrderMessage as _SOM
    from dz_fastapi.models.partner import SupplierResponseConfig as _SRC

    normalized_ids: list[int] = sorted(
        {
            int(raw_id)
            for raw_id in (response_config_ids or [])
            if raw_id not in (None, 0, '')
        }
    )
    if not normalized_ids:
        matched_configs = await _find_matching_supplier_response_configs(
            session,
            inbox_email=inbox_email,
            rule_type=rule_type,
        )
        normalized_ids = sorted({int(cfg.id) for cfg in matched_configs})

    targets: list[tuple[Optional[int], int]] = []
    if normalized_ids:
        cfg_rows = (
            await session.execute(
                select(_SRC).where(_SRC.id.in_(normalized_ids))
            )
        ).scalars().all()
        for cfg in cfg_rows:
            targets.append((int(cfg.id), int(cfg.provider_id)))
    elif provider_id:
        targets.append((None, int(provider_id)))
    else:
        return 0

    created_count = 0
    note = (
        reason_text
        or 'Назначено через Inbox. Ожидает обработки по расписанию.'
    )[:500]
    sender_email = (inbox_email.from_email or '').strip() or None
    subject = (inbox_email.subject or '')[:500] or None
    message_type = 'SHIPPING_DOC' if rule_type == 'document' else 'UNKNOWN'
    received_at = inbox_email.received_at or now_moscow()

    for config_id, target_provider_id in targets:
        marker = (
            f'inbox-setup:{int(inbox_email.id)}:{rule_type}:'
            f'{int(config_id or 0)}'
        )
        existing = (
            await session.execute(
                select(_SOM).where(
                    _SOM.source_message_id == marker,
                    _SOM.provider_id == target_provider_id,
                )
            )
        ).scalar_one_or_none()
        if existing is not None:
            continue

        row = _SOM(
            supplier_order_id=None,
            provider_id=target_provider_id,
            message_type=message_type,
            response_config_id=config_id,
            subject=subject,
            sender_email=sender_email,
            received_at=received_at,
            body_preview=(
                f'Inbox setup ({rule_type}) for email #{inbox_email.id}.'
            ),
            raw_status=None,
            normalized_status=None,
            parse_confidence=None,
            source_uid=None,
            source_message_id=marker,
            import_error_details=note,
            mapping_id=None,
        )
        session.add(row)
        await session.flush()

        for idx, att in enumerate(inbox_email.attachment_info or []):
            filename = str(att.get('name') or '').strip()
            if not filename:
                filename = f'attachment_{idx + 1}'
            file_path = str(att.get('path') or '').strip()
            if not file_path:
                file_path = (
                    f'inbox-setup://email/{int(inbox_email.id)}/'
                    f'attachment/{idx + 1}'
                )
            session.add(
                _SOA(
                    message_id=row.id,
                    filename=filename[:255],
                    mime_type=None,
                    file_path=file_path[:1024],
                    sha256=None,
                    parsed_kind=(
                        'SHIPPING_DOC'
                        if rule_type == 'document'
                        else 'RESPONSE_FILE'
                    ),
                )
            )

        created_count += 1

    return created_count


def _build_supplier_source_uid_for_inbox_email(
    inbox_email: InboxEmail,
) -> Optional[str]:
    uid = getattr(inbox_email, 'uid', None)
    if uid in (None, ''):
        return None
    folder = str(getattr(inbox_email, 'folder', '') or '').strip()
    account_id = int(getattr(inbox_email, 'email_account_id', 0) or 0)
    return f'{account_id}:{folder}:{uid}'[:128]


async def _reset_supplier_message_source_markers(
    session: AsyncSession,
    *,
    source_uid: Optional[str],
    allow_reprocess: bool,
) -> int:
    if not allow_reprocess or not source_uid:
        return 0

    from sqlalchemy import select as _select

    from dz_fastapi.models.partner import SupplierOrderMessage

    rows = (
        await session.execute(
            _select(SupplierOrderMessage).where(
                SupplierOrderMessage.source_uid == source_uid
            )
        )
    ).scalars().all()
    changed = 0
    for row in rows:
        if row.source_uid is None and row.source_message_id is None:
            continue
        row.source_uid = None
        row.source_message_id = None
        row.message_type = 'RETRY_PENDING'
        session.add(row)
        changed += 1
    if changed:
        await session.flush()
    return changed


def _build_force_processing_error_text(
    failed_configs: list[dict],
) -> Optional[str]:
    if not failed_configs:
        return None
    chunks = []
    for item in failed_configs[:3]:
        config_id = item.get('config_id')
        error = str(item.get('error') or 'unknown error')
        chunks.append(f'config {config_id}: {error}')
    extra = len(failed_configs) - len(chunks)
    if extra > 0:
        chunks.append(f'+{extra} more errors')
    return '; '.join(chunks)


async def _create_force_process_audit_record(
    session: AsyncSession,
    *,
    inbox_email_id: int,
    user_id: Optional[int],
    rule_type: str,
    allow_reprocess: bool,
    status: str,
    reason_code: Optional[str],
    reason_text: Optional[str],
    details: dict,
) -> InboxForceProcessAudit:
    mode = 'reprocess' if allow_reprocess else 'check'
    audit = InboxForceProcessAudit(
        inbox_email_id=inbox_email_id,
        requested_by_user_id=user_id,
        rule_type=rule_type,
        mode=mode,
        allow_reprocess=bool(allow_reprocess),
        status=status,
        reason_code=reason_code,
        reason_text=reason_text,
        details=details,
    )
    session.add(audit)
    await session.flush()
    return audit


async def force_process_email(
    session: AsyncSession,
    *,
    email_id: int,
    user_id: Optional[int],
    allow_reprocess: bool = True,
) -> dict:
    """
    Принудительно запускает обработку письма по уже назначенному правилу.
    Поддерживаются:
      - customer_order
      - order_reply
      - document
    """
    inbox_email = await get_inbox_email(session, email_id)
    if inbox_email is None:
        raise LookupError(f'InboxEmail id={email_id} не найдено')

    rule_type = str(inbox_email.rule_type or '').strip()
    if not rule_type:
        raise ValueError('Для письма не назначено правило')
    if rule_type not in FORCE_PROCESSABLE_RULES:
        raise ValueError(
            'Принудительная обработка доступна только для правил: '
            'customer_order, order_reply, document'
        )

    processing_result: dict = {
        'action': 'force_process',
        'status': 'started',
        'reason_code': None,
        'reason': None,
        'reasons': [],
        'rule_type': rule_type,
        'mode': 'reprocess' if allow_reprocess else 'check',
        'allow_reprocess': bool(allow_reprocess),
        'requested_by_user_id': user_id,
        'forced_at': now_moscow().isoformat(),
        'summary': {
            'matched_configs_count': 0,
            'triggered_configs_count': 0,
            'failed_configs_count': 0,
            'reprocess_reset_messages': 0,
        },
    }
    failed_configs: list[dict] = []

    try:
        if rule_type == 'customer_order':
            from dz_fastapi.services.customer_orders import \
                process_customer_orders

            matched_config_ids = await _find_matching_customer_order_configs(
                session,
                from_email=inbox_email.from_email,
                subject=inbox_email.subject,
                email_account_id=inbox_email.email_account_id,
            )
            processing_result['matched_config_ids'] = matched_config_ids
            processing_result['summary']['matched_configs_count'] = len(
                matched_config_ids
            )
            if not matched_config_ids:
                reason = (
                    'Не найден активный CustomerOrderConfig для этого письма'
                )
                processing_result['status'] = 'missing_config'
                processing_result['reason_code'] = 'missing_config'
                processing_result['reason'] = reason
                processing_result['reasons'].append(
                    {
                        'code': 'missing_config',
                        'message': reason,
                    }
                )
            else:
                triggered_config_ids: list[int] = []
                for config_id in matched_config_ids:
                    try:
                        await process_customer_orders(
                            session,
                            config_id=config_id,
                        )
                        triggered_config_ids.append(int(config_id))
                    except Exception as exc:
                        await session.rollback()
                        logger.exception(
                            'Force processing customer_order failed: '
                            'email_id=%s config_id=%s error=%s',
                            inbox_email.id,
                            config_id,
                            exc,
                        )
                        failed_entry = {
                            'config_id': int(config_id),
                            'code': 'config_processing_failed',
                            'error': str(exc),
                        }
                        failed_configs.append(failed_entry)
                        processing_result['reasons'].append(
                            {
                                'code': 'config_processing_failed',
                                'config_id': int(config_id),
                                'message': str(exc),
                            }
                        )
                processing_result['triggered_config_ids'] = (
                    triggered_config_ids
                )
                if failed_configs and triggered_config_ids:
                    processing_result['status'] = 'partially_triggered'
                    processing_result['reason_code'] = 'partial_failure'
                    processing_result['reason'] = (
                        'Часть конфигураций завершилась с ошибкой'
                    )
                elif failed_configs:
                    processing_result['status'] = 'failed'
                    processing_result['reason_code'] = 'processing_failed'
                    processing_result['reason'] = (
                        'Обработка не запустилась ни по одной конфигурации'
                    )
                else:
                    processing_result['status'] = 'triggered'
                    processing_result['reason_code'] = 'triggered'
                    processing_result['reason'] = (
                        'Обработка успешно запущена'
                    )
                processing_result['summary']['triggered_configs_count'] = len(
                    triggered_config_ids
                )
                processing_result['summary']['failed_configs_count'] = len(
                    failed_configs
                )
        else:
            from dz_fastapi.services.supplier_order_responses import (
                process_supplier_response_messages, supplier_response_cutoff)

            matched_configs = await _find_matching_supplier_response_configs(
                session,
                inbox_email=inbox_email,
                rule_type=rule_type,
            )
            matched_config_ids = [
                int(config.id) for config in matched_configs
            ]
            processing_result['matched_config_ids'] = matched_config_ids
            processing_result['summary']['matched_configs_count'] = len(
                matched_config_ids
            )

            if not matched_configs:
                reason = (
                    'Не найден активный SupplierResponseConfig '
                    'для этого письма'
                )
                processing_result['status'] = 'missing_config'
                processing_result['reason_code'] = 'missing_config'
                processing_result['reason'] = reason
                processing_result['reasons'].append(
                    {
                        'code': 'missing_config',
                        'message': reason,
                    }
                )
            else:
                source_uid = _build_supplier_source_uid_for_inbox_email(
                    inbox_email
                )
                reset_messages = await _reset_supplier_message_source_markers(
                    session,
                    source_uid=source_uid,
                    allow_reprocess=allow_reprocess,
                )
                processing_result['source_uid'] = source_uid
                processing_result['reprocess_reset_messages'] = reset_messages
                processing_result['summary']['reprocess_reset_messages'] = (
                    reset_messages
                )
                if allow_reprocess and source_uid and not reset_messages:
                    processing_result['reasons'].append(
                        {
                            'code': 'nothing_to_reprocess',
                            'message': (
                                'Для письма не найдено ранее обработанных '
                                'записей для сброса дедупликации'
                            ),
                        }
                    )
                if reset_messages:
                    await session.commit()

                received_at = getattr(inbox_email, 'received_at', None)
                if isinstance(received_at, datetime):
                    date_from = received_at.date()
                elif isinstance(received_at, date):
                    date_from = received_at
                else:
                    date_from = supplier_response_cutoff()

                triggered_config_ids: list[int] = []
                run_summaries: list[dict] = []
                for config in matched_configs:
                    config_id = int(config.id)
                    try:
                        run_result = await process_supplier_response_messages(
                            session=session,
                            provider_id=int(config.provider_id),
                            supplier_response_config_id=config_id,
                            date_from=date_from,
                            date_to=None,
                        )
                        run_summaries.append(
                            {'config_id': config_id, **run_result}
                        )
                        triggered_config_ids.append(config_id)
                    except Exception as exc:
                        await session.rollback()
                        logger.exception(
                            'Force processing supplier response failed: '
                            'email_id=%s config_id=%s error=%s',
                            inbox_email.id,
                            config_id,
                            exc,
                        )
                        failed_entry = {
                            'config_id': config_id,
                            'code': 'config_processing_failed',
                            'error': str(exc),
                        }
                        failed_configs.append(failed_entry)
                        processing_result['reasons'].append(
                            {
                                'code': 'config_processing_failed',
                                'config_id': config_id,
                                'message': str(exc),
                            }
                        )

                processing_result['triggered_config_ids'] = (
                    triggered_config_ids
                )
                processing_result['runs'] = run_summaries
                if failed_configs and triggered_config_ids:
                    processing_result['status'] = 'partially_triggered'
                    processing_result['reason_code'] = 'partial_failure'
                    processing_result['reason'] = (
                        'Часть конфигураций завершилась с ошибкой'
                    )
                elif failed_configs:
                    processing_result['status'] = 'failed'
                    processing_result['reason_code'] = 'processing_failed'
                    processing_result['reason'] = (
                        'Обработка не запустилась ни по одной конфигурации'
                    )
                else:
                    processing_result['status'] = 'triggered'
                    processing_result['reason_code'] = 'triggered'
                    processing_result['reason'] = (
                        'Обработка успешно запущена'
                    )
                processing_result['summary']['triggered_configs_count'] = len(
                    triggered_config_ids
                )
                processing_result['summary']['failed_configs_count'] = len(
                    failed_configs
                )
    except Exception as exc:
        await session.rollback()
        logger.exception(
            'Unexpected force processing error: email_id=%s error=%s',
            inbox_email.id,
            exc,
        )
        processing_result['status'] = 'failed'
        processing_result['reason_code'] = 'unexpected_error'
        processing_result['reason'] = (
            'Непредвиденная ошибка во время принудительной обработки'
        )
        processing_result['reasons'].append(
            {
                'code': 'unexpected_error',
                'message': str(exc),
            }
        )
        failed_configs.append(
            {
                'config_id': None,
                'code': 'unexpected_error',
                'error': str(exc),
            }
        )

    processing_result['failed_configs'] = failed_configs
    processing_result['summary']['failed_configs_count'] = len(failed_configs)
    if processing_result.get('reason_code') is None:
        processing_result['reason_code'] = 'unknown'
    if processing_result.get('reason') is None:
        processing_result['reason'] = 'Статус обработки обновлён'

    audit = await _create_force_process_audit_record(
        session,
        inbox_email_id=inbox_email.id,
        user_id=user_id,
        rule_type=rule_type,
        allow_reprocess=allow_reprocess,
        status=str(processing_result.get('status') or 'unknown'),
        reason_code=str(processing_result.get('reason_code') or '') or None,
        reason_text=str(processing_result.get('reason') or '') or None,
        details=processing_result,
    )
    processing_result['audit_id'] = int(audit.id)
    audit.details = processing_result

    error_text = _build_force_processing_error_text(failed_configs)
    await mark_processed(
        session,
        email=inbox_email,
        result=processing_result,
        error=error_text,
    )
    await session.commit()
    await session.refresh(inbox_email)
    return {
        'id': inbox_email.id,
        'rule_type': inbox_email.rule_type,
        'processed': inbox_email.processed,
        'processing_result': inbox_email.processing_result,
        'processing_error': inbox_email.processing_error,
    }


# ---------------------------------------------------------------------------
# Диспетчер правил
# ---------------------------------------------------------------------------

async def _process_email_by_rule(
    session: AsyncSession,
    *,
    inbox_email: InboxEmail,
    rule_type: str,
    fetched_msg: Optional[_FetchedInboxMessage] = None,
    account=None,
) -> Tuple[Optional[dict], Optional[str]]:
    """
    Вызывает нужный обработчик в зависимости от rule_type.
    Возвращает (result_dict, error_str).
    """
    try:
        if rule_type == 'price_list':
            return await _process_price_list(
                session, inbox_email=inbox_email, fetched_msg=fetched_msg,
            )
        elif rule_type == 'order_reply':
            return await _process_order_reply(
                session, inbox_email=inbox_email, fetched_msg=fetched_msg,
            )
        elif rule_type == 'customer_order':
            return await _process_customer_order(
                session, inbox_email=inbox_email, fetched_msg=fetched_msg,
            )
        elif rule_type == 'document':
            return await _process_document(
                session, inbox_email=inbox_email, fetched_msg=fetched_msg,
            )
        elif rule_type == 'shipment_notice':
            return await _process_shipment_notice(
                session, inbox_email=inbox_email, fetched_msg=fetched_msg,
            )
        elif rule_type == 'claim':
            return await _process_notify_manager(
                session, inbox_email=inbox_email, rule_type='claim',
                title='Претензия / рекламация',
                level='warning',
            )
        elif rule_type in ('error_report', 'inquiry', 'proposal'):
            label = _rule_label(rule_type)
            return await _process_notify_manager(
                session, inbox_email=inbox_email, rule_type=rule_type,
                title=label,
                level='info',
            )
        elif rule_type == 'spam':
            logger.info('Письмо id=%s помечено как спам', inbox_email.id)
            return {'action': 'spam', 'hidden': True}, None
        elif rule_type == 'ignore':
            logger.info(
                'Письмо id=%s помечено как "игнорировать"', inbox_email.id
            )
            return {'action': 'ignored'}, None
        else:
            return None, f'Неизвестный тип правила: {rule_type}'
    except Exception as e:
        logger.exception(
            'Ошибка обработки письма id=%s rule=%s: %s',
            inbox_email.id, rule_type, e,
        )
        return None, str(e)


# ---------------------------------------------------------------------------
# Обработчики по типам
# ---------------------------------------------------------------------------

async def _process_price_list(
    session: AsyncSession,
    *,
    inbox_email: InboxEmail,
    fetched_msg: Optional[_FetchedInboxMessage] = None,
) -> Tuple[Optional[dict], Optional[str]]:
    """Прайс-лист поставщика → скачать и обработать."""
    from dz_fastapi.crud.partner import (crud_provider,
                                         crud_provider_pricelist_config)
    from dz_fastapi.services.email import (_message_matches_provider_config,
                                           download_new_price_provider)
    from dz_fastapi.services.process import process_provider_pricelist

    if fetched_msg is None:
        return {'action': 'price_list', 'note': 'no fetched_msg'}, None

    provider = await crud_provider.get_by_email_incoming_price(
        session=session, email=inbox_email.from_email
    )
    if provider is None:
        return {
            'action': 'price_list',
            'status': 'provider_not_found',
            'from_email': inbox_email.from_email,
        }, None

    configs = await crud_provider_pricelist_config.get_configs(
        provider_id=provider.id, session=session, only_active=True
    )
    processed_configs = []
    for config in configs:
        if not _message_matches_provider_config(fetched_msg, config):
            continue
        filepath = await download_new_price_provider(
            msg=fetched_msg, provider=provider,
            provider_conf=config, session=session,
        )
        if filepath:
            try:
                await process_provider_pricelist(
                    provider=provider,
                    provider_conf=config,
                    filepath=filepath,
                    session=session,
                )
                processed_configs.append(config.id)
            except Exception as e:
                logger.error(
                    'Ошибка обработки прайса provider=%s: %s',
                    provider.id, e
                )

    return {
        'action': 'price_list',
        'provider_id': provider.id,
        'provider_name': provider.name,
        'configs_processed': processed_configs,
    }, None


async def _process_order_reply(
    session: AsyncSession,
    *,
    inbox_email: InboxEmail,
    fetched_msg: Optional[_FetchedInboxMessage] = None,
) -> Tuple[Optional[dict], Optional[str]]:
    """
    Ответ поставщика на заказ (подтверждение / отказ / частично).

    При назначении правила вручную из Inbox — только тегируем и уведомляем.
    Полная обработка (обновление статусов заказов) произойдёт на следующем
    запуске шедулера process_supplier_response_messages, который увидит
    это письмо уже с правилом и обработает его.
    """
    from dz_fastapi.services.notifications import create_admin_notifications

    try:
        await create_admin_notifications(
            session=session,
            title=f'Ответ поставщика: {inbox_email.from_email}',
            message=(
                f'Получен ответ на заказ от {inbox_email.from_email}.\n'
                f'Тема: {inbox_email.subject or "(без темы)"}\n'
                'Будет обработан автоматически при следующем запуске шедулера.'
            ),
            level='info',
        )
    except Exception as e:
        logger.warning('Не удалось создать уведомление order_reply: %s', e)

    return {
        'action': 'order_reply',
        'from_email': inbox_email.from_email,
        'subject': inbox_email.subject,
        'status': 'queued',
        'note': 'Будет обработан шедулером при следующем запуске',
    }, None


async def _process_customer_order(
    session: AsyncSession,
    *,
    inbox_email: InboxEmail,
    fetched_msg: Optional[_FetchedInboxMessage] = None,
) -> Tuple[Optional[dict], Optional[str]]:
    """
    Входящий заказ от клиента.

    При назначении правила вручную из Inbox — только тегируем и уведомляем.
    Полная обработка (разбор файла, создание CustomerOrder) произойдёт на
    следующем запуске шедулера download_customer_orders_task, который
    найдёт это письмо по EmailRulePattern / order_email в конфиге.
    """
    from dz_fastapi.services.notifications import create_admin_notifications

    matched_config_ids = await _find_matching_customer_order_configs(
        session,
        from_email=inbox_email.from_email,
        subject=inbox_email.subject,
        email_account_id=inbox_email.email_account_id,
    )
    if not matched_config_ids:
        reason = (
            'Не найден активный конфиг заказа клиента '
            'для отправителя/темы письма'
        )
        try:
            await create_admin_notifications(
                session=session,
                title='Заказ клиента без активного конфига',
                message=(
                    f'Письмо от {inbox_email.from_email} '
                    'помечено как customer_order, но '
                    'подходящий CustomerOrderConfig не найден.\n'
                    f'Тема: {inbox_email.subject or "(без темы)"}\n'
                    'Проверьте раздел "Заказы клиентов → Конфигурация".'
                ),
                level='warning',
            )
        except Exception as e:
            logger.warning(
                'Не удалось создать уведомление о missing config: %s', e
            )
        return {
            'action': 'customer_order',
            'from_email': inbox_email.from_email,
            'subject': inbox_email.subject,
            'status': 'missing_config',
            'reason': reason,
            'note': (
                'Письмо не будет загружено в заказы, '
                'пока не добавлен активный конфиг.'
            ),
        }, None

    try:
        att_names = [
            a.get('name', '') for a in (inbox_email.attachment_info or [])
        ]
        await create_admin_notifications(
            session=session,
            title=f'Заказ от клиента: {inbox_email.from_email}',
            message=(
                f'Получен заказ от {inbox_email.from_email}.\n'
                f'Тема: {inbox_email.subject or "(без темы)"}\n'
                + (f'Вложения: {", ".join(att_names)}\n' if att_names else '')
                + (
                    'Будет обработан автоматически при следующем запуске '
                    'шедулера.'
                )
            ),
            level='info',
        )
    except Exception as e:
        logger.warning('Не удалось создать уведомление customer_order: %s', e)

    return {
        'action': 'customer_order',
        'from_email': inbox_email.from_email,
        'subject': inbox_email.subject,
        'matched_config_ids': matched_config_ids,
        'status': 'queued',
        'note': 'Будет обработан шедулером при следующем запуске',
    }, None


async def _process_document(
    session: AsyncSession,
    *,
    inbox_email: InboxEmail,
    fetched_msg: Optional[_FetchedInboxMessage] = None,
) -> Tuple[Optional[dict], Optional[str]]:
    """
    Документ от поставщика (накладная / счёт / акт / счёт-фактура).
    Ищет поставщика по email и создаёт черновик SupplierReceipt для ручного
    подтверждения менеджером.
    """
    from dz_fastapi.crud.partner import crud_provider
    from dz_fastapi.services.notifications import create_admin_notifications

    provider = await crud_provider.get_by_email_incoming_price(
        session=session, email=inbox_email.from_email
    )

    # Пытаемся извлечь номер и дату документа из темы письма
    doc_number = _extract_doc_number(inbox_email.subject or '')
    doc_date = _extract_doc_date(inbox_email.subject or '')

    result = {
        'action': 'document',
        'from_email': inbox_email.from_email,
        'subject': inbox_email.subject,
        'provider_id': provider.id if provider else None,
        'provider_name': provider.name if provider else None,
        'doc_number': doc_number,
        'doc_date': doc_date.isoformat() if doc_date else None,
        'attachments': [
            att.get('name') for att in (inbox_email.attachment_info or [])
        ],
        'status': 'pending_manual_review',
        'note': (
            'Документ готов к ручному оформлению. '
            'Откройте раздел "Документы → Входящие".'
        ),
    }

    # Уведомляем менеджера
    try:
        provider_name = provider.name if provider else inbox_email.from_email
        await create_admin_notifications(
            session=session,
            title=f'Документ от {provider_name}',
            message=(
                f'Получен документ от {inbox_email.from_email}.\n'
                f'Тема: {inbox_email.subject}\n'
                f'Вложений: {len(inbox_email.attachment_info or [])}\n'
                + (f'Номер документа: {doc_number}\n' if doc_number else '')
                + 'Требует ручного оформления в разделе '
                '"Документы → Входящие".'
            ),
            level='info',
        )
    except Exception as e:
        logger.warning('Не удалось создать уведомление для документа: %s', e)

    return result, None


async def _process_shipment_notice(
    session: AsyncSession,
    *,
    inbox_email: InboxEmail,
    fetched_msg: Optional[_FetchedInboxMessage] = None,
) -> Tuple[Optional[dict], Optional[str]]:
    """
    Уведомление об отгрузке / трекинг-номер от поставщика.
    Извлекает трекинг-номер из темы/тела, уведомляет менеджера.
    """
    from dz_fastapi.crud.partner import crud_provider
    from dz_fastapi.services.notifications import create_admin_notifications

    provider = await crud_provider.get_by_email_incoming_price(
        session=session, email=inbox_email.from_email
    )

    # Пытаемся найти трекинг-номер в теме письма
    tracking_number = _extract_tracking_number(inbox_email.subject or '')

    result = {
        'action': 'shipment_notice',
        'from_email': inbox_email.from_email,
        'subject': inbox_email.subject,
        'provider_id': provider.id if provider else None,
        'provider_name': provider.name if provider else None,
        'tracking_number': tracking_number,
        'status': 'notified',
    }

    try:
        provider_name = provider.name if provider else inbox_email.from_email
        tracking_info = (
            f'\nТрекинг: {tracking_number}' if tracking_number else ''
        )
        await create_admin_notifications(
            session=session,
            title=f'Отгрузка от {provider_name}',
            message=(
                f'Поставщик {inbox_email.from_email} сообщил об отгрузке.\n'
                f'Тема: {inbox_email.subject}'
                + tracking_info
            ),
            level='info',
        )
    except Exception as e:
        logger.warning('Не удалось создать уведомление об отгрузке: %s', e)

    return result, None


async def _process_notify_manager(
    session: AsyncSession,
    *,
    inbox_email: InboxEmail,
    rule_type: str,
    title: str,
    level: str = 'info',
) -> Tuple[Optional[dict], Optional[str]]:
    """
    Общий обработчик для типов, требующих только уведомления менеджера:
    claim, error_report, inquiry, proposal.
    """
    from dz_fastapi.services.notifications import create_admin_notifications

    try:
        await create_admin_notifications(
            session=session,
            title=f'{title}: {inbox_email.from_email}',
            message=(
                f'Тип: {title}\n'
                f'От: {inbox_email.from_email}\n'
                f'Тема: {inbox_email.subject or "(без темы)"}\n'
                f'Вложений: {len(inbox_email.attachment_info or [])}'
            ),
            level=level,
        )
    except Exception as e:
        logger.warning(
            'Не удалось создать уведомление rule=%s: %s',
            rule_type,
            e,
        )

    return {
        'action': rule_type,
        'from_email': inbox_email.from_email,
        'subject': inbox_email.subject,
        'status': 'manager_notified',
    }, None


# ---------------------------------------------------------------------------
# Вспомогательные парсеры для извлечения данных из темы письма
# ---------------------------------------------------------------------------

def _extract_doc_number(subject: str) -> Optional[str]:
    """Пытается извлечь номер документа из темы письма."""
    patterns = [
        r'№\s*([А-Яа-яA-Za-z0-9/-]+)',
        r'[Нн]омер[:\s]+([А-Яа-яA-Za-z0-9/-]+)',
        r'\bN[o°]?\s*([A-Za-z0-9/-]+)',
        r'[Сс]чёт[:\s#№]+([А-Яа-яA-Za-z0-9/-]+)',
        r'[Нн]акладная[:\s#№]+([А-Яа-яA-Za-z0-9/-]+)',
    ]
    for pattern in patterns:
        m = re.search(pattern, subject)
        if m:
            return m.group(1).strip()
    return None


def _extract_doc_date(subject: str) -> Optional[date]:
    """Пытается извлечь дату документа из темы письма."""
    patterns = [
        r'(\d{2})[.\-/](\d{2})[.\-/](\d{4})',   # 01.04.2026
        r'(\d{4})[.\-/](\d{2})[.\-/](\d{2})',   # 2026-04-01
    ]
    for pattern in patterns:
        m = re.search(pattern, subject)
        if m:
            try:
                groups = m.groups()
                if len(groups[0]) == 4:
                    return date(int(groups[0]), int(groups[1]), int(groups[2]))
                else:
                    return date(int(groups[2]), int(groups[1]), int(groups[0]))
            except (ValueError, IndexError):
                continue
    return None


def _extract_tracking_number(subject: str) -> Optional[str]:
    """Пытается извлечь трекинг-номер из темы письма."""
    patterns = [
        r'[Тт]рек[инг]*[:\s#№]+([A-Za-z0-9]{6,30})',
        r'[Тт]рекинг[:\s]+([A-Za-z0-9]{6,30})',
        r'[Оо]тслеж[:\s]+([A-Za-z0-9]{6,30})',
        r'\b(RU\d{13}RU)\b',          # Почта России
        r'\b([A-Z]{2}\d{9}[A-Z]{2})\b',  # международный формат
        r'[Tt]rack(?:ing)?[:\s#]+([A-Za-z0-9]{6,30})',
    ]
    for pattern in patterns:
        m = re.search(pattern, subject)
        if m:
            return m.group(1).strip()
    return None


# ---------------------------------------------------------------------------
# Мастер настройки: привязка к реальным конфигам системы
# ---------------------------------------------------------------------------

async def setup_email_rule(
    session: AsyncSession,
    *,
    email_id: int,
    rule_type: str,
    user_id: Optional[int],
    save_pattern: bool = True,
    provider_config=None,   # ProviderSetupConfig | None
    customer_config=None,   # CustomerSetupConfig | None
) -> dict:
    """
    Полная настройка правила через мастер:
      1. Связывает отправителя письма с поставщиком или клиентом
         (обновляет email_incoming_price / order_email
         в реальных таблицах системы)
      2. Сохраняет EmailRulePattern для будущей авто-разметки
      3. Запускает обработку письма по выбранному правилу
         (или ставит в очередь для тяжёлых сценариев)

    Благодаря шагу 1 будущие письма от того же отправителя будут
    auto-detected через _detect_rule_from_existing_configs
    — без повторной настройки.
    """
    from dz_fastapi.crud.partner import crud_customer, crud_provider

    inbox_email = await get_inbox_email(session, email_id)
    if inbox_email is None:
        raise ValueError(f'InboxEmail id={email_id} не найдено')

    configs_set: list[dict] = []
    supplier_registry_provider_id: Optional[int] = None
    supplier_registry_config_ids: list[int] = []

    PROVIDER_RULES = {
        'price_list',
        'order_reply',
        'document',
        'shipment_notice',
    }

    # ------------------------------------------------------------------
    # 1a. Привязка к поставщику + обновление конфигурации файла
    # ------------------------------------------------------------------
    if (
        rule_type in PROVIDER_RULES
        and provider_config
        and provider_config.provider_id
    ):
        cfg_mode = getattr(provider_config, 'config_mode', 'skip')
        price_cfg_name = str(
            getattr(provider_config, 'config_name', '') or ''
        ).strip()
        price_filename_pattern = str(
            getattr(provider_config, 'filename_pattern', '') or ''
        ).strip()
        if (
            rule_type == 'price_list'
            and cfg_mode == 'existing'
            and not getattr(provider_config, 'config_id', None)
        ):
            raise ValueError(
                'Выберите конфигурацию прайс-листа для обновления'
            )
        if rule_type == 'price_list' and cfg_mode == 'skip':
            raise ValueError(
                'Для прайс-листа выберите режим: '
                '«Обновить готовую» или «Создать новую»'
            )
        if rule_type == 'price_list' and cfg_mode == 'new':
            if not price_cfg_name:
                raise ValueError(
                    'Для новой конфигурации прайс-листа '
                    'укажите имя конфигурации'
                )
            if not getattr(provider_config, 'oem_col', None):
                raise ValueError(
                    'Для новой конфигурации прайс-листа укажите колонку OEM'
                )
            if not getattr(provider_config, 'qty_col', None):
                raise ValueError(
                    'Для новой конфигурации прайс-листа укажите колонку Кол-во'
                )
            if not getattr(provider_config, 'price_col', None):
                raise ValueError(
                    'Для новой конфигурации прайс-листа укажите колонку Цена'
                )
        try:
            from dz_fastapi.crud.partner import (
                crud_provider_pricelist_config, crud_supplier_response_config)
            from dz_fastapi.models.partner import ProviderConfigLastEmailUID
            from dz_fastapi.schemas.partner import (
                ProviderPriceListConfigCreate, SupplierResponseConfigCreate)

            provider = await crud_provider.get_by_id(
                provider_id=provider_config.provider_id, session=session
            )
            if provider:
                supplier_registry_provider_id = int(provider.id)
                mailbox_account_id = (
                    int(inbox_email.email_account_id)
                    if inbox_email.email_account_id is not None
                    else None
                )
                sender_email = str(
                    inbox_email.from_email or ''
                ).strip().lower()
                previous_sender = str(
                    provider.email_incoming_price or ''
                ).strip().lower()

                # Для мастера из Inbox отправитель всегда становится
                # актуальным email поставщика для прайсов.
                if sender_email and previous_sender != sender_email:
                    provider.email_incoming_price = sender_email
                    session.add(provider)
                    if previous_sender:
                        action = 'relinked'
                        note = (
                            f'Поставщик «{provider.name}» перепривязан: '
                            f'{previous_sender} → {sender_email}.'
                        )
                    else:
                        action = 'linked'
                        note = (
                            f'Email {sender_email} привязан к '
                            f'поставщику «{provider.name}».'
                        )
                else:
                    action = 'already_linked'
                    note = (
                        f'Поставщик «{provider.name}» уже привязан к '
                        f'{provider.email_incoming_price}.'
                    )

                configs_set.append({
                    'entity_type': 'provider',
                    'entity_id': provider.id,
                    'entity_name': provider.name,
                    'action': action,
                    'note': note,
                })

                # --- price_list: ProviderPriceListConfig ---
                if rule_type == 'price_list' and cfg_mode != 'skip':
                    async def _reset_pricelist_uid_state(
                        config_id: int,
                    ) -> None:
                        uid_state = (
                            await session.execute(
                                select(ProviderConfigLastEmailUID).where(
                                    ProviderConfigLastEmailUID
                                    .provider_config_id == config_id
                                )
                            )
                        ).scalar_one_or_none()
                        if uid_state is None:
                            uid_state = ProviderConfigLastEmailUID(
                                provider_config_id=config_id,
                                last_uid=0,
                                folder_last_uids={},
                            )
                        else:
                            uid_state.last_uid = 0
                            uid_state.folder_last_uids = {}
                        session.add(uid_state)

                    def _to_zero_based_col(v):
                        if v is None or v == '':
                            return None
                        parsed = int(v)
                        if parsed < 1:
                            raise ValueError(
                                'Номера колонок должны быть >= 1'
                            )
                        return parsed - 1

                    def _to_zero_based_row(v):
                        if v is None or v == '':
                            return None
                        parsed = int(v)
                        if parsed < 1:
                            raise ValueError(
                                'Номер строки начала должен быть >= 1'
                            )
                        return parsed - 1

                    if cfg_mode == 'existing' and provider_config.config_id:
                        pl_cfg = await (
                            crud_provider_pricelist_config.get_config(
                                provider_id=provider.id,
                                config_id=provider_config.config_id,
                                session=session,
                            )
                        )
                        if pl_cfg:
                            upd: dict = {}
                            if provider_config.subject_pattern:
                                upd['name_mail'] = (
                                    provider_config.subject_pattern
                                )
                            if price_cfg_name:
                                upd['name_price'] = price_cfg_name
                            if price_filename_pattern:
                                upd['filename_pattern'] = (
                                    price_filename_pattern
                                )
                            if (
                                mailbox_account_id is not None
                                and pl_cfg.incoming_email_account_id
                                != mailbox_account_id
                            ):
                                upd['incoming_email_account_id'] = (
                                    mailbox_account_id
                                )
                            for fld in (
                                'start_row', 'oem_col', 'qty_col',
                                'price_col', 'brand_col', 'name_col',
                                'multiplicity_col',
                            ):
                                v = getattr(provider_config, fld, None)
                                if v is not None:
                                    if fld == 'start_row':
                                        upd[fld] = _to_zero_based_row(v)
                                    else:
                                        upd[fld] = _to_zero_based_col(v)
                            if upd:
                                await crud_provider_pricelist_config.update(
                                    db_obj=pl_cfg,
                                    obj_in=upd,
                                    session=session,
                                )
                            await _reset_pricelist_uid_state(pl_cfg.id)
                            configs_set.append({
                                'entity_type': 'pricelist_config',
                                'entity_id': pl_cfg.id,
                                'entity_name': (
                                    pl_cfg.name_price or f'#{pl_cfg.id}'
                                ),
                                'action': 'updated',
                                'note': (
                                    f'Конфигурация прайс-листа '
                                    f'#{pl_cfg.id} обновлена. '
                                    f'Источник: ящик '
                                    f'#{mailbox_account_id or "?"}, '
                                    'указатель UID сброшен для повторной '
                                    'проверки свежих писем.'
                                ),
                            })

                    elif cfg_mode == 'new':
                        oem = getattr(provider_config, 'oem_col', None)
                        qty = getattr(provider_config, 'qty_col', None)
                        prc = getattr(provider_config, 'price_col', None)
                        new_pl = ProviderPriceListConfigCreate(
                            start_row=(
                                _to_zero_based_row(
                                    provider_config.start_row or 1
                                )
                                or 0
                            ),
                            oem_col=_to_zero_based_col(oem),
                            qty_col=_to_zero_based_col(qty),
                            price_col=_to_zero_based_col(prc),
                            brand_col=_to_zero_based_col(
                                getattr(provider_config, 'brand_col', None)
                            ),
                            name_col=_to_zero_based_col(
                                getattr(provider_config, 'name_col', None)
                            ),
                            multiplicity_col=_to_zero_based_col(
                                getattr(
                                    provider_config, 'multiplicity_col', None
                                )
                            ),
                            filename_pattern=(
                                price_filename_pattern or None
                            ),
                            name_mail=provider_config.subject_pattern,
                            name_price=price_cfg_name,
                            incoming_email_account_id=mailbox_account_id,
                        )
                        created_pl = await (
                            crud_provider_pricelist_config.create(
                                session=session,
                                provider_id=provider.id,
                                config_in=new_pl,
                            )
                        )
                        await _reset_pricelist_uid_state(created_pl.id)
                        configs_set.append({
                            'entity_type': 'pricelist_config',
                            'entity_id': created_pl.id,
                            'entity_name': (
                                created_pl.name_price
                                or f'#{created_pl.id}'
                            ),
                            'action': 'created',
                            'note': (
                                f'Создана конфигурация прайс-листа '
                                f'#{created_pl.id} для '
                                f'«{provider.name}». '
                                f'Источник: ящик '
                                f'#{mailbox_account_id or "?"}, '
                                'указатель UID сброшен для старта импорта.'
                            ),
                        })

                # --- order_reply / document: SupplierResponseConfig ---
                elif (
                    rule_type in ('order_reply', 'document')
                    and cfg_mode != 'skip'
                ):
                    payload_type = (
                        'response'
                        if rule_type == 'order_reply'
                        else 'document'
                    )
                    response_type = getattr(
                        provider_config, 'response_type', None
                    )
                    response_type = str(
                        getattr(response_type, 'value', response_type) or ''
                    ).strip().lower()
                    if response_type not in {'file', 'text'}:
                        response_type = None
                    # В мастере тип ответа применяем только для order_reply.
                    if rule_type != 'order_reply':
                        response_type = None

                    def _sr_one_based_col(v):
                        if v is None or v == '':
                            return None
                        parsed = int(v)
                        if parsed < 1:
                            raise ValueError(
                                'Номера колонок и строк должны быть >= 1'
                            )
                        return parsed

                    numeric_col_fields = (
                        'start_row', 'oem_col', 'qty_col', 'price_col',
                        'brand_col', 'name_col', 'status_col', 'comment_col',
                        'document_number_col', 'document_date_col',
                        'gtd_col', 'country_code_col', 'country_name_col',
                        'total_price_with_vat_col',
                    )
                    passthrough_fields = (
                        'fixed_brand_name',
                        'brand_priority_list',
                        'brand_from_name_regex',
                        'document_number_cell',
                        'document_date_cell',
                        'document_meta_cell',
                    )

                    if (
                        cfg_mode == 'existing'
                        and provider_config.config_id
                    ):
                        sr_cfg = await (
                            crud_supplier_response_config.get_by_id(
                                config_id=provider_config.config_id,
                                session=session,
                            )
                        )
                        if sr_cfg:
                            sr_upd: dict = {}
                            if response_type is not None:
                                sr_upd['response_type'] = response_type
                            if response_type == 'text':
                                # Для текстового ответа файл/колонки не нужны.
                                sr_upd['file_format'] = None
                                sr_upd['file_payload_type'] = 'response'
                                sr_upd['filename_pattern'] = None
                                sr_upd['start_row'] = 1
                                for fld in (
                                    'oem_col', 'qty_col', 'price_col',
                                    'brand_col', 'name_col', 'status_col',
                                    'comment_col',
                                    'document_number_col',
                                    'document_date_col', 'gtd_col',
                                    'country_code_col', 'country_name_col',
                                    'total_price_with_vat_col',
                                ):
                                    sr_upd[fld] = None
                                for fld in passthrough_fields:
                                    if fld == 'brand_priority_list':
                                        sr_upd[fld] = []
                                    else:
                                        sr_upd[fld] = None
                                if (
                                    getattr(
                                        provider_config,
                                        'confirm_keywords',
                                        None
                                    ) is not None
                                ):
                                    sr_upd['confirm_keywords'] = list(
                                        provider_config.confirm_keywords or []
                                    )
                                if (
                                    getattr(
                                        provider_config,
                                        'reject_keywords',
                                        None
                                    ) is not None
                                ):
                                    sr_upd['reject_keywords'] = list(
                                        provider_config.reject_keywords or []
                                    )
                                value_after_article_type = getattr(
                                    provider_config,
                                    'value_after_article_type',
                                    None,
                                )
                                if value_after_article_type in {
                                    'number',
                                    'text',
                                    'both',
                                }:
                                    sr_upd['value_after_article_type'] = (
                                        value_after_article_type
                                    )
                            else:
                                sr_upd['file_payload_type'] = payload_type
                                if provider_config.filename_pattern:
                                    sr_upd['filename_pattern'] = (
                                        provider_config.filename_pattern
                                    )
                                for fld in numeric_col_fields:
                                    v = getattr(provider_config, fld, None)
                                    if v is not None:
                                        sr_upd[fld] = _sr_one_based_col(v)
                                for fld in passthrough_fields:
                                    v = getattr(provider_config, fld, None)
                                    if v is None:
                                        continue
                                    if fld == 'brand_priority_list':
                                        sr_upd[fld] = list(v or [])
                                    else:
                                        sr_upd[fld] = v
                            # Добавляем email отправителя
                            sender_emails = list(
                                getattr(sr_cfg, 'sender_emails', None)
                                or []
                            )
                            if (
                                inbox_email.from_email
                                and inbox_email.from_email
                                not in sender_emails
                            ):
                                sender_emails.append(inbox_email.from_email)
                                sr_upd['sender_emails'] = sender_emails
                            if (
                                inbox_email.email_account_id
                                and sr_cfg.inbox_email_account_id
                                != inbox_email.email_account_id
                            ):
                                sr_upd['inbox_email_account_id'] = (
                                    inbox_email.email_account_id
                                )
                            if sr_upd:
                                await crud_supplier_response_config.update(
                                    db_obj=sr_cfg,
                                    obj_in=sr_upd,
                                    session=session,
                                )
                            configs_set.append({
                                'entity_type': 'response_config',
                                'entity_id': sr_cfg.id,
                                'entity_name': sr_cfg.name or f'#{sr_cfg.id}',
                                'action': 'updated',
                                'note': (
                                    f'Конфигурация «{sr_cfg.name}» '
                                    f'обновлена.'
                                ),
                            })
                            supplier_registry_config_ids.append(int(sr_cfg.id))

                    elif cfg_mode == 'new':
                        cfg_name = (
                            getattr(provider_config, 'config_name', None)
                            or f'{provider.name} — {rule_type}'
                        )
                        create_payload: dict = {
                            'name': cfg_name,
                            'file_payload_type': payload_type,
                            'sender_emails': [inbox_email.from_email],
                            'inbox_email_account_id': (
                                inbox_email.email_account_id or None
                            ),
                        }
                        create_response_type = (
                            response_type
                            if response_type in {'file', 'text'}
                            else 'file'
                        )
                        create_payload['response_type'] = create_response_type
                        if create_response_type == 'text':
                            create_payload['file_format'] = None
                            create_payload['filename_pattern'] = None
                            confirm_keywords = getattr(
                                provider_config,
                                'confirm_keywords',
                                None,
                            )
                            reject_keywords = getattr(
                                provider_config,
                                'reject_keywords',
                                None,
                            )
                            value_after_article_type = getattr(
                                provider_config,
                                'value_after_article_type',
                                None,
                            )
                            if confirm_keywords is not None:
                                create_payload['confirm_keywords'] = list(
                                    confirm_keywords or []
                                )
                            if reject_keywords is not None:
                                create_payload['reject_keywords'] = list(
                                    reject_keywords or []
                                )
                            if value_after_article_type in {
                                'number',
                                'text',
                                'both',
                            }:
                                create_payload['value_after_article_type'] = (
                                    value_after_article_type
                                )
                        else:
                            create_payload['filename_pattern'] = (
                                provider_config.filename_pattern or None
                            )
                            create_payload.update(
                                {
                                    fld: _sr_one_based_col(
                                        getattr(provider_config, fld, None)
                                    )
                                    for fld in numeric_col_fields
                                    if getattr(provider_config, fld, None)
                                    is not None
                                }
                            )
                            for fld in passthrough_fields:
                                v = getattr(provider_config, fld, None)
                                if v is None:
                                    continue
                                if fld == 'brand_priority_list':
                                    create_payload[fld] = list(v or [])
                                else:
                                    create_payload[fld] = v
                        create_data = SupplierResponseConfigCreate(
                            **create_payload
                        )
                        created_sr = await (
                            crud_supplier_response_config.create(
                                provider_id=provider.id,
                                config_in=create_data,
                                session=session,
                            )
                        )
                        configs_set.append({
                            'entity_type': 'response_config',
                            'entity_id': created_sr.id,
                            'entity_name': created_sr.name,
                            'action': 'created',
                            'note': (
                                f'Создана конфигурация «{created_sr.name}» '
                                f'для «{provider.name}».'
                            ),
                        })
                        supplier_registry_config_ids.append(int(created_sr.id))

        except ValueError:
            raise
        except Exception as e:
            logger.exception(
                'Ошибка привязки поставщика id=%s: %s',
                provider_config.provider_id,
                e,
            )
            if rule_type == 'price_list' and cfg_mode != 'skip':
                raise ValueError(
                    'Не удалось сохранить конфигурацию прайс-листа. '
                    'Проверьте поля и повторите.'
                )

    # ------------------------------------------------------------------
    # 1b. Привязка к клиенту — обновляем order_email в CustomerOrderConfig
    # ------------------------------------------------------------------
    elif (
        rule_type == 'customer_order'
        and customer_config
        and customer_config.customer_id
    ):
        target_customer_config_id: Optional[int] = None
        target_customer_id = int(customer_config.customer_id)
        try:
            from dz_fastapi.crud.customer_order import \
                crud_customer_order_config

            customer = await crud_customer.get_by_id(
                customer_config.customer_id, session
            )
            customer_name = (
                customer.name if customer else str(customer_config.customer_id)
            )

            selected_config_id = getattr(customer_config, 'config_id', None)
            config_mode = getattr(customer_config, 'config_mode', 'existing')
            order_config = getattr(customer_config, 'order_config', None) or {}

            def _to_zero_based_column(value):
                if value is None or value == '':
                    return None
                parsed = int(value)
                if parsed < 1:
                    raise ValueError('Номера колонок должны быть >= 1')
                return parsed - 1

            def _to_optional_int(value):
                if value is None or value == '':
                    return None
                return int(value)

            if config_mode == 'new':
                if not order_config:
                    action = 'no_config'
                    note = (
                        'Выбран режим создания новой конфигурации, но не '
                        'переданы настройки файла заказа.'
                    )
                else:
                    if not order_config.get('pricelist_config_id'):
                        raise ValueError(
                            'Для новой конфигурации укажите прайс клиента'
                        )
                    if not order_config.get('oem_col'):
                        raise ValueError(
                            'Для новой конфигурации укажите колонку OEM'
                        )
                    if not order_config.get('brand_col'):
                        raise ValueError(
                            'Для новой конфигурации укажите колонку Бренд'
                        )
                    if not order_config.get('qty_col'):
                        raise ValueError(
                            'Для новой конфигурации укажите колонку Кол-во'
                        )

                    create_data = {
                        'order_email': inbox_email.from_email,
                        'order_subject_pattern': (
                            customer_config.subject_pattern or None
                        ),
                        'order_filename_pattern': (
                            customer_config.filename_pattern or None
                        ),
                        'email_account_id': inbox_email.email_account_id,
                        'pricelist_config_id': int(
                            order_config['pricelist_config_id']
                        ),
                        'order_start_row': int(
                            order_config.get('order_start_row') or 1
                        ),
                        'order_number_row': _to_optional_int(
                            order_config.get('order_number_row')
                        ),
                        'order_date_row': _to_optional_int(
                            order_config.get('order_date_row')
                        ),
                        'order_number_source': order_config.get(
                            'order_number_source'
                        ),
                        'order_number_regex_subject': order_config.get(
                            'order_number_regex_subject'
                        ),
                        'order_number_regex_filename': order_config.get(
                            'order_number_regex_filename'
                        ),
                        'order_number_regex_body': order_config.get(
                            'order_number_regex_body'
                        ),
                        'order_number_prefix': order_config.get(
                            'order_number_prefix'
                        ),
                        'order_number_suffix': order_config.get(
                            'order_number_suffix'
                        ),
                        'oem_col': _to_zero_based_column(
                            order_config.get('oem_col')
                        ),
                        'brand_col': _to_zero_based_column(
                            order_config.get('brand_col')
                        ),
                        'name_col': _to_zero_based_column(
                            order_config.get('name_col')
                        ),
                        'qty_col': _to_zero_based_column(
                            order_config.get('qty_col')
                        ),
                        'price_col': _to_zero_based_column(
                            order_config.get('price_col')
                        ),
                        'ship_qty_col': _to_zero_based_column(
                            order_config.get('ship_qty_col')
                        ),
                        'ship_price_col': _to_zero_based_column(
                            order_config.get('ship_price_col')
                        ),
                        'reject_qty_col': _to_zero_based_column(
                            order_config.get('reject_qty_col')
                        ),
                        'order_number_column': _to_zero_based_column(
                            order_config.get('order_number_column')
                        ),
                        'order_date_column': _to_zero_based_column(
                            order_config.get('order_date_column')
                        ),
                    }
                    if order_config.get('ship_mode'):
                        create_data['ship_mode'] = order_config['ship_mode']
                    if order_config.get('price_tolerance_pct') is not None:
                        create_data['price_tolerance_pct'] = float(
                            order_config['price_tolerance_pct']
                        )
                    if order_config.get('price_warning_pct') is not None:
                        create_data['price_warning_pct'] = float(
                            order_config['price_warning_pct']
                        )
                    if order_config.get('is_active') is not None:
                        create_data['is_active'] = bool(
                            order_config['is_active']
                        )

                    created_cfg = await crud_customer_order_config.create(
                        session=session,
                        customer_id=target_customer_id,
                        data=create_data,
                    )
                    target_customer_config_id = int(created_cfg.id)
                    action = 'created'
                    note = (
                        f'Создана новая конфигурация заказов клиента '
                        f'«{customer_name}» (ID {created_cfg.id}).'
                    )
            else:
                existing_cfg = None
                if selected_config_id:
                    existing_cfg = await crud_customer_order_config.get_by_id(
                        session, int(selected_config_id)
                    )
                if (
                    existing_cfg is None
                    and (selected_config_id is None)
                ):
                    existing_cfg = (
                        await crud_customer_order_config.get_by_customer_id(
                            session, target_customer_id
                        )
                    )

                if (
                    existing_cfg
                    and existing_cfg.customer_id == target_customer_id
                ):
                    target_customer_config_id = int(existing_cfg.id)
                    update_data: dict = {}
                    email_being_set = not existing_cfg.order_email
                    if email_being_set:
                        update_data['order_email'] = inbox_email.from_email
                    if (
                        customer_config.subject_pattern
                        and not existing_cfg.order_subject_pattern
                    ):
                        update_data['order_subject_pattern'] = (
                            customer_config.subject_pattern
                        )
                    if (
                        customer_config.filename_pattern
                        and not existing_cfg.order_filename_pattern
                    ):
                        update_data['order_filename_pattern'] = (
                            customer_config.filename_pattern
                        )
                    if (
                        inbox_email.email_account_id
                        and not existing_cfg.email_account_id
                    ):
                        update_data['email_account_id'] = (
                            inbox_email.email_account_id
                        )
                    if email_being_set or update_data.get('email_account_id'):
                        update_data['last_uid'] = 0
                        update_data['folder_last_uids'] = {}

                    # Применяем изменения столбцов если менеджер
                    # скорректировал их в мастере
                    if order_config:
                        col_fields = [
                            'order_start_row', 'oem_col', 'brand_col',
                            'qty_col', 'name_col', 'price_col',
                            'ship_qty_col', 'reject_qty_col',
                        ]
                        for field in col_fields:
                            if order_config.get(field) is not None:
                                if field == 'order_start_row':
                                    update_data[field] = int(
                                        order_config[field]
                                    )
                                else:
                                    update_data[field] = (
                                        _to_zero_based_column(
                                            order_config[field]
                                        )
                                    )

                    if update_data:
                        await crud_customer_order_config.update(
                            session,
                            existing_cfg,
                            update_data
                        )
                        action = 'updated'
                        note = (
                            f'Обновлена конфигурация заказов клиента '
                            f'«{customer_name}» (ID {existing_cfg.id}).'
                        )
                    else:
                        action = 'already_linked'
                        note = (
                            f'Конфигурация клиента «{customer_name}» '
                            f'(ID {existing_cfg.id}) уже актуальна.'
                        )
                else:
                    action = 'no_config'
                    if selected_config_id:
                        note = (
                            f'Конфигурация ID {selected_config_id} не найдена '
                            'или не принадлежит выбранному клиенту.'
                        )
                    else:
                        note = (
                            f'У клиента «{customer_name}» нет конфигурации '
                            'заказов. Выберите "Создать новую".'
                        )

            config_entry = {
                'entity_type': 'customer',
                'entity_id': customer_config.customer_id,
                'entity_name': customer_name,
                'action': action,
                'note': note,
            }
            if target_customer_config_id is not None:
                config_entry['entity_id'] = target_customer_config_id
            configs_set.append(config_entry)

            if target_customer_config_id is not None:
                configs_set.append({
                    'entity_type': 'customer',
                    'entity_id': target_customer_config_id,
                    'entity_name': customer_name,
                    'action': 'queued',
                    'note': (
                        'Конфигурация сохранена. Обработка письма запустится '
                        'чуть позже по расписанию шедулера.'
                    ),
                })
        except Exception as e:
            logger.warning(
                'Ошибка привязки клиента id=%s: %s',
                customer_config.customer_id, e
            )

    # ------------------------------------------------------------------
    # 2. Назначаем правило + паттерн + запускаем обработку
    # ------------------------------------------------------------------
    process_now = rule_type not in {
        'customer_order',
        'order_reply',
        'document'
    }
    if rule_type in {'order_reply', 'document'}:
        try:
            created_registry_rows = await (
                _create_supplier_response_registry_stub(
                    session,
                    inbox_email=inbox_email,
                    rule_type=rule_type,
                    provider_id=supplier_registry_provider_id,
                    response_config_ids=supplier_registry_config_ids,
                    reason_text=(
                        'Назначено через Inbox. '
                        'Обработка письма запустится по расписанию.'
                    ),
                )
            )
            if created_registry_rows > 0:
                configs_set.append({
                    'entity_type': 'response_config',
                    'entity_id': (
                        supplier_registry_config_ids[0]
                        if supplier_registry_config_ids
                        else supplier_registry_provider_id
                    ),
                    'entity_name': (
                        'Supplier response registry'
                    ),
                    'action': 'queued',
                    'note': (
                        'Письмо сразу добавлено в реестр обработанных писем '
                        f'({created_registry_rows} шт.) с причиной ожидания.'
                    ),
                })
        except Exception as exc:
            logger.warning(
                'Не удалось создать служебную запись реестра '
                'для inbox email id=%s: %s',
                email_id,
                exc,
            )

    updated = await assign_rule(
        session,
        email_id=email_id,
        rule_type=rule_type,
        user_id=user_id,
        save_pattern=save_pattern,
        process_now=process_now,
        queued_note='Настройки приняты. '
                    'Обработка письма запустится чуть позже.',
    )

    return {
        'email_id': updated.id,
        'rule_type': updated.rule_type,
        'processed': updated.processed,
        'processing_result': updated.processing_result,
        'processing_error': updated.processing_error,
        'configs_set': configs_set,
    }


# ---------------------------------------------------------------------------
# Очистка старых писем (вызывается из шедулера)
# ---------------------------------------------------------------------------

async def cleanup_inbox_emails(
    session: AsyncSession, max_days: int = 7
) -> int:
    """
    Удаляет письма старше max_days дней.
    Перед удалением из БД удаляет файлы вложений с диска.
    Файлы хранятся не дольше max_days дней (по умолчанию 7).
    """
    from datetime import timedelta

    from sqlalchemy.future import select as sa_select

    cutoff = now_moscow() - timedelta(days=max_days)

    # Получаем пути файлов перед удалением из БД
    result = await session.execute(
        sa_select(InboxEmail.attachment_info).where(
            InboxEmail.fetched_at < cutoff
        )
    )
    files_deleted = 0
    for (att_info,) in result:
        for att in (att_info or []):
            path = att.get('path')
            fs_path = resolve_inbox_attachment_fs_path(path)
            if fs_path and os.path.exists(fs_path):
                try:
                    os.remove(fs_path)
                    files_deleted += 1
                    # Удаляем пустые родительские папки
                    parent = os.path.dirname(fs_path)
                    if os.path.isdir(parent) and not os.listdir(parent):
                        os.rmdir(parent)
                except Exception as e:
                    logger.warning(
                        'Не удалось удалить файл %s: %s', fs_path, e
                    )

    deleted = await cleanup_old_inbox_emails(session, max_days=max_days)
    orphan_files_deleted, orphan_dirs_deleted = await (
        _cleanup_orphan_inbox_attachment_files(
            session=session,
            cutoff=cutoff,
        )
    )
    logger.info(
        'Удалено %d устаревших писем из InboxEmail, '
        '%d файлов вложений с диска. '
        'Дополнительно удалено сиротских файлов: %d, папок: %d',
        deleted, files_deleted, orphan_files_deleted, orphan_dirs_deleted,
    )
    return deleted


async def _cleanup_orphan_inbox_attachment_files(
    *,
    session: AsyncSession,
    cutoff: datetime,
) -> tuple[int, int]:
    """
    Удаляет сиротские файлы в uploads/inbox_attachments:
      - отсутствуют в attachment_info любой записи InboxEmail
      - и старше cutoff (по mtime)
    """
    from sqlalchemy.future import select as sa_select

    referenced_paths: set[str] = set()
    result = await session.execute(
        sa_select(InboxEmail.attachment_info).where(
            InboxEmail.has_attachments.is_(True)
        )
    )
    for (att_info,) in result:
        for att in (att_info or []):
            path = att.get('path')
            fs_path = resolve_inbox_attachment_fs_path(path)
            if fs_path:
                referenced_paths.add(os.path.realpath(fs_path))

    root_dir = resolve_inbox_attachment_fs_path(
        os.path.join('uploads', 'inbox_attachments')
    )
    cutoff_ts = cutoff.timestamp()
    return await asyncio.to_thread(
        _cleanup_orphan_inbox_attachment_files_sync,
        root_dir=root_dir,
        referenced_paths=referenced_paths,
        cutoff_ts=cutoff_ts,
    )


def _cleanup_orphan_inbox_attachment_files_sync(
    *,
    root_dir: str,
    referenced_paths: set[str],
    cutoff_ts: float,
) -> tuple[int, int]:
    if not root_dir or not os.path.isdir(root_dir):
        return 0, 0

    removed_files = 0
    removed_dirs = 0
    for dirpath, _dirnames, filenames in os.walk(root_dir, topdown=False):
        for filename in filenames:
            file_path = os.path.realpath(os.path.join(dirpath, filename))
            if file_path in referenced_paths:
                continue
            try:
                mtime = os.path.getmtime(file_path)
            except OSError:
                continue
            if mtime > cutoff_ts:
                continue
            try:
                os.remove(file_path)
                removed_files += 1
            except Exception as exc:
                logger.warning(
                    'Не удалось удалить сиротский файл %s: %s',
                    file_path,
                    exc,
                )

        if dirpath == root_dir:
            continue
        try:
            if os.path.isdir(dirpath) and not os.listdir(dirpath):
                os.rmdir(dirpath)
                removed_dirs += 1
        except Exception:
            continue

    return removed_files, removed_dirs
