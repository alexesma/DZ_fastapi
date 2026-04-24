import asyncio
import base64
import hashlib
import logging
import os
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from email import message_from_bytes
from email.header import decode_header
from html import escape
from io import BytesIO
from typing import Dict, List, Optional, Tuple

import aiofiles
import httpx
import pandas as pd

try:
    from imap_tools import AND, MailBox, MailBoxSsl
    from imap_tools.errors import MailboxFolderSelectError
except ImportError:  # pragma: no cover - fallback for older imap_tools
    from imap_tools import AND, MailBox
    MailBoxSsl = None
    try:
        from imap_tools.errors import MailboxFolderSelectError
    except ImportError:
        MailboxFolderSelectError = Exception
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload, selectinload
from sqlalchemy.sql import func

from dz_fastapi.api.validators import normalize_brand_name
from dz_fastapi.core.constants import IMAP_SERVER
from dz_fastapi.core.email_folders import (DEFAULT_IMAP_FOLDER,
                                           normalize_imap_folder,
                                           resolve_imap_folders)
from dz_fastapi.core.time import now_moscow
from dz_fastapi.crud.brand import brand_crud
from dz_fastapi.crud.customer_order import (crud_customer_order,
                                            crud_customer_order_config)
from dz_fastapi.crud.email_account import crud_email_account
from dz_fastapi.crud.partner import (crud_customer_pricelist,
                                     crud_customer_pricelist_config,
                                     crud_customer_pricelist_source,
                                     crud_pricelist)
from dz_fastapi.crud.settings import crud_customer_order_inbox_settings
from dz_fastapi.models.autopart import AutoPart, preprocess_oem_number
from dz_fastapi.models.brand import Brand
from dz_fastapi.models.notification import AppNotificationLevel
from dz_fastapi.models.partner import (CUSTOMER_ORDER_ITEM_STATUS,
                                       CUSTOMER_ORDER_SHIP_MODE,
                                       CUSTOMER_ORDER_STATUS,
                                       ORDER_TRACKING_SOURCE,
                                       STOCK_ORDER_STATUS,
                                       SUPPLIER_ORDER_STATUS, Customer,
                                       CustomerOrder, CustomerOrderConfig,
                                       CustomerOrderItem, CustomerPriceList,
                                       CustomerPriceListAutoPartAssociation,
                                       CustomerPriceListConfig, PriceList,
                                       PriceListAutoPartAssociation, Provider,
                                       StockOrder, StockOrderItem,
                                       SupplierOrder, SupplierOrderItem)
from dz_fastapi.services.email import (build_email_delivery_kwargs,
                                       send_email_with_attachment)
from dz_fastapi.services.google_oauth import refresh_google_access_token
from dz_fastapi.services.notifications import create_admin_notifications
from dz_fastapi.services.process import (_apply_source_filters,
                                         _apply_source_markups)
from dz_fastapi.services.resend_api import fetch_received_emails_for_address

logger = logging.getLogger('dz_fastapi')

EMAIL_NAME_ORDER = os.getenv('EMAIL_NAME_ORDERS')
EMAIL_PASSWORD_ORDER = os.getenv('EMAIL_PASSWORD_ORDERS')
EMAIL_HOST_ORDER = os.getenv('EMAIL_HOST_ORDERS')
EMAIL_FOLDER_ORDER = os.getenv('EMAIL_FOLDER_ORDERS', 'INBOX')

ORDERS_UPLOAD_DIR = os.getenv('CUSTOMER_ORDERS_UPLOAD_DIR', 'uploads/orders')
ORDERS_RESPONSE_DIR = os.getenv(
    'CUSTOMER_ORDERS_RESPONSE_DIR', 'uploads/orders/responses'
)
ORDERS_REPORT_DIR = os.getenv(
    'CUSTOMER_ORDERS_REPORT_DIR', 'uploads/orders/reports'
)
ORDERS_ERROR_DIR = os.getenv(
    'CUSTOMER_ORDERS_ERROR_DIR', 'uploads/orders/errors'
)
ORDERS_RETENTION_DAYS = int(os.getenv('CUSTOMER_ORDERS_REPORT_DAYS', 7))
ORDER_ERROR_DETAIL_MAX_LEN = 500
CUSTOMER_ORDERS_FETCH_LIMIT = int(
    os.getenv('CUSTOMER_ORDERS_FETCH_LIMIT') or '0'
)
CUSTOMER_ORDERS_IMAP_TIMEOUT_SEC = max(
    5,
    int(os.getenv('CUSTOMER_ORDERS_IMAP_TIMEOUT_SEC', '30')),
)
CUSTOMER_ORDERS_IMAP_RETRIES = max(
    1,
    int(os.getenv('CUSTOMER_ORDERS_IMAP_RETRIES', '3')),
)
CUSTOMER_ORDERS_IMAP_RETRY_DELAY_SEC = max(
    1,
    int(os.getenv('CUSTOMER_ORDERS_IMAP_RETRY_DELAY_SEC', '5')),
)


def _customer_order_auto_reply_enabled() -> bool:
    if _customer_order_reply_override_email():
        return True
    return str(
        os.getenv('CUSTOMER_ORDER_AUTO_REPLY_ENABLED', '0')
    ).strip().lower() in {'1', 'true', 'yes', 'on'}


def _customer_order_reply_override_email() -> Optional[str]:
    value = str(
        os.getenv(
            'CUSTOMER_ORDER_REPLY_OVERRIDE_EMAIL',
            'info@dragonzap.ru',
        )
    ).strip()
    return value or None


def _supplier_order_override_email() -> Optional[str]:
    value = str(
        os.getenv(
            'SUPPLIER_ORDER_OVERRIDE_EMAIL',
            'info@dragonzap.ru',
        )
    ).strip()
    return value or None


async def _supplier_order_override_email_from_settings(
    session: AsyncSession,
) -> Optional[str]:
    try:
        inbox_settings = (
            await crud_customer_order_inbox_settings.get_or_create(
                session
            )
        )
    except Exception:
        return _supplier_order_override_email()
    if not bool(getattr(inbox_settings, 'supplier_order_stub_enabled', True)):
        return None
    configured_email = str(
        getattr(inbox_settings, 'supplier_order_stub_email', '') or ''
    ).strip()
    if configured_email:
        return configured_email
    return _supplier_order_override_email()


async def _notify_admins(
    session: AsyncSession,
    *,
    title: str,
    message: str,
    level: str = AppNotificationLevel.INFO,
    link: str | None = None,
    commit: bool = False,
) -> None:
    try:
        await create_admin_notifications(
            session=session,
            title=title,
            message=message,
            level=level,
            link=link,
            commit=commit,
        )
    except Exception as exc:
        logger.error(
            'Failed to create admin notification: %s',
            exc,
            exc_info=True,
        )


def _create_mailbox(
        server_mail: str,
        port: int,
        ssl: bool = True,
        timeout: int = 30
):
    if ssl and MailBoxSsl is not None:
        return MailBoxSsl(server_mail, port, timeout=timeout)
    return MailBox(server_mail, port, timeout=timeout)


@dataclass
class ParsedOrderRow:
    row_index: int
    oem: str
    brand: str
    name: Optional[str]
    requested_qty: int
    requested_price: Optional[float]


@dataclass
class OfferRow:
    autopart_id: int
    provider_id: int
    provider_config_id: Optional[int]
    quantity: int
    price: float
    supplier_price: float
    is_own_price: bool


@dataclass
class SimpleAttachment:
    filename: Optional[str]
    payload: bytes


@dataclass
class SimpleMessage:
    uid: Optional[str]
    from_: str
    subject: str
    attachments: List[SimpleAttachment]
    text: Optional[str]
    html: Optional[str]
    external_id: Optional[str] = None
    received_at: Optional[datetime] = None
    folder_name: Optional[str] = None


async def _fetch_order_messages(
    server_mail: str,
    email_account: str,
    email_password: str,
    folder: str,
    date_from: date,
    mark_seen: bool,
    last_uid: Optional[int] = None,
    port: int = 993,
    ssl: bool = True,
    from_email: Optional[str] = None,
) -> list:
    def _fetch():
        with _create_mailbox(
                server_mail,
                port,
                ssl,
                timeout=CUSTOMER_ORDERS_IMAP_TIMEOUT_SEC
        ).login(
            email_account, email_password
        ) as mailbox:
            mailbox.folder.set(folder)
            fetched = None
            if last_uid and int(last_uid) > 0:
                uid_criteria = f'UID {int(last_uid) + 1}:*'
                try:
                    fetched = mailbox.fetch(
                        uid_criteria,
                        mark_seen=mark_seen,
                    )
                except Exception as uid_exc:
                    logger.warning(
                        'UID fetch failed for order inbox %s '
                        '(folder=%s, uid>%s). Fallback to date search: %s',
                        email_account,
                        folder,
                        last_uid,
                        uid_exc,
                    )
            if fetched is None:
                query_kwargs = {
                    'date_gte': date_from,
                    'all': True,
                }
                normalized_from = str(from_email or '').strip()
                if normalized_from:
                    query_kwargs['from_'] = normalized_from
                fetched = mailbox.fetch(
                    AND(**query_kwargs),
                    mark_seen=mark_seen,
                    charset='utf-8',
                )
            messages = list(fetched)
            folder_name = normalize_imap_folder(folder)
            for message in messages:
                setattr(message, 'folder_name', folder_name)
            return messages

    for attempt in range(1, CUSTOMER_ORDERS_IMAP_RETRIES + 1):
        try:
            return await asyncio.to_thread(_fetch)
        except Exception as exc:
            if (
                not _is_too_many_connections_error(exc)
                or attempt >= CUSTOMER_ORDERS_IMAP_RETRIES
            ):
                raise
            delay = CUSTOMER_ORDERS_IMAP_RETRY_DELAY_SEC * attempt
            logger.warning(
                'IMAP connection limit for %s. '
                'Retry %s/%s in %ss.',
                email_account,
                attempt,
                CUSTOMER_ORDERS_IMAP_RETRIES,
                delay,
            )
            await asyncio.sleep(delay)
    return []


def _decode_header_value(value: Optional[str]) -> str:
    if not value:
        return ''
    parts = decode_header(value)
    decoded = ''
    for part, encoding in parts:
        if isinstance(part, bytes):
            try:
                decoded += part.decode(encoding or 'utf-8', errors='ignore')
            except Exception:
                decoded += part.decode('utf-8', errors='ignore')
        else:
            decoded += str(part)
    return decoded


def _parse_raw_email(raw_bytes: bytes) -> SimpleMessage:
    msg = message_from_bytes(raw_bytes)
    subject = _decode_header_value(msg.get('Subject'))
    from_value = _decode_header_value(msg.get('From'))
    text = None
    html = None
    attachments: List[SimpleAttachment] = []
    for part in msg.walk():
        content_disposition = part.get_content_disposition()
        filename = part.get_filename()
        content_type = part.get_content_type()
        if content_disposition == 'attachment' or filename:
            payload = part.get_payload(decode=True) or b''
            attachments.append(SimpleAttachment(filename, payload))
            continue
        if content_type == 'text/plain' and text is None:
            payload = part.get_payload(decode=True) or b''
            charset = part.get_content_charset() or 'utf-8'
            text = payload.decode(charset, errors='ignore')
        if content_type == 'text/html' and html is None:
            payload = part.get_payload(decode=True) or b''
            charset = part.get_content_charset() or 'utf-8'
            html = payload.decode(charset, errors='ignore')
    return SimpleMessage(
        uid=None,
        from_=from_value,
        subject=subject,
        attachments=attachments,
        text=text,
        html=html,
        external_id=None,
        received_at=None,
    )


async def _fetch_gmail_messages(
    account,
    date_from: date,
    label: Optional[str] = None,
) -> List[SimpleMessage]:
    if not account.oauth_refresh_token:
        return []
    token_data = await refresh_google_access_token(
        account.oauth_refresh_token
    )
    access_token = token_data.get('access_token')
    if not access_token:
        raise RuntimeError('Google OAuth access token not returned')
    headers = {'Authorization': f'Bearer {access_token}'}
    query = f'after:{date_from.strftime("%Y/%m/%d")}'
    if label and label.upper() != 'INBOX':
        query = f'{query} label:{label}'
    url = 'https://gmail.googleapis.com/gmail/v1/users/me/messages'
    messages: List[SimpleMessage] = []
    page_token = None
    async with httpx.AsyncClient(timeout=20) as client:
        while True:
            max_results = 200
            params = {
                'q': query,
                'maxResults': max_results,
            }
            if page_token:
                params['pageToken'] = page_token
            response = await client.get(url, headers=headers, params=params)
            response.raise_for_status()
            payload = response.json()
            for item in payload.get('messages', []):
                message_id = item.get('id')
                if not message_id:
                    continue
                raw_response = await client.get(
                    f'{url}/{message_id}',
                    headers=headers,
                    params={'format': 'raw'},
                )
                raw_response.raise_for_status()
                raw_payload = raw_response.json()
                raw_data = raw_payload.get('raw')
                if not raw_data:
                    continue
                raw_bytes = base64.urlsafe_b64decode(raw_data + '==')
                messages.append(_parse_raw_email(raw_bytes))
                messages[-1].external_id = str(message_id)
                messages[-1].folder_name = normalize_imap_folder(label)
            page_token = payload.get('nextPageToken')
            if not page_token:
                break
    return messages


def _safe_uid_as_int(value: Optional[str]) -> Optional[int]:
    if value is None:
        return None
    text = str(value).strip()
    if not text.isdigit():
        return None
    try:
        return int(text)
    except ValueError:
        return None


async def _fetch_resend_messages(
    account,
    date_from: date,
) -> List[SimpleMessage]:
    if not account.resend_api_key:
        return []
    emails = await fetch_received_emails_for_address(
        api_key=account.resend_api_key,
        email=account.email,
        date_from=date_from,
        received_after=getattr(account, 'resend_last_received_at', None),
        timeout=getattr(account, 'resend_timeout', None),
    )
    messages: List[SimpleMessage] = []
    for item in emails:
        messages.append(
            SimpleMessage(
                uid=None,
                from_=_extract_email(item.get('from')),
                subject=str(item.get('subject') or ''),
                attachments=[
                    SimpleAttachment(
                        att.get('filename'),
                        att.get('payload') or b'',
                    )
                    for att in (item.get('attachments') or [])
                ],
                text=item.get('text'),
                html=item.get('html'),
                external_id=str(item.get('id') or ''),
                received_at=item.get('created_at'),
            )
        )
    return messages


def _match_pattern(pattern: Optional[str], value: Optional[str]) -> bool:
    if not pattern:
        return True
    if not value:
        return False
    try:
        return re.search(pattern, value, flags=re.IGNORECASE) is not None
    except re.error:
        return pattern.lower() in value.lower()


def _extract_email(value: Optional[str]) -> str:
    if not value:
        return ''
    match = re.search(r'[\\w\\.-]+@[\\w\\.-]+\\.[\\w]+', value)
    return match.group(0).lower() if match else value.lower()


def _is_too_many_connections_error(exc: Exception) -> bool:
    text = str(exc).lower()
    return (
        'too many simultaneous connections' in text
        or 'too many connections' in text
    )


async def _get_out_account(
    session: AsyncSession, purpose: str
):
    accounts = await crud_email_account.get_active_by_purpose(
        session, purpose
    )
    return accounts[0] if accounts else None


def _normalize_email_list(values: Optional[List[str]]) -> List[str]:
    if not values:
        return []
    return [str(v).strip().lower() for v in values if str(v).strip()]


def _normalize_email_account_ids(values) -> list[int]:
    if not values:
        return []
    result: list[int] = []
    seen: set[int] = set()
    for value in values:
        try:
            parsed = int(value)
        except (TypeError, ValueError):
            continue
        if parsed < 1 or parsed in seen:
            continue
        seen.add(parsed)
        result.append(parsed)
    return result


def _get_config_email_account_ids(config) -> list[int]:
    scoped_ids = _normalize_email_account_ids(
        getattr(config, 'email_account_ids', None)
    )
    if scoped_ids:
        return scoped_ids
    account_id = getattr(config, 'email_account_id', None)
    if account_id is None:
        return []
    try:
        parsed = int(account_id)
    except (TypeError, ValueError):
        return []
    return [parsed] if parsed > 0 else []


def _config_has_email_account_scope(config) -> bool:
    return bool(_get_config_email_account_ids(config))


def _config_uses_multiple_email_accounts(config) -> bool:
    return len(_get_config_email_account_ids(config)) > 1


def _pick_configs_for_account(configs, account_id: Optional[int]):
    if not configs:
        return []
    if account_id is not None:
        matched = [
            cfg
            for cfg in configs
            if account_id in _get_config_email_account_ids(cfg)
        ]
        if matched:
            return matched
        return [
            cfg for cfg in configs if not _config_has_email_account_scope(cfg)
        ]
    return [cfg for cfg in configs if not _config_has_email_account_scope(cfg)]


def _filter_messages_by_senders(
    messages: List[object],
    allowed_senders: set[str],
) -> List[object]:
    if not allowed_senders:
        return list(messages)
    return [
        msg
        for msg in messages
        if _extract_email(getattr(msg, 'from_', None)) in allowed_senders
    ]


def _message_sort_key(item: tuple[object, Optional[object]]) -> tuple:
    msg, inbox_account = item
    account_id = inbox_account.id if inbox_account else 0
    uid = _safe_uid_as_int(getattr(msg, 'uid', None))
    message_dt = (
        getattr(msg, 'received_at', None)
        or getattr(msg, 'date', None)
        or datetime.min
    )
    if hasattr(message_dt, 'isoformat'):
        message_dt = message_dt.isoformat()
    else:
        message_dt = str(message_dt)
    external_id = str(getattr(msg, 'external_id', '') or '')
    return (
        account_id,
        0 if uid is not None else 1,
        uid if uid is not None else 0,
        message_dt,
        external_id,
    )


def _message_identity_key(item: tuple[object, Optional[object]]) -> tuple:
    msg, inbox_account = item
    account_id = inbox_account.id if inbox_account else 0
    external_id = str(getattr(msg, 'external_id', '') or '').strip()
    if external_id:
        return ('external', account_id, external_id)
    message_dt = (
        getattr(msg, 'received_at', None)
        or getattr(msg, 'date', None)
        or ''
    )
    if hasattr(message_dt, 'isoformat'):
        message_dt = message_dt.isoformat()
    attachments = tuple(
        sorted(
            str(getattr(att, 'filename', '') or '').strip().lower()
            for att in (getattr(msg, 'attachments', None) or [])
        )
    )
    return (
        'fallback',
        account_id,
        _extract_email(getattr(msg, 'from_', None)).lower(),
        str(getattr(msg, 'subject', '') or '').strip(),
        str(message_dt),
        attachments,
    )


def _dedupe_order_messages(
    items: list[tuple[object, Optional[object]]]
) -> list[tuple[object, Optional[object]]]:
    deduped: dict[tuple, tuple[object, Optional[object]]] = {}
    for item in items:
        key = _message_identity_key(item)
        current = deduped.get(key)
        if current is None or _message_sort_key(item) > _message_sort_key(
            current
        ):
            deduped[key] = item
    return list(deduped.values())


def _strip_html(text: str) -> str:
    return re.sub(r'<[^>]+>', ' ', text)


def _find_order_number_in_text(
    text: Optional[str],
    prefix: Optional[str],
    suffix: Optional[str],
) -> Optional[str]:
    if not text:
        return None
    if prefix or suffix:
        prefix_re = re.escape(prefix) if prefix else ''
        suffix_re = re.escape(suffix) if suffix else ''
        pattern = rf'{prefix_re}\\s*([\\w\\-\\/]+)\\s*{suffix_re}'
        match = re.search(pattern, text)
        if match:
            return match.group(1)
    return None


def _normalize_oem_key(oem: Optional[str]) -> str:
    value = str(oem or '').strip()
    if not value:
        return ''
    return preprocess_oem_number(value)


def _canonicalize_brand_key(
    brand: Optional[str],
    brand_aliases: Optional[Dict[str, str]] = None,
) -> str:
    normalized = normalize_brand_name(str(brand or ''))
    if not normalized:
        return ''
    if not brand_aliases:
        return normalized
    return brand_aliases.get(normalized, normalized)


def _normalize_key(
    oem: Optional[str],
    brand: Optional[str],
    brand_aliases: Optional[Dict[str, str]] = None,
) -> Tuple[str, str]:
    return (
        _normalize_oem_key(oem),
        _canonicalize_brand_key(brand, brand_aliases),
    )


def _repair_cp1251_mojibake(
    value: Optional[object],
) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if re.search(r'[А-Яа-яЁё]', text):
        return text

    latin1_like_count = len(re.findall(r'[À-ÿ]', text))
    if latin1_like_count < 3:
        return text

    for source_encoding in ('latin1', 'cp1252'):
        try:
            repaired = text.encode(source_encoding).decode('cp1251')
        except (UnicodeEncodeError, UnicodeDecodeError):
            continue
        if len(re.findall(r'[А-Яа-яЁё]', repaired)) >= latin1_like_count:
            return repaired
    return text


async def _load_brand_alias_map(
    session: AsyncSession,
) -> Dict[str, str]:
    alias_map: Dict[str, str] = {}
    brands = await brand_crud.get_multi_with_synonyms(session)
    canonical_map = brand_crud.build_canonical_brand_map(brands)
    for brand in brands:
        canonical = canonical_map.get(brand.id, brand)
        canonical_name = normalize_brand_name(canonical.name)
        if not canonical_name:
            continue
        alias_map[canonical_name] = canonical_name
        alias = normalize_brand_name(brand.name)
        if alias:
            alias_map[alias] = canonical_name
    return alias_map


def _safe_int(value: Optional[object]) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            return int(value)
        except (TypeError, ValueError):
            return None
    try:
        value_str = str(value).strip()
        if value_str == '':
            return None
        return int(float(value_str.replace(',', '.')))
    except (TypeError, ValueError):
        return None


def _safe_float(value: Optional[object]) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        value_str = str(value).strip()
        if value_str == '':
            return None
        return float(value_str.replace(',', '.'))
    except (TypeError, ValueError):
        return None


def _parse_date(value: Optional[object]) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    try:
        return pd.to_datetime(value).date()
    except (TypeError, ValueError):
        return None


def _extract_order_number(
    config: CustomerOrderConfig,
    subject: Optional[str],
    filename: Optional[str],
    body: Optional[str],
) -> Optional[str]:
    if config.order_number_regex_subject and subject:
        match = re.search(config.order_number_regex_subject, subject)
        if match:
            return match.group(0)
    if config.order_number_regex_body and body:
        match = re.search(config.order_number_regex_body, body)
        if match:
            return match.group(0)
    if config.order_number_regex_filename and filename:
        match = re.search(config.order_number_regex_filename, filename)
        if match:
            return match.group(0)
    source = (config.order_number_source or '').lower()
    sources = []
    if source in ('subject', 'filename', 'body'):
        sources = [source]
    else:
        sources = ['subject', 'body', 'filename']

    for item in sources:
        if item == 'subject':
            found = _find_order_number_in_text(
                subject, config.order_number_prefix, config.order_number_suffix
            )
        elif item == 'body':
            found = _find_order_number_in_text(
                body, config.order_number_prefix, config.order_number_suffix
            )
        else:
            found = _find_order_number_in_text(
                filename,
                config.order_number_prefix,
                config.order_number_suffix,
            )
        if found:
            return found
    return None


def _parse_excel_order(
    file_bytes: bytes,
    config: CustomerOrderConfig,
) -> Tuple[List[ParsedOrderRow], Optional[date], Optional[str], BytesIO]:
    wb = load_workbook(BytesIO(file_bytes))
    ws = wb.active

    parsed_rows: List[ParsedOrderRow] = []
    order_date: Optional[date] = None
    order_number: Optional[str] = None

    oem_col = config.oem_col + 1
    brand_col = config.brand_col + 1
    name_col = config.name_col + 1 if config.name_col is not None else None
    qty_col = config.qty_col + 1
    price_col = config.price_col + 1 if config.price_col is not None else None
    order_date_col = (
        config.order_date_column + 1
        if config.order_date_column is not None
        else None
    )
    order_date_row = (
        int(config.order_date_row)
        if config.order_date_row is not None
        else None
    )
    order_number_col = (
        config.order_number_column + 1
        if config.order_number_column is not None
        else None
    )
    order_number_row = (
        int(config.order_number_row)
        if config.order_number_row is not None
        else None
    )

    start_row = max(1, int(config.order_start_row or 1))
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if (
            order_date is None
            and order_date_col is not None
            and (order_date_row is None or row_idx == order_date_row)
            and order_date_col - 1 < len(row)
        ):
            order_date = _parse_date(row[order_date_col - 1])
        if (
            order_number is None
            and order_number_col is not None
            and (order_number_row is None or row_idx == order_number_row)
            and order_number_col - 1 < len(row)
        ):
            value = row[order_number_col - 1]
            if value is not None and str(value).strip():
                order_number = str(value).strip()
        if row_idx < start_row:
            continue
        oem = row[oem_col - 1] if oem_col - 1 < len(row) else None
        brand = row[brand_col - 1] if brand_col - 1 < len(row) else None
        qty = row[qty_col - 1] if qty_col - 1 < len(row) else None
        if not oem or not brand or qty is None:
            continue
        requested_qty = _safe_int(qty)
        if requested_qty is None:
            continue
        name = None
        if name_col is not None and name_col - 1 < len(row):
            name = row[name_col - 1]
            name = _repair_cp1251_mojibake(name)
        requested_price = None
        if price_col is not None and price_col - 1 < len(row):
            requested_price = _safe_float(row[price_col - 1])
        parsed_rows.append(
            ParsedOrderRow(
                row_index=row_idx,
                oem=str(oem).strip(),
                brand=str(brand).strip(),
                name=name,
                requested_qty=requested_qty,
                requested_price=requested_price,
            )
        )

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return parsed_rows, order_date, order_number, output


def _parse_csv_order(
    file_bytes: bytes,
    config: CustomerOrderConfig,
) -> Tuple[List[ParsedOrderRow], Optional[date], Optional[str], BytesIO]:
    df = pd.read_csv(BytesIO(file_bytes), header=None)
    parsed_rows: List[ParsedOrderRow] = []
    order_date: Optional[date] = None
    order_number: Optional[str] = None

    start_row = max(1, int(config.order_start_row or 1))
    order_date_row = (
        int(config.order_date_row)
        if config.order_date_row is not None
        else None
    )
    order_number_row = (
        int(config.order_number_row)
        if config.order_number_row is not None
        else None
    )
    for idx, row in df.iterrows():
        row_num = idx + 1
        if (
            order_date is None
            and config.order_date_column is not None
            and (order_date_row is None or row_num == order_date_row)
        ):
            order_date = _parse_date(row[config.order_date_column])
        if (
            order_number is None
            and config.order_number_column is not None
            and (order_number_row is None or row_num == order_number_row)
        ):
            value = row[config.order_number_column]
            if not pd.isna(value) and str(value).strip():
                order_number = str(value).strip()
        if idx + 1 < start_row:
            continue
        oem = row[config.oem_col]
        brand = row[config.brand_col]
        qty = row[config.qty_col]
        if pd.isna(oem) or pd.isna(brand) or pd.isna(qty):
            continue
        requested_qty = _safe_int(qty)
        if requested_qty is None:
            continue
        name = None
        if config.name_col is not None:
            value = row[config.name_col]
            if not pd.isna(value):
                name = _repair_cp1251_mojibake(value)
        requested_price = None
        if config.price_col is not None:
            value = row[config.price_col]
            if not pd.isna(value):
                requested_price = _safe_float(value)
        parsed_rows.append(
            ParsedOrderRow(
                row_index=idx,
                oem=str(oem).strip(),
                brand=str(brand).strip(),
                name=name,
                requested_qty=requested_qty,
                requested_price=requested_price,
            )
        )

    output = BytesIO()
    df.to_csv(output, index=False, header=False)
    output.seek(0)
    return parsed_rows, order_date, order_number, output


def _parse_xls_order(
    file_bytes: bytes,
    config: CustomerOrderConfig,
) -> Tuple[List[ParsedOrderRow], Optional[date], Optional[str], BytesIO]:
    df = pd.read_excel(BytesIO(file_bytes), header=None, engine='xlrd')
    parsed_rows: List[ParsedOrderRow] = []
    order_date: Optional[date] = None
    order_number: Optional[str] = None

    start_row = max(1, int(config.order_start_row or 1))
    order_date_row = (
        int(config.order_date_row)
        if config.order_date_row is not None
        else None
    )
    order_number_row = (
        int(config.order_number_row)
        if config.order_number_row is not None
        else None
    )
    for idx, row in df.iterrows():
        row_num = idx + 1
        if (
            order_date is None
            and config.order_date_column is not None
            and (order_date_row is None or row_num == order_date_row)
        ):
            order_date = _parse_date(row[config.order_date_column])
        if (
            order_number is None
            and config.order_number_column is not None
            and (order_number_row is None or row_num == order_number_row)
        ):
            value = row[config.order_number_column]
            if not pd.isna(value) and str(value).strip():
                order_number = str(value).strip()
        if idx + 1 < start_row:
            continue
        oem = row[config.oem_col]
        brand = row[config.brand_col]
        qty = row[config.qty_col]
        if pd.isna(oem) or pd.isna(brand) or pd.isna(qty):
            continue
        requested_qty = _safe_int(qty)
        if requested_qty is None:
            continue
        name = None
        if config.name_col is not None:
            value = row[config.name_col]
            if not pd.isna(value):
                name = _repair_cp1251_mojibake(value)
        requested_price = None
        if config.price_col is not None:
            value = row[config.price_col]
            if not pd.isna(value):
                requested_price = _safe_float(value)
        parsed_rows.append(
            ParsedOrderRow(
                row_index=idx,
                oem=str(oem).strip(),
                brand=str(brand).strip(),
                name=name,
                requested_qty=requested_qty,
                requested_price=requested_price,
            )
        )

    output = BytesIO()
    df.to_excel(output, index=False, header=False)
    output.seek(0)
    return parsed_rows, order_date, order_number, output


def _parse_order_attachment(
    file_bytes: bytes,
    filename: str,
    config: CustomerOrderConfig,
) -> Tuple[List[ParsedOrderRow], Optional[date], Optional[str], BytesIO, str]:
    file_ext = (filename.rsplit('.', 1)[-1] if '.' in filename else '').lower()
    if file_ext == 'csv':
        parsed_rows, order_date, order_number, file_buffer = (
            _parse_csv_order(file_bytes, config)
        )
    elif file_ext == 'xls':
        parsed_rows, order_date, order_number, file_buffer = (
            _parse_xls_order(file_bytes, config)
        )
    elif file_ext == 'xlsx':
        parsed_rows, order_date, order_number, file_buffer = (
            _parse_excel_order(file_bytes, config)
        )
    else:
        raise ValueError(
            f'Неподдерживаемый тип файла: {file_ext or "unknown"}'
        )
    return parsed_rows, order_date, order_number, file_buffer, file_ext


def _apply_response_updates_excel(
    file_bytes: BytesIO,
    config: CustomerOrderConfig,
    items: List[CustomerOrderItem],
):
    wb = load_workbook(file_bytes)
    ws = wb.active
    qty_col = config.qty_col + 1
    ship_col = (
        config.ship_qty_col + 1
        if config.ship_qty_col is not None
        else None
    )
    ship_price_col = (
        config.ship_price_col + 1
        if config.ship_price_col is not None
        else None
    )
    reject_col = (
        config.reject_qty_col + 1
        if config.reject_qty_col is not None
        else None
    )

    for item in items:
        row_index = item.row_index
        if row_index is None:
            continue
        if config.ship_mode == CUSTOMER_ORDER_SHIP_MODE.REPLACE_QTY:
            ws.cell(row=row_index, column=qty_col).value = (
                item.ship_qty or 0
            )
        elif config.ship_mode == CUSTOMER_ORDER_SHIP_MODE.WRITE_SHIP_QTY:
            if ship_col is None:
                raise ValueError('ship_qty_col is required for WRITE_SHIP_QTY')
            ws.cell(row=row_index, column=ship_col).value = (
                item.ship_qty or 0
            )
        elif config.ship_mode == CUSTOMER_ORDER_SHIP_MODE.WRITE_REJECT_QTY:
            if reject_col is None:
                raise ValueError(
                    'reject_qty_col is required for WRITE_REJECT_QTY'
                )
            ws.cell(row=row_index, column=reject_col).value = (
                item.reject_qty or 0
            )
        if ship_price_col is not None:
            ws.cell(row=row_index, column=ship_price_col).value = (
                _get_response_ship_price_value(item)
            )

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _apply_response_updates_csv(
    file_bytes: BytesIO,
    config: CustomerOrderConfig,
    items: List[CustomerOrderItem],
):
    file_bytes.seek(0)
    df = pd.read_csv(file_bytes, header=None)
    if config.ship_price_col is not None:
        df[config.ship_price_col] = df[config.ship_price_col].astype(object)
    for item in items:
        row_index = item.row_index
        if row_index is None:
            continue
        if config.ship_mode == CUSTOMER_ORDER_SHIP_MODE.REPLACE_QTY:
            df.iat[row_index, config.qty_col] = item.ship_qty or 0
        elif config.ship_mode == CUSTOMER_ORDER_SHIP_MODE.WRITE_SHIP_QTY:
            if config.ship_qty_col is None:
                raise ValueError('ship_qty_col is required for WRITE_SHIP_QTY')
            df.iat[row_index, config.ship_qty_col] = item.ship_qty or 0
        elif config.ship_mode == CUSTOMER_ORDER_SHIP_MODE.WRITE_REJECT_QTY:
            if config.reject_qty_col is None:
                raise ValueError(
                    'reject_qty_col is required for WRITE_REJECT_QTY'
                )
            df.iat[row_index, config.reject_qty_col] = item.reject_qty or 0
        if config.ship_price_col is not None:
            df.iat[row_index, config.ship_price_col] = (
                _get_response_ship_price_value(item, blank_value='')
            )

    output = BytesIO()
    df.to_csv(output, index=False, header=False)
    output.seek(0)
    return output


async def _load_latest_customer_pricelist(
    session: AsyncSession, customer_id: int
) -> Optional[CustomerPriceList]:
    stmt = (
        select(CustomerPriceList)
        .where(CustomerPriceList.customer_id == customer_id)
        .order_by(CustomerPriceList.date.desc(), CustomerPriceList.id.desc())
        .options(
            joinedload(CustomerPriceList.autopart_associations)
            .joinedload(CustomerPriceListAutoPartAssociation.autopart)
            .joinedload(AutoPart.brand)
        )
        .limit(1)
    )
    result = await session.execute(stmt)
    return result.scalars().first()


def _build_expected_price_map(
    pricelist: CustomerPriceList,
    brand_aliases: Optional[Dict[str, str]] = None,
) -> Dict:
    expected = {}
    for assoc in pricelist.autopart_associations:
        autopart = assoc.autopart
        if not autopart or not autopart.brand:
            continue
        key = _normalize_key(
            autopart.oem_number, autopart.brand.name, brand_aliases
        )
        expected[key] = float(assoc.price or 0)
    return expected


async def _build_current_offers(
    session: AsyncSession,
    config: CustomerPriceListConfig,
    brand_aliases: Optional[Dict[str, str]] = None,
) -> Dict[Tuple[str, str], OfferRow]:
    sources = await crud_customer_pricelist_source.get_by_config_id(
        config_id=config.id, session=session
    )
    combined_data = []
    for source in sources:
        if not source.enabled:
            continue
        latest_pl = await crud_pricelist.get_latest_pricelist_by_config(
            session=session, provider_config_id=source.provider_config_id
        )
        if not latest_pl:
            continue
        associations = await crud_pricelist.fetch_pricelist_data(
            latest_pl.id, session
        )
        if not associations:
            continue
        df = await crud_pricelist.transform_to_dataframe(
            associations=associations, session=session
        )
        # For order matching we ignore price/quantity thresholds from the
        # outbound pricelist. A valid offer should still match even if it
        # would be hidden from the mailed pricelist by stock/price limits.
        df = _apply_source_filters(
            df, source, ignore_price_quantity_filters=True
        )
        if df.empty:
            continue
        # Keep the original supplier price from the provider price list.
        # Later we apply customer-facing markups to `price`, but supplier
        # orders must use this raw source price.
        df = df.copy()
        df['price'] = pd.to_numeric(df['price'], errors='coerce')
        df['supplier_price'] = df['price']
        df = crud_customer_pricelist.apply_coefficient(
            df,
            config,
            apply_general_markup=False,
            ignore_price_quantity_filters=True,
        )
        df = _apply_source_markups(df, config, source)
        combined_data.append(df)

    if not combined_data:
        return {}

    final_df = pd.concat(combined_data, ignore_index=True)
    final_df['__normalized_oem'] = final_df['oem_number'].map(
        _normalize_oem_key
    )
    final_df['__normalized_brand'] = final_df['brand'].map(
        lambda brand: _canonicalize_brand_key(brand, brand_aliases)
    )

    if 'is_own_price' in final_df.columns:
        final_df['__own_rank'] = final_df['is_own_price'].astype(int)
        final_df = (
            final_df.sort_values(
                by=[
                    '__normalized_oem',
                    '__normalized_brand',
                    '__own_rank',
                    'price',
                ],
                ascending=[True, True, False, True],
            )
            .drop_duplicates(
                subset=['__normalized_oem', '__normalized_brand'],
                keep='first',
            )
            .drop(columns=['__own_rank'])
        )
    else:
        final_df = final_df.sort_values(
            by=['__normalized_oem', '__normalized_brand', 'price']
        ).drop_duplicates(
            subset=['__normalized_oem', '__normalized_brand'], keep='first'
        )

    offers = {}
    for _, row in final_df.iterrows():
        key = (
            str(row.get('__normalized_oem') or ''),
            str(row.get('__normalized_brand') or ''),
        )
        supplier_price_raw = row.get('supplier_price')
        if pd.isna(supplier_price_raw):
            supplier_price_raw = row.get('price')
        supplier_price = (
            float(supplier_price_raw)
            if pd.notna(supplier_price_raw)
            else 0.0
        )
        offers[key] = OfferRow(
            autopart_id=int(row.get('autopart_id')),
            provider_id=int(row.get('provider_id')),
            provider_config_id=row.get('provider_config_id'),
            quantity=int(row.get('quantity') or 0),
            price=float(row.get('price') or 0),
            supplier_price=supplier_price,
            is_own_price=bool(row.get('is_own_price')),
        )
    return offers


def _resolve_customer_target_price(
    expected_price: Optional[float],
    requested_price: Optional[float],
    offer: Optional[OfferRow],
) -> Optional[float]:
    if requested_price is not None and requested_price > 0:
        return float(requested_price)
    if expected_price is not None and expected_price > 0:
        return float(expected_price)
    if offer and offer.price and offer.price > 0:
        return float(offer.price)
    return None


def _compute_price_diff_pct(
    customer_price: float, offered_price: float
) -> float:
    if customer_price <= 0 or offered_price <= 0:
        return 0.0
    if customer_price >= offered_price:
        return 0.0
    return ((offered_price - customer_price) / offered_price) * 100


def _compute_order_requested_total(
    parsed_rows: List[ParsedOrderRow],
) -> Optional[float]:
    total = 0.0
    has_price = False
    for row in parsed_rows:
        if row.requested_price is None:
            continue
        has_price = True
        total += float(row.requested_price) * int(row.requested_qty or 0)
    if not has_price:
        return None
    return round(total, 2)


def _get_response_ship_price_value(
    item: CustomerOrderItem,
    blank_value=None,
):
    if not (item.ship_qty and item.ship_qty > 0):
        return blank_value
    if item.requested_price is None:
        return blank_value
    return float(item.requested_price)


def _set_reject_reason(
    item: CustomerOrderItem,
    code: Optional[str],
    text: Optional[str],
) -> None:
    item.reject_reason_code = (code or None)
    item.reject_reason_text = _truncate_error_details(text)


def _clear_reject_reason(item: CustomerOrderItem) -> None:
    _set_reject_reason(item, None, None)


def _source_display_name(source) -> str:
    provider_config = getattr(source, 'provider_config', None)
    provider = getattr(provider_config, 'provider', None)
    provider_name = getattr(provider, 'name', None) or (
        f'provider_config_id={source.provider_config_id}'
    )
    config_name = getattr(provider_config, 'name_price', None)
    if config_name:
        return f'{provider_name} / {config_name}'
    return provider_name


def _describe_source_filter_rules(source) -> str:
    parts: List[str] = []
    if source.brand_filters:
        parts.append('brand_filters')
    if source.position_filters:
        parts.append('position_filters')
    if source.min_price is not None:
        parts.append(f'min_price={float(source.min_price):.2f}')
    if source.max_price is not None:
        parts.append(f'max_price={float(source.max_price):.2f}')
    if source.min_quantity is not None:
        parts.append(f'min_quantity={int(source.min_quantity)}')
    if source.max_quantity is not None:
        parts.append(f'max_quantity={int(source.max_quantity)}')
    if source.additional_filters:
        parts.append('additional_filters')
    return ', '.join(parts) if parts else 'неизвестные правила'


def _merge_pricelist_filter_blocks(
    base: dict,
    override: Optional[dict],
) -> dict:
    merged = dict(base)
    if not override:
        return merged
    for key, value in override.items():
        merged[key] = value
    return merged


def _resolve_pricelist_filters_for_offer(
    config: CustomerPriceListConfig,
    provider_id: Optional[int],
    is_own_price: bool,
) -> dict:
    if config.default_filters:
        base = dict(config.default_filters)
    else:
        base = {
            'brand_filters': config.brand_filters,
            'category_filter': config.category_filter,
            'price_intervals': config.price_intervals,
            'position_filters': config.position_filters,
            'supplier_quantity_filters': (
                config.supplier_quantity_filters
            ),
            'additional_filters': config.additional_filters,
        }

    if is_own_price:
        return _merge_pricelist_filter_blocks(base, config.own_filters)

    supplier_filters = config.supplier_filters or {}
    if provider_id is not None:
        override = supplier_filters.get(provider_id)
        if override is None:
            override = supplier_filters.get(str(provider_id))
        if override:
            return _merge_pricelist_filter_blocks(base, override)
    return _merge_pricelist_filter_blocks(base, config.other_filters)


def _describe_pricelist_filter_rules(filters_cfg: dict) -> str:
    parts: List[str] = []
    if filters_cfg.get('brand_filters'):
        parts.append('brand_filters')
    if filters_cfg.get('position_filters'):
        parts.append('position_filters')
    if filters_cfg.get('price_intervals'):
        parts.append('price_intervals')
    if filters_cfg.get('supplier_quantity_filters'):
        parts.append('supplier_quantity_filters')
    min_price = filters_cfg.get('min_price')
    if min_price is not None:
        parts.append(f'min_price={float(min_price):.2f}')
    max_price = filters_cfg.get('max_price')
    if max_price is not None:
        parts.append(f'max_price={float(max_price):.2f}')
    min_qty = filters_cfg.get('min_quantity')
    if min_qty is not None:
        parts.append(f'min_quantity={int(min_qty)}')
    max_qty = filters_cfg.get('max_quantity')
    if max_qty is not None:
        parts.append(f'max_quantity={int(max_qty)}')
    if filters_cfg.get('additional_filters'):
        parts.append('additional_filters')
    return ', '.join(parts) if parts else 'неизвестные правила'


def _normalize_offer_dataframe_keys(
    df: pd.DataFrame,
    brand_aliases: Optional[Dict[str, str]] = None,
) -> pd.DataFrame:
    if df.empty:
        return df.copy()
    normalized = df.copy()
    normalized['__normalized_oem'] = normalized['oem_number'].map(
        _normalize_oem_key
    )
    normalized['__normalized_brand'] = normalized['brand'].map(
        lambda brand: _canonicalize_brand_key(brand, brand_aliases)
    )
    return normalized


async def _diagnose_missing_offer_reason(
    session: AsyncSession,
    pricelist_config: Optional[CustomerPriceListConfig],
    row: ParsedOrderRow,
    brand_aliases: Optional[Dict[str, str]] = None,
) -> tuple[str, str]:
    if not pricelist_config:
        return (
            'NO_LINKED_PRICELIST',
            'К заказу не привязана конфигурация клиентского прайса.',
        )

    sources = await crud_customer_pricelist_source.get_by_config_id(
        config_id=pricelist_config.id,
        session=session,
    )
    enabled_sources = [source for source in sources if source.enabled]
    if not enabled_sources:
        return (
            'NO_LINKED_SOURCE',
            'У клиентского прайса нет активных источников предложений.',
        )

    key = _normalize_key(row.oem, row.brand, brand_aliases)
    diagnostics: List[str] = []
    filtered_out = False
    nonpositive_offer = False

    for source in enabled_sources:
        latest_pl = await crud_pricelist.get_latest_pricelist_by_config(
            session=session,
            provider_config_id=source.provider_config_id,
        )
        if not latest_pl:
            continue
        associations = await crud_pricelist.fetch_pricelist_data(
            latest_pl.id, session
        )
        if not associations:
            continue

        raw_df = await crud_pricelist.transform_to_dataframe(
            associations=associations,
            session=session,
        )
        if raw_df.empty:
            continue

        raw_df = _normalize_offer_dataframe_keys(raw_df, brand_aliases)
        raw_match = raw_df[
            (raw_df['__normalized_oem'] == key[0])
            & (raw_df['__normalized_brand'] == key[1])
        ]
        if raw_match.empty:
            continue

        source_name = _source_display_name(source)
        numeric_raw = raw_match.copy()
        numeric_raw['price'] = pd.to_numeric(
            numeric_raw['price'], errors='coerce'
        )
        numeric_raw['quantity'] = pd.to_numeric(
            numeric_raw['quantity'], errors='coerce'
        )
        positive_raw = numeric_raw[
            (numeric_raw['price'] > 0) & (numeric_raw['quantity'] > 0)
        ]
        if positive_raw.empty:
            diagnostics.append(
                f'{source_name}: предложение найдено, но цена или остаток '
                'неположительные.'
            )
            nonpositive_offer = True
            continue

        filtered_df = _apply_source_filters(
            raw_df, source, ignore_price_quantity_filters=True
        )
        filtered_df = _normalize_offer_dataframe_keys(
            filtered_df, brand_aliases
        )
        filtered_match = filtered_df[
            (filtered_df['__normalized_oem'] == key[0])
            & (filtered_df['__normalized_brand'] == key[1])
        ]
        if filtered_match.empty:
            diagnostics.append(
                f'{source_name}: позиция исключена фильтрами источника '
                f'({_describe_source_filter_rules(source)}).'
            )
            filtered_out = True
            continue

        first_match = positive_raw.iloc[0]
        provider_id_value = first_match.get('provider_id')
        own_flag_value = first_match.get('is_own_price')
        provider_id = (
            int(provider_id_value)
            if pd.notna(provider_id_value)
            else None
        )
        own_flag = (
            bool(own_flag_value)
            if pd.notna(own_flag_value)
            else False
        )

        config_filtered_df = crud_customer_pricelist.apply_coefficient(
            filtered_df,
            pricelist_config,
            apply_general_markup=False,
            provider_id=provider_id,
            is_own_price=own_flag,
            ignore_price_quantity_filters=True,
        )
        config_filtered_df = _normalize_offer_dataframe_keys(
            config_filtered_df, brand_aliases
        )
        config_match = config_filtered_df[
            (config_filtered_df['__normalized_oem'] == key[0])
            & (config_filtered_df['__normalized_brand'] == key[1])
        ]
        if config_match.empty:
            pricelist_filters = _resolve_pricelist_filters_for_offer(
                pricelist_config,
                provider_id=provider_id,
                is_own_price=own_flag,
            )
            diagnostics.append(
                f'{source_name}: позиция исключена фильтрами клиентского '
                'прайса '
                f'({_describe_pricelist_filter_rules(pricelist_filters)}).'
            )
            continue

        final_df = _apply_source_markups(
            config_filtered_df, pricelist_config, source
        )
        final_df = _normalize_offer_dataframe_keys(final_df, brand_aliases)
        final_match = final_df[
            (final_df['__normalized_oem'] == key[0])
            & (final_df['__normalized_brand'] == key[1])
        ]
        if final_match.empty:
            diagnostics.append(
                f'{source_name}: предложение найдено, но не попало в '
                'итоговый набор офферов после обработки.'
            )
            continue

        diagnostics.append(
            f'{source_name}: предложение найдено, но не было выбрано '
            'для строки заказа.'
        )

    if filtered_out:
        return 'FILTERED_BY_SOURCE_RULE', ' '.join(diagnostics[:2])
    if nonpositive_offer:
        return 'NONPOSITIVE_OFFER', ' '.join(diagnostics[:2])
    if diagnostics:
        if any('фильтрами клиентского прайса' in item for item in diagnostics):
            return 'FILTERED_BY_PRICE_CONFIG', ' '.join(diagnostics[:2])
        return 'OFFER_MATCH_DIAGNOSTIC', ' '.join(diagnostics[:2])
    return 'NO_OFFER', 'Нет предложения в подключенных источниках клиента.'


def _format_order_amount(value: Optional[float]) -> str:
    if value is None:
        return 'не определена'
    return f'{value:,.2f}'.replace(',', ' ')


def _truncate_error_details(reason: Optional[str]) -> Optional[str]:
    if not reason:
        return None
    reason = str(reason).strip()
    if not reason:
        return None
    return reason[:ORDER_ERROR_DETAIL_MAX_LEN]


def _order_source_storage_name(order_id: int, filename: Optional[str]) -> str:
    safe_name = os.path.basename(filename or 'order.xlsx') or 'order.xlsx'
    return f'{order_id}_{safe_name}'


def _order_source_storage_path(order: CustomerOrder) -> Optional[str]:
    if not order.id or not order.source_filename:
        return None
    return os.path.join(
        ORDERS_ERROR_DIR,
        _order_source_storage_name(order.id, order.source_filename),
    )


async def _save_order_source_file(
    order: CustomerOrder,
    payload: bytes,
) -> Optional[str]:
    path = _order_source_storage_path(order)
    if not path:
        return None
    os.makedirs(ORDERS_ERROR_DIR, exist_ok=True)
    async with aiofiles.open(path, 'wb') as f:
        await f.write(payload)
    return path


async def _load_order_source_file(order: CustomerOrder) -> Optional[bytes]:
    path = _order_source_storage_path(order)
    if not path or not os.path.isfile(path):
        return None
    async with aiofiles.open(path, 'rb') as f:
        return await f.read()


def _mark_inbox_account_received_at(
    inbox_account,
    received_at: Optional[datetime],
) -> bool:
    if (
        not inbox_account
        or (inbox_account.transport or '').strip().lower() != 'resend_api'
        or not received_at
    ):
        return False
    inbox_account.resend_last_received_at = received_at
    return True


def _advance_config_last_uid(
    config: CustomerOrderConfig,
    uid: Optional[object],
    folder_name: Optional[str] = None,
    account_id: Optional[int] = None,
) -> bool:
    msg_uid_int = _safe_uid_as_int(uid)
    if msg_uid_int is None:
        return False
    config.last_uid = max(int(config.last_uid or 0), msg_uid_int)
    if folder_name:
        normalized_folder = normalize_imap_folder(folder_name)
        folder_uids = dict(config.folder_last_uids or {})
        parsed_account_id: Optional[int]
        try:
            parsed_account_id = (
                int(account_id) if account_id is not None else None
            )
        except (TypeError, ValueError):
            parsed_account_id = None
        if parsed_account_id and parsed_account_id > 0:
            scoped_key = f'{parsed_account_id}:{normalized_folder}'
            folder_uids[scoped_key] = msg_uid_int
            if not _config_uses_multiple_email_accounts(config):
                folder_uids[normalized_folder] = msg_uid_int
        else:
            folder_uids[normalized_folder] = msg_uid_int
        config.folder_last_uids = folder_uids
    return True


def _get_config_last_uid(
    config: CustomerOrderConfig,
    folder_name: Optional[str] = None,
    account_id: Optional[int] = None,
) -> int:
    if folder_name:
        folder_uids = dict(config.folder_last_uids or {})
        normalized_folder = normalize_imap_folder(folder_name)
        parsed_account_id: Optional[int]
        try:
            parsed_account_id = (
                int(account_id) if account_id is not None else None
            )
        except (TypeError, ValueError):
            parsed_account_id = None
        if parsed_account_id and parsed_account_id > 0:
            scoped_key = f'{parsed_account_id}:{normalized_folder}'
            scoped_uid = folder_uids.get(scoped_key)
            if scoped_uid is not None:
                try:
                    return int(scoped_uid)
                except (TypeError, ValueError):
                    pass
            if _config_uses_multiple_email_accounts(config):
                return 0
        folder_uid = folder_uids.get(normalized_folder)
        if folder_uid is not None:
            try:
                return int(folder_uid)
            except (TypeError, ValueError):
                pass
    return int(config.last_uid or 0)


def _mark_order_error(
    order: CustomerOrder,
    reason: Optional[str],
) -> None:
    order.status = CUSTOMER_ORDER_STATUS.ERROR
    order.processed_at = now_moscow()
    order.error_details = _truncate_error_details(reason)


def _build_order_reply_recipients(
    sender: Optional[str],
    config: CustomerOrderConfig,
    *,
    use_override: bool = True,
) -> str:
    recipients = set()
    if sender:
        recipients.add(sender)
    for email in _normalize_email_list(config.order_reply_emails):
        recipients.add(email)
    if use_override:
        override_email = _customer_order_reply_override_email()
        if override_email:
            return override_email
    return ','.join(sorted(recipients))


def _build_supplier_order_recipient(
    provider: Optional[Provider],
    *,
    use_override: bool = True,
) -> Optional[str]:
    if use_override:
        override_email = _supplier_order_override_email()
        if override_email:
            return override_email
    if provider and provider.email_contact:
        return provider.email_contact
    return None


async def _send_email_attachment_async(
    to_email: str,
    subject: str,
    body: str,
    attachment: bytes,
    filename: str,
    use_tls: bool,
    **kwargs,
):
    await asyncio.to_thread(
        send_email_with_attachment,
        to_email,
        subject,
        body,
        attachment,
        filename,
        use_tls,
        **kwargs,
    )


def _supplier_order_print_code(order_id: int) -> str:
    return f'A{int(order_id):010d}'


def _supplier_order_attachment_filename(order_id: int) -> str:
    return f'МастерА{int(order_id):010d}.xlsx'


def _supplier_order_provider_alias(provider: Optional[Provider]) -> str:
    if not provider:
        return ''
    for raw_email in (
        getattr(provider, 'email_contact', None),
        getattr(provider, 'email_incoming_price', None),
    ):
        value = str(raw_email or '').strip()
        if '@' in value:
            local = value.split('@', 1)[0].strip()
            if local:
                return local.upper()
    return ''


def _format_numeric_for_excel(value: Optional[float]) -> object:
    if value is None:
        return ''
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return value
    if parsed.is_integer():
        return int(parsed)
    return round(parsed, 2)


def _build_supplier_order_rows(
    order: SupplierOrder,
) -> tuple[list[dict[str, object]], int, float]:
    rows: list[dict[str, object]] = []
    total_qty = 0
    total_sum = 0.0
    for index, item in enumerate(order.items or [], start=1):
        autopart = item.autopart
        brand_name = (
            autopart.brand.name if autopart and autopart.brand else ''
        )
        oem_number = (
            item.oem_number
            or (autopart.oem_number if autopart else '')
            or ''
        )
        part_name = (
            item.autopart_name
            or (autopart.name if autopart else '')
            or ''
        )
        quantity = int(item.quantity or 0)
        price = float(item.price) if item.price is not None else 0.0
        line_sum = round(quantity * price, 2)
        total_qty += quantity
        total_sum += line_sum
        rows.append(
            {
                'index': index,
                'brand': brand_name,
                'oem': oem_number,
                'comment': (
                    int(item.customer_order_item_id)
                    if item.customer_order_item_id is not None
                    else ''
                ),
                'name': part_name,
                'note': '',
                'qty': quantity,
                'price': _format_numeric_for_excel(price),
                'sum': _format_numeric_for_excel(line_sum),
                'price_code': '',
            }
        )
    return rows, total_qty, round(total_sum, 2)


def _build_supplier_order_attachment_bytes(
    *,
    order: SupplierOrder,
    rows: list[dict[str, object]],
    total_qty: int,
    total_sum: float,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'TDSheet'

    column_widths = [
        6.0,
        15.5,
        13.83,
        10.5,
        23.33,
        10.5,
        8.0,
        10.5,
        10.5,
        10.5,
    ]
    for index, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    row_heights = {
        1: 15.75,
        3: 15.75,
        5: 15.75,
        6: 11.25,
        8: 12.75,
        9: 12.75,
    }
    for row_idx, height in row_heights.items():
        sheet.row_dimensions[row_idx].height = height

    thin_side = Side(style='thin', color='000000')
    no_side = Side(style=None)

    def _make_border(
        *,
        left: bool = False,
        right: bool = False,
        top: bool = False,
        bottom: bool = False,
    ) -> Border:
        return Border(
            left=thin_side if left else no_side,
            right=thin_side if right else no_side,
            top=thin_side if top else no_side,
            bottom=thin_side if bottom else no_side,
        )

    font_title = Font(name='Arial', size=12)
    font_main = Font(name='Arial', size=10)
    font_small = Font(name='Arial', size=8)
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    align_general = Alignment(horizontal='general', vertical='center')
    align_right_wrap = Alignment(
        horizontal='right',
        vertical='center',
        wrap_text=True,
    )

    order_datetime = order.created_at or now_moscow()
    order_datetime_text = order_datetime.strftime('%d.%m.%Y %H:%M:%S')
    provider_name = str(
        getattr(getattr(order, 'provider', None), 'name', '') or ''
    ).strip()
    provider_alias = _supplier_order_provider_alias(order.provider)
    provider_line = ' '.join(
        part for part in [provider_name, provider_alias] if part
    ).strip()

    sheet.merge_cells('H1:J1')
    cell_h1 = sheet['H1']
    cell_h1.value = 'Заказ поставщику'
    cell_h1.font = font_title
    cell_h1.alignment = align_right_wrap

    sheet['C3'].value = 'Дата'
    sheet['C3'].font = font_main
    sheet['C3'].alignment = align_left
    sheet['C3'].border = _make_border(left=True, top=True, bottom=True)

    sheet.merge_cells('D3:E3')
    sheet['D3'].value = order_datetime_text
    sheet['D3'].font = font_title
    sheet['D3'].alignment = align_left
    sheet['D3'].border = _make_border(top=True, right=True, bottom=True)

    sheet['C5'].value = int(order.id)
    sheet['C5'].font = font_title
    sheet['C5'].alignment = align_left
    sheet['C5'].border = _make_border(
        left=True,
        right=True,
        top=True,
        bottom=True,
    )

    sheet['E5'].value = _supplier_order_print_code(order.id)
    sheet['E5'].font = font_title
    sheet['E5'].alignment = align_left
    sheet['E5'].border = _make_border(
        left=True,
        right=True,
        top=True,
        bottom=True,
    )

    sheet['E6'].value = provider_line
    sheet['E6'].font = font_small
    sheet['E6'].alignment = align_left
    sheet['E6'].border = _make_border(
        left=True,
        right=True,
        top=True,
        bottom=True,
    )

    sheet['F8'].value = 'Итого:'
    sheet['F8'].font = font_main
    sheet['F8'].alignment = align_left
    sheet['G8'].value = total_qty
    sheet['G8'].font = font_main
    sheet['G8'].alignment = align_right
    sheet['G8'].number_format = '0.00'
    sheet['H8'].value = 'RUR'
    sheet['H8'].font = font_main
    sheet['H8'].alignment = align_left
    sheet['I8'].value = round(total_sum, 2)
    sheet['I8'].font = font_main
    sheet['I8'].alignment = align_right
    sheet['I8'].number_format = '0.00'

    headers = [
        '№',
        'Производитель',
        'Номер детали',
        'Комментарий',
        'Наименование',
        'Примечание',
        'Кол-во',
        'Цена',
        'Сумма',
        'Прайс',
    ]
    for col_idx, title in enumerate(headers, start=1):
        cell = sheet.cell(row=9, column=col_idx)
        cell.value = title
        cell.font = font_main
        cell.alignment = align_left
        cell.border = _make_border(left=True, right=True, top=True)

    data_row_start = 10
    for index, row in enumerate(rows):
        row_idx = data_row_start + index
        sheet.row_dimensions[row_idx].height = 12.75
        values = [
            row['index'],
            row['brand'],
            row['oem'],
            row['comment'],
            row['name'],
            row['note'],
            row['qty'],
            row['price'],
            row['sum'],
            row['price_code'],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.font = font_main
            if col_idx in {1, 7, 8, 9}:
                cell.alignment = align_right
            elif col_idx == 3:
                cell.alignment = align_general
            else:
                cell.alignment = align_left
            cell.border = _make_border(left=True, right=True, bottom=True)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _build_supplier_order_body_html(
    *,
    order: SupplierOrder,
    rows: list[dict[str, object]],
    total_qty: int,
    total_sum: float,
) -> str:
    order_datetime = order.created_at or now_moscow()
    order_datetime_text = order_datetime.strftime('%d.%m.%Y %H:%M:%S')
    provider_name = str(
        getattr(getattr(order, 'provider', None), 'name', '') or ''
    ).strip()
    provider_alias = _supplier_order_provider_alias(order.provider)
    provider_line = ' '.join(
        part for part in [provider_name, provider_alias] if part
    ).strip()
    row_lines = ''.join(
        (
            '<tr>'
            f'<td>{escape(str(row["index"]))}</td>'
            f'<td>{escape(str(row["brand"] or ""))}</td>'
            f'<td>{escape(str(row["oem"] or ""))}</td>'
            f'<td>{escape(str(row["comment"] or ""))}</td>'
            f'<td>{escape(str(row["name"] or ""))}</td>'
            f'<td>{escape(str(row["note"] or ""))}</td>'
            f'<td>{escape(str(row["qty"] or ""))}</td>'
            f'<td>{escape(str(row["price"] or ""))}</td>'
            f'<td>{escape(str(row["sum"] or ""))}</td>'
            f'<td>{escape(str(row["price_code"] or ""))}</td>'
            '</tr>'
        )
        for row in rows
    )
    return (
        '<div>'
        f'<p><b>Заказ поставщику № {escape(str(order.id))}</b></p>'
        '<table cellspacing="0" cellpadding="2">'
        f'<tr><td><b>Дата</b></td><td>{escape(order_datetime_text)}</td></tr>'
        f'<tr><td><b>{escape(str(order.id))}</b></td>'
        f'<td><b>{escape(_supplier_order_print_code(order.id))}</b></td></tr>'
        f'<tr><td></td><td>{escape(provider_line)}</td></tr>'
        '</table>'
        '<br/>'
        '<table border="1" cellpadding="4" cellspacing="0" '
        'style="border-collapse: collapse;">'
        '<thead>'
        '<tr>'
        '<th>№</th>'
        '<th>Производитель</th>'
        '<th>Номер детали</th>'
        '<th>Комментарий</th>'
        '<th>Наименование</th>'
        '<th>Примечание</th>'
        '<th>Кол-во</th>'
        '<th>Цена</th>'
        '<th>Сумма</th>'
        '<th>Прайс</th>'
        '</tr>'
        '</thead>'
        f'<tbody>{row_lines}</tbody>'
        '<tfoot>'
        f'<tr><td colspan="6"><b>Итого:</b></td>'
        f'<td>{escape(str(total_qty))}</td>'
        '<td>RUR</td>'
        f'<td>{escape(f"{total_sum:.2f}")}</td>'
        '<td></td></tr>'
        '</tfoot>'
        '</table>'
        '</div>'
    )


async def _send_order_import_notification(
    session: AsyncSession,
    config: CustomerOrderConfig,
    sender: str,
    subject: Optional[str],
    filename: Optional[str],
    *,
    success: bool,
    reason: Optional[str] = None,
    order_number: Optional[str] = None,
    total_amount: Optional[float] = None,
    rows_count: Optional[int] = None,
):
    customer_name = (
        getattr(getattr(config, 'customer', None), 'name', None)
        or str(config.customer_id)
    )
    lines = [
        'Заказ загружен' if success else 'Заказ не загружен',
        f'Клиент: {customer_name}',
        f'Конфиг заказа: {config.id}',
        f'Отправитель: {sender}',
    ]
    if subject:
        lines.append(f'Тема: {subject}')
    if filename:
        lines.append(f'Файл: {filename}')
    if order_number:
        lines.append(f'Номер заказа: {order_number}')
    if rows_count is not None:
        lines.append(f'Строк заказа: {rows_count}')
    lines.append(f'Сумма заказа: {_format_order_amount(total_amount)}')
    if reason:
        lines.append(f'Причина: {reason}')
    await _notify_admins(
        session,
        title=(
            'Импорт заказа клиента'
            if success
            else 'Ошибка импорта заказа клиента'
        ),
        message='\n'.join(lines),
        level=(
            AppNotificationLevel.SUCCESS
            if success
            else AppNotificationLevel.ERROR
        ),
        link='/customer-orders',
        commit=True,
    )


async def _send_price_warning(
    session: AsyncSession,
    order: CustomerOrder,
    item: CustomerOrderItem,
    customer_price: float,
    offered_price: float,
    diff_pct: float,
    critical: bool,
):
    label = '!!!' if critical else '!'
    text = (
        f'{label} Отклонение цены по заказу клиента {order.customer_id} '
        f'({order.order_number or order.id})\n'
        f'Позиция: {item.oem} / {item.brand}\n'
        f'Цена клиента/прайса: {customer_price:.2f}, текущая: '
        f'{offered_price:.2f}, '
        f'отклонение: {diff_pct:.2f}%'
    )
    await _notify_admins(
        session,
        title='Отклонение цены по заказу клиента',
        message=text,
        level=(
            AppNotificationLevel.ERROR
            if critical
            else AppNotificationLevel.WARNING
        ),
        link=f'/customer-orders/{order.id}',
    )


async def _send_reject_report(
    session: AsyncSession,
    order: CustomerOrder,
    rejected_items: List[CustomerOrderItem],
):
    if not rejected_items:
        return
    result = await session.execute(
        select(Customer.name).where(Customer.id == order.customer_id)
    )
    customer_name = result.scalar()
    lines = [
        f'Клиент: {customer_name or order.customer_id}',
        f'Заказ: {order.order_number or order.id}',
        'Отказы:',
    ]
    for item in rejected_items:
        price = (
            item.requested_price
            if item.requested_price is not None
            else item.matched_price
        )
        price_value = float(price) if price is not None else 0.0
        price_text = f'{price_value:.2f}'
        qty = item.reject_qty or item.requested_qty
        name = item.name or ''
        lines.append(
            f'- {item.oem} / {item.brand} / {name} — '
            f'{qty} шт, {price_text}'
        )
    await _notify_admins(
        session,
        title='Отказы по заказу клиента',
        message='\n'.join(lines),
        level=AppNotificationLevel.WARNING,
        link=f'/customer-orders/{order.id}',
        commit=True,
    )


async def _resolve_pricelist_config(
    session: AsyncSession,
    config: CustomerOrderConfig,
) -> CustomerPriceListConfig | None:
    if not config.pricelist_config_id:
        return None
    return await crud_customer_pricelist_config.get_by_id(
        customer_id=config.customer_id,
        config_id=config.pricelist_config_id,
        session=session,
    )


async def _prepare_customer_order_context(
    session: AsyncSession,
    config: CustomerOrderConfig,
):
    brand_aliases = await _load_brand_alias_map(session)
    last_pricelist = await _load_latest_customer_pricelist(
        session, config.customer_id
    )
    expected_prices = (
        _build_expected_price_map(last_pricelist, brand_aliases)
        if last_pricelist
        else {}
    )
    pricelist_config = await _resolve_pricelist_config(session, config)
    offers = (
        await _build_current_offers(session, pricelist_config, brand_aliases)
        if pricelist_config
        else {}
    )
    return expected_prices, offers, brand_aliases, pricelist_config


async def _process_manual_rows(
    session: AsyncSession,
    config: CustomerOrderConfig,
    order: CustomerOrder,
    parsed_rows: List[ParsedOrderRow],
):
    (
        expected_prices,
        offers,
        brand_aliases,
        pricelist_config,
    ) = await _prepare_customer_order_context(session, config)

    order_items: List[CustomerOrderItem] = []
    supplier_items: Dict[int, List[CustomerOrderItem]] = {}
    stock_items: List[CustomerOrderItem] = []
    rejected_items: List[CustomerOrderItem] = []

    for row in parsed_rows:
        key = _normalize_key(row.oem, row.brand, brand_aliases)
        expected_price = expected_prices.get(key)
        offer = offers.get(key)
        requested_price = row.requested_price
        customer_price = _resolve_customer_target_price(
            expected_price, requested_price, offer
        )

        item = CustomerOrderItem(
            order_id=order.id,
            row_index=row.row_index,
            oem=row.oem,
            brand=row.brand,
            name=row.name,
            requested_qty=row.requested_qty,
            requested_price=requested_price,
            status=CUSTOMER_ORDER_ITEM_STATUS.NEW,
        )
        if not offer:
            item.status = CUSTOMER_ORDER_ITEM_STATUS.REJECTED
            item.ship_qty = 0
            item.reject_qty = row.requested_qty
            reason_code, reason_text = await _diagnose_missing_offer_reason(
                session,
                pricelist_config,
                row,
                brand_aliases,
            )
            _set_reject_reason(item, reason_code, reason_text)
        elif not customer_price or customer_price <= 0:
            item.status = CUSTOMER_ORDER_ITEM_STATUS.REJECTED
            item.ship_qty = 0
            item.reject_qty = row.requested_qty
            _set_reject_reason(
                item,
                'NO_TARGET_PRICE',
                'Не удалось определить целевую цену для строки заказа: '
                'нет цены в заказе и нет цены в клиентском прайсе.',
            )
        else:
            offered_price = offer.price
            diff_pct = _compute_price_diff_pct(
                customer_price, offered_price
            )
            item.price_diff_pct = diff_pct
            item.matched_price = offer.supplier_price

            warning_pct = config.price_warning_pct or 5.0
            tolerance_pct = config.price_tolerance_pct or 2.0
            if diff_pct > warning_pct:
                if offer.is_own_price:
                    item.status = CUSTOMER_ORDER_ITEM_STATUS.OWN_STOCK
                    await _send_price_warning(
                        session,
                        order,
                        item,
                        customer_price,
                        offered_price,
                        diff_pct,
                        critical=True,
                    )
                else:
                    item.status = CUSTOMER_ORDER_ITEM_STATUS.REJECTED
                    item.ship_qty = 0
                    item.reject_qty = row.requested_qty
                    _set_reject_reason(
                        item,
                        'PRICE_TOO_HIGH',
                        'Цена предложения выше допустимой для клиента: '
                        f'{offered_price:.2f} против {customer_price:.2f} '
                        f'({diff_pct:.2f}%).',
                    )
                    await _send_price_warning(
                        session,
                        order,
                        item,
                        customer_price,
                        offered_price,
                        diff_pct,
                        critical=True,
                    )
            else:
                if diff_pct > tolerance_pct:
                    await _send_price_warning(
                        session,
                        order,
                        item,
                        customer_price,
                        offered_price,
                        diff_pct,
                        critical=False,
                    )
                item.status = (
                    CUSTOMER_ORDER_ITEM_STATUS.OWN_STOCK
                    if offer.is_own_price
                    else CUSTOMER_ORDER_ITEM_STATUS.SUPPLIER
                )

            if item.status != CUSTOMER_ORDER_ITEM_STATUS.REJECTED:
                available_qty = int(offer.quantity or 0)
                if available_qty < 0:
                    available_qty = 0
                ship_qty = min(row.requested_qty, available_qty)
                reject_qty = max(row.requested_qty - ship_qty, 0)
                item.ship_qty = ship_qty
                item.reject_qty = reject_qty
                item.autopart_id = offer.autopart_id
                if offer.is_own_price:
                    item.supplier_id = None
                else:
                    item.supplier_id = offer.provider_id
                if ship_qty == 0:
                    item.status = CUSTOMER_ORDER_ITEM_STATUS.REJECTED
                    _set_reject_reason(
                        item,
                        'ZERO_STOCK',
                        'Предложение найдено, но доступный остаток равен 0.',
                    )
                elif reject_qty > 0:
                    _set_reject_reason(
                        item,
                        'PARTIAL_STOCK',
                        'Частичная отгрузка: доступно '
                        f'{ship_qty} из {row.requested_qty}.',
                    )
                else:
                    _clear_reject_reason(item)

        order_items.append(item)
        session.add(item)
        await session.flush()

        if (
            item.status == CUSTOMER_ORDER_ITEM_STATUS.OWN_STOCK
            and item.ship_qty
        ):
            stock_items.append(item)
        elif (
            item.status == CUSTOMER_ORDER_ITEM_STATUS.SUPPLIER
            and item.ship_qty
        ):
            supplier_items.setdefault(item.supplier_id, []).append(item)
        elif item.status == CUSTOMER_ORDER_ITEM_STATUS.REJECTED:
            rejected_items.append(item)

    if stock_items:
        stock_order = StockOrder(
            customer_id=config.customer_id,
            status=STOCK_ORDER_STATUS.NEW,
        )
        session.add(stock_order)
        await session.flush()
        for item in stock_items:
            session.add(
                StockOrderItem(
                    stock_order_id=stock_order.id,
                    customer_order_item_id=item.id,
                    autopart_id=item.autopart_id,
                    quantity=item.ship_qty or 0,
                )
            )

    for provider_id, items in supplier_items.items():
        order_stmt = (
            select(SupplierOrder)
            .where(
                SupplierOrder.provider_id == provider_id,
                SupplierOrder.status.in_(
                    [
                        SUPPLIER_ORDER_STATUS.NEW,
                        SUPPLIER_ORDER_STATUS.SCHEDULED,
                    ]
                ),
            )
            .order_by(SupplierOrder.created_at.asc())
            .limit(1)
        )
        supplier_order = (
            await session.execute(order_stmt)
        ).scalar_one_or_none()
        if not supplier_order:
            supplier_order = SupplierOrder(
                provider_id=provider_id,
                status=SUPPLIER_ORDER_STATUS.NEW,
            )
            session.add(supplier_order)
            await session.flush()
        for item in items:
            session.add(
                SupplierOrderItem(
                    supplier_order_id=supplier_order.id,
                    customer_order_item_id=item.id,
                    autopart_id=item.autopart_id,
                    oem_number=item.oem,
                    brand_name=item.brand,
                    autopart_name=item.name,
                    quantity=item.ship_qty or 0,
                    price=item.matched_price,
                )
            )

    order.status = CUSTOMER_ORDER_STATUS.PROCESSED
    order.processed_at = now_moscow()
    await session.commit()
    return order_items, rejected_items


def _build_order_response_buffer(
    file_ext: str,
    file_buffer,
    config: CustomerOrderConfig,
    order_items: List[CustomerOrderItem],
):
    if file_ext == 'csv':
        return _apply_response_updates_csv(file_buffer, config, order_items)
    return _apply_response_updates_excel(file_buffer, config, order_items)


async def _write_order_response_file(
    order: CustomerOrder,
    filename: str,
    file_ext: str,
    response_buffer: BytesIO,
) -> tuple[str, str]:
    os.makedirs(ORDERS_RESPONSE_DIR, exist_ok=True)
    if file_ext == 'xls':
        base = filename.rsplit('.', 1)[0]
        response_name = f'{base}.xlsx'
    else:
        response_name = filename
    response_path = os.path.join(
        ORDERS_RESPONSE_DIR,
        f'{order.id}_{response_name}',
    )
    async with aiofiles.open(response_path, 'wb') as f:
        await f.write(response_buffer.getvalue())
    order.response_file_path = response_path
    order.response_file_name = response_name
    return response_path, response_name


async def _send_order_response_email(
    session: AsyncSession,
    config: CustomerOrderConfig,
    order: CustomerOrder,
    attachment_bytes: Optional[bytes] = None,
) -> None:
    if not _customer_order_auto_reply_enabled():
        logger.info(
            'Skipping automatic order response email for order_id=%s '
            'because CUSTOMER_ORDER_AUTO_REPLY_ENABLED is disabled',
            order.id,
        )
        order.status = CUSTOMER_ORDER_STATUS.PROCESSED
        order.error_details = None
        await session.commit()
        return

    if attachment_bytes is None:
        if not order.response_file_path or not os.path.isfile(
            order.response_file_path
        ):
            raise ValueError('Response file is missing')
        async with aiofiles.open(order.response_file_path, 'rb') as f:
            attachment_bytes = await f.read()
    if not attachment_bytes:
        raise ValueError('Response file is empty')
    if not order.response_file_name:
        raise ValueError('Response filename is missing')

    original_recipients = _build_order_reply_recipients(
        order.source_email,
        config,
        use_override=False,
    )
    to_email = _build_order_reply_recipients(order.source_email, config)
    if not to_email:
        raise ValueError('No recipients for order response')
    override_email = _customer_order_reply_override_email()

    try:
        out_account = await _get_out_account(session, 'orders_out')
        kwargs = {}
        if out_account:
            kwargs = build_email_delivery_kwargs(out_account)
        body = 'Во вложении файл с подтвержденными количествами.'
        if override_email:
            original_recipients_label = (
                original_recipients or 'не определены'
            )
            body = (
                'Заглушка ответа по заказу. Письмо отправлено только на '
                f'{override_email} для ручной сверки.\n'
                f'Исходные адресаты: {original_recipients_label}\n'
                f'Письмо-источник: {order.source_email or "не определено"}\n\n'
                'Во вложении файл с подтвержденными количествами.'
            )
        await _send_email_attachment_async(
            to_email,
            f'Ответ по заказу {order.order_number or order.id}',
            body,
            attachment_bytes,
            order.response_file_name,
            False,
            **kwargs,
        )
        order.status = (
            CUSTOMER_ORDER_STATUS.PROCESSED
            if override_email
            else CUSTOMER_ORDER_STATUS.SENT
        )
        order.error_details = None
    except Exception as exc:
        logger.error(
            'Failed to send order response: %s', exc, exc_info=True
        )
        _mark_order_error(order, f'Ошибка отправки ответа: {exc}')
    await session.commit()


async def _is_customer_order_ready_for_response(
    session: AsyncSession,
    order: CustomerOrder,
) -> bool:
    if not order.items:
        return False

    customer_order_item_ids = [int(item.id) for item in (order.items or [])]
    if not customer_order_item_ids:
        return False

    supplier_rows = (
        await session.execute(
            select(SupplierOrderItem).where(
                SupplierOrderItem.customer_order_item_id.in_(
                    customer_order_item_ids
                )
            )
        )
    ).scalars().all()
    stock_rows = (
        await session.execute(
            select(StockOrderItem).where(
                StockOrderItem.customer_order_item_id.in_(
                    customer_order_item_ids
                )
            )
        )
    ).scalars().all()

    supplier_by_customer_item = {
        int(row.customer_order_item_id): row
        for row in supplier_rows
        if row.customer_order_item_id is not None
    }
    stock_by_customer_item = {
        int(row.customer_order_item_id): row
        for row in stock_rows
        if row.customer_order_item_id is not None
    }

    for item in order.items or []:
        requested_qty = max(int(item.requested_qty or 0), 0)
        status_value = item.status

        if status_value == CUSTOMER_ORDER_ITEM_STATUS.REJECTED:
            continue

        if status_value == CUSTOMER_ORDER_ITEM_STATUS.OWN_STOCK:
            stock_item = stock_by_customer_item.get(int(item.id))
            if stock_item is None:
                return False
            expected_qty = max(
                int(stock_item.quantity or requested_qty),
                requested_qty,
            )
            picked_qty = int(stock_item.picked_quantity or 0)
            if picked_qty < expected_qty:
                return False
            continue

        if status_value == CUSTOMER_ORDER_ITEM_STATUS.SUPPLIER:
            supplier_item = supplier_by_customer_item.get(int(item.id))
            if supplier_item is None:
                return False
            received_qty = int(supplier_item.received_quantity or 0)
            confirmed_qty = supplier_item.confirmed_quantity
            if received_qty >= requested_qty:
                continue
            if confirmed_qty is not None:
                continue
            return False

        return False

    return True


async def try_finalize_customer_order_response(
    session: AsyncSession,
    *,
    order_id: int,
) -> bool:
    order = (
        await session.execute(
            select(CustomerOrder)
            .options(selectinload(CustomerOrder.items))
            .where(CustomerOrder.id == order_id)
        )
    ).scalar_one_or_none()
    if not order:
        return False
    if order.status == CUSTOMER_ORDER_STATUS.SENT:
        return False
    if not order.order_config_id:
        return False
    if not order.response_file_path or not os.path.isfile(
        order.response_file_path
    ):
        return False
    if not _customer_order_auto_reply_enabled():
        return False
    if not await _is_customer_order_ready_for_response(session, order):
        return False

    config = await crud_customer_order_config.get_by_id(
        session=session,
        config_id=order.order_config_id,
    )
    if not config:
        return False

    await _send_order_response_email(session, config, order)
    return order.status in (
        CUSTOMER_ORDER_STATUS.SENT,
        CUSTOMER_ORDER_STATUS.PROCESSED,
    )


async def _complete_imported_order_processing(
    session: AsyncSession,
    config: CustomerOrderConfig,
    order: CustomerOrder,
    parsed_rows: List[ParsedOrderRow],
    file_buffer,
    file_ext: str,
    filename: str,
    total_amount: Optional[float],
) -> CustomerOrder:
    order_items, rejected_items = await _process_manual_rows(
        session, config, order, parsed_rows
    )
    response_buffer = _build_order_response_buffer(
        file_ext, file_buffer, config, order_items
    )
    await _write_order_response_file(
        order, filename, file_ext, response_buffer
    )
    order.status = CUSTOMER_ORDER_STATUS.PROCESSED
    order.error_details = None
    await session.commit()

    await _send_order_import_notification(
        session,
        config,
        order.source_email or '',
        order.source_subject,
        filename,
        success=True,
        order_number=order.order_number,
        total_amount=total_amount,
        rows_count=len(parsed_rows),
    )
    await _send_reject_report(session, order, rejected_items)
    if _customer_order_auto_reply_enabled():
        sent = await try_finalize_customer_order_response(
            session,
            order_id=order.id,
        )
        if not sent:
            logger.info(
                (
                    "Automatic customer order response deferred: "
                    "order_id=%s not ready yet"
                ),
                order.id,
            )
    else:
        logger.info(
            'Automatic customer order response email disabled; '
            'response file kept for order_id=%s',
            order.id,
        )
    return order


def _apply_matched_email_state_for_configs(
    session: AsyncSession,
    configs: list[CustomerOrderConfig],
    msg,
    inbox_account,
) -> None:
    account_id = inbox_account.id if inbox_account else None
    seen_ids: set[int] = set()
    for config in configs:
        config_id = int(getattr(config, 'id', 0) or 0)
        if config_id and config_id in seen_ids:
            continue
        if config_id:
            seen_ids.add(config_id)
        _advance_config_last_uid(
            config,
            getattr(msg, 'uid', None),
            getattr(msg, 'folder_name', None),
            account_id,
        )
        session.add(config)

    if _mark_inbox_account_received_at(
        inbox_account, getattr(msg, 'received_at', None)
    ):
        session.add(inbox_account)


def _apply_matched_email_state(
    session: AsyncSession,
    config: CustomerOrderConfig,
    msg,
    inbox_account,
) -> None:
    _apply_matched_email_state_for_configs(
        session,
        [config],
        msg,
        inbox_account,
    )


async def _create_import_order_stub(
    session: AsyncSession,
    config: CustomerOrderConfig,
    sender: str,
    msg,
    filename: str,
    file_hash: str,
    order_number: Optional[str] = None,
    order_date: Optional[date] = None,
) -> CustomerOrder:
    order = CustomerOrder(
        customer_id=config.customer_id,
        order_config_id=config.id,
        status=CUSTOMER_ORDER_STATUS.NEW,
        received_at=getattr(msg, 'received_at', None) or now_moscow(),
        source_email=sender,
        source_uid=_safe_uid_as_int(getattr(msg, 'uid', None)),
        source_subject=getattr(msg, 'subject', None),
        source_filename=filename,
        order_number=order_number,
        order_date=order_date,
        file_hash=file_hash,
    )
    session.add(order)
    await session.commit()
    await session.refresh(order)
    return order


async def _ensure_import_order_stub(
    session: AsyncSession,
    *,
    order_id: Optional[int],
    config: CustomerOrderConfig,
    sender: str,
    msg,
    filename: str,
    file_hash: str,
    order_number: Optional[str] = None,
    order_date: Optional[date] = None,
) -> CustomerOrder:
    if order_id is not None:
        order = await session.get(CustomerOrder, order_id)
        if order:
            if order_number and not order.order_number:
                order.order_number = order_number
            if order_date and not order.order_date:
                order.order_date = order_date
            return order
    return await _create_import_order_stub(
        session,
        config,
        sender,
        msg,
        filename,
        file_hash,
        order_number=order_number,
        order_date=order_date,
    )


async def _store_import_error(
    session: AsyncSession,
    config: CustomerOrderConfig,
    order: CustomerOrder,
    msg,
    inbox_account,
    file_bytes: bytes,
    configs_for_uid_update: Optional[list[CustomerOrderConfig]] = None,
    *,
    reason: str,
    total_amount: Optional[float] = None,
    rows_count: Optional[int] = None,
) -> CustomerOrder:
    _mark_order_error(order, reason)
    await _save_order_source_file(order, file_bytes)
    _apply_matched_email_state_for_configs(
        session,
        configs_for_uid_update or [config],
        msg,
        inbox_account,
    )
    session.add(order)
    await session.commit()
    await _send_order_import_notification(
        session,
        config,
        order.source_email or '',
        order.source_subject,
        order.source_filename,
        success=False,
        reason=reason,
        order_number=order.order_number,
        total_amount=total_amount,
        rows_count=rows_count,
    )
    return order


async def create_manual_customer_order(
    session: AsyncSession,
    customer_id: int,
    order_number: str | None,
    order_date,
    items: List[dict],
    auto_process: bool,
    order_config_id: int | None = None,
) -> CustomerOrder:
    cleaned_items: list[dict] = []
    for item in items or []:
        oem = (item.get('oem') or '').strip()
        brand = (item.get('brand') or '').strip()
        name = _repair_cp1251_mojibake(item.get('name'))
        try:
            quantity = int(item.get('quantity') or 0)
        except (TypeError, ValueError):
            quantity = 0
        price = item.get('price')
        if not oem or not brand or quantity <= 0:
            continue
        cleaned_items.append(
            {
                'oem': oem,
                'brand': brand,
                'name': name,
                'quantity': quantity,
                'price': price,
            }
        )
    if not cleaned_items:
        raise ValueError('Items list is empty')

    config = None
    if auto_process:
        if order_config_id is not None:
            config = await crud_customer_order_config.get_by_id(
                session=session, config_id=order_config_id
            )
            if not config or config.customer_id != customer_id:
                raise ValueError('Customer order config not found')
        else:
            configs = await crud_customer_order_config.list_by_customer_id(
                session=session, customer_id=customer_id
            )
            if not configs:
                raise ValueError(
                    'Customer order config not found for auto processing'
                )
            if len(configs) > 1:
                raise ValueError(
                    'Multiple configs found, choose one for processing'
                )
            config = configs[0]
        if not config.pricelist_config_id:
            raise ValueError(
                'Order config must be linked to a pricelist config'
            )

    order = CustomerOrder(
        customer_id=customer_id,
        status=CUSTOMER_ORDER_STATUS.NEW,
        received_at=now_moscow(),
        source_email='manual',
        order_number=order_number,
        order_date=order_date or now_moscow().date(),
    )
    session.add(order)
    await session.flush()

    if auto_process:
        parsed_rows = []
        for idx, item in enumerate(cleaned_items, start=1):
            parsed_rows.append(
                ParsedOrderRow(
                    row_index=idx,
                    oem=item['oem'].strip(),
                    brand=item['brand'].strip(),
                    name=(item.get('name') or '').strip() or None,
                    requested_qty=int(item['quantity']),
                    requested_price=item.get('price'),
                )
            )
        order_items, rejected_items = await _process_manual_rows(
            session, config, order, parsed_rows
        )
        if rejected_items:
            await _send_reject_report(session, order, rejected_items)
        await session.refresh(order)
        return order

    for idx, item in enumerate(cleaned_items, start=1):
        session.add(
            CustomerOrderItem(
                order_id=order.id,
                row_index=idx,
                oem=item['oem'].strip(),
                brand=item['brand'].strip(),
                name=(item.get('name') or '').strip() or None,
                requested_qty=int(item['quantity']),
                requested_price=item.get('price'),
                status=CUSTOMER_ORDER_ITEM_STATUS.NEW,
            )
        )
    await session.commit()
    await session.refresh(order)
    return order


async def process_manual_customer_order(
    session: AsyncSession,
    order_id: int,
) -> CustomerOrder:
    order = await crud_customer_order.get_by_id(
        session=session, order_id=order_id
    )
    if not order:
        raise LookupError('Order not found')
    if order.status != CUSTOMER_ORDER_STATUS.NEW:
        raise ValueError('Order already processed')
    config = await crud_customer_order_config.get_by_customer_id(
        session=session, customer_id=order.customer_id
    )
    if not config:
        raise ValueError('Customer order config not found')
    if not config.pricelist_config_id:
        raise ValueError('Order config must be linked to a pricelist config')

    existing_link = await session.execute(
        select(SupplierOrderItem.id)
        .join(CustomerOrderItem)
        .where(CustomerOrderItem.order_id == order.id)
        .limit(1)
    )
    if existing_link.scalar_one_or_none() is not None:
        raise ValueError('Order already has supplier items')

    if not order.items:
        raise ValueError('Order has no items')

    parsed_rows = []
    for idx, item in enumerate(order.items, start=1):
        parsed_rows.append(
            ParsedOrderRow(
                row_index=item.row_index or idx,
                oem=item.oem,
                brand=item.brand,
                name=item.name,
                requested_qty=item.requested_qty,
                requested_price=item.requested_price,
            )
        )

    await session.execute(
        delete(CustomerOrderItem).where(
            CustomerOrderItem.order_id == order.id
        )
    )
    await session.flush()
    order_items, rejected_items = await _process_manual_rows(
        session, config, order, parsed_rows
    )
    if rejected_items:
        await _send_reject_report(session, order, rejected_items)
    await session.refresh(order)
    return order


async def retry_customer_order(
    session: AsyncSession,
    order_id: int,
) -> CustomerOrder:
    order = await crud_customer_order.get_by_id(
        session=session, order_id=order_id
    )
    if not order:
        raise LookupError('Order not found')
    if order.status != CUSTOMER_ORDER_STATUS.ERROR:
        raise ValueError('Only errored orders can be retried')
    if not order.order_config_id:
        raise ValueError('Order is not linked to an order config')

    config = await crud_customer_order_config.get_by_id(
        session=session, config_id=order.order_config_id
    )
    if not config:
        raise ValueError('Order config not found')

    if order.response_file_path and os.path.isfile(order.response_file_path):
        if _customer_order_auto_reply_enabled():
            sent = await try_finalize_customer_order_response(
                session,
                order_id=order.id,
            )
            if not sent:
                order.status = CUSTOMER_ORDER_STATUS.PROCESSED
                order.error_details = None
                session.add(order)
                await session.commit()
        else:
            order.status = CUSTOMER_ORDER_STATUS.PROCESSED
            order.error_details = None
            session.add(order)
            await session.commit()
        await session.refresh(order)
        if order.status == CUSTOMER_ORDER_STATUS.ERROR:
            raise ValueError(
                order.error_details or 'Не удалось повторно отправить ответ'
            )
        return order

    source_bytes = await _load_order_source_file(order)
    if not source_bytes:
        reason = (
            'Исходный файл заказа больше недоступен. '
            'Проверьте письмо вручную и загрузите заказ заново.'
        )
        _mark_order_error(order, reason)
        session.add(order)
        await session.commit()
        raise ValueError(reason)

    filename = order.source_filename or 'order.xlsx'
    try:
        (
            parsed_rows,
            order_date,
            order_number_file,
            file_buffer,
            file_ext,
        ) = _parse_order_attachment(source_bytes, filename, config)
    except Exception as exc:
        reason = f'Ошибка повторного разбора файла: {exc}'
        _mark_order_error(order, reason)
        session.add(order)
        await session.commit()
        raise ValueError(reason) from exc

    if order.order_date is None and order_date:
        order.order_date = order_date
    if not order.order_number and order_number_file:
        order.order_number = order_number_file

    if order.items:
        try:
            response_buffer = _build_order_response_buffer(
                file_ext, file_buffer, config, order.items
            )
            await _write_order_response_file(
                order, filename, file_ext, response_buffer
            )
            order.status = CUSTOMER_ORDER_STATUS.PROCESSED
            order.error_details = None
            session.add(order)
            await session.commit()
            if _customer_order_auto_reply_enabled():
                sent = await try_finalize_customer_order_response(
                    session,
                    order_id=order.id,
                )
                if not sent:
                    order.status = CUSTOMER_ORDER_STATUS.PROCESSED
                    order.error_details = None
                    session.add(order)
                    await session.commit()
            else:
                order.status = CUSTOMER_ORDER_STATUS.PROCESSED
                order.error_details = None
                session.add(order)
                await session.commit()
        except Exception as exc:
            reason = f'Ошибка повторной отправки ответа: {exc}'
            _mark_order_error(order, reason)
            session.add(order)
            await session.commit()
            raise ValueError(reason) from exc
        await session.refresh(order)
        if order.status == CUSTOMER_ORDER_STATUS.ERROR:
            raise ValueError(
                order.error_details or 'Не удалось повторно отправить ответ'
            )
        return order

    requested_total = _compute_order_requested_total(parsed_rows)
    if not parsed_rows:
        reason = 'Не удалось распознать строки заказа'
        _mark_order_error(order, reason)
        session.add(order)
        await session.commit()
        raise ValueError(reason)

    order.status = CUSTOMER_ORDER_STATUS.NEW
    order.processed_at = None
    order.error_details = None
    session.add(order)
    await session.commit()

    try:
        await _complete_imported_order_processing(
            session,
            config,
            order,
            parsed_rows,
            file_buffer,
            file_ext,
            filename,
            requested_total,
        )
    except Exception as exc:
        await session.rollback()
        order = await session.get(CustomerOrder, order.id)
        if not order:
            raise ValueError('Order not found after retry rollback') from exc
        reason = f'Ошибка повторной обработки: {exc}'
        _mark_order_error(order, reason)
        session.add(order)
        await session.commit()
        raise ValueError(reason) from exc

    await session.refresh(order)
    if order.status == CUSTOMER_ORDER_STATUS.ERROR:
        raise ValueError(
            order.error_details or 'Не удалось завершить повторную обработку'
        )
    return order


async def retry_customer_order_errors_for_config(
    session: AsyncSession,
    config_id: int,
) -> Dict[str, int]:
    config = await crud_customer_order_config.get_by_id(
        session=session, config_id=config_id
    )
    if not config:
        raise LookupError('Config not found')

    result = await session.execute(
        select(CustomerOrder.id)
        .where(
            CustomerOrder.order_config_id == config_id,
            CustomerOrder.status == CUSTOMER_ORDER_STATUS.ERROR,
        )
        .order_by(CustomerOrder.received_at.asc(), CustomerOrder.id.asc())
    )
    order_ids = list(result.scalars().all())
    stats = {
        'config_id': config_id,
        'total': len(order_ids),
        'retried': 0,
        'succeeded': 0,
        'failed': 0,
    }
    for order_id in order_ids:
        stats['retried'] += 1
        try:
            await retry_customer_order(session, order_id)
            stats['succeeded'] += 1
        except Exception as exc:
            logger.warning(
                'Retry failed for customer order %s: %s',
                order_id,
                exc,
            )
            stats['failed'] += 1
    return stats


async def create_manual_supplier_order(
    session: AsyncSession,
    provider_id: int,
    items: List[dict],
    created_by_user_id: int | None = None,
) -> SupplierOrder:
    provider = await session.get(Provider, provider_id)
    if not provider:
        raise ValueError('Supplier not found')
    cleaned_items: list[dict] = []
    for item in items or []:
        autopart_id = item.get('autopart_id')
        oem = (item.get('oem') or '').strip()
        brand = (item.get('brand') or '').strip()
        try:
            quantity = int(item.get('quantity') or 0)
        except (TypeError, ValueError):
            quantity = 0
        price_raw = item.get('price')
        try:
            price_value = float(price_raw) if price_raw is not None else None
        except (TypeError, ValueError):
            price_value = None
        if not oem or not brand or quantity <= 0:
            continue
        cleaned_items.append(
            {
                'autopart_id': autopart_id,
                'oem': oem,
                'brand': brand,
                'name': (item.get('name') or '').strip() or None,
                'quantity': quantity,
                'price': price_value,
                'min_delivery_day': item.get('min_delivery_day'),
                'max_delivery_day': item.get('max_delivery_day'),
            }
        )
    if not cleaned_items:
        raise ValueError('Items list is empty')

    supplier_order = SupplierOrder(
        provider_id=provider_id,
        status=SUPPLIER_ORDER_STATUS.NEW,
        source_type=ORDER_TRACKING_SOURCE.SEARCH_OFFERS.value,
        created_by_user_id=created_by_user_id,
    )
    session.add(supplier_order)
    await session.flush()

    for item in cleaned_items:
        autopart_id = item.get('autopart_id')
        oem = item['oem']
        brand = item['brand']
        name = item.get('name')
        brand_key = brand.lower()
        quantity = item['quantity']
        autopart = None
        if autopart_id is not None:
            autopart = await session.get(AutoPart, autopart_id)
        if autopart is None:
            autopart_stmt = (
                select(AutoPart)
                .join(Brand)
                .where(
                    AutoPart.oem_number == oem,
                    func.lower(Brand.name) == brand_key,
                )
                .limit(1)
            )
            autopart = (
                await session.execute(autopart_stmt)
            ).scalar_one_or_none()
        price_value = item.get('price')
        if price_value is None and autopart:
            price_stmt = (
                select(PriceListAutoPartAssociation.price)
                .join(PriceList)
                .where(
                    PriceList.provider_id == provider_id,
                    PriceListAutoPartAssociation.autopart_id == autopart.id,
                )
                .order_by(
                    PriceList.date.desc().nullslast(),
                    PriceList.id.desc(),
                )
                .limit(1)
            )
            price_value = (
                (await session.execute(price_stmt)).scalar_one_or_none()
            ) or 0.0
        elif price_value is None:
            price_value = 0.0

        session.add(
            SupplierOrderItem(
                supplier_order_id=supplier_order.id,
                customer_order_item_id=None,
                autopart_id=autopart.id if autopart else None,
                oem_number=oem,
                brand_name=brand,
                autopart_name=name or (autopart.name if autopart else None),
                quantity=quantity,
                price=price_value,
                min_delivery_day=item.get('min_delivery_day'),
                max_delivery_day=item.get('max_delivery_day'),
            )
        )

    await session.commit()
    await session.refresh(supplier_order)
    return supplier_order


async def process_customer_orders(
    session: AsyncSession,
    customer_id: Optional[int] = None,
    config_id: Optional[int] = None,
) -> None:
    order_accounts = await crud_email_account.get_active_by_purpose(
        session, 'orders_in'
    )
    if not order_accounts and (
        not EMAIL_NAME_ORDER
        or not EMAIL_PASSWORD_ORDER
        or not EMAIL_HOST_ORDER
    ):
        logger.warning('Order email credentials are not configured.')
        return

    config_stmt = (
        select(CustomerOrderConfig)
        .where(CustomerOrderConfig.is_active.is_(True))
        .options(joinedload(CustomerOrderConfig.customer))
    )
    if customer_id is not None:
        config_stmt = config_stmt.where(
            CustomerOrderConfig.customer_id == customer_id
        )
    if config_id is not None:
        config_stmt = config_stmt.where(CustomerOrderConfig.id == config_id)
    configs = await session.execute(config_stmt)
    configs = configs.scalars().all()
    if not configs:
        logger.info('No active customer order configs found.')
        return

    specific_account_ids: set[int] = set()
    for cfg in configs:
        specific_account_ids.update(_get_config_email_account_ids(cfg))
    if order_accounts and specific_account_ids:
        order_accounts = [
            account
            for account in order_accounts
            if account.id in specific_account_ids
        ]

    config_by_email: dict[str, list[CustomerOrderConfig]] = {}
    global_sender_filter: set[str] = set()
    account_sender_filter: dict[int, set[str]] = {}
    global_configs_for_uid: list[CustomerOrderConfig] = []
    account_configs_for_uid: dict[int, list[CustomerOrderConfig]] = {}
    for config in configs:
        emails = _normalize_email_list(config.order_emails)
        if config.order_email:
            emails.append(config.order_email.lower())
        config_account_ids = _get_config_email_account_ids(config)
        for email in emails:
            config_by_email.setdefault(email, []).append(config)
            if not config_account_ids:
                global_sender_filter.add(email)
            else:
                for account_id in config_account_ids:
                    account_sender_filter.setdefault(
                        account_id, set()
                    ).add(email)
        if not config_account_ids:
            global_configs_for_uid.append(config)
        else:
            for account_id in config_account_ids:
                account_configs_for_uid.setdefault(account_id, []).append(
                    config
                )

    # Для автоматического шедулера ускоряем IMAP-запросы по UID.
    # В ручных запусках (customer_id/config_id) сохраняем старую
    # дату-поиска, чтобы можно было импортировать исторические письма.
    use_uid_optimization = customer_id is None and config_id is None

    inbox_settings = await crud_customer_order_inbox_settings.get_or_create(
        session
    )
    lookback_days = max(1, int(inbox_settings.lookback_days or 1))
    mark_seen = bool(inbox_settings.mark_seen)
    date_from = now_moscow().date() - timedelta(days=lookback_days - 1)

    messages: list[tuple[object, Optional[object]]] = []
    if order_accounts:
        unique_accounts = {}
        for account in order_accounts:
            host = (
                account.imap_host or EMAIL_HOST_ORDER or ''
            ).strip().lower()
            folders = tuple(
                folder.casefold()
                for folder in resolve_imap_folders(
                    account.imap_folder,
                    getattr(account, 'imap_additional_folders', None),
                    default=EMAIL_FOLDER_ORDER or DEFAULT_IMAP_FOLDER,
                )
            )
            port = account.imap_port or IMAP_SERVER
            key = (account.email.strip().lower(), host, folders, port)
            if key not in unique_accounts:
                unique_accounts[key] = account
        order_accounts = list(unique_accounts.values())

        for account in order_accounts:
            host = account.imap_host or EMAIL_HOST_ORDER
            transport = (account.transport or 'smtp').strip().lower()
            folders = resolve_imap_folders(
                account.imap_folder,
                getattr(account, 'imap_additional_folders', None),
                default=EMAIL_FOLDER_ORDER or DEFAULT_IMAP_FOLDER,
            )
            allowed_senders = set(global_sender_filter)
            allowed_senders.update(
                account_sender_filter.get(int(account.id), set())
            )
            folder_uid_floor: dict[str, int] = {}
            if use_uid_optimization:
                uid_configs = list(global_configs_for_uid)
                uid_configs.extend(
                    account_configs_for_uid.get(int(account.id), [])
                )
                for folder in folders:
                    normalized_folder = normalize_imap_folder(folder)
                    floor_uid = 0
                    for cfg in uid_configs:
                        floor_uid = max(
                            floor_uid,
                            _get_config_last_uid(
                                cfg,
                                normalized_folder,
                                account_id=account.id,
                            ),
                        )
                    if floor_uid > 0:
                        folder_uid_floor[folder] = floor_uid
            if transport == 'resend_api':
                try:
                    account_messages = await _fetch_resend_messages(
                        account,
                        date_from,
                    )
                    fetched_count = len(account_messages)
                    account_messages = _filter_messages_by_senders(
                        account_messages, allowed_senders
                    )
                    logger.debug(
                        'Order inbox %s transport=%s fetched=%s '
                        'matched_sender=%s',
                        account.email,
                        transport,
                        fetched_count,
                        len(account_messages),
                    )
                    messages.extend(
                        [(msg, account) for msg in account_messages]
                    )
                except Exception as exc:
                    logger.error(
                        'Order inbox fetch failed for Resend %s: %s',
                        account.email,
                        exc,
                        exc_info=True,
                    )
                continue
            if account.oauth_provider == 'google':
                try:
                    account_messages = []
                    for label in folders:
                        account_messages.extend(
                            await _fetch_gmail_messages(
                                account, date_from, label=label
                            )
                        )
                    fetched_count = len(account_messages)
                    account_messages = _filter_messages_by_senders(
                        account_messages, allowed_senders
                    )
                    logger.debug(
                        'Order inbox %s transport=%s fetched=%s '
                        'matched_sender=%s',
                        account.email,
                        transport,
                        fetched_count,
                        len(account_messages),
                    )
                    messages.extend(
                        [(msg, account) for msg in account_messages]
                    )
                except Exception as exc:
                    logger.error(
                        'Order inbox fetch failed for %s: %s',
                        account.email,
                        exc,
                        exc_info=True,
                    )
                continue
            if not host:
                continue
            try:
                account_messages = []
                for folder in folders:
                    try:
                        account_messages.extend(
                            await _fetch_order_messages(
                                host,
                                account.email,
                                account.password,
                                folder,
                                date_from,
                                mark_seen,
                                last_uid=folder_uid_floor.get(folder),
                                port=account.imap_port or IMAP_SERVER,
                                ssl=True,
                            )
                        )
                    except MailboxFolderSelectError as folder_exc:
                        logger.warning(
                            'IMAP папка "%s" не найдена для %s — пропускаем. '
                            'Проверьте название папки в настройках аккаунта. '
                            'Ошибка: %s',
                            folder,
                            account.email,
                            folder_exc,
                        )
                fetched_count = len(account_messages)
                account_messages = _filter_messages_by_senders(
                    account_messages, allowed_senders
                )
                logger.debug(
                    'Order inbox %s transport=imap fetched=%s '
                    'matched_sender=%s',
                    account.email,
                    fetched_count,
                    len(account_messages),
                )
                messages.extend([(msg, account) for msg in account_messages])
            except Exception as exc:
                if _is_too_many_connections_error(exc):
                    logger.warning(
                        'Order inbox fetch throttled for %s: %s',
                        account.email,
                        exc,
                    )
                else:
                    logger.error(
                        'Order inbox fetch failed for %s: %s',
                        account.email,
                        exc,
                        exc_info=True,
                    )
                continue
    else:
        try:
            fallback_last_uid = None
            if use_uid_optimization:
                normalized_folder = normalize_imap_folder(EMAIL_FOLDER_ORDER)
                uid_floor = 0
                for cfg in configs:
                    uid_floor = max(
                        uid_floor,
                        _get_config_last_uid(cfg, normalized_folder),
                    )
                if uid_floor > 0:
                    fallback_last_uid = uid_floor
            fallback_messages = await _fetch_order_messages(
                EMAIL_HOST_ORDER,
                EMAIL_NAME_ORDER,
                EMAIL_PASSWORD_ORDER,
                EMAIL_FOLDER_ORDER,
                date_from,
                mark_seen,
                last_uid=fallback_last_uid,
                port=IMAP_SERVER,
                ssl=True,
            )
            fetched_count = len(fallback_messages)
            fallback_messages = _filter_messages_by_senders(
                fallback_messages, global_sender_filter
            )
            logger.debug(
                'Fallback order inbox fetched=%s matched_sender=%s',
                fetched_count,
                len(fallback_messages),
            )
            messages = [(msg, None) for msg in fallback_messages]
        except Exception as exc:
            logger.error(
                'Order inbox fetch failed for fallback mailbox: %s',
                exc,
                exc_info=True,
            )
            return

    if not messages:
        logger.info('No order emails found.')
        return

    messages = _dedupe_order_messages(messages)
    messages.sort(key=_message_sort_key)

    logger.debug(
        'Получено %d писем-кандидатов после фильтра по отправителю',
        len(messages),
    )

    for msg, inbox_account in messages:
        config = None
        attachment = None
        order = None
        order_id = None
        sender = _extract_email(msg.from_)
        filename = None
        file_bytes = b''
        file_hash = ''
        requested_total = None
        parsed_rows: List[ParsedOrderRow] | None = None
        order_number_hint = None
        try:
            configs_for_sender = config_by_email.get(sender) or []
            account_id = inbox_account.id if inbox_account else None
            logger.debug(
                'Processing order email sender=%s subject=%s account_id=%s '
                'attachments=%s uid=%s',
                sender,
                getattr(msg, 'subject', ''),
                account_id,
                len(getattr(msg, 'attachments', []) or []),
                getattr(msg, 'uid', None),
            )
            if not configs_for_sender:
                logger.debug(
                    'No order config found for sender=%s account_id=%s',
                    sender,
                    account_id,
                )
            candidate_configs = _pick_configs_for_account(
                configs_for_sender, account_id
            )
            configs_for_uid_update = list(candidate_configs)
            if configs_for_sender and not candidate_configs:
                logger.debug(
                    'Sender=%s matched configs=%s but none for account_id=%s',
                    sender,
                    [cfg.id for cfg in configs_for_sender],
                    account_id,
                )
            for candidate in candidate_configs:
                if not candidate.pricelist_config_id:
                    logger.warning(
                        'Order config %s has no pricelist_config_id; skip',
                        candidate.id,
                    )
                    continue
                msg_uid_int = _safe_uid_as_int(msg.uid)
                folder_last_uid = _get_config_last_uid(
                    candidate,
                    getattr(msg, 'folder_name', None),
                    account_id=account_id,
                )
                if (
                    msg_uid_int is not None
                    and msg_uid_int <= folder_last_uid
                ):
                    logger.debug(
                        'Skip order config %s for sender=%s: '
                        'msg_uid=%s <= last_uid=%s folder=%s',
                        candidate.id,
                        sender,
                        msg_uid_int,
                        folder_last_uid,
                        getattr(msg, 'folder_name', None),
                    )
                    continue
                if not _match_pattern(
                    candidate.order_subject_pattern, msg.subject
                ):
                    logger.debug(
                        'Skip order config %s for sender=%s: '
                        'subject mismatch pattern=%r subject=%r',
                        candidate.id,
                        sender,
                        candidate.order_subject_pattern,
                        getattr(msg, 'subject', None),
                    )
                    continue
                candidate_attachment = None
                for att in msg.attachments:
                    if _match_pattern(
                        candidate.order_filename_pattern, att.filename
                    ):
                        candidate_attachment = att
                        break
                if candidate_attachment is None and msg.attachments:
                    candidate_attachment = msg.attachments[0]
                if candidate_attachment is None:
                    logger.debug(
                        'Skip order config %s for sender=%s: '
                        'no suitable attachment filename_pattern=%r',
                        candidate.id,
                        sender,
                        candidate.order_filename_pattern,
                    )
                    continue
                config = candidate
                attachment = candidate_attachment
                break

            if not config or not attachment:
                if not msg.attachments:
                    logger.info(
                        'No attachment for order email uid=%s', msg.uid
                    )
                if (
                    inbox_account
                    and (inbox_account.transport or '').strip().lower()
                    == 'resend_api'
                    and msg.received_at
                ):
                    inbox_account.resend_last_received_at = msg.received_at
                    session.add(inbox_account)
                    await session.commit()
                continue

            filename = attachment.filename or 'order.xlsx'
            file_bytes = attachment.payload or b''
            file_hash = hashlib.sha256(file_bytes).hexdigest()
            body_text = msg.text or ''
            if not body_text and msg.html:
                body_text = _strip_html(msg.html)
            order_number_hint = _extract_order_number(
                config, msg.subject, filename, body_text
            )

            existing = await session.execute(
                select(CustomerOrder).where(
                    CustomerOrder.customer_id == config.customer_id,
                    CustomerOrder.file_hash == file_hash,
                )
            )
            existing_order = existing.scalars().first()
            if existing_order:
                _apply_matched_email_state_for_configs(
                    session,
                    configs_for_uid_update or [config],
                    msg,
                    inbox_account,
                )
                await _send_order_import_notification(
                    session,
                    config,
                    sender,
                    getattr(msg, 'subject', None),
                    filename,
                    success=False,
                    reason='Дубликат файла: заказ уже загружен ранее',
                    order_number=(
                        existing_order.order_number or order_number_hint
                    ),
                )
                await session.commit()
                continue

            order = await _create_import_order_stub(
                session,
                config,
                sender,
                msg,
                filename,
                file_hash,
                order_number=order_number_hint,
            )
            order_id = order.id

            try:
                (
                    parsed_rows,
                    order_date,
                    order_number_file,
                    file_buffer,
                    file_ext,
                ) = _parse_order_attachment(file_bytes, filename, config)
            except Exception as exc:
                logger.error(
                    'Failed to parse order email uid=%s: %s',
                    getattr(msg, 'uid', None),
                    exc,
                    exc_info=True,
                )
                order = await _ensure_import_order_stub(
                    session,
                    order_id=order_id,
                    config=config,
                    sender=sender,
                    msg=msg,
                    filename=filename,
                    file_hash=file_hash,
                    order_number=order_number_hint,
                )
                order_id = order.id
                await _store_import_error(
                    session,
                    config,
                    order,
                    msg,
                    inbox_account,
                    file_bytes,
                    configs_for_uid_update,
                    reason=(
                        str(exc)
                        if str(exc).startswith('Неподдерживаемый тип файла:')
                        else f'Ошибка разбора файла: {exc}'
                    ),
                )
                continue

            if order.order_date is None and order_date:
                order.order_date = order_date
            if not order.order_number:
                order.order_number = order_number_file or order_number_hint
            session.add(order)
            await session.commit()

            requested_total = _compute_order_requested_total(parsed_rows)
            if not parsed_rows:
                logger.info('No order rows found in %s', filename)
                await _store_import_error(
                    session,
                    config,
                    order,
                    msg,
                    inbox_account,
                    file_bytes,
                    configs_for_uid_update,
                    reason='Не удалось распознать строки заказа',
                    total_amount=requested_total,
                    rows_count=0,
                )
                continue

            try:
                await _complete_imported_order_processing(
                    session,
                    config,
                    order,
                    parsed_rows,
                    file_buffer,
                    file_ext,
                    filename,
                    requested_total,
                )
                _apply_matched_email_state_for_configs(
                    session,
                    configs_for_uid_update or [config],
                    msg,
                    inbox_account,
                )
                await session.commit()
            except Exception as exc:
                await session.rollback()
                logger.error(
                    'Failed to process order email uid=%s: %s',
                    getattr(msg, 'uid', None),
                    exc,
                    exc_info=True,
                )
                order = await _ensure_import_order_stub(
                    session,
                    order_id=order_id,
                    config=config,
                    sender=sender,
                    msg=msg,
                    filename=filename,
                    file_hash=file_hash,
                    order_number=order_number_file or order_number_hint,
                    order_date=order_date,
                )
                order_id = order.id
                await _store_import_error(
                    session,
                    config,
                    order,
                    msg,
                    inbox_account,
                    file_bytes,
                    configs_for_uid_update,
                    reason=f'Ошибка обработки: {exc}',
                    total_amount=requested_total,
                    rows_count=len(parsed_rows),
                )
        except Exception as exc:
            await session.rollback()
            if config and attachment and filename is not None:
                try:
                    order = await _ensure_import_order_stub(
                        session,
                        order_id=order_id,
                        config=config,
                        sender=sender,
                        msg=msg,
                        filename=filename,
                        file_hash=(
                            file_hash
                            or hashlib.sha256(file_bytes).hexdigest()
                        ),
                        order_number=order_number_hint,
                    )
                    order_id = order.id
                    await _store_import_error(
                        session,
                        config,
                        order,
                        msg,
                        inbox_account,
                        file_bytes,
                        configs_for_uid_update if config else None,
                        reason=f'Ошибка обработки: {exc}',
                        total_amount=requested_total,
                        rows_count=len(parsed_rows or []),
                    )
                except Exception as persist_exc:
                    await session.rollback()
                    logger.error(
                        'Failed to persist order import error uid=%s: %s',
                        getattr(msg, 'uid', None),
                        persist_exc,
                        exc_info=True,
                    )
            logger.error(
                'Failed to process order email uid=%s: %s',
                getattr(msg, 'uid', None),
                exc,
                exc_info=True,
            )


async def send_supplier_orders(
    session: AsyncSession,
    supplier_order_ids: List[int],
) -> Dict[str, int]:
    if not supplier_order_ids:
        return {'sent': 0, 'failed': 0}

    stmt = (
        select(SupplierOrder)
        .where(SupplierOrder.id.in_(supplier_order_ids))
        .options(
            joinedload(SupplierOrder.items)
            .joinedload(SupplierOrderItem.autopart)
            .joinedload(AutoPart.brand),
            joinedload(SupplierOrder.provider),
        )
    )
    result = await session.execute(stmt)
    orders = result.unique().scalars().all()
    sent = 0
    failed = 0

    account = await _get_out_account(session, 'orders_out')
    smtp_kwargs = {}
    if account:
        smtp_kwargs = build_email_delivery_kwargs(account)
    override_email = await _supplier_order_override_email_from_settings(
        session
    )

    for order in orders:
        provider = order.provider
        original_recipient = _build_supplier_order_recipient(
            provider,
            use_override=False,
        )
        to_email = override_email or original_recipient
        if not to_email:
            failed += 1
            continue

        rows, total_qty, total_sum = _build_supplier_order_rows(order)
        attachment_bytes = _build_supplier_order_attachment_bytes(
            order=order,
            rows=rows,
            total_qty=total_qty,
            total_sum=total_sum,
        )
        try:
            provider_name = (
                provider.name if provider and provider.name else 'Поставщик'
            )
            body = _build_supplier_order_body_html(
                order=order,
                rows=rows,
                total_qty=total_qty,
                total_sum=total_sum,
            )
            subject = f'Заказ поставщику № {order.id}'
            if override_email:
                original_recipient_label = (
                    original_recipient or 'не указан'
                )
                subject = f'[STUB] {subject}'
                body = (
                    '<p><b>Заглушка отправки заказа поставщику.</b></p>'
                    f'<p>Письмо отправлено только на '
                    f'{escape(override_email)} для ручной сверки.<br/>'
                    f'Поставщик: {escape(provider_name)}<br/>'
                    f'Исходный адресат: '
                    f'{escape(original_recipient_label)}</p>'
                    f'{body}'
                )
            await _send_email_attachment_async(
                to_email,
                subject,
                body,
                attachment_bytes,
                _supplier_order_attachment_filename(order.id),
                True,
                **smtp_kwargs,
            )
            order.status = SUPPLIER_ORDER_STATUS.SENT
            order.sent_at = now_moscow()
            sent += 1
        except Exception as exc:
            logger.error(
                'Failed to send supplier order %s: %s',
                order.id,
                exc,
                exc_info=True,
            )
            order.status = SUPPLIER_ORDER_STATUS.ERROR
            failed += 1

    await session.commit()
    return {'sent': sent, 'failed': failed}


async def send_scheduled_supplier_orders(
    session: AsyncSession,
    *,
    use_provider_schedule: bool = True,
) -> Dict[str, int]:
    if not use_provider_schedule:
        orders_stmt = select(SupplierOrder.id).where(
            SupplierOrder.status == SUPPLIER_ORDER_STATUS.NEW,
        )
        order_ids = [row[0] for row in (
            await session.execute(orders_stmt)
        ).all()]
        if not order_ids:
            return {'sent': 0, 'failed': 0}
        return await send_supplier_orders(session, order_ids)

    now = now_moscow()
    day_key = {
        0: 'mon',
        1: 'tue',
        2: 'wed',
        3: 'thu',
        4: 'fri',
        5: 'sat',
        6: 'sun',
    }[now.weekday()]
    time_key = now.strftime('%H:%M')

    providers_stmt = select(Provider).where(
        Provider.order_schedule_enabled.is_(True)
    )
    providers = (await session.execute(providers_stmt)).scalars().all()

    eligible_provider_ids = []
    for provider in providers:
        days = provider.order_schedule_days or []
        times = provider.order_schedule_times or []
        if days and day_key not in days:
            continue
        if times and time_key not in times:
            continue
        eligible_provider_ids.append(provider.id)

    if not eligible_provider_ids:
        return {'sent': 0, 'failed': 0}

    orders_stmt = (
        select(SupplierOrder.id)
        .where(
            SupplierOrder.provider_id.in_(eligible_provider_ids),
            SupplierOrder.status == SUPPLIER_ORDER_STATUS.NEW,
        )
    )
    order_ids = [row[0] for row in (await session.execute(orders_stmt)).all()]
    return await send_supplier_orders(session, order_ids)


async def update_customer_order_item_manual(
    session: AsyncSession,
    item_id: int,
    status: CUSTOMER_ORDER_ITEM_STATUS | None,
    supplier_id: int | None,
) -> CustomerOrderItem:
    stmt = (
        select(CustomerOrderItem)
        .options(joinedload(CustomerOrderItem.order))
        .where(CustomerOrderItem.id == item_id)
    )
    item = (await session.execute(stmt)).scalar_one_or_none()
    if not item:
        raise LookupError('Order item not found')

    target_status = status
    if target_status is None and supplier_id is not None:
        target_status = CUSTOMER_ORDER_ITEM_STATUS.SUPPLIER

    if target_status not in (
        CUSTOMER_ORDER_ITEM_STATUS.REJECTED,
        CUSTOMER_ORDER_ITEM_STATUS.SUPPLIER,
        CUSTOMER_ORDER_ITEM_STATUS.OWN_STOCK,
    ):
        raise ValueError('Unsupported status')

    if target_status == CUSTOMER_ORDER_ITEM_STATUS.SUPPLIER:
        if not supplier_id:
            raise ValueError('supplier_id is required')
        provider = await session.get(Provider, supplier_id)
        if not provider:
            raise ValueError('Supplier not found')
        if provider.is_own_price:
            target_status = CUSTOMER_ORDER_ITEM_STATUS.OWN_STOCK
            supplier_id = None
    if target_status == CUSTOMER_ORDER_ITEM_STATUS.OWN_STOCK:
        supplier_id = None

    existing_item = None
    existing_order = None
    existing_stmt = (
        select(SupplierOrderItem, SupplierOrder)
        .join(SupplierOrder)
        .where(SupplierOrderItem.customer_order_item_id == item.id)
    )
    existing_row = (await session.execute(existing_stmt)).first()
    if existing_row:
        existing_item, existing_order = existing_row

    existing_stock_item = None
    existing_stock_order = None
    existing_stock_stmt = (
        select(StockOrderItem, StockOrder)
        .join(StockOrder)
        .where(StockOrderItem.customer_order_item_id == item.id)
    )
    existing_stock_row = (await session.execute(existing_stock_stmt)).first()
    if existing_stock_row:
        existing_stock_item, existing_stock_order = existing_stock_row

    def _is_modifiable(order: SupplierOrder) -> bool:
        return order.status in (
            SUPPLIER_ORDER_STATUS.NEW,
            SUPPLIER_ORDER_STATUS.SCHEDULED,
        )

    def _is_stock_modifiable(order: StockOrder) -> bool:
        return order.status == STOCK_ORDER_STATUS.NEW

    if target_status == CUSTOMER_ORDER_ITEM_STATUS.REJECTED:
        if existing_stock_item and existing_stock_order:
            if not _is_stock_modifiable(existing_stock_order):
                raise ValueError('Stock order already closed')
            await session.delete(existing_stock_item)
            await session.flush()
            remaining_stmt = (
                select(StockOrderItem.id)
                .where(
                    StockOrderItem.stock_order_id == existing_stock_order.id
                )
                .limit(1)
            )
            remaining = (
                await session.execute(remaining_stmt)
            ).scalar_one_or_none()
            if remaining is None:
                await session.delete(existing_stock_order)
        if existing_item and existing_order:
            if not _is_modifiable(existing_order):
                raise ValueError('Supplier order already sent')
            await session.delete(existing_item)
            await session.flush()
            remaining_stmt = (
                select(SupplierOrderItem.id)
                .where(
                    SupplierOrderItem.supplier_order_id == existing_order.id
                )
                .limit(1)
            )
            remaining = (
                await session.execute(remaining_stmt)
            ).scalar_one_or_none()
            if remaining is None:
                await session.delete(existing_order)
        item.status = CUSTOMER_ORDER_ITEM_STATUS.REJECTED
        item.supplier_id = None
        item.ship_qty = 0
        item.reject_qty = item.requested_qty
        _set_reject_reason(
            item,
            'MANUAL_REJECT',
            'Позиция отклонена вручную пользователем.',
        )
        await session.commit()
        await session.refresh(item)
        try:
            await try_finalize_customer_order_response(
                session,
                order_id=int(item.order_id),
            )
        except Exception:
            await session.rollback()
        return item

    if target_status == CUSTOMER_ORDER_ITEM_STATUS.OWN_STOCK:
        if existing_item and existing_order:
            if not _is_modifiable(existing_order):
                raise ValueError('Supplier order already sent')
            await session.delete(existing_item)
            await session.flush()
            remaining_stmt = (
                select(SupplierOrderItem.id)
                .where(
                    SupplierOrderItem.supplier_order_id == existing_order.id
                )
                .limit(1)
            )
            remaining = (
                await session.execute(remaining_stmt)
            ).scalar_one_or_none()
            if remaining is None:
                await session.delete(existing_order)
            existing_item = None
            existing_order = None

        stock_order = existing_stock_order
        if not stock_order:
            order_stmt = (
                select(StockOrder)
                .where(
                    StockOrder.customer_id == item.order.customer_id,
                    StockOrder.status == STOCK_ORDER_STATUS.NEW,
                )
                .order_by(StockOrder.created_at.asc())
                .limit(1)
            )
            stock_order = (
                await session.execute(order_stmt)
            ).scalar_one_or_none()
            if not stock_order:
                stock_order = StockOrder(
                    customer_id=item.order.customer_id,
                    status=STOCK_ORDER_STATUS.NEW,
                )
                session.add(stock_order)
                await session.flush()

        if (
            existing_stock_item
            and stock_order
            and existing_stock_order
            and stock_order.id == existing_stock_order.id
        ):
            existing_stock_item.quantity = item.requested_qty
            existing_stock_item.autopart_id = item.autopart_id
        else:
            session.add(
                StockOrderItem(
                    stock_order_id=stock_order.id,
                    customer_order_item_id=item.id,
                    autopart_id=item.autopart_id,
                    quantity=item.requested_qty,
                )
            )

        item.status = CUSTOMER_ORDER_ITEM_STATUS.OWN_STOCK
        item.supplier_id = None
        item.ship_qty = item.requested_qty
        item.reject_qty = 0
        _clear_reject_reason(item)

        await session.commit()
        await session.refresh(item)
        try:
            await try_finalize_customer_order_response(
                session,
                order_id=int(item.order_id),
            )
        except Exception:
            await session.rollback()
        return item

    if existing_item and existing_order:
        if not _is_modifiable(existing_order):
            raise ValueError('Supplier order already sent')
        if existing_order.provider_id != supplier_id:
            await session.delete(existing_item)
            await session.flush()
            remaining_stmt = (
                select(SupplierOrderItem.id)
                .where(
                    SupplierOrderItem.supplier_order_id == existing_order.id
                )
                .limit(1)
            )
            remaining = (
                await session.execute(remaining_stmt)
            ).scalar_one_or_none()
            if remaining is None:
                await session.delete(existing_order)
            existing_item = None
            existing_order = None

    if existing_stock_item and existing_stock_order:
        if not _is_stock_modifiable(existing_stock_order):
            raise ValueError('Stock order already closed')
        await session.delete(existing_stock_item)
        await session.flush()
        remaining_stmt = (
            select(StockOrderItem.id)
            .where(
                StockOrderItem.stock_order_id == existing_stock_order.id
            )
            .limit(1)
        )
        remaining = (
            await session.execute(remaining_stmt)
        ).scalar_one_or_none()
        if remaining is None:
            await session.delete(existing_stock_order)

    supplier_order = existing_order
    if not supplier_order:
        order_stmt = (
            select(SupplierOrder)
            .where(
                SupplierOrder.provider_id == supplier_id,
                SupplierOrder.status.in_(
                    [
                        SUPPLIER_ORDER_STATUS.NEW,
                        SUPPLIER_ORDER_STATUS.SCHEDULED,
                    ]
                ),
            )
            .order_by(SupplierOrder.created_at.asc())
            .limit(1)
        )
        supplier_order = (
            await session.execute(order_stmt)
        ).scalar_one_or_none()
        if not supplier_order:
            supplier_order = SupplierOrder(
                provider_id=supplier_id, status=SUPPLIER_ORDER_STATUS.NEW
            )
            session.add(supplier_order)
            await session.flush()

    # Resolve supplier price: always use the supplier's own price list price,
    # never fall back to the customer's requested_price (that is the customer-
    # facing sale price, not what we pay the supplier).
    # Priority:
    #   1. Current supplier's price list (PriceListAutoPartAssociation)
    #   2. matched_price already on the item
    #      (may be from a previous pricelist run)
    #   3. None (leave blank rather than use customer price)
    supplier_price: Optional[float] = None
    if item.autopart_id and supplier_id:
        price_stmt = (
            select(PriceListAutoPartAssociation.price)
            .join(PriceList)
            .where(
                PriceList.provider_id == supplier_id,
                PriceListAutoPartAssociation.autopart_id == item.autopart_id,
            )
            .order_by(
                PriceList.date.desc().nullslast(),
                PriceList.id.desc(),
            )
            .limit(1)
        )
        supplier_price = (
            await session.execute(price_stmt)
        ).scalar_one_or_none()
    if not supplier_price and item.matched_price:
        supplier_price = float(item.matched_price)

    if (
        existing_item
        and supplier_order
        and supplier_order.id == existing_order.id
    ):
        existing_item.quantity = item.requested_qty
        existing_item.price = supplier_price
        existing_item.autopart_id = item.autopart_id
    else:
        session.add(
            SupplierOrderItem(
                supplier_order_id=supplier_order.id,
                customer_order_item_id=item.id,
                autopart_id=item.autopart_id,
                oem_number=item.oem,
                brand_name=item.brand,
                autopart_name=item.name,
                quantity=item.requested_qty,
                price=supplier_price,
            )
        )

    item.status = CUSTOMER_ORDER_ITEM_STATUS.SUPPLIER
    item.supplier_id = supplier_id
    item.ship_qty = item.requested_qty
    item.reject_qty = 0
    _clear_reject_reason(item)

    await session.commit()
    await session.refresh(item)
    try:
        await try_finalize_customer_order_response(
            session,
            order_id=int(item.order_id),
        )
    except Exception:
        await session.rollback()
    return item


def cleanup_order_reports(days: int = ORDERS_RETENTION_DAYS) -> int:
    if days <= 0:
        return 0
    cutoff = now_moscow().timestamp() - days * 86400
    if not os.path.isdir(ORDERS_REPORT_DIR):
        return 0
    removed = 0
    for root, _, files in os.walk(ORDERS_REPORT_DIR):
        for name in files:
            path = os.path.join(root, name)
            try:
                if os.path.getmtime(path) < cutoff:
                    os.remove(path)
                    removed += 1
            except Exception as exc:
                logger.error('Failed to remove report %s: %s', path, exc)
    return removed


def cleanup_order_error_files(days: int) -> int:
    if days <= 0:
        return 0
    cutoff = now_moscow().timestamp() - days * 86400
    if not os.path.isdir(ORDERS_ERROR_DIR):
        return 0
    removed = 0
    for root, _, files in os.walk(ORDERS_ERROR_DIR):
        for name in files:
            path = os.path.join(root, name)
            try:
                if os.path.getmtime(path) < cutoff:
                    os.remove(path)
                    removed += 1
            except Exception as exc:
                logger.error(
                    'Failed to remove order error file %s: %s',
                    path,
                    exc,
                )
    return removed
