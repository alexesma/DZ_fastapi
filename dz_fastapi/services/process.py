import asyncio
import copy
import hashlib
import logging
import os
import re
from datetime import date, datetime
from functools import partial
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional

import numpy as np
import pandas as pd
from fastapi import HTTPException
from libarchive import memory_reader
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import func, insert, select, text
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import aliased, selectinload

from dz_fastapi.analytics.price_history import analyze_new_pricelist
from dz_fastapi.core.constants import (
    BRILLIANCE_OEM,
    CUMMINS_OEM,
    FAW_OEM,
    GEELY_NOT_OEM,
    INDICATOR_BYD,
    INDICATOR_BYD_FIRST_FIVE,
    INDICATOR_BYD_FIRST_THREE,
    INDICATOR_CHANGAN_END_THREE,
    INDICATOR_CHANGAN_FIRST_FOUR,
    INDICATOR_CHANGAN_FIRST_SEVEN,
    INDICATOR_CHANGAN_FIRST_THREE,
    INDICATOR_CHANGAN_FIRST_TWO,
    INDICATOR_CHERY_10_11_POSITION,
    INDICATOR_CHERY_FIRST_THREE,
    INDICATOR_CHERY_FIRST_THREE_LEN_10,
    INDICATOR_CHERY_FULL,
    INDICATOR_CHERY_GW_FIRST_THREE,
    INDICATOR_CHERY_GW_FIRST_TWO,
    INDICATOR_CHERY_GW_FULL,
    INDICATOR_DONGFENG_FULL,
    INDICATOR_END_IS_NOT_LIFAN,
    INDICATOR_FAW_OTHER_PATTERNS,
    INDICATOR_FAW_PREFIXES,
    INDICATOR_FOTON,
    INDICATOR_GEELY_FIRST_THREE,
    INDICATOR_GEELY_FIRST_TWO,
    INDICATOR_HAIMA_FULL,
    INDICATOR_HAVAL,
    INDICATOR_JAC,
    INDICATOR_LIFAN_END_FIVE,
    INDICATOR_LIFAN_END_FOUR,
    INDICATOR_LIFAN_END_THREE,
    INDICATOR_LIFAN_END_TWO,
    INDICATOR_LIFAN_FIRST_THREE,
    INDICATOR_LIFAN_FIRST_THREE_2,
    INDICATOR_LIFAN_LEN_NINE,
    INDICATOR_LIFAN_LEN_SEVEN,
    INDICATOR_LIFAN_LEN_TEN,
    INDICATOR_LIFAN_WHISOUT,
    INDICATOR_LIFAN_WHISOUT_FIRST,
    MAX_PRICE_LISTS,
    ORIGINAL_BRANDS,
)
from dz_fastapi.core.time import now_moscow
from dz_fastapi.crud.email_account import crud_email_account
from dz_fastapi.crud.partner import (
    crud_customer_pricelist,
    crud_customer_pricelist_config,
    crud_customer_pricelist_source,
    crud_pricelist,
    crud_provider,
)
from dz_fastapi.crud.price_control import crud_customer_pricelist_override
from dz_fastapi.models.autopart import AutoPart, preprocess_oem_number
from dz_fastapi.models.brand import Brand
from dz_fastapi.models.cross import AutoPartCross
from dz_fastapi.models.partner import (
    Customer,
    CustomerPriceList,
    CustomerPriceListConfig,
    CustomerPriceListExportRow,
    CustomerPriceListPublicationRule,
    CustomerPriceListPublicationRuleTarget,
    CustomerPriceListPublishedAlias,
    PriceListAutoPartAssociation,
    Provider,
    ProviderPriceListConfig,
)
from dz_fastapi.schemas.autopart import AutoPartResponse
from dz_fastapi.schemas.partner import (
    AutoPartInPricelist,
    CustomerPriceListCreate,
    CustomerPriceListResponse,
    PriceListCreate,
)
from dz_fastapi.services.email import (
    EMAIL_NAME,
    EMAIL_TRANSPORT,
    SMTP_PORT,
    SMTP_SERVER,
    build_email_delivery_kwargs,
    describe_email_delivery,
    send_email_with_attachment,
)
from dz_fastapi.services.pricelist_guard import guard_automatic_provider_pricelist
from dz_fastapi.services.utils import (
    brand_filters,
    normalize_markup,
    normalize_mixed_cyrillic,
    position_exclude,
    position_filters,
    prepare_excel_data_from_records,
)
from dz_fastapi.services.watchlist import handle_provider_pricelist_watch

logger = logging.getLogger("dz_fastapi")

DEFAULT_CUSTOMER_PRICELIST_FILE_NAME = "zzap_kross"
DEFAULT_CUSTOMER_PRICELIST_OUTBOX_EMAIL = (
    os.getenv(
        "CUSTOMER_PRICELIST_OUTBOX_EMAIL",
        "price@dragonzap.online",
    )
    .strip()
    .lower()
)
CUSTOMER_PRICELIST_ARTIFACT_ROOT = Path(
    os.getenv("CUSTOMER_PRICELIST_ARTIFACT_ROOT", "uploads/customer_pricelists")
)

CUSTOMER_PRICELIST_PIPELINE_DEFAULT = [
    "source_filters",
    "price_control_before",
    "dragonzap_crosses",
    "dragonzap_transform",
    "product_labels",
    "price_control_after",
    "publication_rules",
    "final_filters",
    "deduplication",
    "quality_control",
]


def _customer_pricelist_setting(
    config: CustomerPriceListConfig,
    key: str,
    default: Any,
) -> Any:
    settings = getattr(config, "additional_filters", None) or {}
    return settings.get(key, default) if isinstance(settings, dict) else default


def customer_pricelist_pipeline(config: CustomerPriceListConfig) -> list[str]:
    raw = _customer_pricelist_setting(
        config,
        "PIPELINE_ORDER",
        CUSTOMER_PRICELIST_PIPELINE_DEFAULT,
    )
    requested = [str(value) for value in raw] if isinstance(raw, list) else []
    result = [value for value in requested if value in CUSTOMER_PRICELIST_PIPELINE_DEFAULT]
    for value in CUSTOMER_PRICELIST_PIPELINE_DEFAULT:
        if value not in result:
            result.append(value)
    dependencies = {
        "price_control_before": {"source_filters"},
        "dragonzap_crosses": {"source_filters", "price_control_before"},
        "dragonzap_transform": {"dragonzap_crosses"},
        "product_labels": {"dragonzap_transform"},
        "price_control_after": {"dragonzap_transform"},
        "publication_rules": {
            "price_control_before",
            "dragonzap_crosses",
            "dragonzap_transform",
            "product_labels",
            "price_control_after",
        },
        "deduplication": {
            "price_control_before",
            "dragonzap_crosses",
            "dragonzap_transform",
            "product_labels",
            "price_control_after",
            "publication_rules",
            "final_filters",
        },
        "final_filters": {
            "price_control_before",
            "dragonzap_crosses",
            "dragonzap_transform",
            "product_labels",
            "price_control_after",
            "publication_rules",
        },
        "quality_control": set(CUSTOMER_PRICELIST_PIPELINE_DEFAULT) - {"quality_control"},
    }
    ordered: list[str] = []
    pending = list(result)
    while pending:
        available = [
            value
            for value in pending
            if dependencies.get(value, set()).issubset(set(ordered))
        ]
        value = available[0] if available else pending[0]
        ordered.append(value)
        pending.remove(value)
    return ordered


def _customer_pricelist_v2_enabled(config: CustomerPriceListConfig) -> bool:
    return bool(_customer_pricelist_setting(config, "PIPELINE_V2_ENABLED", False))


def customer_pricelist_requires_draft(
    config: CustomerPriceListConfig,
) -> bool:
    return bool(_customer_pricelist_setting(config, "REQUIRE_DRAFT_APPROVAL", False))


def _dataframe_summary(df: pd.DataFrame, label: str) -> str:
    if df is None:
        return f"{label}: <none>"
    columns = [str(col) for col in list(df.columns[:10])]
    extra_columns = max(len(df.columns) - len(columns), 0)
    return (
        f"{label}: rows={len(df)} cols={len(df.columns)} "
        f"columns={columns}"
        f"{' ...' if extra_columns else ''}"
    )


def _is_pricelist_out_account_eligible(account) -> bool:
    purposes = [str(p).lower() for p in (account.purposes or [])]
    return account.is_active and (
        "prices_out" in purposes or "orders_out" in purposes or "orders_in" in purposes
    )


async def _get_preferred_pricelist_out_account(
    session: AsyncSession,
):
    if not DEFAULT_CUSTOMER_PRICELIST_OUTBOX_EMAIL:
        return None
    account = await crud_email_account.get_by_email(
        session,
        DEFAULT_CUSTOMER_PRICELIST_OUTBOX_EMAIL,
    )
    if not account:
        return None
    if not _is_pricelist_out_account_eligible(account):
        logger.warning(
            "Preferred customer pricelist outbox is inactive "
            "or missing required purpose: email=%s",
            DEFAULT_CUSTOMER_PRICELIST_OUTBOX_EMAIL,
        )
        return None
    return account


def _resolve_customer_pricelist_export_format(
    config: CustomerPriceListConfig,
) -> str:
    export_format = str(getattr(config, "export_file_format", None) or "xlsx").strip().lower()
    if export_format not in {"xlsx", "csv"}:
        return "xlsx"
    return export_format


def _build_customer_pricelist_attachment_filename(
    config: CustomerPriceListConfig,
) -> str:
    base_name = str(
        getattr(config, "export_file_name", None) or DEFAULT_CUSTOMER_PRICELIST_FILE_NAME
    ).strip()
    if not base_name:
        base_name = DEFAULT_CUSTOMER_PRICELIST_FILE_NAME
    base_name = re.sub(r'[\\/:*?"<>|]+', "_", base_name).strip(" .")
    if not base_name:
        base_name = DEFAULT_CUSTOMER_PRICELIST_FILE_NAME

    export_format = _resolve_customer_pricelist_export_format(config)
    extension = (
        str(getattr(config, "export_file_extension", None) or export_format)
        .strip()
        .lstrip(".")
        .lower()
    )
    extension = re.sub(r"[^a-z0-9_]+", "", extension) or export_format
    return f"{base_name}.{extension}"


def _build_customer_pricelist_attachment_bytes(
    df_excel: pd.DataFrame,
    config: CustomerPriceListConfig,
) -> bytes:
    """CPU-тяжёлая генерация вложения — вызывать через asyncio.to_thread.

    Используем write-only режим openpyxl и общие объекты стилей:
    классический режим с созданием Font на каждую ячейку на прайсе
    в 100к строк работал минуты и держал весь файл в памяти.
    """
    export_format = _resolve_customer_pricelist_export_format(config)
    if export_format == "csv":
        return df_excel.to_csv(index=False).encode("utf-8-sig")

    output = BytesIO()
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()

    note_font = Font(name="Arial", size=7)
    header_font = Font(name="Arial", size=10, bold=True)
    header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    data_font = Font(name="Arial", size=10)

    current_time = now_moscow().strftime("%Y-%m-%d %H:%M:%S")
    note_cell = WriteOnlyCell(ws, value=f"Сформирован {current_time}")
    note_cell.font = note_font
    note_cell.alignment = center_alignment
    ws.append([None, None, None, None, note_cell])

    logger.debug("Write headers on the second row")
    header_cells = []
    for column_title in df_excel.columns:
        cell = WriteOnlyCell(ws, value=column_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        header_cells.append(cell)
    ws.append(header_cells)

    # Regex that matches characters illegal in Excel worksheets
    # (control characters except tab \x09, newline \x0A, carriage return \x0D)
    _illegal_chars_re = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]")

    def _sanitize_cell(value: Any) -> Any:
        if isinstance(value, str):
            return _illegal_chars_re.sub("", value)
        return value

    logger.debug("Write data rows starting from the third row")
    for row_data in df_excel.itertuples(index=False):
        row_cells = []
        for cell_value in row_data:
            cell = WriteOnlyCell(ws, value=_sanitize_cell(cell_value))
            cell.font = data_font
            row_cells.append(cell)
        ws.append(row_cells)

    wb.save(output)
    attachment_bytes = output.getvalue()
    logger.debug(
        "Workbook saved successfully. Size: %s bytes",
        len(attachment_bytes),
    )
    return attachment_bytes


def deduplicate_autoparts_data(autoparts_data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    unique_map = {}
    for row in autoparts_data:
        key = (
            row.get("brand", "").strip().lower(),
            row["oem_number"].strip().lower(),
        )
        if key not in unique_map:
            unique_map[key] = copy.deepcopy(row)
        else:
            if unique_map[key]["price"] < row["price"]:
                continue
            unique_map[key]["quantity"] = row["quantity"]
            unique_map[key]["price"] = row["price"]
            if "multiplicity" in row:
                unique_map[key]["multiplicity"] = row.get("multiplicity")
    return list(unique_map.values())


def extract_first_file_from_archive(file_content: bytes) -> (str, bytes):
    extracted_content = None
    extracted_extension = None
    try:
        with memory_reader(file_content) as entries:
            for entry in entries:
                if entry.isfile:
                    extracted_content = b"".join(list(entry.get_blocks()))
                    extracted_extension = entry.pathname.split(".")[-1].lower()
                    break
    except Exception as e:
        raise Exception(f"Error reading archive: {e}")
    if extracted_content is None:
        raise Exception("Archive is empty")
    return extracted_extension, extracted_content


def _sanitize_positive_price_quantity(df: pd.DataFrame, context: str = "") -> pd.DataFrame:
    if df.empty:
        return df
    df = df.copy()
    original_len = len(df)
    df["price"] = pd.to_numeric(df["price"], errors="coerce")
    df["quantity"] = pd.to_numeric(df["quantity"], errors="coerce")
    df = df[df["price"].notna() & df["quantity"].notna()]
    df = df[(df["price"] > 0) & (df["quantity"] > 0)]
    dropped = original_len - len(df)
    if dropped > 0:
        logger.debug(
            "Dropped %s rows with non-positive price/quantity%s",
            dropped,
            f" ({context})" if context else "",
        )
    return df


def _apply_source_filters(
    df: pd.DataFrame,
    source,
    *,
    ignore_price_quantity_filters: bool = False,
    dragonzap_mode: str = "normal",
) -> pd.DataFrame:
    df = _sanitize_positive_price_quantity(df, context="source_filters")
    dragonzap_mode = str(dragonzap_mode or "normal").strip().lower()
    if dragonzap_mode not in {"normal", "exclude", "transform_only", "auto"}:
        dragonzap_mode = "normal"
    if "__transform_only" not in df.columns:
        df["__transform_only"] = False

    def _to_int_list(values):
        cleaned = []
        for value in values or []:
            try:
                cleaned.append(int(value))
            except (TypeError, ValueError):
                continue
        return cleaned

    dragonzap_rows = df[df["brand"].map(_is_dragonzap_brand)].copy()
    if source.brand_filters:
        normalized = dict(source.brand_filters)
        if "brands" in normalized:
            normalized["brands"] = _to_int_list(normalized.get("brands"))
        df = brand_filters(brand_filters=normalized, df=df)
    if dragonzap_mode == "exclude":
        df = df[~df["brand"].map(_is_dragonzap_brand)].copy()
    elif dragonzap_mode == "transform_only":
        df = df[~df["brand"].map(_is_dragonzap_brand)].copy()
        if not dragonzap_rows.empty:
            dragonzap_rows["__transform_only"] = True
            df = pd.concat([df, dragonzap_rows], ignore_index=True)
    elif dragonzap_mode == "auto" and not dragonzap_rows.empty:
        remaining_ids = set(pd.to_numeric(df.get("autopart_id"), errors="coerce").dropna())
        removed = dragonzap_rows[
            ~pd.to_numeric(dragonzap_rows.get("autopart_id"), errors="coerce").isin(
                remaining_ids
            )
        ].copy()
        if not removed.empty:
            removed["__transform_only"] = True
            df = pd.concat([df, removed], ignore_index=True)
    if source.position_filters:
        normalized = dict(source.position_filters)
        if "autoparts" in normalized:
            normalized["autoparts"] = _to_int_list(normalized.get("autoparts"))
        df = position_filters(position_filters=normalized, df=df)

    if not ignore_price_quantity_filters:
        if source.min_price is not None:
            df = df[df["price"] >= float(source.min_price)]
        if source.max_price is not None:
            df = df[df["price"] <= float(source.max_price)]
        if source.min_quantity is not None:
            df = df[df["quantity"] >= int(source.min_quantity)]
        if source.max_quantity is not None:
            df = df[df["quantity"] <= int(source.max_quantity)]

        source_settings = getattr(source, "additional_filters", None) or {}
        brand_rules = source_settings.get("BRAND_FILTER_RULES", [])
        if isinstance(brand_rules, list) and "brand_id" in df.columns:
            for rule in brand_rules:
                if not isinstance(rule, dict):
                    continue
                brand_ids = {
                    int(value)
                    for value in rule.get("brand_ids", [])
                    if str(value).isdigit()
                }
                if not brand_ids:
                    continue
                brand_mask = pd.to_numeric(df["brand_id"], errors="coerce").isin(brand_ids)
                condition = brand_mask.copy()
                if rule.get("min_price") is not None:
                    condition &= df["price"] >= float(rule["min_price"])
                if rule.get("max_price") is not None:
                    condition &= df["price"] <= float(rule["max_price"])
                if rule.get("min_quantity") is not None:
                    condition &= df["quantity"] >= int(rule["min_quantity"])
                if rule.get("max_quantity") is not None:
                    condition &= df["quantity"] <= int(rule["max_quantity"])
                action = str(rule.get("action") or "include").strip().lower()
                if action == "exclude":
                    df = df[~condition]
                else:
                    df = df[~brand_mask | condition]

    return _sanitize_positive_price_quantity(df, context="source_filters_after_limits")


def _normalize_source_brand_markup_key(value: object) -> str:
    normalized = normalize_mixed_cyrillic(str(value or "")).strip()
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized.upper()


def _resolve_source_brand_markup_multipliers(source) -> dict[str, float]:
    raw_map = getattr(source, "brand_markups", None) or {}
    if not isinstance(raw_map, dict):
        return {}
    result: dict[str, float] = {}
    for raw_brand, raw_markup in raw_map.items():
        key = _normalize_source_brand_markup_key(raw_brand)
        if not key:
            continue
        result[key] = normalize_markup(raw_markup)
    return result


def _stable_unit_interval(*parts: object) -> float:
    seed = "|".join(str(part or "") for part in parts)
    digest = hashlib.blake2b(seed.encode("utf-8"), digest_size=8).digest()
    return int.from_bytes(digest, "big") / float(2**64 - 1)


def _customer_pricelist_mask_week_key() -> str:
    year, week, _ = now_moscow().isocalendar()
    return f"{year}-W{week:02d}"


def _masked_markup_multiplier(base_multiplier: float, unit: float) -> float:
    markup = max(float(base_multiplier or 1.0) - 1.0, 0.0)
    if markup <= 0:
        return 1.0
    min_markup = markup * 0.9
    max_markup = markup * 1.2
    return 1.0 + min_markup + (max_markup - min_markup) * unit


def _mask_supplier_quantity(quantity: object, unit: float) -> int:
    try:
        qty = int(float(quantity or 0))
    except (TypeError, ValueError):
        return 0
    if qty <= 0:
        return 0
    if qty <= 2:
        return qty
    if qty <= 5:
        return max(qty - 1, 1)
    if qty <= 10:
        return max(qty - 2, 1)
    if qty <= 30:
        masked = round(qty * (0.65 + 0.25 * unit))
    elif qty <= 100:
        masked = round(qty * (0.50 + 0.30 * unit))
    else:
        masked = round(qty * (0.25 + 0.30 * unit))
        masked = min(masked, 180)
        masked = max(10, int(round(masked / 5) * 5))
    return max(1, min(int(masked), qty))


def _apply_supplier_price_quantity_mask(
    df: pd.DataFrame,
    config: CustomerPriceListConfig,
    source,
) -> pd.DataFrame:
    if df.empty or not bool(getattr(source, "mask_price_quantity", False)):
        return df

    df = df.copy()
    week_key = _customer_pricelist_mask_week_key()
    source_id = getattr(source, "id", None) or getattr(source, "provider_config_id", "")
    customer_id = getattr(config, "customer_id", "")

    def _row_key(row: pd.Series) -> tuple[object, ...]:
        return (
            customer_id,
            source_id,
            row.get("provider_config_id", ""),
            row.get("brand", ""),
            row.get("oem_number", ""),
            week_key,
        )

    def _masked_quantity(row: pd.Series) -> int:
        if bool(row.get("is_own_price")):
            return int(float(row.get("quantity") or 0))
        unit = _stable_unit_interval(*_row_key(row), "qty")
        return _mask_supplier_quantity(row.get("quantity"), unit)

    def _masked_price(row: pd.Series) -> float:
        price = float(row.get("price") or 0)
        if price <= 0 or bool(row.get("is_own_price")):
            return price
        source_multiplier = float(row.get("__source_multiplier") or 1.0)
        if source_multiplier <= 1.0:
            return max(1, round(price))
        unit = _stable_unit_interval(*_row_key(row), "price")
        masked_source_multiplier = _masked_markup_multiplier(source_multiplier, unit)
        return max(1, round(price / source_multiplier * masked_source_multiplier))

    df["quantity"] = df.apply(_masked_quantity, axis=1)
    df["price"] = df.apply(_masked_price, axis=1)
    return _sanitize_positive_price_quantity(df, context="source_masking")


def _apply_source_markups(
    df: pd.DataFrame,
    config: CustomerPriceListConfig,
    source,
) -> pd.DataFrame:
    df = df.copy()
    general_multiplier = normalize_markup(config.general_markup)
    default_source_multiplier = normalize_markup(source.markup)
    brand_markup_multipliers = _resolve_source_brand_markup_multipliers(source)

    own_multiplier = normalize_markup(config.own_price_list_markup)
    third_multiplier = normalize_markup(config.third_party_markup)

    df["price"] = pd.to_numeric(df["price"], errors="coerce")
    df["is_own_price"] = df.get("is_own_price", False)
    df["__brand_markup_key"] = (
        df.get("brand", "").fillna("").astype(str).map(_normalize_source_brand_markup_key)
    )
    df["__source_multiplier"] = (
        df["__brand_markup_key"].map(brand_markup_multipliers).fillna(default_source_multiplier)
    )

    def _row_multiplier(is_own: bool) -> float:
        return own_multiplier if is_own else third_multiplier

    df["price"] = df.apply(
        lambda row: (
            row["price"]
            * general_multiplier
            * float(row.get("__source_multiplier") or 1.0)
            * _row_multiplier(bool(row.get("is_own_price")))
        ),
        axis=1,
    )
    df = _apply_supplier_price_quantity_mask(df, config, source)
    df = df.drop(columns=["__brand_markup_key", "__source_multiplier"])
    return _sanitize_positive_price_quantity(df, context="source_markups")


def apply_price_overrides(df: pd.DataFrame, overrides: dict[int, float]) -> pd.DataFrame:
    if df.empty or not overrides:
        return df
    df = df.copy()
    df["price"] = df["autopart_id"].map(overrides).fillna(df["price"])
    return _sanitize_positive_price_quantity(df, context="price_overrides")


def _normalize_dedup_oem_key(value: object) -> str:
    normalized = preprocess_oem_number(str(value or "")).strip()
    return normalized.upper()


def _normalize_dedup_brand_key(value: object) -> str:
    normalized = normalize_mixed_cyrillic(str(value or "")).strip()
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized.upper()


def _collapse_duplicate_rows(
    df: pd.DataFrame,
    *,
    prefer_min_price: bool,
) -> pd.DataFrame:
    if df.empty:
        return df

    collapsed = df.copy()
    collapsed["price"] = pd.to_numeric(collapsed["price"], errors="coerce")
    oem_series = (
        collapsed["oem_number"]
        if "oem_number" in collapsed.columns
        else pd.Series("", index=collapsed.index)
    )
    brand_series = (
        collapsed["brand"] if "brand" in collapsed.columns else pd.Series("", index=collapsed.index)
    )
    collapsed["__dedup_oem"] = oem_series.map(_normalize_dedup_oem_key)
    collapsed["__dedup_brand"] = brand_series.map(_normalize_dedup_brand_key)

    if prefer_min_price:
        if "is_own_price" in collapsed.columns:
            collapsed["__dz_own_rank"] = (
                collapsed["__dedup_brand"].eq("DRAGONZAP")
                & collapsed["is_own_price"].fillna(False).astype(bool)
            ).astype(int)
        else:
            collapsed["__dz_own_rank"] = 0
        collapsed = collapsed.sort_values(
            by=[
                "__dedup_oem",
                "__dedup_brand",
                "__dz_own_rank",
                "price",
            ],
            ascending=[True, True, False, True],
        )
    elif "is_own_price" in collapsed.columns:
        collapsed["__own_rank"] = collapsed["is_own_price"].astype(bool).astype(int)
        collapsed = collapsed.sort_values(
            by=["__dedup_oem", "__dedup_brand", "__own_rank", "price"],
            ascending=[True, True, False, True],
        )
    else:
        collapsed = collapsed.sort_values(
            by=["__dedup_oem", "__dedup_brand", "price"],
            ascending=[True, True, True],
        )

    collapsed = collapsed.drop_duplicates(
        subset=["__dedup_oem", "__dedup_brand"],
        keep="first",
    )
    return collapsed.drop(
        columns=[
            "__dedup_oem",
            "__dedup_brand",
            "__own_rank",
            "__dz_own_rank",
        ],
        errors="ignore",
    )


def _collapse_duplicate_excel_rows(df_excel: pd.DataFrame) -> pd.DataFrame:
    if df_excel.empty:
        return df_excel
    if not {"Производитель", "Артикул"}.issubset(df_excel.columns):
        return df_excel

    collapsed = df_excel.copy()
    collapsed["__dedup_oem"] = collapsed["Артикул"].map(_normalize_dedup_oem_key)
    collapsed["__dedup_brand"] = collapsed["Производитель"].map(_normalize_dedup_brand_key)

    if "Цена" in collapsed.columns:
        collapsed["__dedup_price"] = pd.to_numeric(collapsed["Цена"], errors="coerce").fillna(
            float("inf")
        )
        collapsed = collapsed.sort_values(
            by=["__dedup_oem", "__dedup_brand", "__dedup_price"],
            ascending=[True, True, True],
            kind="stable",
        )
    else:
        collapsed = collapsed.sort_values(
            by=["__dedup_oem", "__dedup_brand"],
            ascending=[True, True],
            kind="stable",
        )

    collapsed = collapsed.drop_duplicates(
        subset=["__dedup_oem", "__dedup_brand"],
        keep="first",
    )
    return collapsed.drop(
        columns=["__dedup_oem", "__dedup_brand", "__dedup_price"],
        errors="ignore",
    )


def _is_dragonzap_brand(value: object) -> bool:
    return _normalize_dedup_brand_key(value) == "DRAGONZAP"


def _prefixed_name(name: object, prefix: str) -> str:
    value = str(name or "").strip()
    marker = str(prefix or "").strip()
    if not marker or value.startswith(marker):
        return value
    return f"{marker} {value}".strip()


def _transform_dragonzap_records(
    records: list[dict[str, Any]],
    *,
    keep_source: bool,
) -> list[dict[str, Any]]:
    """Create client-facing original offers while retaining physical ids."""

    result: list[dict[str, Any]] = []
    for source in records:
        row = dict(source)
        row.setdefault("__row_type", "direct")
        if not _is_dragonzap_brand(row.get("brand")):
            row.setdefault("__origin_type", "original_source")
            result.append(row)
            continue

        source_oem = str(row.get("oem_number") or "").strip()
        source_brand = str(row.get("brand") or "").strip()
        advertised_oem = source_oem[2:] if source_oem.upper().startswith("DZ") else source_oem
        assigned_brands = assign_brand(advertised_oem)
        for assigned_brand in assigned_brands:
            transformed = dict(row)
            transformed_row_type = (
                "transformed_cross"
                if row.get("__row_type") == "automatic_cross"
                else "zzap_transform"
            )
            transformed.update(
                {
                    "brand": assigned_brand,
                    "oem_number": advertised_oem,
                    "__source_oem": source_oem,
                    "__source_brand": source_brand,
                    "__origin_type": "dragonzap_transform",
                    "__row_type": transformed_row_type,
                }
            )
            result.append(transformed)

        if keep_source and not bool(row.get("__transform_only")):
            row["__origin_type"] = "dragonzap_source"
            result.append(row)
    return result


def _apply_product_labels(
    records: list[dict[str, Any]],
    *,
    label_original: bool,
    label_transformed: bool,
    original_label: str,
    transformed_label: str,
) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    original_keys = {_normalize_dedup_brand_key(value) for value in ORIGINAL_BRANDS}
    for source in records:
        row = dict(source)
        origin_type = row.get("__origin_type")
        if label_transformed and origin_type == "dragonzap_transform":
            row["name"] = _prefixed_name(row.get("name"), transformed_label)
        elif (
            label_original
            and origin_type == "original_source"
            and _normalize_dedup_brand_key(row.get("brand")) in original_keys
        ):
            row["name"] = _prefixed_name(row.get("name"), original_label)
        result.append(row)
    return result


def _collapse_output_records(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Choose cheapest Brand+OEM, then biggest stock, then real original."""

    origin_rank = {
        "original_source": 0,
        "dragonzap_transform": 1,
        "dragonzap_source": 2,
    }
    selected: dict[tuple[str, str], dict[str, Any]] = {}
    for source in records:
        row = dict(source)
        key = (
            _normalize_dedup_brand_key(row.get("brand")),
            _normalize_dedup_oem_key(row.get("oem_number")),
        )
        candidate_rank = (
            float(row.get("price") or float("inf")),
            -int(float(row.get("quantity") or 0)),
            origin_rank.get(str(row.get("__origin_type") or ""), 3),
            int(row.get("autopart_id") or 0),
        )
        current = selected.get(key)
        if current is None:
            selected[key] = row
            continue
        current_rank = (
            float(current.get("price") or float("inf")),
            -int(float(current.get("quantity") or 0)),
            origin_rank.get(str(current.get("__origin_type") or ""), 3),
            int(current.get("autopart_id") or 0),
        )
        if candidate_rank < current_rank:
            selected[key] = row
    return list(selected.values())


def _final_filter_rule_matches(row: dict[str, Any], rule: dict[str, Any]) -> bool:
    """Match a generated client-facing row against one final filter rule."""

    conditions = 0
    brand_values = {
        _normalize_dedup_brand_key(value)
        for value in rule.get("brands", [])
        if str(value or "").strip()
    }
    if brand_values:
        conditions += 1
        if _normalize_dedup_brand_key(row.get("brand")) not in brand_values:
            return False

    oem_value = _normalize_dedup_oem_key(rule.get("oem"))
    if oem_value:
        conditions += 1
        row_oem = _normalize_dedup_oem_key(row.get("oem_number"))
        match_mode = str(rule.get("oem_match") or "exact").strip().lower()
        if match_mode == "contains":
            matched = oem_value in row_oem
        elif match_mode == "prefix":
            matched = row_oem.startswith(oem_value)
        else:
            matched = row_oem == oem_value
        if not matched:
            return False

    name_value = str(rule.get("name_contains") or "").strip().casefold()
    if name_value:
        conditions += 1
        if name_value not in str(row.get("name") or "").casefold():
            return False

    row_types = {
        str(value).strip().lower()
        for value in rule.get("row_types", [])
        if str(value or "").strip()
    }
    if row_types:
        conditions += 1
        if str(row.get("__row_type") or "direct").strip().lower() not in row_types:
            return False

    origin_types = {
        str(value).strip().lower()
        for value in rule.get("origin_types", [])
        if str(value or "").strip()
    }
    if origin_types:
        conditions += 1
        if str(row.get("__origin_type") or "").strip().lower() not in origin_types:
            return False

    provider_config_ids = {
        int(value)
        for value in rule.get("provider_config_ids", [])
        if str(value).isdigit()
    }
    if provider_config_ids:
        conditions += 1
        try:
            row_provider_config_id = int(row.get("provider_config_id") or 0)
        except (TypeError, ValueError):
            return False
        if row_provider_config_id not in provider_config_ids:
            return False

    numeric_rules = (
        ("min_price", "price", float, lambda current, limit: current >= limit),
        ("max_price", "price", float, lambda current, limit: current <= limit),
        ("min_quantity", "quantity", int, lambda current, limit: current >= limit),
        ("max_quantity", "quantity", int, lambda current, limit: current <= limit),
    )
    for rule_key, row_key, converter, comparator in numeric_rules:
        if rule.get(rule_key) in (None, ""):
            continue
        conditions += 1
        try:
            if not comparator(converter(float(row.get(row_key) or 0)), converter(rule[rule_key])):
                return False
        except (TypeError, ValueError):
            return False
    return conditions > 0


def _apply_final_output_filters(
    records: list[dict[str, Any]],
    rules: Any,
    *,
    enabled: bool,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    criteria_keys = {
        "brands",
        "oem",
        "name_contains",
        "row_types",
        "origin_types",
        "provider_config_ids",
        "min_price",
        "max_price",
        "min_quantity",
        "max_quantity",
    }

    def has_criteria(rule: dict[str, Any]) -> bool:
        return any(rule.get(key) not in (None, "", []) for key in criteria_keys)

    normalized_rules = [
        dict(rule)
        for rule in (rules if isinstance(rules, list) else [])
        if isinstance(rule, dict) and rule.get("enabled", True) and has_criteria(rule)
    ]
    if not enabled or not normalized_rules:
        return records, {
            "enabled": bool(enabled),
            "rules_count": len(normalized_rules),
            "input_count": len(records),
            "output_count": len(records),
            "excluded_count": 0,
            "examples": [],
        }

    include_rules = [
        rule for rule in normalized_rules if str(rule.get("action")).lower() == "include"
    ]
    exclude_rules = [
        rule for rule in normalized_rules if str(rule.get("action")).lower() != "include"
    ]
    output: list[dict[str, Any]] = []
    examples: list[dict[str, Any]] = []
    for row in records:
        excluded_by = next(
            (rule for rule in exclude_rules if _final_filter_rule_matches(row, rule)),
            None,
        )
        include_match = next(
            (rule for rule in include_rules if _final_filter_rule_matches(row, rule)),
            None,
        )
        manual_override = bool(row.get("__manual_publication_rule"))
        reason = None
        matched_rule = excluded_by
        if excluded_by is not None:
            reason = "Совпало с запрещающим финальным правилом."
        elif include_rules and include_match is None and not manual_override:
            reason = "Не совпало ни с одним разрешающим финальным правилом."
        if reason is None:
            output.append(row)
            continue
        if len(examples) < 10:
            examples.append(
                {
                    "brand": str(row.get("brand") or ""),
                    "oem": str(row.get("oem_number") or ""),
                    "name": str(row.get("name") or ""),
                    "quantity": int(float(row.get("quantity") or 0)),
                    "price": float(row.get("price") or 0),
                    "source_autopart_id": row.get("autopart_id"),
                    "rule_id": matched_rule.get("id") if matched_rule else None,
                    "rule_name": matched_rule.get("name") if matched_rule else None,
                    "reason": reason,
                }
            )
    return output, {
        "enabled": True,
        "rules_count": len(normalized_rules),
        "input_count": len(records),
        "output_count": len(output),
        "excluded_count": len(records) - len(output),
        "examples": examples,
    }


async def _load_customer_benchmark_prices(
    session: AsyncSession,
    provider_config_ids: list[int],
) -> tuple[dict[tuple[str, str], float], list[int]]:
    latest_ids: list[int] = []
    for provider_config_id in provider_config_ids:
        pricelist = await crud_pricelist.get_latest_pricelist_by_config(
            session=session,
            provider_config_id=int(provider_config_id),
        )
        if pricelist is not None:
            latest_ids.append(int(pricelist.id))
    if not latest_ids:
        return {}, []

    rows = (
        await session.execute(
            select(
                Brand.name,
                AutoPart.oem_number,
                func.min(PriceListAutoPartAssociation.price),
            )
            .join(
                AutoPart,
                AutoPart.id == PriceListAutoPartAssociation.autopart_id,
            )
            .join(Brand, Brand.id == AutoPart.brand_id)
            .where(
                PriceListAutoPartAssociation.pricelist_id.in_(latest_ids),
                PriceListAutoPartAssociation.quantity > 0,
                PriceListAutoPartAssociation.price > 0,
            )
            .group_by(Brand.name, AutoPart.oem_number)
        )
    ).all()
    prices: dict[tuple[str, str], float] = {}
    for brand_name, oem_number, price in rows:
        if price is None:
            continue
        key = (
            _normalize_dedup_brand_key(brand_name),
            _normalize_dedup_oem_key(oem_number),
        )
        numeric_price = float(price)
        current = prices.get(key)
        if current is None or numeric_price < current:
            prices[key] = numeric_price
    return prices, latest_ids


def _apply_benchmark_floor_records(
    records: list[dict[str, Any]],
    benchmark_prices: dict[tuple[str, str], float],
    *,
    multiplier: float,
    rounding_step: float,
    stage: str,
) -> tuple[list[dict[str, Any]], int]:
    result: list[dict[str, Any]] = []
    changed = 0
    multiplier = max(float(multiplier or 1.0), 0.0)
    rounding_step = max(float(rounding_step or 1.0), 0.01)
    for source in records:
        row = dict(source)
        key = (
            _normalize_dedup_brand_key(row.get("brand")),
            _normalize_dedup_oem_key(row.get("oem_number")),
        )
        benchmark = benchmark_prices.get(key)
        if benchmark is not None:
            before = float(row.get("price") or 0)
            floor = float(np.ceil(benchmark * multiplier / rounding_step) * rounding_step)
            if floor > before:
                row["price"] = floor
                row[f"__price_before_{stage}"] = before
                row[f"__benchmark_price_{stage}"] = benchmark
                changed += 1
        result.append(row)
    return result, changed


async def _build_dragonzap_cross_alias_records(
    session: AsyncSession,
    *,
    customer_id: int,
    source_df: pd.DataFrame,
) -> list[dict[str, Any]]:
    """Build DZ-only aliases while preserving the selected physical item id."""

    if source_df.empty or "is_own_price" not in source_df.columns:
        return []
    own_rows = source_df[
        source_df["is_own_price"].fillna(False).astype(bool)
        & source_df["brand"].map(_is_dragonzap_brand)
        & (pd.to_numeric(source_df["quantity"], errors="coerce").fillna(0) > 0)
    ].copy()
    if own_rows.empty:
        return []

    own_rows["autopart_id"] = pd.to_numeric(own_rows["autopart_id"], errors="coerce")
    own_rows = own_rows.dropna(subset=["autopart_id"])
    own_rows["autopart_id"] = own_rows["autopart_id"].astype(int)
    seed_ids = set(own_rows["autopart_id"].tolist())

    source_part = aliased(AutoPart)
    target_part = aliased(AutoPart)
    source_brand = aliased(Brand)
    target_brand = aliased(Brand)
    edge_rows = (
        await session.execute(
            select(
                AutoPartCross.source_autopart_id,
                AutoPartCross.cross_autopart_id,
            )
            .join(source_part, source_part.id == AutoPartCross.source_autopart_id)
            .join(source_brand, source_brand.id == source_part.brand_id)
            .join(target_part, target_part.id == AutoPartCross.cross_autopart_id)
            .join(target_brand, target_brand.id == target_part.brand_id)
            .where(
                AutoPartCross.is_bidirectional.is_(True),
                AutoPartCross.cross_autopart_id.is_not(None),
                func.upper(source_brand.name) == "DRAGONZAP",
                func.upper(target_brand.name) == "DRAGONZAP",
            )
        )
    ).all()

    parent: dict[int, int] = {}

    def find(value: int) -> int:
        parent.setdefault(value, value)
        while parent[value] != value:
            parent[value] = parent[parent[value]]
            value = parent[value]
        return value

    def union(left: int, right: int) -> None:
        left_root, right_root = find(left), find(right)
        if left_root != right_root:
            parent[right_root] = left_root

    for source_id, target_id in edge_rows:
        if source_id is not None and target_id is not None:
            union(int(source_id), int(target_id))
    for seed_id in seed_ids:
        find(seed_id)

    relevant_roots = {find(seed_id) for seed_id in seed_ids}
    relevant_ids = {autopart_id for autopart_id in parent if find(autopart_id) in relevant_roots}
    member_rows = (
        await session.execute(
            select(
                AutoPart.id,
                Brand.name,
                AutoPart.oem_number,
                AutoPart.name,
            )
            .join(Brand, Brand.id == AutoPart.brand_id)
            .where(
                AutoPart.id.in_(relevant_ids),
                func.upper(Brand.name) == "DRAGONZAP",
            )
        )
    ).all()

    candidates_by_root: dict[int, list[dict[str, Any]]] = {}
    for record in own_rows.to_dict("records"):
        candidates_by_root.setdefault(find(int(record["autopart_id"])), []).append(record)
    for candidates in candidates_by_root.values():
        candidates.sort(
            key=lambda row: (
                float(row.get("price") or float("inf")),
                -int(float(row.get("quantity") or 0)),
                int(row["autopart_id"]),
            )
        )

    direct_keys = {
        (
            _normalize_dedup_oem_key(row.get("oem_number")),
            _normalize_dedup_brand_key(row.get("brand")),
        )
        for row in own_rows.to_dict("records")
    }
    direct_min_prices: dict[tuple[str, str], float] = {}
    for row in source_df.to_dict("records"):
        key = (
            _normalize_dedup_oem_key(row.get("oem_number")),
            _normalize_dedup_brand_key(row.get("brand")),
        )
        price = float(row.get("price") or 0)
        if price > 0:
            direct_min_prices[key] = min(direct_min_prices.get(key, price), price)

    week_key = _customer_pricelist_mask_week_key()
    aliases_by_key: dict[tuple[str, str], dict[str, Any]] = {}
    for member_id, brand_name, oem_number, part_name in member_rows:
        key = (
            _normalize_dedup_oem_key(oem_number),
            _normalize_dedup_brand_key(brand_name),
        )
        if key in direct_keys:
            continue
        candidates = candidates_by_root.get(find(int(member_id))) or []
        if not candidates:
            continue
        selected = candidates[0]
        selected_price = float(selected.get("price") or 0)
        if direct_min_prices.get(key, selected_price) < selected_price:
            continue
        quantity = int(float(selected.get("quantity") or 0))
        unit = _stable_unit_interval(
            customer_id,
            int(selected["autopart_id"]),
            int(member_id),
            week_key,
            "dragonzap_cross_qty",
        )
        alias = dict(selected)
        alias.update(
            {
                "autopart_id": int(selected["autopart_id"]),
                "__source_oem": selected.get("oem_number"),
                "__source_brand": selected.get("brand"),
                "brand": "DRAGONZAP",
                "oem_number": str(oem_number or "").strip(),
                "name": str(part_name or selected.get("name") or "").strip(),
                "quantity": _mask_supplier_quantity(quantity, unit),
                "price": selected_price,
                "__dragonzap_alias": True,
            }
        )
        current = aliases_by_key.get(key)
        if current is None or (float(alias["price"]), int(alias["autopart_id"])) < (
            float(current["price"]),
            int(current["autopart_id"]),
        ):
            aliases_by_key[key] = alias
    return list(aliases_by_key.values())


async def _apply_customer_publication_rules(
    session: AsyncSession,
    *,
    config_id: int,
    customer_id: int,
    source_df: pd.DataFrame,
    automatic_aliases: list[dict[str, Any]],
) -> tuple[pd.DataFrame, list[dict[str, Any]], dict[str, Any]]:
    """Apply explicit per-client publication rules after price selection."""

    rules = list(
        (
            await session.execute(
                select(CustomerPriceListPublicationRule)
                .options(
                    selectinload(CustomerPriceListPublicationRule.source_autopart)
                    .selectinload(AutoPart.brand),
                    selectinload(CustomerPriceListPublicationRule.target_autopart)
                    .selectinload(AutoPart.brand),
                    selectinload(CustomerPriceListPublicationRule.targets)
                    .selectinload(CustomerPriceListPublicationRuleTarget.target_autopart)
                    .selectinload(AutoPart.brand),
                )
                .where(
                    CustomerPriceListPublicationRule.config_id == config_id,
                    CustomerPriceListPublicationRule.is_active.is_(True),
                )
                .order_by(CustomerPriceListPublicationRule.id.asc())
            )
        )
        .scalars()
        .all()
    )
    if not rules:
        return (
            source_df,
            automatic_aliases,
            {
                "publication_rules": 0,
                "manual_aliases": 0,
                "hidden_positions": 0,
                "publication_rule_warnings": [],
            },
        )

    working = source_df.copy()
    working["autopart_id"] = pd.to_numeric(working.get("autopart_id"), errors="coerce")
    suppressed_source_ids: set[int] = set()
    manual_aliases: list[dict[str, Any]] = []
    warnings: list[str] = []
    week_key = _customer_pricelist_mask_week_key()

    for rule in rules:
        source_part = rule.source_autopart
        source_brand_name = (
            source_part.brand.name if source_part and source_part.brand else ""
        )
        source_oem = source_part.oem_number if source_part else ""
        source_id = int(rule.source_autopart_id)
        candidates = working[working["autopart_id"] == source_id]
        if candidates.empty:
            warnings.append(
                f"{source_brand_name} {source_oem}: фактическая позиция "
                "не попала в прайс после фильтров"
            )
            continue
        selected = (
            candidates.sort_values(by=["price", "quantity"], ascending=[True, False], kind="stable")
            .iloc[0]
            .to_dict()
        )
        mode = str(rule.mode or "only_cross").lower()
        if mode in {"hide", "only_cross"}:
            suppressed_source_ids.add(source_id)
        if mode == "hide":
            continue
        target_parts = [
            item.target_autopart
            for item in (rule.targets or [])
            if item.target_autopart is not None
        ]
        if not target_parts and rule.target_autopart is not None:
            target_parts = [rule.target_autopart]
        if not target_parts:
            warnings.append(
                f"{source_brand_name} {source_oem}: выбранные кроссы больше не существуют"
            )
            continue
        for target_part in target_parts:
            target_brand_name = (
                target_part.brand.name if target_part.brand is not None else ""
            )
            target_oem = target_part.oem_number
            if not target_brand_name or not target_oem:
                warnings.append(
                    f"{source_brand_name} {source_oem}: один из выбранных кроссов "
                    "не содержит бренд или артикул"
                )
                continue
            quantity = int(float(selected.get("quantity") or 0))
            unit = _stable_unit_interval(
                customer_id,
                source_id,
                int(target_part.id),
                week_key,
                "manual_customer_cross_qty",
            )
            alias = dict(selected)
            alias.update(
                {
                    "autopart_id": source_id,
                    "__source_oem": selected.get("oem_number"),
                    "brand": str(target_brand_name).strip(),
                    "oem_number": str(target_oem).strip(),
                    "name": str(target_part.name or selected.get("name") or "").strip(),
                    "quantity": _mask_supplier_quantity(quantity, unit),
                    "price": float(selected.get("price") or 0),
                    "__dragonzap_alias": True,
                    "__manual_publication_rule": True,
                    "__publication_rule_id": int(rule.id),
                    "__row_type": "manual_cross",
                }
            )
            manual_aliases.append(alias)

    if suppressed_source_ids:
        working = working[~working["autopart_id"].isin(suppressed_source_ids)].copy()
        automatic_aliases = [
            alias
            for alias in automatic_aliases
            if int(alias.get("autopart_id") or 0) not in suppressed_source_ids
        ]

    aliases_by_key: dict[tuple[str, str], dict[str, Any]] = {}
    for alias in [*automatic_aliases, *manual_aliases]:
        key = (
            _normalize_dedup_brand_key(alias.get("brand")),
            _normalize_dedup_oem_key(alias.get("oem_number")),
        )
        current = aliases_by_key.get(key)
        if current is None or alias.get("__manual_publication_rule"):
            aliases_by_key[key] = alias

    return (
        working,
        list(aliases_by_key.values()),
        {
            "publication_rules": len(rules),
            "manual_aliases": len(manual_aliases),
            "hidden_positions": len(suppressed_source_ids),
            "publication_rule_warnings": warnings,
        },
    )


def _sync_dragonzap_alias_prices(
    aliases: list[dict[str, Any]],
    df_excel: pd.DataFrame,
) -> None:
    """Keep alias price identical to its source after optional ZZAP floors."""

    prices: dict[tuple[str, str], float] = {}
    if {"Производитель", "Артикул", "Цена"}.issubset(df_excel.columns):
        for row in df_excel.to_dict("records"):
            key = (
                _normalize_dedup_brand_key(row.get("Производитель")),
                _normalize_dedup_oem_key(row.get("Артикул")),
            )
            value = pd.to_numeric(row.get("Цена"), errors="coerce")
            if pd.notna(value):
                prices[key] = float(value)
    for alias in aliases:
        if alias.get("__manual_publication_rule"):
            continue
        source_key = (
            _normalize_dedup_brand_key(alias.get("brand")),
            _normalize_dedup_oem_key(alias.get("__source_oem") or ""),
        )
        if source_key in prices:
            alias["price"] = prices[source_key]


def open_csv(file: bytes) -> pd.DataFrame:
    encodings = [
        "utf-8-sig",
        "utf-8",
        "cp1251",
        "windows-1251",
        "koi8-r",
        "cp866",
        "latin1",
    ]
    separators = [",", ";", "\t", "|"]
    for encoding in encodings:
        for sep in separators:
            try:
                df = pd.read_csv(
                    BytesIO(file),
                    sep=sep,
                    engine="python",
                    header=None,
                    encoding=encoding,
                )
                if df.shape[1] > 1:
                    logger.debug(
                        "CSV detected with encoding=%s, separator=%s",
                        encoding,
                        sep,
                    )
                    return df
            except (UnicodeDecodeError, pd.errors.ParserError):
                continue
    raise HTTPException(status_code=400, detail="Invalid CSV file.")


def process_download_pricelist(file_extension: str, file_content: bytes) -> pd.DataFrame:
    """
    Функция принимает файл (архив или обычный файл) и возвращает DataFrame.
    Поддерживает форматы: zip, rar, xls, xlsx, csv.
    """
    try:
        # Разархивируем ZIP
        if file_extension in ["zip", "rar"]:
            logger.debug(f"File is an archive ({file_extension}), " f"attempting extraction...")
            file_extension, file_content = extract_first_file_from_archive(file_content)
            logger.debug(f"Extracted file extension: {file_extension}")

        if file_extension in ["xls", "xlsx"]:
            try:
                df = pd.read_excel(
                    BytesIO(file_content),
                    header=None,
                    engine="xlrd" if file_extension == "xls" else "openpyxl",
                )
            except Exception as e:
                logger.error(f"Error reading Excel file: {e}")
                raise HTTPException(status_code=400, detail="Invalid Excel file.")
        elif file_extension == "csv":
            try:
                df = open_csv(file_content)
            except Exception as e:
                logger.error(f"Error reading CSV file: {e}")
                raise HTTPException(status_code=400, detail="Invalid CSV file.")
        else:
            raise HTTPException(
                status_code=400,
                detail=f"Unsupported file type: {file_extension}",
            )
        return df
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid format file:{e}")


def _prepare_pricelist_data(
    file_extension: str,
    file_content: bytes,
    start_row: int,
    oem_col: int,
    brand_col: Optional[int],
    name_col: Optional[int],
    multiplicity_col: Optional[int],
    qty_col: int,
    price_col: int,
):
    df = process_download_pricelist(file_extension=file_extension, file_content=file_content)
    data_df = df.iloc[start_row:]
    required_columns = {
        "oem_number": oem_col,
        "brand": brand_col,
        "name": name_col,
        "multiplicity": multiplicity_col,
        "quantity": qty_col,
        "price": price_col,
    }
    required_columns = {k: v for k, v in required_columns.items() if v is not None}

    data_df = data_df.loc[:, list(required_columns.values())]
    data_df.columns = list(required_columns.keys())

    total_rows = len(data_df)
    data_df.dropna(subset=["oem_number", "quantity", "price"], inplace=True)
    data_df["oem_number"] = (
        data_df["oem_number"].astype(str).str.strip().apply(preprocess_oem_number)
    )
    if "name" in data_df.columns:
        data_df["name"] = data_df["name"].astype(str).str.strip().apply(normalize_mixed_cyrillic)
    if "brand" in data_df.columns:
        data_df["brand"] = data_df["brand"].astype(str).str.strip()
    data_df["quantity"] = (
        data_df["quantity"]
        .astype(str)
        .str.replace(",", ".", regex=False)
        .str.replace(r"[^\d\.]", "", regex=True)
    )
    if "multiplicity" in data_df.columns:
        data_df["multiplicity"] = (
            data_df["multiplicity"]
            .astype(str)
            .str.replace(",", ".", regex=False)
            .str.replace(r"[^\d\.]", "", regex=True)
        )
    else:
        # Если колонка кратности не задана в конфигурации, используем 1.
        data_df["multiplicity"] = 1
    data_df["price"] = (
        data_df["price"]
        .astype(str)
        .str.replace(",", ".", regex=False)
        .str.replace(r"[^\d\.]", "", regex=True)
    )
    data_df["quantity"] = pd.to_numeric(data_df["quantity"], errors="coerce")
    data_df["multiplicity"] = pd.to_numeric(data_df["multiplicity"], errors="coerce")
    data_df.loc[data_df["multiplicity"] <= 0, "multiplicity"] = None
    data_df["multiplicity"] = data_df["multiplicity"].fillna(1).apply(int)
    data_df["price"] = pd.to_numeric(data_df["price"], errors="coerce")
    data_df.dropna(subset=["quantity", "price"], inplace=True)
    MAX_PRICE = 99999999.99
    before_count = len(data_df)
    data_df = data_df[data_df["price"] <= MAX_PRICE]
    after_count = len(data_df)
    logger.debug(
        f"Removed {before_count - after_count} " f"rows due to exceeding price {MAX_PRICE}"
    )
    data_df = data_df[data_df["price"] >= 0]
    clean_rows = len(data_df)

    autoparts_data = data_df.to_dict(orient="records")
    deduplicated_data = deduplicate_autoparts_data(autoparts_data)
    dedup_rows = len(deduplicated_data)
    del autoparts_data
    del data_df

    stats = {
        "rows_total": int(total_rows),
        "rows_clean": int(clean_rows),
        "rows_deduplicated": int(dedup_rows),
        "rows_removed": int(max(total_rows - clean_rows, 0)),
        "rows_dedup_removed": int(max(clean_rows - dedup_rows, 0)),
    }
    return deduplicated_data, stats


def _normalize_exclude_positions(exclude_positions):
    if not exclude_positions:
        return set()
    normalized = set()
    for item in exclude_positions:
        if not isinstance(item, dict):
            continue
        brand = str(item.get("brand", "")).strip().upper()
        oem = str(item.get("oem", "")).strip().upper()
        if brand and oem:
            normalized.add((brand, oem))
    return normalized


def _apply_provider_filters(items, provider_list_conf):
    if not items:
        return items
    min_price = provider_list_conf.min_price
    max_price = provider_list_conf.max_price
    min_quantity = provider_list_conf.min_quantity
    max_quantity = provider_list_conf.max_quantity
    exclude_positions = _normalize_exclude_positions(provider_list_conf.exclude_positions)

    filtered = []
    removed = 0
    for item in items:
        price = float(item.get("price", 0))
        quantity = int(item.get("quantity", 0))
        brand = str(item.get("brand", "")).strip().upper()
        oem = str(item.get("oem_number", "")).strip().upper()

        if min_price is not None and price < min_price:
            removed += 1
            continue
        if max_price is not None and price > max_price:
            removed += 1
            continue
        if min_quantity is not None and quantity < min_quantity:
            removed += 1
            continue
        if max_quantity is not None and quantity > max_quantity:
            removed += 1
            continue
        if exclude_positions and (brand, oem) in exclude_positions:
            removed += 1
            continue

        filtered.append(item)

    if removed:
        logger.debug(f"Removed {removed} items by provider filters")
    return filtered


def parse_exclude_positions_file(file_extension, file_content):
    buffer = BytesIO(file_content)
    ext = file_extension.lower()
    if ext in ("xlsx", "xls"):
        df = pd.read_excel(buffer, header=None, dtype=str)
    elif ext in ("csv", "txt"):
        df = pd.read_csv(buffer, header=None, dtype=str)
    else:
        raise ValueError("Unsupported file extension")

    df = df.dropna(how="all")
    items = []
    for _, row in df.iterrows():
        brand = str(row.iloc[0]).strip() if len(row) > 0 else ""
        oem = str(row.iloc[1]).strip() if len(row) > 1 else ""
        if not brand or not oem:
            continue
        items.append({"brand": brand, "oem": oem})
    return items


async def process_provider_pricelist(
    provider: Provider,
    file_content: bytes,
    file_extension: str,
    provider_list_conf: ProviderPriceListConfig,
    use_stored_params: bool,
    start_row: Optional[int],
    oem_col: Optional[int],
    brand_col: Optional[int],
    name_col: Optional[int],
    multiplicity_col: Optional[int],
    qty_col: Optional[int],
    price_col: Optional[int],
    session: AsyncSession,
    return_stats: bool = False,
    include_autoparts_response: bool = True,
    enforce_anomaly_guard: bool = True,
    source_filename: str | None = None,
):
    logger.debug(
        f"Зашли в process_provider_pricelist "
        f"provider name = {provider.name} "
        f"file_extension = {file_extension} "
        f"use_stored_params = {use_stored_params}"
    )
    if not provider_list_conf:
        raise HTTPException(status_code=404, detail="Configuration not transferred")

    if session.get_bind().dialect.name == "postgresql":
        lock_key = 4_450_000_000 + int(provider_list_conf.id)
        lock_acquired = (
            await session.execute(
                text("SELECT pg_try_advisory_xact_lock(:lock_key)"),
                {"lock_key": lock_key},
            )
        ).scalar_one()
        if not lock_acquired:
            raise HTTPException(
                status_code=409,
                detail=(
                    "Этот прайс уже обрабатывается. Дождитесь завершения "
                    "текущей загрузки и не запускайте её повторно."
                ),
            )

    if use_stored_params:
        start_row = provider_list_conf.start_row
        oem_col = provider_list_conf.oem_col
        brand_col = provider_list_conf.brand_col
        name_col = provider_list_conf.name_col
        multiplicity_col = provider_list_conf.multiplicity_col
        qty_col = provider_list_conf.qty_col
        price_col = provider_list_conf.price_col
    else:
        if None in (start_row, oem_col, qty_col, price_col):
            raise HTTPException(status_code=400, detail="Missing required parameters.")

    try:
        deduplicated_data, stats = await asyncio.to_thread(
            _prepare_pricelist_data,
            file_extension,
            file_content,
            start_row,
            oem_col,
            brand_col,
            name_col,
            multiplicity_col,
            qty_col,
            price_col,
        )
    except KeyError as e:
        raise HTTPException(status_code=422, detail=f"Invalid column indices provided: {e}")
    except Exception as e:
        logger.error(f"Error during data cleaning: {e}")
        raise HTTPException(status_code=400, detail="Error during data cleaning.")

    deduplicated_data = _apply_provider_filters(deduplicated_data, provider_list_conf)
    stats["rows_after_filters"] = int(len(deduplicated_data))
    logger.info(
        "Prepared provider pricelist payload: provider_id=%s "
        "config_id=%s rows_total=%s rows_clean=%s "
        "rows_deduplicated=%s rows_removed=%s "
        "rows_dedup_removed=%s rows_after_filters=%s",
        provider.id,
        provider_list_conf.id,
        stats.get("rows_total"),
        stats.get("rows_clean"),
        stats.get("rows_deduplicated"),
        stats.get("rows_removed"),
        stats.get("rows_dedup_removed"),
        stats.get("rows_after_filters"),
    )

    if enforce_anomaly_guard:
        anomaly = await guard_automatic_provider_pricelist(
            session=session,
            provider=provider,
            provider_config=provider_list_conf,
            items=deduplicated_data,
            source_filename=source_filename,
            file_content=file_content,
            file_extension=file_extension,
        )
        if anomaly.blocked:
            raise HTTPException(
                status_code=409,
                detail=(
                    "Подозрительное обновление прайса заблокировано. "
                    "Администратору отправлено предупреждение. После проверки "
                    "файл можно принять через ручную загрузку."
                ),
            )

    pricelist_in = PriceListCreate(
        provider_id=provider.id,
        provider_config_id=provider_list_conf.id,
        autoparts=[],
    )

    # Передаём строки прайса обычными dict-ами: построение и валидация
    # десятков тысяч вложенных pydantic-моделей с последующим model_dump()
    # занимали десятки секунд CPU прямо в event loop.
    autoparts_payload: list[dict] = []
    for item in deduplicated_data:
        try:
            autoparts_payload.append(
                {
                    "autopart": {
                        "oem_number": item["oem_number"],
                        "brand": item.get("brand"),
                        "name": item.get("name"),
                    },
                    "quantity": int(item["quantity"]),
                    "price": float(item["price"]),
                    "multiplicity": int(item.get("multiplicity") or 1),
                }
            )
        except KeyError as ke:
            logger.error(f"Missing key in item: {ke}")
            raise HTTPException(status_code=400, detail=f"Missing key in item: {ke}")

    # Create the price list
    try:
        pricelist = await crud_pricelist.create(
            obj_in=pricelist_in,
            session=session,
            include_autoparts_response=include_autoparts_response,
            autoparts_payload=autoparts_payload,
        )
        await handle_provider_pricelist_watch(
            session=session,
            provider=provider,
            provider_config=provider_list_conf,
            pricelist_id=pricelist.id,
            items=deduplicated_data,
        )
        # Получили Pydantic-ответ с .id
        created_id = pricelist.id
        # А теперь достаём полноценный ORM-объект (со всеми relationships)
        pl_orm = await crud_pricelist.get(session=session, obj_id=created_id)

        await analyze_new_pricelist(pl_orm, session=session)
        if return_stats:
            return pricelist, stats
        return pricelist
    except HTTPException as e:
        raise e
    except Exception as e:
        logger.exception(f"Unexpected error occurred while creating PriceList: {e}")
        raise HTTPException(
            status_code=500,
            detail="Unexpected error during PriceList creation",
        )


def starts_with_any(s, prefixes):
    return any(s.startswith(pref) for pref in prefixes)


def contains_any(s, substrings):
    return any(sub in s for sub in substrings)


def is_chery_haval_gw(oem_original):
    return (
        oem_original[:2] in INDICATOR_CHERY_GW_FIRST_TWO
        or oem_original[:3] in INDICATOR_CHERY_GW_FIRST_THREE
        or oem_original in INDICATOR_CHERY_GW_FULL
    )


def is_faw(oem_original):
    # FAW
    # Условие для определения FAW:
    # - начинается на один из INDICATOR_FAW_PREFIXES или
    # - попадает в FAW_OEM или
    # - содержит любой из INDICATOR_FAW_OTHER_PATTERNS
    return (
        starts_with_any(oem_original, INDICATOR_FAW_PREFIXES)
        or oem_original in FAW_OEM
        or contains_any(oem_original, INDICATOR_FAW_OTHER_PATTERNS)
    )


def is_dongfeng(oem_original):
    return oem_original in INDICATOR_DONGFENG_FULL


def is_haima(oem_original):
    return oem_original in INDICATOR_HAIMA_FULL


def is_lifan_simple(oem_original):
    # Простое условие для Лифан по одному из критериев
    # Второй критерий для LIFAN: Если длина 8 и
    # первые 3 символа в INDICATOR_LIFAN_FIRST_THREE_2
    return len(oem_original) == 8 and oem_original[:3] in INDICATOR_LIFAN_FIRST_THREE_2


def is_changan(oem_original):
    return (
        oem_original[:3] in INDICATOR_CHANGAN_FIRST_THREE
        or (len(oem_original) == 15 and oem_original[:4] in INDICATOR_CHANGAN_FIRST_FOUR)
        or (len(oem_original) == 8 and oem_original[:2] in INDICATOR_CHANGAN_FIRST_TWO)
        or (len(oem_original) == 14 and oem_original[:7] in INDICATOR_CHANGAN_FIRST_SEVEN)
        or (len(oem_original) == 10 and oem_original[-3:] in INDICATOR_CHANGAN_END_THREE)
    )


def is_chery(oem_original):
    # CHERY Определяется сложными условиями
    # Разбиваем на несколько отдельных проверок:
    cond1 = oem_original[:3] in INDICATOR_CHERY_FIRST_THREE and len(oem_original) > 8
    cond2 = oem_original in INDICATOR_CHERY_FULL
    cond3 = len(oem_original) >= 11 and oem_original[9:11] in INDICATOR_CHERY_10_11_POSITION
    cond4 = (
        len(oem_original) == 10
        and oem_original[:3] in INDICATOR_CHERY_FIRST_THREE_LEN_10
        and oem_original[7:] not in INDICATOR_HAVAL
    )
    return cond1 or cond2 or cond3 or cond4


def is_lifan(oem_original):
    # LIFAN Определяется набором сложных условий
    cond1 = (
        len(oem_original) == 8
        and (oem_original not in INDICATOR_LIFAN_WHISOUT)
        and (oem_original[-1] not in INDICATOR_END_IS_NOT_LIFAN)
        and (oem_original[:1] not in INDICATOR_LIFAN_WHISOUT_FIRST)
    )
    cond2 = len(oem_original) == 10 and oem_original[-2:] in INDICATOR_LIFAN_END_TWO
    cond3 = len(oem_original) == 10 and oem_original[-3:] in INDICATOR_LIFAN_END_THREE
    cond4 = (
        len(oem_original) == 7 and oem_original[:3] in INDICATOR_LIFAN_LEN_SEVEN
    )  # Исправлено [:2] на [:3]
    cond5 = len(oem_original) == 9 and oem_original[:4] in INDICATOR_LIFAN_LEN_NINE
    cond6 = oem_original[:3] in INDICATOR_LIFAN_FIRST_THREE
    cond7 = len(oem_original) == 12 and oem_original[-5:] in INDICATOR_LIFAN_END_FIVE
    cond8 = len(oem_original) == 11 and oem_original[-4:] in INDICATOR_LIFAN_END_FOUR
    cond9 = len(oem_original) == 13 and oem_original[-5:] in INDICATOR_LIFAN_END_FIVE
    cond10 = (
        len(oem_original) == 10 and oem_original[:3] in INDICATOR_LIFAN_LEN_TEN
    )  # Исправил на [:3] для единообразия,
    # хотя можно [:2], но в списке по 3 символа.
    return cond1 or cond2 or cond3 or cond4 or cond5 or cond6 or cond7 or cond8 or cond9 or cond10


def is_byd(oem_original):
    cond1 = oem_original[:3] in INDICATOR_BYD_FIRST_THREE and len(oem_original) != 11
    cond2 = oem_original in INDICATOR_BYD
    cond3 = oem_original[:5] in INDICATOR_BYD_FIRST_FIVE and len(oem_original) == 10
    return cond1 or cond2 or cond3


def is_geely(oem_original):
    cond1 = oem_original[:3] in INDICATOR_GEELY_FIRST_THREE
    cond2 = (
        len(oem_original) == 10 and oem_original.isdigit() and (oem_original not in GEELY_NOT_OEM)
    )
    cond3 = (len(oem_original) in [11, 12, 13]) and oem_original.isdigit()
    cond4 = len(oem_original) == 11 and oem_original[:2] in INDICATOR_GEELY_FIRST_TWO
    return cond1 or cond2 or cond3 or cond4


def is_jac(oem_original):
    return oem_original in INDICATOR_JAC


def is_foton(oem_original):
    return oem_original in INDICATOR_FOTON


def is_brilliance(oem_original):
    return oem_original in BRILLIANCE_OEM


def is_cummins(oem_original):
    return oem_original in CUMMINS_OEM


def assign_brand(oem_original):
    # 1. CHERY & HAVAL GW
    if is_chery_haval_gw(oem_original):
        return ["CHERY", "HAVAL"]

    # 2. FAW
    if is_faw(oem_original):
        return ["FAW"]

    # 3. DONGFENG
    if is_dongfeng(oem_original):
        return ["DONGFENG"]

    # 4. HAIMA
    if is_haima(oem_original):
        return ["HAIMA"]

    # 5. Часть логики LIFAN (простое правило)
    if is_lifan_simple(oem_original):
        return ["LIFAN"]

    # 6. CHANGAN
    if is_changan(oem_original):
        return ["CHANGAN"]

    # 7. CHERY
    if is_chery(oem_original):
        return ["CHERY"]

    # 8. LIFAN
    if is_lifan(oem_original):
        return ["LIFAN"]

    # 9. BYD
    if is_byd(oem_original):
        return ["BYD"]

    # 10. GEELY
    if is_geely(oem_original):
        return ["GEELY"]

    # 11. JAC
    if is_jac(oem_original):
        return ["JAC"]

    # 12. FOTON
    if is_foton(oem_original):
        return ["FOTON"]

    # 13. BRILLIANCE
    if is_brilliance(oem_original):
        return ["BRILLIANCE"]

    # 14. CUMMINS
    if is_cummins(oem_original):
        return ["CUMMINS"]

    # 15. Если ни одно условие не выполнилось - HAVAL
    return ["HAVAL"]


# def assign_brand(oem_original):
#
#     if oem_original[:2] in (
#             INDICATOR_CHERY_GW_FIRST_TWO
#     ) or oem_original[:3] in (
#         INDICATOR_CHERY_GW_FIRST_THREE
#     ) or oem_original in (
#         INDICATOR_CHERY_GW_FULL
#     ):
#         return ['CHERY', 'HAVAL']
#     elif any(
#             oem_original.startswith(
#                 prefix
#             ) for prefix in INDICATOR_FAW_PREFIXES
#     ) or (
#             oem_original in FAW_OEM
#     ):
#         return ['FAW']
#     elif any(
#             pattern in oem_original
#             for pattern in INDICATOR_FAW_OTHER_PATTERNS
#     ):
#         return ['FAW']
#     elif oem_original in INDICATOR_DONGFENG_FULL:
#         return ['DONGFENG']
#     elif oem_original in INDICATOR_HAIMA_FULL:
#         return ['HAIMA']
#     elif (
#     len(oem_original) == 8 and oem_original[:3]
#     in INDICATOR_LIFAN_FIRST_THREE_2
#     ):
#         return ['LIFAN']
#     elif (oem_original[:3] in INDICATOR_CHANGAN_FIRST_THREE) or (
#         len(
#             oem_original
#         ) == 15 and oem_original[:4] in INDICATOR_CHANGAN_FIRST_FOUR
#     ) or (
#             len(
#             oem_original
#             ) == 8 and oem_original[:2] in INDICATOR_CHANGAN_FIRST_TWO
#     ) or (
#             len(
#                 oem_original
#             ) == 14 and oem_original[:7] in INDICATOR_CHANGAN_FIRST_SEVEN
#     ) or (
#         len(oem_original) == 10 and oem_original[-3:]
#         in INDICATOR_CHANGAN_END_THREE
#     ):
#         return ['CHANGAN']
#     elif oem_original[:3] in (
#             INDICATOR_CHERY_FIRST_THREE
#     ) and len(oem_original) > 8 or oem_original in (
#         INDICATOR_CHERY_FULL
#     ) or (
#         len(
#             oem_original
#         ) >= 11 and oem_original[9:11] in INDICATOR_CHERY_10_11_POSITION
#     ) or (
#         len(
#             oem_original
#         ) == 10 and oem_original[:3]
#         in INDICATOR_CHERY_FIRST_THREE_LEN_10
#         and oem_original[7:] not in INDICATOR_HAVAL
#     ):
#         return ['CHERY']
#     elif (
#             len(oem_original) == 8 and (oem_original not in (
#         INDICATOR_LIFAN_WHISOUT
#     ) and oem_original[-1] not in INDICATOR_END_IS_NOT_LIFAN
#     ) and (oem_original[:1] not in INDICATOR_LIFAN_WHISOUT_FIRST)
#     ) or (len(oem_original) == 10 and oem_original[-2:] in (
#             INDICATOR_LIFAN_END_TWO
#     )) or (len(oem_original) == 10 and oem_original[-3:] in (
#             INDICATOR_LIFAN_END_THREE
#     )) or (len(oem_original) == 7 and oem_original[:2] in (
#             INDICATOR_LIFAN_LEN_SEVEN
#     )) or (len(oem_original) == 9 and oem_original[:4] in (
#             INDICATOR_LIFAN_LEN_NINE
#     )) or (oem_original[:3] in INDICATOR_LIFAN_FIRST_THREE
#     ) or (len(oem_original) == 12 and oem_original[-5:] in (
#             INDICATOR_LIFAN_END_FIVE
#     )) or (len(oem_original) == 11 and oem_original[-4:] in
#            INDICATOR_LIFAN_END_FOUR
#     ) or (len(oem_original) == 13 and oem_original[-5:] in
#           INDICATOR_LIFAN_END_FIVE
#     ) or (
#         len(oem_original) == 10 and oem_original[:2]
#         in INDICATOR_LIFAN_LEN_TEN
#     ):
#         return ['LIFAN']
#     elif (
#             oem_original[:3]
#             in INDICATOR_BYD_FIRST_THREE and len(oem_original) != 11
#     ) or (
#             oem_original in INDICATOR_BYD
#     ) or (
#             oem_original[:5]
#             in INDICATOR_BYD_FIRST_FIVE and len(oem_original) == 10
#     ):
#         return ['BYD']
#     elif oem_original[:3] in (
#             INDICATOR_GEELY_FIRST_THREE
#     ) or (len(oem_original) == 10
#     and oem_original.isdigit()
#     and (oem_original not in GEELY_NOT_OEM)) or (
#             (len(oem_original) == 12
#             or len(oem_original) == 11
#             or len(oem_original) == 13)
#             and oem_original.isdigit()
#     ) or (len(oem_original) == 11
#     and oem_original[:2] in INDICATOR_GEELY_FIRST_TWO):
#         return ['GEELY']
#     elif oem_original in INDICATOR_JAC:
#         return ['JAC']
#     elif oem_original in INDICATOR_FOTON:
#         return ['FOTON']
#     elif oem_original in BRILLIANCE_OEM:
#         return ['BRILLIANCE']
#     elif oem_original in CUMMINS_OEM:
#         return ['CUMMINS']
#     else:
#         return ['HAVAL']


async def add_origin_brand_from_dz(
    price_zzap: pd.DataFrame,
    session: AsyncSession,
    *,
    label_products: bool = True,
) -> pd.DataFrame:
    # Создаем копию DataFrame для предотвращения изменения оригинала
    price_zzap = price_zzap.copy()

    # Добавляем префикс 'Оригинал ' к названию для оригинальных брендов
    mask_original = price_zzap["Производитель"].isin(ORIGINAL_BRANDS)
    if label_products:
        price_zzap.loc[mask_original, "Наименование"] = (
            ">>Оригинал<< " + price_zzap.loc[mask_original, "Наименование"]
        )

    # Обработка записей с брендом 'DRAGONZAP'
    dz_items = price_zzap.loc[price_zzap["Производитель"] == "DRAGONZAP"].copy()

    # Добавляем префикс 'Неоригинал ' к названию для новых брендов
    if label_products:
        dz_items["Наименование"] = ">>Неоригинал<< " + dz_items["Наименование"]
    dz_items["Артикул"] = dz_items["Артикул"].apply(lambda x: x[2:] if "DZ" in x else x)

    # Применяем функцию assign_brand для получения новых брендов
    dz_items["assigned_brands"] = dz_items["Артикул"].apply(assign_brand)

    # Разворачиваем список брендов в отдельные строки
    dz_items = dz_items.explode("assigned_brands")

    # Обновляем поле 'brand' с новым брендом
    dz_items["Производитель"] = dz_items["assigned_brands"]

    # Удаляем временное поле 'assigned_brands'
    dz_items = dz_items.drop(columns=["assigned_brands"])

    # Получаем уникальные названия новых брендов
    # new_brands = dz_items['Brand'].unique().tolist()
    #
    # # Получаем записи брендов из базы данных
    # brand_records = await brand_crud.get_brands_by_names(new_brands, session)
    #
    # # Создаем словарь соответствия названий брендов и их ID
    # brand_id_map = {brand.name: brand.id for brand in brand_records}
    #
    # # Присваиваем brand_id новым записям
    # dz_items['brand_id'] = dz_items['Brand'].map(brand_id_map)
    #
    # # Проверяем, есть ли бренды без brand_id
    # missing_brands = dz_items[dz_items['brand_id'].isna()]['brand'].unique()
    # if len(missing_brands) > 0:
    #     logger.warning(f"Missing brand_id for brands: {missing_brands}")

    # # Automatically create missing brands
    # await create_missing_brands(missing_brands, session)

    # # Fetch brand records again
    # brand_records = await brand_crud.get_brands_by_names(new_brands, session)
    # brand_id_map = {brand.name: brand.id for brand in brand_records}
    #
    # # Update brand_id for dz_items
    # dz_items['brand_id'] = dz_items['brand'].map(brand_id_map)

    # Объединяем оригинальный DataFrame с новыми записями
    price_zzap = pd.concat([price_zzap, dz_items], ignore_index=True)

    return price_zzap


def expand_dz_brands(df: pd.DataFrame) -> pd.DataFrame:
    """
    Expand DRAGONZAP positions into separate rows per assigned brand,
    WITHOUT adding any label prefixes to names.

    Works on the EXCEL DataFrame format (columns: 'Производитель', 'Артикул').
    Must be called AFTER prepare_excel_data() —
    same stage as add_origin_brand_from_dz.

    Logic:
    - Positions with Производитель == 'DRAGONZAP':
        * Strip leading 'DZ' prefix from oem_number
        * Determine target brand(s) via assign_brand()
        * Explode into one row per assigned brand
    - All other positions pass through unchanged.
    """
    df = df.copy()

    mask_dz = df["Производитель"].str.upper() == "DRAGONZAP"
    dz_items = df.loc[mask_dz].copy()

    if not dz_items.empty:
        # Strip 'DZ' prefix from OEM number
        dz_items["Артикул"] = dz_items["Артикул"].apply(
            lambda x: (x[2:] if isinstance(x, str) and x.upper().startswith("DZ") else x)
        )
        # Determine brand(s) per OEM
        dz_items["assigned_brands"] = dz_items["Артикул"].apply(assign_brand)
        # One row per assigned brand
        dz_items = dz_items.explode("assigned_brands")
        dz_items["Производитель"] = dz_items["assigned_brands"]
        dz_items = dz_items.drop(columns=["assigned_brands"])

    # Replace DRAGONZAP rows with expanded brand rows
    df_result = pd.concat([df[~mask_dz], dz_items], ignore_index=True)
    return df_result


async def send_pricelist(
    session: AsyncSession,
    df_excel: pd.DataFrame | None,
    customer: Customer,
    config: CustomerPriceListConfig,
    to_emails: Optional[List[str]],
    subject: str,
    body: str,
    attachment_filename: str | None = None,
    attachment_bytes: bytes | None = None,
):
    logger.debug("Build customer pricelist attachment")
    to_email = None
    if to_emails:
        to_email = ",".join([email for email in to_emails if email])
    if not to_email:
        to_email = customer.email_outgoing_price
    if not to_email:
        raise ValueError(
            "Не указан получатель прайс-листа: заполните получателей "
            "в конфигурации или email для исходящего прайса в карточке клиента"
        )
    subject = subject
    body = body

    # CPU-тяжёлая генерация файла — в отдельном потоке, чтобы не
    # блокировать event loop (а с ним и все остальные запросы) на минуты.
    if attachment_bytes is None:
        if df_excel is None:
            raise ValueError("df_excel or attachment_bytes is required")
        attachment_bytes = await asyncio.to_thread(
            _build_customer_pricelist_attachment_bytes,
            df_excel,
            config,
        )
    attachment_filename = attachment_filename or _build_customer_pricelist_attachment_filename(
        config
    )

    # Send the email asynchronously
    logger.debug("Send the email asynchronously")
    loop = asyncio.get_running_loop()
    account = None
    if config.outgoing_email_account_id:
        selected = await crud_email_account.get(session, config.outgoing_email_account_id)
        if not selected:
            logger.warning(
                "Configured outgoing mailbox not found: id=%s",
                config.outgoing_email_account_id,
            )
        elif _is_pricelist_out_account_eligible(selected):
            account = selected
            logger.info(
                "Using config outgoing email account for pricelist send: "
                "config=%s account_id=%s email=%s",
                config.id,
                account.id,
                account.email,
            )
        else:
            logger.warning(
                "Configured outgoing mailbox is inactive or "
                "missing prices_out/orders_out/orders_in purpose: id=%s",
                config.outgoing_email_account_id,
            )
    if account is None:
        account = await _get_preferred_pricelist_out_account(session)
        if account:
            logger.info(
                "Using preferred outgoing email account for pricelist send: "
                "config=%s account_id=%s email=%s",
                config.id,
                account.id,
                account.email,
            )
    if account is None:
        accounts = await crud_email_account.get_active_by_purpose(
            session=session,
            purpose="prices_out",
        )
        if not accounts:
            accounts = await crud_email_account.get_active_by_purpose(
                session=session,
                purpose="orders_out",
            )
        if not accounts:
            accounts = await crud_email_account.get_active_by_purpose(
                session=session,
                purpose="orders_in",
            )
        if accounts:
            account = accounts[0]

    kwargs = {}
    if account:
        logger.info(
            "Using configured outgoing email account for pricelist send: "
            "config=%s account_id=%s email=%s %s",
            config.id,
            account.id,
            account.email,
            describe_email_delivery(account),
        )
        kwargs = build_email_delivery_kwargs(account)
    else:
        logger.info(
            "Using ENV email settings for pricelist send: config=%s "
            "transport=%s host=%s port=%s user=%s",
            config.id,
            EMAIL_TRANSPORT,
            SMTP_SERVER,
            SMTP_PORT,
            EMAIL_NAME,
        )

    await loop.run_in_executor(
        None,
        partial(
            send_email_with_attachment,
            to_email=to_email,
            subject=subject,
            body=body,
            attachment_bytes=attachment_bytes,
            attachment_filename=attachment_filename,
            **kwargs,
        ),
    )
    logger.debug("Final send email")


async def _persist_customer_pricelist_artifact(
    *,
    customer_pricelist: CustomerPriceList,
    customer: Customer,
    config: CustomerPriceListConfig,
    df_excel: pd.DataFrame,
    direct_records: list[dict[str, Any]],
    alias_records: list[dict[str, Any]],
    session: AsyncSession,
) -> bytes:
    """Persist the exact attachment and searchable rows before delivery."""

    attachment_filename = _build_customer_pricelist_attachment_filename(config)
    attachment_bytes = await asyncio.to_thread(
        _build_customer_pricelist_attachment_bytes,
        df_excel,
        config,
    )
    artifact_dir = (
        CUSTOMER_PRICELIST_ARTIFACT_ROOT
        / str(customer.id)
        / str(config.id)
        / str(customer_pricelist.id)
    )
    artifact_path = artifact_dir / attachment_filename

    def _write_artifact() -> None:
        artifact_dir.mkdir(parents=True, exist_ok=True)
        temporary = artifact_path.with_suffix(f"{artifact_path.suffix}.tmp")
        temporary.write_bytes(attachment_bytes)
        temporary.replace(artifact_path)

    await asyncio.to_thread(_write_artifact)

    direct_lookup: dict[tuple[str, str], dict[str, Any]] = {}
    oem_lookup: dict[str, dict[str, Any]] = {}
    for record in direct_records:
        key = (
            _normalize_dedup_brand_key(record.get("brand")),
            _normalize_dedup_oem_key(record.get("oem_number")),
        )
        direct_lookup[key] = record
        oem_lookup.setdefault(key[1], record)
        if key[0] == "DRAGONZAP" and key[1].startswith("DZ"):
            oem_lookup.setdefault(key[1][2:], record)
    alias_lookup: dict[tuple[str, str], dict[str, Any]] = {}
    for record in alias_records:
        key = (
            _normalize_dedup_brand_key(record.get("brand")),
            _normalize_dedup_oem_key(record.get("oem_number")),
        )
        alias_lookup[key] = record

    insert_rows: list[dict[str, Any]] = []
    for row in df_excel.itertuples(index=False, name=None):
        values = dict(zip(df_excel.columns, row))
        brand = str(values.get("Производитель") or "").strip()
        oem = str(values.get("Артикул") or "").strip()
        key = (
            _normalize_dedup_brand_key(brand),
            _normalize_dedup_oem_key(oem),
        )
        source = alias_lookup.get(key)
        if source is not None:
            row_type = str(
                source.get("__row_type")
                or (
                    "manual_cross"
                    if source.get("__manual_publication_rule")
                    else "automatic_cross"
                )
            )
        else:
            source = direct_lookup.get(key)
            row_type = str((source or {}).get("__row_type") or "direct")
        if source is None:
            source = oem_lookup.get(key[1])
            row_type = "zzap_transform"
        quantity_value = pd.to_numeric(values.get("Количество"), errors="coerce")
        price_value = pd.to_numeric(values.get("Цена"), errors="coerce")
        if pd.isna(quantity_value) or pd.isna(price_value):
            continue
        insert_rows.append(
            {
                "customer_pricelist_id": customer_pricelist.id,
                "source_autopart_id": (
                    int(source.get("autopart_id")) if source and source.get("autopart_id") else None
                ),
                "advertised_brand": brand,
                "advertised_oem": oem,
                "advertised_name": (str(values.get("Наименование") or "").strip() or None),
                "normalized_brand": key[0],
                "normalized_oem": key[1],
                "quantity": int(float(quantity_value)),
                "price": float(price_value),
                "row_type": row_type,
            }
        )
        if len(insert_rows) >= 5000:
            await session.execute(insert(CustomerPriceListExportRow), insert_rows)
            insert_rows.clear()
    if insert_rows:
        await session.execute(insert(CustomerPriceListExportRow), insert_rows)

    customer_pricelist.artifact_path = str(artifact_path)
    customer_pricelist.artifact_filename = attachment_filename
    customer_pricelist.artifact_content_type = (
        "text/csv"
        if _resolve_customer_pricelist_export_format(config) == "csv"
        else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    customer_pricelist.positions_count = len(df_excel)
    customer_pricelist.generated_at = now_moscow()
    session.add(customer_pricelist)
    return attachment_bytes


async def process_customer_pricelist(
    customer: Customer,
    request: CustomerPriceListCreate,
    session: AsyncSession,
    include_autoparts_response: bool = True,
    delivery_mode: str = "auto",
) -> CustomerPriceListResponse:

    config = await crud_customer_pricelist_config.get_by_id(
        config_id=request.config_id, customer_id=customer.id, session=session
    )
    if not config:
        raise HTTPException(
            status_code=400,
            detail="No pricelist configuration found for the customer",
        )
    # Удаляем старую историю пакетно по id. Нельзя вызывать здесь
    # get_all_pricelist(): он подгружает все позиции всех старых прайсов и
    # на крупных клиентах занимал несколько гигабайт памяти scheduler.
    await crud_customer_pricelist.delete_older_pricelists(
        session=session,
        customer_id=customer.id,
        max_count=MAX_PRICE_LISTS,
    )

    if delivery_mode not in {"auto", "draft", "send"}:
        raise ValueError("delivery_mode must be auto, draft or send")

    combined_data = []
    dz_expand_enabled = False
    pipeline_v2 = _customer_pricelist_v2_enabled(config)
    pipeline_order = customer_pricelist_pipeline(config)
    transform_enabled = pipeline_v2 and bool(
        _customer_pricelist_setting(config, "DZ_ORIGINAL_TRANSFORM_ENABLED", False)
    )
    source_pricelist_ids: list[int] = []
    source_filter_summary: list[dict[str, Any]] = []

    if request.items:
        for pricelist_id in request.items:
            associations = await crud_pricelist.fetch_pricelist_data(pricelist_id, session)
            if not associations:
                continue

            df = await crud_pricelist.transform_to_dataframe(
                associations=associations, session=session
            )
            logger.debug(_dataframe_summary(df, "customer_pricelist_source_df"))

            df = crud_customer_pricelist.apply_coefficient(df, config, apply_general_markup=True)
            combined_data.append(df)
    else:
        sources = await crud_customer_pricelist_source.get_by_config_id(
            config_id=config.id, session=session
        )
        if not sources:
            raise HTTPException(
                status_code=400,
                detail="No autoparts to include in the pricelist",
            )
        dz_expand_enabled = any(
            s.enabled and (s.additional_filters or {}).get("DZ_EXPAND_BRANDS") for s in sources
        )
        for source in sources:
            if not source.enabled:
                continue

            latest_pl = await crud_pricelist.get_latest_pricelist_by_config(
                session=session, provider_config_id=source.provider_config_id
            )
            if not latest_pl:
                continue

            associations = await crud_pricelist.fetch_pricelist_data(latest_pl.id, session)
            if not associations:
                continue
            source_pricelist_ids.append(int(latest_pl.id))

            df = await crud_pricelist.transform_to_dataframe(
                associations=associations, session=session
            )
            logger.debug(_dataframe_summary(df, "customer_pricelist_latest_df"))

            source_rows_before = len(df)
            source_settings = source.additional_filters or {}
            dragonzap_mode = str(source_settings.get("DRAGONZAP_MODE") or "").strip().lower()
            if pipeline_v2 and transform_enabled and not dragonzap_mode:
                dragonzap_mode = "auto"
            df = _apply_source_filters(
                df,
                source,
                dragonzap_mode=dragonzap_mode or "normal",
            )
            source_filter_summary.append(
                {
                    "source_id": int(source.id),
                    "provider_config_id": int(source.provider_config_id),
                    "pricelist_id": int(latest_pl.id),
                    "rows_before": source_rows_before,
                    "rows_after": len(df),
                    "excluded": max(source_rows_before - len(df), 0),
                    "transform_only": int(
                        df.get("__transform_only", pd.Series(dtype=bool))
                        .fillna(False)
                        .astype(bool)
                        .sum()
                    ),
                    "dragonzap_mode": dragonzap_mode or "normal",
                }
            )
            if df.empty:
                continue

            df = crud_customer_pricelist.apply_coefficient(df, config, apply_general_markup=False)
            df = _apply_source_markups(df, config, source)
            combined_data.append(df)

    if combined_data:
        final_df = pd.concat(combined_data, ignore_index=True)
    else:
        final_df = pd.DataFrame()

    logger.debug(_dataframe_summary(final_df, "customer_pricelist_final_df"))
    if not final_df.empty:
        overrides = await crud_customer_pricelist_override.get_for_config(
            session=session, config_id=config.id
        )
        if overrides:
            final_df = apply_price_overrides(final_df, overrides)
    # Apply exclusions

    if not final_df.empty:
        if request.excluded_supplier_positions:
            for (
                provider_id,
                excluded_autoparts,
            ) in request.excluded_supplier_positions.items():
                excluded_autoparts = [
                    int(v) for v in (excluded_autoparts or []) if str(v).isdigit()
                ]
                final_df = position_exclude(
                    provider_id=provider_id,
                    excluded_autoparts=excluded_autoparts,
                    df=final_df,
                )
        # Сворачивание дубликатов на больших прайсах — чистый CPU (pandas),
        # выносим из event loop.
        final_df = await asyncio.to_thread(
            _collapse_duplicate_rows,
            final_df,
            prefer_min_price=bool(getattr(config, "collapse_duplicates_by_min_price", True)),
        )
        if pipeline_v2:
            physical_records = final_df.to_dict("records")
            direct_output_records = []
            for source_row in physical_records:
                output_row = dict(source_row)
                output_row.setdefault("__row_type", "direct")
                output_row.setdefault(
                    "__origin_type",
                    (
                        "dragonzap_source"
                        if _is_dragonzap_brand(output_row.get("brand"))
                        else "original_source"
                    ),
                )
                direct_output_records.append(output_row)
            dragonzap_alias_records: list[dict[str, Any]] = []
            rule_summary: dict[str, Any] = {
                "publication_rules": 0,
                "manual_aliases": 0,
                "hidden_positions": 0,
                "publication_rule_warnings": [],
                "final_filters": {
                    "enabled": False,
                    "rules_count": 0,
                    "input_count": 0,
                    "output_count": 0,
                    "excluded_count": 0,
                    "examples": [],
                },
            }
            publish_dz_crosses = bool(
                _customer_pricelist_setting(config, "PUBLISH_CONFIRMED_DZ_CROSSES", False)
            )
            transform_crosses = bool(
                _customer_pricelist_setting(config, "DZ_TRANSFORM_INCLUDE_CROSSES", True)
            )
            keep_dragonzap = bool(
                _customer_pricelist_setting(config, "DZ_TRANSFORM_KEEP_DRAGONZAP", False)
            )
            labels_enabled = bool(
                _customer_pricelist_setting(config, "PRODUCT_LABELS_ENABLED", False)
            )
            price_control_enabled = bool(
                _customer_pricelist_setting(config, "PRICE_CONTROL_ENABLED", False)
            )
            price_control_stages = {
                str(value)
                for value in _customer_pricelist_setting(
                    config,
                    "PRICE_CONTROL_STAGES",
                    [],
                )
                if str(value) in {"before", "after"}
            }
            benchmark_config_ids = [
                int(value)
                for value in _customer_pricelist_setting(
                    config,
                    "PRICE_CONTROL_PROVIDER_CONFIG_IDS",
                    [],
                )
                if str(value).isdigit()
            ]
            benchmark_prices, benchmark_pricelist_ids = await _load_customer_benchmark_prices(
                session,
                benchmark_config_ids if price_control_enabled else [],
            )
            price_multiplier = float(
                _customer_pricelist_setting(config, "PRICE_CONTROL_MULTIPLIER", 1.2)
            )
            price_rounding = float(
                _customer_pricelist_setting(config, "PRICE_CONTROL_ROUNDING_STEP", 10)
            )
            price_changes = {"before": 0, "after": 0}
            raw_cross_records: list[dict[str, Any]] = []

            for pipeline_step in pipeline_order:
                if (
                    pipeline_step == "price_control_before"
                    and price_control_enabled
                    and "before" in price_control_stages
                ):
                    direct_output_records, price_changes["before"] = (
                        _apply_benchmark_floor_records(
                            direct_output_records,
                            benchmark_prices,
                            multiplier=price_multiplier,
                            rounding_step=price_rounding,
                            stage="before",
                        )
                    )
                    final_df = pd.DataFrame(direct_output_records)
                elif pipeline_step == "dragonzap_crosses" and (
                    publish_dz_crosses or (transform_enabled and transform_crosses)
                ):
                    raw_cross_records = await _build_dragonzap_cross_alias_records(
                        session,
                        customer_id=customer.id,
                        source_df=final_df,
                    )
                    for row in raw_cross_records:
                        row["__row_type"] = "automatic_cross"
                        row["__origin_type"] = "dragonzap_source"
                elif pipeline_step == "dragonzap_transform" and transform_enabled:
                    direct_output_records = _transform_dragonzap_records(
                        direct_output_records,
                        keep_source=keep_dragonzap,
                    )
                    transformed_crosses = (
                        _transform_dragonzap_records(raw_cross_records, keep_source=False)
                        if transform_crosses
                        else []
                    )
                    dragonzap_alias_records = [
                        *([dict(row) for row in raw_cross_records] if publish_dz_crosses else []),
                        *transformed_crosses,
                    ]
                elif pipeline_step == "product_labels" and labels_enabled:
                    label_kwargs = {
                        "label_original": bool(
                            _customer_pricelist_setting(config, "LABEL_ORIGINAL_ENABLED", True)
                        ),
                        "label_transformed": bool(
                            _customer_pricelist_setting(config, "LABEL_TRANSFORMED_ENABLED", True)
                        ),
                        "original_label": str(
                            _customer_pricelist_setting(
                                config,
                                "LABEL_ORIGINAL_TEXT",
                                ">>Оригинал<<",
                            )
                        ),
                        "transformed_label": str(
                            _customer_pricelist_setting(
                                config,
                                "LABEL_TRANSFORMED_TEXT",
                                ">>Неоригинал<<",
                            )
                        ),
                    }
                    direct_output_records = _apply_product_labels(
                        direct_output_records,
                        **label_kwargs,
                    )
                    dragonzap_alias_records = _apply_product_labels(
                        dragonzap_alias_records,
                        **label_kwargs,
                    )
                elif (
                    pipeline_step == "price_control_after"
                    and price_control_enabled
                    and "after" in price_control_stages
                ):
                    direct_output_records, direct_changed = _apply_benchmark_floor_records(
                        direct_output_records,
                        benchmark_prices,
                        multiplier=price_multiplier,
                        rounding_step=price_rounding,
                        stage="after",
                    )
                    dragonzap_alias_records, alias_changed = _apply_benchmark_floor_records(
                        dragonzap_alias_records,
                        benchmark_prices,
                        multiplier=price_multiplier,
                        rounding_step=price_rounding,
                        stage="after",
                    )
                    price_changes["after"] = direct_changed + alias_changed
                elif pipeline_step == "publication_rules":
                    output_df = pd.DataFrame(direct_output_records)
                    output_df, dragonzap_alias_records, rule_summary = (
                        await _apply_customer_publication_rules(
                            session,
                            config_id=config.id,
                            customer_id=customer.id,
                            source_df=output_df,
                            automatic_aliases=dragonzap_alias_records,
                        )
                    )
                    direct_output_records = output_df.to_dict("records")
                elif pipeline_step == "final_filters":
                    combined_output = [
                        *({**row, "__output_group": "direct"} for row in direct_output_records),
                        *(
                            {**row, "__output_group": "alias"}
                            for row in dragonzap_alias_records
                        ),
                    ]
                    combined_output, final_filter_summary = _apply_final_output_filters(
                        combined_output,
                        _customer_pricelist_setting(config, "FINAL_FILTER_RULES", []),
                        enabled=bool(
                            _customer_pricelist_setting(
                                config,
                                "FINAL_FILTER_ENABLED",
                                False,
                            )
                        ),
                    )
                    rule_summary["final_filters"] = final_filter_summary
                    direct_output_records = [
                        row
                        for row in combined_output
                        if row.get("__output_group") == "direct"
                    ]
                    dragonzap_alias_records = [
                        row
                        for row in combined_output
                        if row.get("__output_group") == "alias"
                    ]
                elif pipeline_step == "deduplication":
                    combined_output = [
                        *({**row, "__output_group": "direct"} for row in direct_output_records),
                        *({**row, "__output_group": "alias"} for row in dragonzap_alias_records),
                    ]
                    combined_output = _collapse_output_records(combined_output)
                    direct_output_records = [
                        row for row in combined_output if row.get("__output_group") == "direct"
                    ]
                    dragonzap_alias_records = [
                        row for row in combined_output if row.get("__output_group") == "alias"
                    ]

            direct_output_records = [
                row
                for row in direct_output_records
                if not row.get("__transform_only")
                or row.get("__origin_type") == "dragonzap_transform"
            ]
            dragonzap_alias_records = [
                row
                for row in dragonzap_alias_records
                if not row.get("__transform_only")
                or row.get("__origin_type") == "dragonzap_transform"
            ]
            transformed_direct = [
                row
                for row in direct_output_records
                if str(row.get("__row_type") or "direct") != "direct"
            ]
            direct_output_records = [
                row
                for row in direct_output_records
                if str(row.get("__row_type") or "direct") == "direct"
            ]
            dragonzap_alias_records.extend(transformed_direct)
            final_source_ids = {
                int(row.get("autopart_id") or 0)
                for row in [*direct_output_records, *dragonzap_alias_records]
                if row.get("autopart_id")
            }
            customer_autoparts_data = [
                dict(row)
                for row in physical_records
                if int(row.get("autopart_id") or 0) in final_source_ids
            ]
            if not direct_output_records and not dragonzap_alias_records:
                final_filter_summary = rule_summary.get("final_filters") or {}
                raise HTTPException(
                    status_code=400,
                    detail=(
                        "Финальные фильтры исключили все позиции прайса"
                        if final_filter_summary.get("excluded_count")
                        else "Publication rules excluded all pricelist positions"
                    ),
                )
            v2_summary = {
                "pipeline_v2": True,
                "pipeline_order": pipeline_order,
                "source_pricelist_ids": source_pricelist_ids,
                "source_filters": source_filter_summary,
                "benchmark_pricelist_ids": benchmark_pricelist_ids,
                "benchmark_positions": len(benchmark_prices),
                "price_changes": price_changes,
                "transform_enabled": transform_enabled,
                "duplicate_policy": "cheapest_then_stock_then_original",
            }
            dragonzap_source_records = []
        else:
            if bool(_customer_pricelist_setting(config, "PUBLISH_CONFIRMED_DZ_CROSSES", True)):
                dragonzap_alias_records = await _build_dragonzap_cross_alias_records(
                    session,
                    customer_id=customer.id,
                    source_df=final_df,
                )
            else:
                dragonzap_alias_records = []
            final_df, dragonzap_alias_records, rule_summary = (
                await _apply_customer_publication_rules(
                    session,
                    config_id=config.id,
                    customer_id=customer.id,
                    source_df=final_df,
                    automatic_aliases=dragonzap_alias_records,
                )
            )
            customer_autoparts_data = final_df.to_dict("records")
            if not customer_autoparts_data and not dragonzap_alias_records:
                raise HTTPException(
                    status_code=400,
                    detail="Publication rules excluded all pricelist positions",
                )
            dragonzap_source_records = [
                dict(row)
                for row in customer_autoparts_data
                if bool(row.get("is_own_price")) and _is_dragonzap_brand(row.get("brand"))
            ]
            direct_output_records = customer_autoparts_data
            v2_summary = {"pipeline_v2": False}
        del final_df
    else:
        raise HTTPException(status_code=400, detail="No autoparts to include in the pricelist")

    customer_pricelist = CustomerPriceList(
        customer_id=customer.id,
        customer_config_id=config.id,
        date=request.date or date.today(),
        is_active=True,
        generation_status="generating",
        generation_summary=rule_summary,
    )
    session.add(customer_pricelist)
    await session.flush()

    # Полная перезагрузка ассоциаций нужна только когда вызывающему коду
    # требуется развёрнутый ответ (API). Регламенты передают False и
    # экономят повторный SELECT на ~100к строк с autopart+brand.
    associations = await crud_customer_pricelist.create_associations(
        customer_pricelist_id=customer_pricelist.id,
        autoparts_data=customer_autoparts_data,
        session=session,
        load_associations=include_autoparts_response,
    )
    # Prepare data for Excel file: строим из уже готовых записей,
    # без зависимости от перезагруженных ассоциаций.
    df_excel = await asyncio.to_thread(
        prepare_excel_data_from_records,
        direct_output_records if pipeline_v2 else customer_autoparts_data,
    )

    # DZ brand expansion (without name labels) — applied at Excel level
    # so that the brand column in output reflects assigned brands,
    # not DRAGONZAP.
    # dz_expand_enabled is set when iterating sources above.
    if not pipeline_v2 and dz_expand_enabled:
        logger.debug("DZ_EXPAND_BRANDS: expanding DRAGONZAP positions in Excel DF")
        df_excel = expand_dz_brands(df_excel)

    if not pipeline_v2 and bool(_customer_pricelist_setting(config, "ZZAP", False)):
        logger.debug("Зашел в get additional_filters")
        benchmark_config_id = _customer_pricelist_setting(
            config, "ZZAP_BENCHMARK_PROVIDER_CONFIG_ID", None
        )
        benchmark_pricelist = None
        if benchmark_config_id:
            benchmark_pricelist = await crud_pricelist.get_latest_pricelist_by_config(
                session=session,
                provider_config_id=int(benchmark_config_id),
            )
        if benchmark_pricelist is None:
            provider_diller = await crud_provider.get_provider_or_none(
                provider="AVTODIN KAMA", session=session
            )
            if not provider_diller:
                raise ValueError(
                    "Не выбран контрольный прайс ZZap и поставщик " "AVTODIN KAMA не найден"
                )
            pricelist_ids = await crud_pricelist.get_pricelist_ids_by_provider(
                provider_id=provider_diller.id, session=session
            )
            if not pricelist_ids:
                raise ValueError(f"No pricelists found for provider {provider_diller.name}.")
            benchmark_pricelist_id = pricelist_ids[-1]
        else:
            benchmark_pricelist_id = benchmark_pricelist.id
        benchmark_associations = await crud_pricelist.fetch_pricelist_data(
            benchmark_pricelist_id, session
        )
        df_diller = await crud_pricelist.transform_to_dataframe(
            associations=benchmark_associations, session=session
        )
        logger.debug(_dataframe_summary(df_diller, "zzap_diller_df"))
        df_diller_rename = df_diller.rename(
            columns={
                "brand": "Производитель",
                "name": "Наименование",
                "oem_number": "Артикул",
                "quantity": "Количество",
                "price": "Цена",
            }
        )
        # 1. Предположим, что и df_excel, и
        # df_diller имеют колонку "brand_id" и "price".
        #    Если у вас "brand" (строка), замените ниже 'brand_id' на 'brand'.

        # 2. Соединяем df_excel и df_diller по "brand_id".
        #    how='left' чтобы к df_excel присоединить цены диллера (если есть).
        min_multiplier = max(
            0.0,
            float(_customer_pricelist_setting(config, "ZZAP_MIN_PRICE_MULTIPLIER", 1.2)),
        )
        rounding_step = max(
            0.01,
            float(_customer_pricelist_setting(config, "ZZAP_ROUNDING_STEP", 10)),
        )

        def _apply_diller_floor(
            df_excel_local: pd.DataFrame,
            df_diller_local: pd.DataFrame,
        ) -> pd.DataFrame:
            df_merged = pd.merge(
                df_excel_local,
                df_diller_local,
                on=["Производитель", "Артикул"],
                how="left",
                suffixes=("", "_diller"),  # Чтобы колонки не конфликтовали
            )
            # Поднимаем цену до настроенного порога и шага округления.
            mask = (df_merged["Цена_diller"].notna()) & (
                df_merged["Цена"] < df_merged["Цена_diller"] * min_multiplier
            )
            df_merged.loc[mask, "Цена"] = (
                np.ceil(df_merged.loc[mask, "Цена_diller"] * min_multiplier / rounding_step)
                * rounding_step
            )
            # Возвращаемся к исходному набору колонок df_excel.
            return df_merged[df_excel_local.columns]

        # Merge на 100к+ строк — CPU-тяжёлый pandas, выносим из event loop.
        df_excel = await asyncio.to_thread(_apply_diller_floor, df_excel, df_diller_rename)

        dealer_price_by_key: dict[tuple[str, str], float] = {}
        for dealer_row in df_diller_rename.to_dict("records"):
            dealer_price = pd.to_numeric(dealer_row.get("Цена"), errors="coerce")
            if pd.isna(dealer_price):
                continue
            dealer_key = (
                _normalize_dedup_brand_key(dealer_row.get("Производитель")),
                _normalize_dedup_oem_key(dealer_row.get("Артикул")),
            )
            current = dealer_price_by_key.get(dealer_key)
            if current is None or float(dealer_price) < current:
                dealer_price_by_key[dealer_key] = float(dealer_price)
        for alias in dragonzap_alias_records:
            alias_key = (
                _normalize_dedup_brand_key(alias.get("brand")),
                _normalize_dedup_oem_key(alias.get("oem_number")),
            )
            dealer_price = dealer_price_by_key.get(alias_key)
            if dealer_price is None:
                continue
            floor_price = np.ceil(dealer_price * min_multiplier / rounding_step) * rounding_step
            alias["price"] = max(float(alias.get("price") or 0), float(floor_price))

        df_excel = await add_origin_brand_from_dz(
            price_zzap=df_excel,
            session=session,
            label_products=bool(_customer_pricelist_setting(config, "ZZAP_LABEL_PRODUCTS", True)),
        )
        logger.debug(_dataframe_summary(df_excel, "zzap_excel_df"))

    if not pipeline_v2:
        _sync_dragonzap_alias_prices(dragonzap_alias_records, df_excel)
    if not pipeline_v2 and dz_expand_enabled and dragonzap_source_records:
        source_excel = await asyncio.to_thread(
            prepare_excel_data_from_records, dragonzap_source_records
        )
        df_excel = pd.concat([df_excel, source_excel], ignore_index=True)
    if dragonzap_alias_records:
        for alias in dragonzap_alias_records:
            session.add(
                CustomerPriceListPublishedAlias(
                    customer_pricelist_id=customer_pricelist.id,
                    source_autopart_id=int(alias["autopart_id"]),
                    advertised_oem=str(alias.get("oem_number") or "").strip(),
                    advertised_brand=str(alias.get("brand") or "DRAGONZAP").strip(),
                    advertised_name=(str(alias.get("name") or "").strip() or None),
                    normalized_oem=_normalize_dedup_oem_key(alias.get("oem_number")),
                    normalized_brand=_normalize_dedup_brand_key(alias.get("brand")),
                    quantity=int(alias.get("quantity") or 0),
                    price=float(alias.get("price") or 0),
                )
            )
        alias_excel = await asyncio.to_thread(
            prepare_excel_data_from_records, dragonzap_alias_records
        )
        df_excel = pd.concat([df_excel, alias_excel], ignore_index=True)
        logger.info(
            "Added Dragonzap stock aliases to customer pricelist: "
            "customer_id=%s config_id=%s aliases=%s",
            customer.id,
            config.id,
            len(dragonzap_alias_records),
        )

    if bool(getattr(config, "collapse_duplicates_by_min_price", True)):
        df_excel = await asyncio.to_thread(_collapse_duplicate_excel_rows, df_excel)

    if {"Производитель", "Наименование"}.issubset(df_excel.columns):
        df_excel = await asyncio.to_thread(
            lambda df: df.sort_values(
                by=["Производитель", "Наименование"], kind="stable"
            ).reset_index(drop=True),
            df_excel,
        )
    generation_summary = dict(rule_summary)
    generation_summary.update(
        {
            "base_positions": len(customer_autoparts_data),
            "published_crosses": len(dragonzap_alias_records),
            "final_positions": len(df_excel),
            "zzap_mode": bool(_customer_pricelist_setting(config, "ZZAP", False)),
            "dz_expand_brands": bool(dz_expand_enabled),
            **v2_summary,
        }
    )
    duplicate_count = 0
    if {"Производитель", "Артикул"}.issubset(df_excel.columns):
        quality_keys = pd.DataFrame(
            {
                "brand": df_excel["Производитель"].map(_normalize_dedup_brand_key),
                "oem": df_excel["Артикул"].map(_normalize_dedup_oem_key),
            }
        )
        duplicate_count = int(quality_keys.duplicated(subset=["brand", "oem"]).sum())

    quality_checks = [
        {
            "key": "non_empty",
            "status": "passed" if len(df_excel) > 0 else "failed",
            "message": f"В итоговом файле {len(df_excel)} строк.",
        },
        {
            "key": "source_mapping",
            "status": (
                "passed"
                if all(
                    row.get("autopart_id")
                    for row in [*direct_output_records, *dragonzap_alias_records]
                )
                else "failed"
            ),
            "message": "Все опубликованные строки связаны с фактической номенклатурой.",
        },
        {
            "key": "positive_values",
            "status": (
                "passed"
                if all(
                    float(row.get("price") or 0) > 0
                    and int(float(row.get("quantity") or 0)) > 0
                    for row in [*direct_output_records, *dragonzap_alias_records]
                )
                else "failed"
            ),
            "message": "Цена и количество проверены.",
        },
        {
            "key": "duplicates",
            "status": "passed" if duplicate_count == 0 else "failed",
            "message": (
                "Совпадений Бренд + Артикул нет."
                if duplicate_count == 0
                else f"Осталось совпадающих строк: {duplicate_count}."
            ),
        },
    ]
    final_filter_summary = generation_summary.get("final_filters") or {}
    if final_filter_summary.get("enabled"):
        quality_checks.append(
            {
                "key": "final_filters",
                "status": (
                    "passed" if int(final_filter_summary.get("output_count") or 0) > 0 else "failed"
                ),
                "message": (
                    "Финальные фильтры исключили "
                    f"{int(final_filter_summary.get('excluded_count') or 0)} строк; "
                    f"осталось {int(final_filter_summary.get('output_count') or 0)}."
                ),
            }
        )
    if pipeline_v2 and bool(
        _customer_pricelist_setting(config, "PRICE_CONTROL_ENABLED", False)
    ):
        benchmark_ready = bool(v2_summary.get("benchmark_pricelist_ids")) and int(
            v2_summary.get("benchmark_positions") or 0
        ) > 0
        quality_checks.append(
            {
                "key": "benchmark_prices",
                "status": "passed" if benchmark_ready else "failed",
                "message": (
                    "Контрольные прайсы найдены."
                    if benchmark_ready
                    else "Контроль цены включён, но в контрольных прайсах нет предложений."
                ),
            }
        )
    generation_summary["quality_control"] = {
        "enabled": bool(
            _customer_pricelist_setting(config, "QUALITY_CONTROL_ENABLED", pipeline_v2)
        ),
        "status": (
            "failed"
            if any(check["status"] == "failed" for check in quality_checks)
            else "passed"
        ),
        "checks": quality_checks,
    }
    customer_pricelist.generation_summary = generation_summary
    customer_pricelist.generation_status = "draft"
    attachment_bytes = await _persist_customer_pricelist_artifact(
        customer_pricelist=customer_pricelist,
        customer=customer,
        config=config,
        df_excel=df_excel,
        direct_records=(direct_output_records if pipeline_v2 else customer_autoparts_data),
        alias_records=dragonzap_alias_records,
        session=session,
    )
    await session.commit()

    should_send = delivery_mode == "send" or (
        delivery_mode == "auto" and not customer_pricelist_requires_draft(config)
    )
    quality_control = generation_summary.get("quality_control") or {}
    quality_failed = bool(quality_control.get("enabled")) and (
        quality_control.get("status") == "failed"
    )
    if should_send and quality_failed:
        customer_pricelist.generation_status = "draft"
        customer_pricelist.send_error = (
            "Отправка заблокирована: итоговый файл не прошёл контроль качества."
        )
        session.add(customer_pricelist)
        await session.commit()
        if delivery_mode == "send":
            raise HTTPException(status_code=409, detail=customer_pricelist.send_error)
        should_send = False
    recipients = config.emails or (
        [customer.email_outgoing_price] if customer.email_outgoing_price else []
    )
    if should_send:
        logger.debug("Calling send_pricelist")
        try:
            await send_pricelist(
                session=session,
                customer=customer,
                config=config,
                to_emails=recipients,
                df_excel=None,
                attachment_bytes=attachment_bytes,
                attachment_filename=customer_pricelist.artifact_filename,
                subject=f"Прайс лист {customer_pricelist.date}",
                body="Добрый день, высылаем Вам наш прайс-лист",
            )
            sent_at = now_moscow()
            config.last_sent_at = sent_at
            customer_pricelist.sent_at = sent_at
            customer_pricelist.generation_status = "sent"
            customer_pricelist.send_error = None
        except Exception as exc:
            customer_pricelist.generation_status = "send_failed"
            customer_pricelist.send_error = str(exc)
            session.add(customer_pricelist)
            await session.commit()
            raise
        session.add(customer_pricelist)
        session.add(config)
        await session.commit()
        logger.debug("Finished send_pricelist")

    autoparts_response = []
    for assoc in associations:
        autopart = AutoPartResponse.model_validate(assoc.autopart, from_attributes=True)
        autopart_in_pricelist = AutoPartInPricelist(
            autopart_id=assoc.autopart_id,
            quantity=assoc.quantity,
            price=float(assoc.price),
            autopart=autopart,
        )
        autoparts_response.append(autopart_in_pricelist)

    response = CustomerPriceListResponse(
        id=customer_pricelist.id,
        date=customer_pricelist.date,
        customer_id=customer.id,
        autoparts=autoparts_response,
        positions_count=customer_pricelist.positions_count,
        generation_status=customer_pricelist.generation_status,
        generated_at=customer_pricelist.generated_at,
        artifact_filename=customer_pricelist.artifact_filename,
    )
    return response


def write_error_for_bulk(
    problem_items: dict,
    not_found: list,
    error_message: str,
    error: Optional[str] = None,
) -> None:
    record_str = {
        k: (v.decode("utf-8", errors="replace") if isinstance(v, bytes) else str(v))
        for k, v in problem_items.items()
    }
    not_found.append(
        {
            "record": record_str,
            "error": f"{error_message}: {error or 'Unknown error'}",
        }
    )


def check_start_and_finish_date(
    date_start: Optional[str], date_finish: Optional[str]
) -> tuple[datetime, datetime]:
    try:
        start_dt = datetime.fromisoformat(date_start) if date_start else datetime(2020, 1, 1)
        finish_dt = datetime.fromisoformat(date_finish) if date_finish else now_moscow()
        return start_dt, finish_dt
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail="Invalid date format. Use YYYY-MM-DD or ISO format.",
        )
