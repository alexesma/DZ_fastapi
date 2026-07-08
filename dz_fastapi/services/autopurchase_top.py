from __future__ import annotations

import io
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Any, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from dz_fastapi.core.time import now_moscow
from dz_fastapi.models.autopart import AutoPart, AutoPurchaseExcludedItem, AutoPurchaseTopItem
from dz_fastapi.models.brand import Brand
from dz_fastapi.models.partner import (
    CustomerOrder,
    CustomerOrderItem,
    PriceList,
    PriceListAutoPartAssociation,
    Provider,
    ProviderPriceListConfig,
)

TOP_SOURCE_FILE = "file"
TOP_SOURCE_CURRENT = "current"
REPORT_HEADER_FILL = "F6EDC5"
REPORT_BORDER = Side(style="thin", color="000000")


def _normalize_oem(value: Any) -> str:
    return "".join(ch for ch in str(value or "").upper() if ch.isalnum())


def _is_missing_value(value: Any) -> bool:
    if value is None or value == "":
        return True
    try:
        return bool(pd.isna(value))
    except (TypeError, ValueError):
        return False


def _normalize_text(value: Any) -> str:
    if _is_missing_value(value):
        return ""
    return str(value or "").strip()


def _normalize_source(value: str | None) -> str:
    source = _normalize_text(value or TOP_SOURCE_FILE).lower()
    return source or TOP_SOURCE_FILE


def _normalize_brand_filters(value: str | None) -> list[str]:
    raw_value = _normalize_text(value).lower()
    if not raw_value:
        return []
    return [
        item.strip()
        for item in raw_value.split(",")
        if item.strip()
    ]


def _normalize_brand_key(value: Any) -> str:
    return _normalize_text(value).casefold()


def _default_periods(today: date) -> tuple[date, date, date, date]:
    return (
        date(today.year - 1, 6, 1),
        date(today.year - 1, 12, 31),
        date(today.year, 1, 1),
        today,
    )


def _format_report_date(value: date) -> str:
    month_names = [
        "",
        "января",
        "февраля",
        "марта",
        "апреля",
        "мая",
        "июня",
        "июля",
        "августа",
        "сентября",
        "октября",
        "ноября",
        "декабря",
    ]
    return f"{value.day} {month_names[value.month]} {value.year} г."


def _format_report_period(start: date, end: date) -> str:
    month_names = [
        "",
        "Январь",
        "Февраль",
        "Март",
        "Апрель",
        "Май",
        "Июнь",
        "Июль",
        "Август",
        "Сентябрь",
        "Октябрь",
        "Ноябрь",
        "Декабрь",
    ]
    if start.day == 1 and end.day >= 28 and start.year == end.year:
        return (
            f"{month_names[start.month]} {start.year} г. - "
            f"{month_names[end.month]} {end.year} г."
        )
    return f"{start.strftime('%d.%m.%Y')} - {end.strftime('%d.%m.%Y')}"


def _pick_best_name(current: str | None, candidate: Any) -> str | None:
    candidate_text = _normalize_text(candidate) or None
    if not candidate_text:
        return current
    if not current:
        return candidate_text
    # Длинное название обычно информативнее короткого общего описания.
    return candidate_text if len(candidate_text) > len(current) else current


def _exclusion_key(oem_number: Any, brand_name: Any) -> tuple[str, str]:
    return (_normalize_oem(oem_number), _normalize_brand_key(brand_name))


async def _load_active_exclusions(
    session: AsyncSession,
) -> dict[tuple[str, str], AutoPurchaseExcludedItem]:
    rows = (
        await session.execute(
            select(AutoPurchaseExcludedItem).where(
                AutoPurchaseExcludedItem.is_active.is_(True)
            )
        )
    ).scalars().all()
    return {
        _exclusion_key(item.oem_number, item.brand_name): item
        for item in rows
        if _normalize_oem(item.oem_number)
    }


def _pick_column(row: dict[str, Any], names: set[str]) -> Any:
    for key, value in row.items():
        normalized_key = _normalize_text(key).lower().replace(" ", "_")
        if normalized_key in names:
            return value
    return None


def _safe_int(value: Any, default: int = 0) -> int:
    if _is_missing_value(value):
        return default
    try:
        return max(int(float(str(value).replace(",", ".").strip())), 0)
    except (TypeError, ValueError):
        return default


def _json_safe_value(value: Any) -> Any:
    if _is_missing_value(value):
        return None
    if isinstance(value, dict):
        return {str(key): _json_safe_value(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_json_safe_value(item) for item in value]
    if isinstance(value, tuple):
        return [_json_safe_value(item) for item in value]
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if hasattr(value, "item"):
        try:
            return _json_safe_value(value.item())
        except (TypeError, ValueError):
            return str(value)
    return value


def _json_safe_payload(row: dict[str, Any]) -> dict[str, Any]:
    return {
        str(key): _json_safe_value(value)
        for key, value in dict(row or {}).items()
    }


def _serialize_top_item(
    item: AutoPurchaseTopItem,
    *,
    current_quantity: int = 0,
    in_transit_qty: int = 0,
    exclusion: AutoPurchaseExcludedItem | None = None,
) -> dict[str, Any]:
    target_stock_qty = int(item.target_stock_qty or 0)
    available_qty = int(current_quantity or 0) + int(in_transit_qty or 0)
    return {
        "id": int(item.id),
        "source": item.source,
        "autopart_id": item.autopart_id,
        "oem_number": item.oem_number,
        "brand_name": item.brand_name,
        "autopart_name": item.autopart_name,
        "rank": int(item.rank or 0),
        "sold_qty": int(item.sold_qty or 0),
        "target_stock_qty": target_stock_qty,
        "is_active": bool(item.is_active),
        "note": item.note,
        "current_quantity": int(current_quantity or 0),
        "in_transit_qty": int(in_transit_qty or 0),
        "gap_qty": max(target_stock_qty - available_qty, 0),
        "excluded_from_autopurchase": exclusion is not None,
        "exclusion_reason": exclusion.reason if exclusion is not None else None,
        "imported_at": item.imported_at,
        "updated_at": item.updated_at,
    }


async def _load_latest_own_stock_by_oem(
    session: AsyncSession,
) -> dict[str, int]:
    stmt = (
        select(
            AutoPart.oem_number,
            PriceListAutoPartAssociation.quantity,
        )
        .select_from(PriceListAutoPartAssociation)
        .join(PriceList, PriceList.id == PriceListAutoPartAssociation.pricelist_id)
        .join(
            ProviderPriceListConfig,
            ProviderPriceListConfig.id == PriceList.provider_config_id,
        )
        .join(Provider, Provider.id == ProviderPriceListConfig.provider_id)
        .join(AutoPart, AutoPart.id == PriceListAutoPartAssociation.autopart_id)
        .where(
            PriceList.is_active.is_(True),
            Provider.is_own_price.is_(True),
            ProviderPriceListConfig.use_for_order_insights.is_(True),
        )
        .order_by(
            PriceList.date.desc().nulls_last(),
            PriceList.id.desc(),
        )
    )
    result: dict[str, int] = {}
    for row in (await session.execute(stmt)).all():
        normalized_oem = _normalize_oem(row.oem_number)
        if not normalized_oem or normalized_oem in result:
            continue
        result[normalized_oem] = int(row.quantity or 0)
    return result


async def _resolve_autopart_id(
    session: AsyncSession,
    *,
    oem_number: str,
    brand_name: str | None,
) -> Optional[int]:
    normalized_oem = _normalize_oem(oem_number)
    if not normalized_oem:
        return None

    stmt = (
        select(AutoPart.id)
        .join(Brand, Brand.id == AutoPart.brand_id)
        .where(func.upper(AutoPart.oem_number) == normalized_oem)
        .order_by(AutoPart.id.asc())
        .limit(1)
    )
    normalized_brand = _normalize_text(brand_name).lower()
    if normalized_brand:
        stmt = stmt.where(func.lower(Brand.name) == normalized_brand)

    autopart_id = (await session.execute(stmt)).scalar_one_or_none()
    if autopart_id is not None:
        return int(autopart_id)

    fallback_stmt = (
        select(AutoPart.id)
        .where(func.upper(AutoPart.oem_number) == normalized_oem)
        .order_by(AutoPart.id.asc())
        .limit(1)
    )
    fallback_id = (await session.execute(fallback_stmt)).scalar_one_or_none()
    return int(fallback_id) if fallback_id is not None else None


async def list_autopurchase_top_items(
    session: AsyncSession,
    *,
    source: str = TOP_SOURCE_FILE,
    limit: int = 100,
    active_only: bool = True,
    brand: Optional[str] = None,
) -> dict[str, Any]:
    normalized_source = _normalize_source(source)
    normalized_limit = max(min(int(limit or 100), 1000), 1)
    normalized_brands = _normalize_brand_filters(brand)
    stmt = (
        select(AutoPurchaseTopItem)
        .where(AutoPurchaseTopItem.source == normalized_source)
        .order_by(
            AutoPurchaseTopItem.rank.asc(),
            AutoPurchaseTopItem.sold_qty.desc(),
            AutoPurchaseTopItem.id.asc(),
        )
        .limit(normalized_limit)
    )
    if active_only:
        stmt = stmt.where(AutoPurchaseTopItem.is_active.is_(True))
    if normalized_brands:
        stmt = stmt.where(
            or_(*[
                func.lower(AutoPurchaseTopItem.brand_name).contains(item)
                for item in normalized_brands
            ])
        )
    rows = (await session.execute(stmt)).scalars().all()
    stock_by_oem = await _load_latest_own_stock_by_oem(session)
    exclusions = await _load_active_exclusions(session)
    return {
        "source": normalized_source,
        "total_items": len(rows),
        "rows": [
            _serialize_top_item(
                item,
                current_quantity=stock_by_oem.get(_normalize_oem(item.oem_number), 0),
                exclusion=exclusions.get(
                    _exclusion_key(item.oem_number, item.brand_name)
                ),
            )
            for item in rows
        ],
    }


async def create_autopurchase_top_item(
    session: AsyncSession,
    *,
    payload: dict[str, Any],
) -> dict[str, Any]:
    source = _normalize_source(payload.get("source"))
    oem_number = _normalize_oem(payload.get("oem_number"))
    if not oem_number:
        raise ValueError("Укажите OEM/артикул топ-позиции")

    brand_name = _normalize_text(payload.get("brand_name"))
    autopart_id = await _resolve_autopart_id(
        session,
        oem_number=oem_number,
        brand_name=brand_name,
    )
    item = AutoPurchaseTopItem(
        source=source,
        autopart_id=autopart_id,
        oem_number=oem_number,
        brand_name=brand_name,
        autopart_name=_normalize_text(payload.get("autopart_name")) or None,
        rank=_safe_int(payload.get("rank")),
        sold_qty=_safe_int(payload.get("sold_qty")),
        target_stock_qty=_safe_int(payload.get("target_stock_qty")),
        is_active=bool(payload.get("is_active", True)),
        note=_normalize_text(payload.get("note")) or None,
        imported_at=now_moscow(),
        updated_at=now_moscow(),
        raw_payload=_json_safe_payload(payload),
    )
    session.add(item)
    await session.commit()
    await session.refresh(item)
    stock_by_oem = await _load_latest_own_stock_by_oem(session)
    exclusions = await _load_active_exclusions(session)
    return _serialize_top_item(
        item,
        current_quantity=stock_by_oem.get(_normalize_oem(item.oem_number), 0),
        exclusion=exclusions.get(_exclusion_key(item.oem_number, item.brand_name)),
    )


async def update_autopurchase_top_item(
    session: AsyncSession,
    *,
    item_id: int,
    payload: dict[str, Any],
) -> dict[str, Any]:
    item = await session.get(AutoPurchaseTopItem, int(item_id))
    if item is None:
        raise ValueError("Топ-позиция не найдена")

    if "oem_number" in payload and payload["oem_number"] is not None:
        normalized_oem = _normalize_oem(payload["oem_number"])
        if not normalized_oem:
            raise ValueError("Укажите OEM/артикул топ-позиции")
        item.oem_number = normalized_oem
    if "brand_name" in payload:
        item.brand_name = _normalize_text(payload.get("brand_name")) or None
    if "autopart_name" in payload:
        item.autopart_name = _normalize_text(payload.get("autopart_name")) or None
    if "rank" in payload and payload["rank"] is not None:
        item.rank = _safe_int(payload["rank"])
    if "sold_qty" in payload and payload["sold_qty"] is not None:
        item.sold_qty = _safe_int(payload["sold_qty"])
    if "target_stock_qty" in payload and payload["target_stock_qty"] is not None:
        item.target_stock_qty = _safe_int(payload["target_stock_qty"])
    if "is_active" in payload and payload["is_active"] is not None:
        item.is_active = bool(payload["is_active"])
    if "note" in payload:
        item.note = _normalize_text(payload.get("note")) or None

    item.autopart_id = await _resolve_autopart_id(
        session,
        oem_number=item.oem_number,
        brand_name=item.brand_name,
    )
    item.updated_at = now_moscow()
    session.add(item)
    await session.commit()
    await session.refresh(item)
    stock_by_oem = await _load_latest_own_stock_by_oem(session)
    exclusions = await _load_active_exclusions(session)
    return _serialize_top_item(
        item,
        current_quantity=stock_by_oem.get(_normalize_oem(item.oem_number), 0),
        exclusion=exclusions.get(_exclusion_key(item.oem_number, item.brand_name)),
    )


async def exclude_autopurchase_item(
    session: AsyncSession,
    *,
    payload: dict[str, Any],
) -> AutoPurchaseExcludedItem:
    oem_number = _normalize_oem(payload.get("oem_number"))
    if not oem_number:
        raise ValueError("Укажите OEM/артикул для исключения")

    brand_name = _normalize_text(payload.get("brand_name"))
    existing = (
        await session.execute(
            select(AutoPurchaseExcludedItem).where(
                AutoPurchaseExcludedItem.oem_number == oem_number,
                AutoPurchaseExcludedItem.brand_name == brand_name,
            )
        )
    ).scalar_one_or_none()
    now = now_moscow()
    if existing is None:
        existing = AutoPurchaseExcludedItem(
            oem_number=oem_number,
            brand_name=brand_name,
            created_at=now,
        )
    existing.autopart_id = payload.get("autopart_id") or existing.autopart_id
    existing.autopart_name = (
        _normalize_text(payload.get("autopart_name"))
        or existing.autopart_name
    )
    existing.reason = (
        _normalize_text(payload.get("reason"))
        or existing.reason
        or "Исключено вручную из автозаказа"
    )
    existing.is_active = bool(payload.get("is_active", True))
    existing.updated_at = now
    session.add(existing)
    await session.commit()
    await session.refresh(existing)
    return existing


async def restore_autopurchase_item(
    session: AsyncSession,
    *,
    payload: dict[str, Any],
) -> AutoPurchaseExcludedItem:
    oem_number = _normalize_oem(payload.get("oem_number"))
    if not oem_number:
        raise ValueError("Укажите OEM/артикул для возврата")

    brand_name = _normalize_text(payload.get("brand_name")) or None
    item = (
        await session.execute(
            select(AutoPurchaseExcludedItem).where(
                AutoPurchaseExcludedItem.oem_number == oem_number,
                AutoPurchaseExcludedItem.brand_name == brand_name,
            )
        )
    ).scalar_one_or_none()
    if item is None:
        raise ValueError("Исключение для позиции не найдено")
    item.is_active = False
    item.updated_at = now_moscow()
    session.add(item)
    await session.commit()
    await session.refresh(item)
    return item


async def import_autopurchase_top_items(
    session: AsyncSession,
    *,
    content: bytes,
    filename: str,
    source: str = TOP_SOURCE_FILE,
) -> dict[str, Any]:
    normalized_source = _normalize_source(source)
    lower_name = str(filename or "").lower()
    if lower_name.endswith(".csv"):
        df = pd.read_csv(io.BytesIO(content))
    else:
        df = pd.read_excel(io.BytesIO(content))

    imported_count = 0
    updated_count = 0
    skipped_count = 0
    now = now_moscow()

    for idx, raw_row in enumerate(df.to_dict(orient="records"), start=1):
        oem_number = _normalize_oem(
            _pick_column(
                raw_row,
                {"oem", "артикул", "номер", "номер_детали", "part_number"},
            )
        )
        if not oem_number:
            skipped_count += 1
            continue

        brand_name = _normalize_text(
            _pick_column(raw_row, {"brand", "бренд", "make", "марка"})
        ) or None
        autopart_name = _normalize_text(
            _pick_column(raw_row, {"name", "название", "наименование"})
        ) or None
        sold_qty = _safe_int(
            _pick_column(raw_row, {"sold_qty", "sold", "продано", "qty", "количество"})
        )
        target_stock_qty = _safe_int(
            _pick_column(
                raw_row,
                {"target_stock_qty", "target", "цель", "остаток", "заказать"},
            ),
            default=sold_qty,
        )
        rank = _safe_int(_pick_column(raw_row, {"rank", "рейтинг", "место"}), idx)
        note = _normalize_text(_pick_column(raw_row, {"note", "комментарий"})) or None
        autopart_id = await _resolve_autopart_id(
            session,
            oem_number=oem_number,
            brand_name=brand_name,
        )

        existing = (
            await session.execute(
                select(AutoPurchaseTopItem).where(
                    AutoPurchaseTopItem.source == normalized_source,
                    AutoPurchaseTopItem.oem_number == oem_number,
                    AutoPurchaseTopItem.brand_name == brand_name,
                )
            )
        ).scalar_one_or_none()
        if existing:
            existing.autopart_id = autopart_id
            existing.autopart_name = autopart_name or existing.autopart_name
            existing.rank = rank
            existing.sold_qty = sold_qty
            existing.target_stock_qty = target_stock_qty
            existing.is_active = True
            existing.note = note
            existing.updated_at = now
            existing.raw_payload = _json_safe_payload(raw_row)
            updated_count += 1
        else:
            session.add(
                AutoPurchaseTopItem(
                    source=normalized_source,
                    autopart_id=autopart_id,
                    oem_number=oem_number,
                    brand_name=brand_name,
                    autopart_name=autopart_name,
                    rank=rank,
                    sold_qty=sold_qty,
                    target_stock_qty=target_stock_qty,
                    is_active=True,
                    note=note,
                    imported_at=now,
                    updated_at=now,
                    raw_payload=_json_safe_payload(raw_row),
                )
            )
            imported_count += 1

    await session.commit()
    total_active = (
        await session.execute(
            select(func.count(AutoPurchaseTopItem.id)).where(
                AutoPurchaseTopItem.source == normalized_source,
                AutoPurchaseTopItem.is_active.is_(True),
            )
        )
    ).scalar_one()
    return {
        "source": normalized_source,
        "imported_count": imported_count,
        "updated_count": updated_count,
        "skipped_count": skipped_count,
        "total_active": int(total_active or 0),
    }


async def list_current_autopurchase_top_items(
    session: AsyncSession,
    *,
    days: int = 365,
    limit: int = 100,
    brand: Optional[str] = None,
) -> dict[str, Any]:
    normalized_days = max(min(int(days or 365), 730), 1)
    normalized_limit = max(min(int(limit or 100), 1000), 1)
    normalized_brands = _normalize_brand_filters(brand)
    cutoff = now_moscow() - timedelta(days=normalized_days)
    stmt = (
        select(
            CustomerOrderItem.oem,
            CustomerOrderItem.brand,
            CustomerOrderItem.name,
            CustomerOrderItem.requested_qty,
        )
        .join(CustomerOrder, CustomerOrder.id == CustomerOrderItem.order_id)
        .where(
            CustomerOrder.received_at >= cutoff,
            CustomerOrderItem.requested_qty.isnot(None),
            CustomerOrderItem.requested_qty > 0,
        )
    )
    if normalized_brands:
        stmt = stmt.where(
            or_(*[
                func.lower(CustomerOrderItem.brand).contains(item)
                for item in normalized_brands
            ])
        )
    stock_by_oem = await _load_latest_own_stock_by_oem(session)
    exclusions = await _load_active_exclusions(session)
    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    for row in (await session.execute(stmt)).all():
        oem_number = _normalize_oem(row.oem)
        if not oem_number:
            continue
        brand_name = _normalize_text(row.brand) or ""
        key = (oem_number, _normalize_brand_key(brand_name))
        item = grouped.setdefault(
            key,
            {
                "oem_number": oem_number,
                "brand_name": brand_name or None,
                "autopart_name": None,
                "sold_qty": 0,
                "order_count": 0,
            },
        )
        item["autopart_name"] = _pick_best_name(
            item.get("autopart_name"),
            row.name,
        )
        item["sold_qty"] = int(item.get("sold_qty") or 0) + int(
            row.requested_qty or 0
        )
        item["order_count"] = int(item.get("order_count") or 0) + 1

    grouped_rows = sorted(
        grouped.values(),
        key=lambda item: (
            -int(item.get("sold_qty") or 0),
            str(item.get("brand_name") or ""),
            str(item.get("oem_number") or ""),
        ),
    )[:normalized_limit]
    rows: list[dict[str, Any]] = []
    for rank, row in enumerate(grouped_rows, start=1):
        oem_number = str(row.get("oem_number") or "")
        sold_qty = int(row.get("sold_qty") or 0)
        current_quantity = stock_by_oem.get(oem_number, 0)
        target_stock_qty = max(int(round(sold_qty / max(normalized_days, 1) * 45)), 1)
        exclusion = exclusions.get(_exclusion_key(oem_number, row.get("brand_name")))
        rows.append(
            {
                "id": -rank,
                "source": TOP_SOURCE_CURRENT,
                "autopart_id": None,
                "oem_number": oem_number,
                "brand_name": row.get("brand_name"),
                "autopart_name": row.get("autopart_name"),
                "rank": rank,
                "sold_qty": sold_qty,
                "target_stock_qty": target_stock_qty,
                "is_active": True,
                "note": (
                    f"Заказов за {normalized_days} дн: "
                    f"{int(row.get('order_count') or 0)}"
                ),
                "current_quantity": current_quantity,
                "in_transit_qty": 0,
                "gap_qty": max(target_stock_qty - current_quantity, 0),
                "excluded_from_autopurchase": exclusion is not None,
                "exclusion_reason": (
                    exclusion.reason if exclusion is not None else None
                ),
                "imported_at": now_moscow(),
                "updated_at": now_moscow(),
            }
        )
    return {
        "source": TOP_SOURCE_CURRENT,
        "total_items": len(rows),
        "rows": rows,
    }


async def build_customer_order_period_report_data(
    session: AsyncSession,
    *,
    period1_from: date | None = None,
    period1_to: date | None = None,
    period2_from: date | None = None,
    period2_to: date | None = None,
    limit: int = 1000,
    brand: Optional[str] = None,
    min_total_qty: int = 1,
    sort_by: str = "total_desc",
) -> dict[str, Any]:
    today = now_moscow().date()
    default_p1_from, default_p1_to, default_p2_from, default_p2_to = (
        _default_periods(today)
    )
    p1_from = period1_from or default_p1_from
    p1_to = period1_to or default_p1_to
    p2_from = period2_from or default_p2_from
    p2_to = period2_to or default_p2_to
    if p1_from > p1_to:
        raise ValueError("Дата начала периода 1 позже даты окончания")
    if p2_from > p2_to:
        raise ValueError("Дата начала периода 2 позже даты окончания")

    normalized_limit = max(min(int(limit or 1000), 10000), 1)
    normalized_min_total_qty = max(int(min_total_qty or 0), 0)
    normalized_sort_by = _normalize_text(sort_by or "total_desc").lower()
    normalized_brands = _normalize_brand_filters(brand)
    order_date = func.date(CustomerOrder.received_at)
    stmt = (
        select(
            order_date.label("order_date"),
            CustomerOrderItem.oem,
            CustomerOrderItem.brand,
            CustomerOrderItem.name,
            CustomerOrderItem.requested_qty,
            CustomerOrderItem.requested_price,
        )
        .join(CustomerOrder, CustomerOrder.id == CustomerOrderItem.order_id)
        .where(
            CustomerOrderItem.requested_qty.isnot(None),
            CustomerOrderItem.requested_qty > 0,
            or_(
                order_date.between(p1_from, p1_to),
                order_date.between(p2_from, p2_to),
            ),
        )
    )
    if normalized_brands:
        stmt = stmt.where(
            or_(*[
                func.lower(CustomerOrderItem.brand).contains(item)
                for item in normalized_brands
            ])
        )

    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    for row in (await session.execute(stmt)).all():
        oem_number = _normalize_oem(row.oem)
        if not oem_number:
            continue
        brand_name = _normalize_text(row.brand) or ""
        key = (oem_number, _normalize_brand_key(brand_name))
        item = grouped.setdefault(
            key,
            {
                "oem_number": oem_number,
                "brand_name": brand_name or None,
                "autopart_name": None,
                "period1_qty": 0,
                "period2_qty": 0,
                "period1_amount": Decimal("0"),
                "period1_price_qty": 0,
                "period2_amount": Decimal("0"),
                "period2_price_qty": 0,
            },
        )
        item["autopart_name"] = _pick_best_name(
            item.get("autopart_name"),
            row.name,
        )
        qty = max(int(row.requested_qty or 0), 0)
        if not qty:
            continue
        row_date = row.order_date
        if isinstance(row_date, datetime):
            row_date = row_date.date()
        elif isinstance(row_date, str):
            row_date = date.fromisoformat(row_date[:10])
        requested_price = row.requested_price
        if p1_from <= row_date <= p1_to:
            item["period1_qty"] = int(item["period1_qty"]) + qty
            if requested_price is not None:
                item["period1_amount"] = Decimal(
                    item["period1_amount"]
                ) + Decimal(str(requested_price)) * qty
                item["period1_price_qty"] = int(item["period1_price_qty"]) + qty
        elif p2_from <= row_date <= p2_to:
            item["period2_qty"] = int(item["period2_qty"]) + qty
            if requested_price is not None:
                item["period2_amount"] = Decimal(
                    item["period2_amount"]
                ) + Decimal(str(requested_price)) * qty
                item["period2_price_qty"] = int(item["period2_price_qty"]) + qty

    stock_by_oem = await _load_latest_own_stock_by_oem(session)
    exclusions = await _load_active_exclusions(session)
    raw_rows = [
        item
        for item in grouped.values()
        if (
            int(item.get("period1_qty") or 0)
            + int(item.get("period2_qty") or 0)
        )
        >= normalized_min_total_qty
    ]

    def sort_key(item: dict[str, Any]) -> tuple[Any, ...]:
        if normalized_sort_by == "period1_desc":
            return (
                -int(item.get("period1_qty") or 0),
                str(item.get("brand_name") or ""),
                str(item.get("oem_number") or ""),
            )
        if normalized_sort_by == "period2_desc":
            return (
                -int(item.get("period2_qty") or 0),
                str(item.get("brand_name") or ""),
                str(item.get("oem_number") or ""),
            )
        if normalized_sort_by == "brand_oem":
            return (
                str(item.get("brand_name") or ""),
                str(item.get("oem_number") or ""),
            )
        return (
            -(
                int(item.get("period1_qty") or 0)
                + int(item.get("period2_qty") or 0)
            ),
            str(item.get("brand_name") or ""),
            str(item.get("oem_number") or ""),
        )

    raw_rows = sorted(raw_rows, key=sort_key)[:normalized_limit]
    rows: list[dict[str, Any]] = []
    for item in raw_rows:
        period1_qty = int(item.get("period1_qty") or 0)
        period2_qty = int(item.get("period2_qty") or 0)
        price_qty1 = int(item.get("period1_price_qty") or 0)
        avg_price1 = None
        if price_qty1 > 0:
            avg_price1 = float(Decimal(item["period1_amount"]) / price_qty1)
        price_qty2 = int(item.get("period2_price_qty") or 0)
        avg_price2 = None
        if price_qty2 > 0:
            avg_price2 = float(Decimal(item["period2_amount"]) / price_qty2)
        oem_number = str(item.get("oem_number") or "")
        exclusion = exclusions.get(
            _exclusion_key(oem_number, item.get("brand_name"))
        )
        rows.append(
            {
                "oem_number": oem_number,
                "brand_name": item.get("brand_name"),
                "autopart_name": item.get("autopart_name"),
                "current_quantity": stock_by_oem.get(oem_number, 0),
                "period1_qty": period1_qty,
                "period2_qty": period2_qty,
                "total_qty": period1_qty + period2_qty,
                "period1_avg_price": avg_price1,
                "period2_avg_price": avg_price2,
                "excluded_from_autopurchase": exclusion is not None,
            }
        )

    return {
        "period1_from": p1_from,
        "period1_to": p1_to,
        "period2_from": p2_from,
        "period2_to": p2_to,
        "brand": brand,
        "limit": normalized_limit,
        "min_total_qty": normalized_min_total_qty,
        "sort_by": normalized_sort_by,
        "total_items": len(rows),
        "summary": {
            "period1_qty": sum(int(row["period1_qty"]) for row in rows),
            "period2_qty": sum(int(row["period2_qty"]) for row in rows),
            "total_qty": sum(int(row["total_qty"]) for row in rows),
            "stock_qty": sum(int(row["current_quantity"]) for row in rows),
        },
        "rows": rows,
    }


async def build_customer_order_period_report_xlsx(
    session: AsyncSession,
    *,
    period1_from: date | None = None,
    period1_to: date | None = None,
    period2_from: date | None = None,
    period2_to: date | None = None,
    limit: int = 1000,
    brand: Optional[str] = None,
    min_total_qty: int = 1,
    sort_by: str = "total_desc",
) -> bytes:
    data = await build_customer_order_period_report_data(
        session=session,
        period1_from=period1_from,
        period1_to=period1_to,
        period2_from=period2_from,
        period2_to=period2_to,
        limit=limit,
        brand=brand,
        min_total_qty=min_total_qty,
        sort_by=sort_by,
    )
    p1_from = data["period1_from"]
    p1_to = data["period1_to"]
    p2_from = data["period2_from"]
    p2_to = data["period2_to"]
    rows = data["rows"]

    wb = Workbook()
    ws = wb.active
    ws.title = "TDSheet"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    ws.merge_cells("A2:G2")
    ws["A1"] = "Заказы клиентов по периодам с остатками"
    ws["A2"] = (
        f"Остатки на дату {_format_report_date(p2_to)}. "
        f"Период 1: {_format_report_period(p1_from, p1_to)}. "
        f"Период 2: {_format_report_period(p2_from, p2_to)}"
    )
    ws["A1"].font = Font(name="Arial", size=12, bold=True)
    ws["A2"].font = Font(name="Arial", size=10)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")

    headers = [
        "Артикул",
        "Производитель",
        "Наименование",
        "Кол-во на складе(Остаток)",
        "Кол-во заказанных клиентами позиций за период 1",
        "Кол-во заказанных клиентами позиций за период 2",
        "Средняя цена заказа единицы за период 2",
    ]
    ws.append([])
    ws.append(headers)
    header_row = 4
    for cell in ws[header_row]:
        cell.fill = PatternFill("solid", fgColor=REPORT_HEADER_FILL)
        cell.font = Font(name="Arial", bold=True, size=10)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = Border(
            left=REPORT_BORDER,
            right=REPORT_BORDER,
            top=REPORT_BORDER,
            bottom=REPORT_BORDER,
        )
    ws.row_dimensions[header_row].height = 52

    for item in rows:
        ws.append(
            [
                item.get("oem_number") or "",
                item.get("brand_name") or "",
                item.get("autopart_name") or "",
                item.get("current_quantity") or 0,
                item.get("period1_qty") or 0,
                item.get("period2_qty") or 0,
                item.get("period2_avg_price"),
            ]
        )

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.border = Border(
                left=REPORT_BORDER,
                right=REPORT_BORDER,
                top=REPORT_BORDER,
                bottom=REPORT_BORDER,
            )
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for idx in range(4, 8):
            row[idx - 1].alignment = Alignment(horizontal="right", vertical="top")
            row[idx - 1].number_format = "#,##0.00" if idx == 7 else "#,##0"

    widths = [18, 18, 44, 18, 22, 22, 22]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:G{max(ws.max_row, 4)}"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
