from datetime import timedelta
from io import BytesIO
from types import SimpleNamespace

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import select

import dz_fastapi.services.supplier_order_responses as response_service
from dz_fastapi.core.time import now_moscow
from dz_fastapi.models.order_status_mapping import (ExternalStatusMapping,
                                                    ExternalStatusMatchMode,
                                                    ExternalStatusUnmapped,
                                                    SupplierResponseAction)
from dz_fastapi.models.partner import (CUSTOMER_ORDER_SHIP_MODE,
                                       SUPPLIER_ORDER_STATUS, SupplierOrder,
                                       SupplierOrderAttachment,
                                       SupplierOrderItem, SupplierOrderMessage,
                                       SupplierReceipt, SupplierReceiptItem,
                                       SupplierResponseConfig)
from dz_fastapi.models.settings import CustomerOrderInboxSettings
from dz_fastapi.services.customer_orders import (
    _apply_response_updates_csv, _apply_response_updates_excel,
    _build_order_reply_recipients, _build_supplier_order_recipient,
    _customer_order_auto_reply_enabled, _customer_order_reply_override_email,
    _supplier_order_override_email, send_supplier_orders)
from dz_fastapi.services.order_status_mapping import \
    EXTERNAL_STATUS_SOURCE_SUPPLIER_EMAIL
from dz_fastapi.services.supplier_order_responses import (
    _extract_supplier_order_id, _get_message_match_context,
    process_supplier_response_messages)


def test_apply_response_updates_excel_writes_ship_price_when_configured():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=2, column=2).value = 5
    ws.cell(row=2, column=4).value = None

    source = BytesIO()
    wb.save(source)
    source.seek(0)

    config = SimpleNamespace(
        qty_col=1,
        ship_qty_col=None,
        ship_price_col=3,
        reject_qty_col=None,
        ship_mode=CUSTOMER_ORDER_SHIP_MODE.REPLACE_QTY,
    )
    item = SimpleNamespace(
        row_index=2,
        ship_qty=3,
        reject_qty=0,
        requested_price=3131.0,
    )

    output = _apply_response_updates_excel(source, config, [item])
    result = load_workbook(output)
    result_ws = result.active

    assert result_ws.cell(row=2, column=2).value == 3
    assert result_ws.cell(row=2, column=4).value == 3131.0


def test_apply_response_updates_csv_leaves_ship_price_blank_for_reject():
    df = pd.DataFrame(
        [
            ["OEM1", 2, "", ""],
            ["OEM2", 1, "", ""],
        ]
    )
    source = BytesIO()
    df.to_csv(source, index=False, header=False)
    source.seek(0)

    config = SimpleNamespace(
        qty_col=1,
        ship_qty_col=2,
        ship_price_col=3,
        reject_qty_col=None,
        ship_mode=CUSTOMER_ORDER_SHIP_MODE.WRITE_SHIP_QTY,
    )
    item = SimpleNamespace(
        row_index=1,
        ship_qty=0,
        reject_qty=1,
        requested_price=2500.0,
    )

    output = _apply_response_updates_csv(source, config, [item])
    result = pd.read_csv(output, header=None, keep_default_na=False)

    assert float(result.iat[1, 2]) == 0.0
    assert result.iat[1, 3] == ""


def test_customer_order_auto_reply_enabled_with_default_stub(monkeypatch):
    monkeypatch.delenv("CUSTOMER_ORDER_AUTO_REPLY_ENABLED", raising=False)
    monkeypatch.delenv("CUSTOMER_ORDER_REPLY_OVERRIDE_EMAIL", raising=False)
    assert _customer_order_reply_override_email() == "info@dragonzap.ru"
    assert _customer_order_auto_reply_enabled() is True


def test_customer_order_auto_reply_can_be_fully_disabled(monkeypatch):
    monkeypatch.setenv("CUSTOMER_ORDER_AUTO_REPLY_ENABLED", "0")
    monkeypatch.setenv("CUSTOMER_ORDER_REPLY_OVERRIDE_EMAIL", "")
    assert _customer_order_reply_override_email() is None
    assert _customer_order_auto_reply_enabled() is False


def test_build_order_reply_recipients_uses_stub_override(monkeypatch):
    monkeypatch.setenv(
        "CUSTOMER_ORDER_REPLY_OVERRIDE_EMAIL",
        "info@dragonzap.ru",
    )
    config = SimpleNamespace(
        order_reply_emails=["client@example.com", "sales@example.com"]
    )
    recipients = _build_order_reply_recipients(
        "sender@example.com",
        config,
    )
    original = _build_order_reply_recipients(
        "sender@example.com",
        config,
        use_override=False,
    )

    assert recipients == "info@dragonzap.ru"
    assert (
        original
        == "client@example.com,sales@example.com,sender@example.com"
    )


def test_supplier_order_override_email_defaults_to_stub(monkeypatch):
    monkeypatch.delenv("SUPPLIER_ORDER_OVERRIDE_EMAIL", raising=False)

    assert _supplier_order_override_email() == "info@dragonzap.ru"


def test_build_supplier_order_recipient_uses_stub_override(monkeypatch):
    monkeypatch.setenv("SUPPLIER_ORDER_OVERRIDE_EMAIL", "info@dragonzap.ru")
    provider = SimpleNamespace(email_contact="supplier@example.com")

    assert _build_supplier_order_recipient(provider) == "info@dragonzap.ru"
    assert (
        _build_supplier_order_recipient(provider, use_override=False)
        == "supplier@example.com"
    )


def test_extract_supplier_order_id_ignores_int32_overflow():
    assert (
        _extract_supplier_order_id("Заказ поставщику #5709826577")
        is None
    )
    assert (
        _extract_supplier_order_id("supplier_order_2147483647.xlsx")
        == 2147483647
    )


def test_extract_supplier_order_id_from_alnum_code():
    assert (
        _extract_supplier_order_id("МастерА0000002485.xlsx")
        == 2485
    )


@pytest.mark.asyncio
async def test_get_message_match_context_ignores_out_of_range_order_id(
    monkeypatch,
    test_session,
    created_providers,
):
    provider = created_providers[0]
    provider.email_contact = "supplier@example.com"
    await test_session.commit()

    monkeypatch.setattr(
        response_service,
        "_extract_supplier_order_id",
        lambda *args: 5_709_826_577,
    )

    order, matched_provider = await _get_message_match_context(
        test_session,
        sender_email="supplier@example.com",
        subject="Заказ поставщику #5709826577",
        body_preview=None,
        attachments=[],
    )

    assert order is None
    assert matched_provider is not None
    assert matched_provider.id == provider.id


@pytest.mark.asyncio
async def test_send_supplier_orders_uses_stub_override_email(
    monkeypatch,
    test_session,
    created_providers,
    created_autopart,
):
    provider = created_providers[0]
    provider.email_contact = "supplier@example.com"
    order = SupplierOrder(
        provider_id=provider.id,
        status=SUPPLIER_ORDER_STATUS.NEW,
    )
    test_session.add(order)
    await test_session.flush()
    test_session.add(
        SupplierOrderItem(
            supplier_order_id=order.id,
            autopart_id=created_autopart.id,
            quantity=2,
            price=150.0,
        )
    )
    await test_session.commit()

    sent_calls = []

    async def fake_send_email_attachment_async(
        to_email,
        subject,
        body,
        attachment,
        filename,
        use_tls,
        **kwargs,
    ):
        sent_calls.append(
            {
                "to_email": to_email,
                "subject": subject,
                "body": body,
                "filename": filename,
                "use_tls": use_tls,
                "attachment": attachment,
            }
        )

    async def fake_get_out_account(session, purpose):
        return None

    monkeypatch.setenv("SUPPLIER_ORDER_OVERRIDE_EMAIL", "info@dragonzap.ru")
    monkeypatch.setattr(
        "dz_fastapi.services.customer_orders._send_email_attachment_async",
        fake_send_email_attachment_async,
    )
    monkeypatch.setattr(
        "dz_fastapi.services.customer_orders._get_out_account",
        fake_get_out_account,
    )

    result = await send_supplier_orders(test_session, [order.id])
    await test_session.refresh(order)

    assert result == {"sent": 1, "failed": 0}
    assert order.status == SUPPLIER_ORDER_STATUS.SENT
    assert order.sent_at is not None
    assert len(sent_calls) == 1
    sent = sent_calls[0]
    assert sent["to_email"] == "info@dragonzap.ru"
    assert sent["subject"] == f"[STUB] Заказ поставщику № {order.id}"
    assert f"Заказ поставщику № {order.id}" in sent["body"]
    assert "Заглушка отправки заказа поставщику." in sent["body"]
    assert "Исходный адресат: supplier@example.com" in sent["body"]
    assert sent["filename"] == f"МастерА{order.id:010d}.xlsx"
    assert sent["use_tls"] is True
    workbook = load_workbook(BytesIO(sent["attachment"]))
    sheet = workbook.active
    assert sheet.title == "TDSheet"
    assert sheet["H1"].value == "Заказ поставщику"
    assert sheet["C3"].value == "Дата"
    assert sheet["C5"].value == order.id
    assert sheet["E5"].value == f"A{order.id:010d}"
    assert sheet["A9"].value == "№"
    assert sheet["C10"].value == created_autopart.oem_number
    assert sheet["G8"].value == 2
    assert float(sheet["I8"].value) == 300.0


@pytest.mark.asyncio
async def test_send_supplier_orders_uses_provider_email_when_stub_disabled(
    monkeypatch,
    test_session,
    created_providers,
    created_autopart,
):
    provider = created_providers[0]
    provider.email_contact = "supplier@example.com"
    order = SupplierOrder(
        provider_id=provider.id,
        status=SUPPLIER_ORDER_STATUS.NEW,
    )
    test_session.add(order)
    await test_session.flush()
    test_session.add(
        SupplierOrderItem(
            supplier_order_id=order.id,
            autopart_id=created_autopart.id,
            quantity=1,
            price=50.0,
        )
    )
    test_session.add(
        CustomerOrderInboxSettings(
            lookback_days=1,
            mark_seen=False,
            error_file_retention_days=5,
            supplier_response_lookback_days=14,
            supplier_order_stub_enabled=False,
            supplier_order_stub_email="info@dragonzap.ru",
        )
    )
    await test_session.commit()

    sent_calls = []

    async def fake_send_email_attachment_async(
        to_email,
        subject,
        body,
        attachment,
        filename,
        use_tls,
        **kwargs,
    ):
        sent_calls.append(
            {
                "to_email": to_email,
                "subject": subject,
                "body": body,
            }
        )

    async def fake_get_out_account(session, purpose):
        return None

    monkeypatch.setattr(
        "dz_fastapi.services.customer_orders._send_email_attachment_async",
        fake_send_email_attachment_async,
    )
    monkeypatch.setattr(
        "dz_fastapi.services.customer_orders._get_out_account",
        fake_get_out_account,
    )

    result = await send_supplier_orders(test_session, [order.id])

    assert result == {"sent": 1, "failed": 0}
    assert len(sent_calls) == 1
    sent = sent_calls[0]
    assert sent["to_email"] == "supplier@example.com"
    assert sent["subject"] == f"Заказ поставщику № {order.id}"
    assert f"Заказ поставщику № {order.id}" in sent["body"]
    assert (
        f"<b>Заказ поставщику № {order.id}</b>" in sent["body"]
    )
    assert "<table" in sent["body"]


@pytest.mark.asyncio
async def test_process_supplier_response_messages_updates_confirmed_quantities(
    monkeypatch,
    test_session,
    created_providers,
    created_autopart,
    tmp_path,
):
    provider = created_providers[0]
    provider.email_contact = "supplier@example.com"
    order = SupplierOrder(
        provider_id=provider.id,
        status=SUPPLIER_ORDER_STATUS.SENT,
    )
    test_session.add(order)
    await test_session.flush()
    item = SupplierOrderItem(
        supplier_order_id=order.id,
        autopart_id=created_autopart.id,
        oem_number=created_autopart.oem_number,
        brand_name=created_autopart.brand.name,
        autopart_name=created_autopart.name,
        quantity=5,
        price=150.0,
    )
    test_session.add(item)
    test_session.add(
        ExternalStatusMapping(
            source_key=EXTERNAL_STATUS_SOURCE_SUPPLIER_EMAIL,
            provider_id=provider.id,
            raw_status="готово",
            normalized_status="готово",
            match_mode=ExternalStatusMatchMode.EXACT,
            supplier_response_action=SupplierResponseAction.FULL_CONFIRM.value,
            priority=10,
            is_active=True,
        )
    )
    await test_session.commit()

    response_frame = pd.DataFrame(
        [
            {
                "OEM": created_autopart.oem_number,
                "Brand": created_autopart.brand.name,
                "Qty": 2,
                "Price": 77.5,
                "Comment": "Частично подтверждено",
            }
        ]
    )
    buffer = BytesIO()
    response_frame.to_excel(buffer, index=False)
    payload = buffer.getvalue()

    async def fake_fetch_messages(session, *, date_from, date_to=None):
        return [
            (
                SimpleNamespace(
                    from_="supplier@example.com",
                    subject=f"Заказ поставщику #{order.id} готово",
                    text="готово",
                    html=None,
                    attachments=[
                        SimpleNamespace(
                            filename=f"supplier_order_{order.id}.xlsx",
                            payload=payload,
                        )
                    ],
                    received_at=None,
                    date=None,
                    external_id="supplier-msg-1",
                    uid=None,
                    folder_name="INBOX",
                ),
                None,
            )
        ]

    monkeypatch.setattr(
        (
            "dz_fastapi.services.supplier_order_responses."
            "_fetch_supplier_response_messages"
        ),
        fake_fetch_messages,
    )
    monkeypatch.setattr(
        "dz_fastapi.services.supplier_order_responses.SUPPLIER_RESPONSE_DIR",
        str(tmp_path),
    )

    result = await process_supplier_response_messages(test_session)
    await test_session.refresh(order)
    await test_session.refresh(item)

    assert result["processed_messages"] == 1
    assert result["parsed_response_files"] == 1
    assert order.response_status_raw == "готово"
    assert item.confirmed_quantity == 2
    assert float(item.response_price) == 77.5
    assert item.response_comment == "Частично подтверждено"


@pytest.mark.asyncio
async def test_process_supplier_response_creates_draft_receipt(
    monkeypatch,
    test_session,
    created_providers,
    created_autopart,
    tmp_path,
):
    provider = created_providers[0]
    provider.email_contact = "supplier@example.com"
    order = SupplierOrder(
        provider_id=provider.id,
        status=SUPPLIER_ORDER_STATUS.SENT,
    )
    test_session.add(order)
    await test_session.flush()
    item = SupplierOrderItem(
        supplier_order_id=order.id,
        autopart_id=created_autopart.id,
        oem_number=created_autopart.oem_number,
        brand_name=created_autopart.brand.name,
        autopart_name=created_autopart.name,
        quantity=5,
        price=150.0,
    )
    test_session.add(item)
    await test_session.commit()

    response_frame = pd.DataFrame(
        [
            {
                "OEM": created_autopart.oem_number,
                "Brand": created_autopart.brand.name,
                "Qty": 2,
                "Price": 90.0,
            }
        ]
    )
    buffer = BytesIO()
    response_frame.to_excel(buffer, index=False)
    payload = buffer.getvalue()

    async def fake_fetch_messages(session, *, date_from, date_to=None):
        return [
            (
                SimpleNamespace(
                    from_="supplier@example.com",
                    subject=f"Заказ поставщику #{order.id}",
                    text="",
                    html=None,
                    attachments=[
                        SimpleNamespace(
                            filename=f"supplier_order_{order.id}.xlsx",
                            payload=payload,
                        )
                    ],
                    received_at=None,
                    date=None,
                    external_id="supplier-msg-draft-receipt",
                    uid=None,
                    folder_name="INBOX",
                ),
                None,
            )
        ]

    monkeypatch.setattr(
        (
            "dz_fastapi.services.supplier_order_responses."
            "_fetch_supplier_response_messages"
        ),
        fake_fetch_messages,
    )
    monkeypatch.setattr(
        "dz_fastapi.services.supplier_order_responses.SUPPLIER_RESPONSE_DIR",
        str(tmp_path),
    )

    result = await process_supplier_response_messages(test_session)
    await test_session.refresh(item)

    receipt = (
        await test_session.execute(
            select(SupplierReceipt).where(
                SupplierReceipt.provider_id == provider.id
            )
        )
    ).scalar_one()
    receipt_item = (
        await test_session.execute(
            select(SupplierReceiptItem).where(
                SupplierReceiptItem.receipt_id == receipt.id
            )
        )
    ).scalar_one()

    assert result["processed_messages"] == 1
    assert result["created_receipts"] == 1
    assert result["draft_receipts"] == 1
    assert result["receipt_items_added"] == 1
    assert receipt.posted_at is None
    assert receipt_item.received_quantity == 2
    assert int(item.received_quantity or 0) == 0


@pytest.mark.asyncio
async def test_process_supplier_response_without_order_id_matches_positions(
    monkeypatch,
    test_session,
    created_providers,
    created_autopart,
    tmp_path,
):
    provider = created_providers[0]
    provider.email_contact = "supplier@example.com"
    order = SupplierOrder(
        provider_id=provider.id,
        status=SUPPLIER_ORDER_STATUS.SENT,
    )
    test_session.add(order)
    await test_session.flush()
    item = SupplierOrderItem(
        supplier_order_id=order.id,
        autopart_id=created_autopart.id,
        oem_number=created_autopart.oem_number,
        brand_name=created_autopart.brand.name,
        autopart_name=created_autopart.name,
        quantity=4,
        price=120.0,
    )
    test_session.add(item)
    await test_session.commit()

    response_frame = pd.DataFrame(
        [
            {
                "OEM": created_autopart.oem_number,
                "Brand": created_autopart.brand.name,
                "Qty": 3,
                "Price": 80.0,
            }
        ]
    )
    buffer = BytesIO()
    response_frame.to_excel(buffer, index=False)
    payload = buffer.getvalue()

    async def fake_fetch_messages(session, *, date_from, date_to=None):
        return [
            (
                SimpleNamespace(
                    from_="supplier@example.com",
                    subject="Ответ по заявке без номера",
                    text="",
                    html=None,
                    attachments=[
                        SimpleNamespace(
                            filename="supplier_answer.xlsx",
                            payload=payload,
                        )
                    ],
                    received_at=None,
                    date=None,
                    external_id="supplier-msg-no-order-id",
                    uid=None,
                    folder_name="INBOX",
                ),
                None,
            )
        ]

    monkeypatch.setattr(
        (
            "dz_fastapi.services.supplier_order_responses."
            "_fetch_supplier_response_messages"
        ),
        fake_fetch_messages,
    )
    monkeypatch.setattr(
        "dz_fastapi.services.supplier_order_responses.SUPPLIER_RESPONSE_DIR",
        str(tmp_path),
    )

    result = await process_supplier_response_messages(test_session)
    await test_session.refresh(item)
    message_row = (
        await test_session.execute(
            select(SupplierOrderMessage).where(
                SupplierOrderMessage.source_message_id
                == "supplier-msg-no-order-id"
            )
        )
    ).scalar_one()

    assert result["processed_messages"] == 1
    assert result["recognized_positions"] == 1
    assert result["created_receipts"] == 1
    assert result["draft_receipts"] == 1
    assert item.confirmed_quantity == 3
    assert float(item.response_price) == 80.0
    assert message_row.supplier_order_id == order.id


@pytest.mark.asyncio
async def test_process_supplier_response_appends_existing_draft_receipt(
    monkeypatch,
    test_session,
    created_providers,
    created_autopart,
    tmp_path,
):
    provider = created_providers[0]
    provider.email_contact = "supplier@example.com"
    existing_draft = SupplierReceipt(
        provider_id=provider.id,
        document_date=pd.Timestamp.now().date(),
        posted_at=None,
    )
    test_session.add(existing_draft)
    await test_session.flush()
    order = SupplierOrder(
        provider_id=provider.id,
        status=SUPPLIER_ORDER_STATUS.SENT,
    )
    test_session.add(order)
    await test_session.flush()
    item = SupplierOrderItem(
        supplier_order_id=order.id,
        autopart_id=created_autopart.id,
        oem_number=created_autopart.oem_number,
        brand_name=created_autopart.brand.name,
        autopart_name=created_autopart.name,
        quantity=4,
        price=120.0,
    )
    test_session.add(item)
    await test_session.commit()

    response_frame = pd.DataFrame(
        [
            {
                "OEM": created_autopart.oem_number,
                "Qty": 3,
            }
        ]
    )
    buffer = BytesIO()
    response_frame.to_excel(buffer, index=False)
    payload = buffer.getvalue()

    async def fake_fetch_messages(session, *, date_from, date_to=None):
        return [
            (
                SimpleNamespace(
                    from_="supplier@example.com",
                    subject=f"Заказ поставщику #{order.id}",
                    text="",
                    html=None,
                    attachments=[
                        SimpleNamespace(
                            filename=f"supplier_order_{order.id}.xlsx",
                            payload=payload,
                        )
                    ],
                    received_at=None,
                    date=None,
                    external_id="supplier-msg-append-receipt",
                    uid=None,
                    folder_name="INBOX",
                ),
                None,
            )
        ]

    monkeypatch.setattr(
        (
            "dz_fastapi.services.supplier_order_responses."
            "_fetch_supplier_response_messages"
        ),
        fake_fetch_messages,
    )
    monkeypatch.setattr(
        "dz_fastapi.services.supplier_order_responses.SUPPLIER_RESPONSE_DIR",
        str(tmp_path),
    )

    result = await process_supplier_response_messages(test_session)
    receipts = (
        await test_session.execute(
            select(SupplierReceipt).where(
                SupplierReceipt.provider_id == provider.id
            )
        )
    ).scalars().all()
    assert len(receipts) == 1
    assert receipts[0].id == existing_draft.id

    receipt_items = (
        await test_session.execute(
            select(SupplierReceiptItem).where(
                SupplierReceiptItem.receipt_id == existing_draft.id
            )
        )
    ).scalars().all()

    assert result["processed_messages"] == 1
    assert result["created_receipts"] == 0
    assert result["updated_receipts"] == 1
    assert result["receipt_items_added"] == 1
    assert len(receipt_items) == 1
    assert receipt_items[0].received_quantity == 3


@pytest.mark.asyncio
async def test_document_payload_file_posts_receipt_with_doc_fields(
    monkeypatch,
    test_session,
    created_providers,
    created_autopart,
    tmp_path,
):
    provider = created_providers[0]
    provider.email_contact = "supplier@example.com"
    order = SupplierOrder(
        provider_id=provider.id,
        status=SUPPLIER_ORDER_STATUS.SENT,
    )
    test_session.add(order)
    await test_session.flush()
    item = SupplierOrderItem(
        supplier_order_id=order.id,
        autopart_id=created_autopart.id,
        oem_number=created_autopart.oem_number,
        brand_name=created_autopart.brand.name,
        autopart_name=created_autopart.name,
        quantity=3,
        price=150.0,
    )
    test_session.add(item)
    test_session.add(
        SupplierResponseConfig(
            provider_id=provider.id,
            name="Doc parser",
            sender_emails=["supplier@example.com"],
            response_type="file",
            file_payload_type="document",
            file_format="excel",
            start_row=1,
            oem_col=1,
            brand_col=2,
            qty_col=3,
            total_price_with_vat_col=4,
            document_number_col=5,
            document_date_col=6,
            gtd_col=7,
            country_code_col=8,
            country_name_col=9,
            process_shipping_docs=True,
            is_active=True,
        )
    )
    await test_session.commit()

    response_frame = pd.DataFrame(
        [
            [
                created_autopart.oem_number,
                created_autopart.brand.name,
                2,
                240.0,
                "UPD-7788",
                "09.04.2026",
                "1234567890",
                "156",
                "Китай",
            ]
        ]
    )
    buffer = BytesIO()
    response_frame.to_excel(buffer, index=False, header=False)
    payload = buffer.getvalue()

    async def fake_fetch_messages(
            session, *, date_from, date_to=None, **kwargs
    ):
        return [
            (
                SimpleNamespace(
                    from_="supplier@example.com",
                    subject=f"Заказ поставщику #{order.id}: документ поставки",
                    text="",
                    html=None,
                    attachments=[
                        SimpleNamespace(
                            filename="doc_rows.xlsx",
                            payload=payload,
                        )
                    ],
                    received_at=None,
                    date=None,
                    external_id="supplier-msg-doc-file",
                    uid=None,
                    folder_name="INBOX",
                ),
                None,
            )
        ]

    monkeypatch.setattr(
        (
            "dz_fastapi.services.supplier_order_responses."
            "_fetch_supplier_response_messages"
        ),
        fake_fetch_messages,
    )
    monkeypatch.setattr(
        "dz_fastapi.services.supplier_order_responses.SUPPLIER_RESPONSE_DIR",
        str(tmp_path),
    )

    result = await process_supplier_response_messages(test_session)
    await test_session.refresh(item)
    receipt = (
        await test_session.execute(
            select(SupplierReceipt).where(
                SupplierReceipt.provider_id == provider.id
            )
        )
    ).scalar_one()
    receipt_item = (
        await test_session.execute(
            select(SupplierReceiptItem).where(
                SupplierReceiptItem.receipt_id == receipt.id
            )
        )
    ).scalar_one()

    assert result["processed_messages"] == 1
    assert result["posted_receipts"] == 1
    assert receipt.posted_at is not None
    assert receipt.document_number == "UPD-7788"
    assert str(receipt.document_date) == "2026-04-09"
    assert float(receipt_item.price) == 120.0
    assert float(receipt_item.total_price_with_vat) == 240.0
    assert receipt_item.gtd_code == "1234567890"
    assert receipt_item.country_code == "156"
    assert receipt_item.country_name == "Китай"
    assert int(item.received_quantity or 0) == 2


@pytest.mark.asyncio
async def test_shipping_doc_creates_posted_receipt(
    monkeypatch,
    test_session,
    created_providers,
    created_autopart,
    tmp_path,
):
    provider = created_providers[0]
    provider.email_contact = "supplier@example.com"
    provider.supplier_response_allow_text_status = False
    order = SupplierOrder(
        provider_id=provider.id,
        status=SUPPLIER_ORDER_STATUS.SENT,
    )
    test_session.add(order)
    await test_session.flush()
    item = SupplierOrderItem(
        supplier_order_id=order.id,
        autopart_id=created_autopart.id,
        oem_number=created_autopart.oem_number,
        brand_name=created_autopart.brand.name,
        autopart_name=created_autopart.name,
        quantity=2,
        price=110.0,
    )
    test_session.add(item)
    await test_session.commit()

    async def fake_fetch_messages(session, *, date_from, date_to=None):
        return [
            (
                SimpleNamespace(
                    from_="supplier@example.com",
                    subject=f"Заказ поставщику #{order.id} документы",
                    text="Во вложении УПД",
                    html=None,
                    attachments=[
                        SimpleNamespace(
                            filename="upd_20260409.pdf",
                            payload=b"%PDF-upd%",
                        )
                    ],
                    received_at=None,
                    date=None,
                    external_id="supplier-msg-upd-auto-post",
                    uid=None,
                    folder_name="INBOX",
                ),
                None,
            )
        ]

    monkeypatch.setattr(
        (
            "dz_fastapi.services.supplier_order_responses."
            "_fetch_supplier_response_messages"
        ),
        fake_fetch_messages,
    )
    monkeypatch.setattr(
        "dz_fastapi.services.supplier_order_responses.SUPPLIER_RESPONSE_DIR",
        str(tmp_path),
    )

    result = await process_supplier_response_messages(test_session)
    await test_session.refresh(item)
    receipt = (
        await test_session.execute(
            select(SupplierReceipt).where(
                SupplierReceipt.provider_id == provider.id
            )
        )
    ).scalar_one()
    receipt_item = (
        await test_session.execute(
            select(SupplierReceiptItem).where(
                SupplierReceiptItem.receipt_id == receipt.id
            )
        )
    ).scalar_one()

    assert result["processed_messages"] == 1
    assert result["created_receipts"] == 1
    assert result["posted_receipts"] == 1
    assert receipt.posted_at is not None
    assert receipt.document_number == "upd_20260409"
    assert receipt_item.received_quantity == 2
    assert int(item.received_quantity or 0) == 2


@pytest.mark.asyncio
async def test_process_supplier_response_messages_records_unmapped_status(
    monkeypatch,
    test_session,
    created_providers,
    tmp_path,
):
    provider = created_providers[0]
    provider.email_contact = "supplier@example.com"
    order = SupplierOrder(
        provider_id=provider.id,
        status=SUPPLIER_ORDER_STATUS.SENT,
    )
    test_session.add(order)
    await test_session.commit()

    async def fake_fetch_messages(session, *, date_from, date_to=None):
        return [
            (
                SimpleNamespace(
                    from_="supplier@example.com",
                    subject=f"Заказ поставщику #{order.id} manual review",
                    text="manual review",
                    html=None,
                    attachments=[],
                    received_at=None,
                    date=None,
                    external_id="supplier-msg-2",
                    uid=None,
                    folder_name="INBOX",
                ),
                None,
            )
        ]

    monkeypatch.setattr(
        (
            "dz_fastapi.services.supplier_order_responses."
            "_fetch_supplier_response_messages"
        ),
        fake_fetch_messages,
    )
    monkeypatch.setattr(
        "dz_fastapi.services.supplier_order_responses.SUPPLIER_RESPONSE_DIR",
        str(tmp_path),
    )

    result = await process_supplier_response_messages(test_session)

    row = (
        await test_session.execute(
            select(ExternalStatusUnmapped).where(
                ExternalStatusUnmapped.source_key
                == EXTERNAL_STATUS_SOURCE_SUPPLIER_EMAIL
            )
        )
    ).scalar_one()

    assert result["processed_messages"] == 1
    assert result["unmapped_statuses"] == 1
    assert row.sample_payload["supplier_order_id"] == order.id


@pytest.mark.asyncio
async def test_supplier_response_skip_text_status_when_disabled(
    monkeypatch,
    test_session,
    created_providers,
    tmp_path,
):
    provider = created_providers[0]
    provider.email_contact = "supplier@example.com"
    provider.supplier_response_allow_text_status = False
    order = SupplierOrder(
        provider_id=provider.id,
        status=SUPPLIER_ORDER_STATUS.SENT,
    )
    test_session.add(order)
    await test_session.commit()

    async def fake_fetch_messages(session, *, date_from, date_to=None):
        return [
            (
                SimpleNamespace(
                    from_="supplier@example.com",
                    subject=f"Заказ поставщику #{order.id} manual review",
                    text="manual review",
                    html=None,
                    attachments=[],
                    received_at=None,
                    date=None,
                    external_id="supplier-msg-text-off",
                    uid=None,
                    folder_name="INBOX",
                ),
                None,
            )
        ]

    monkeypatch.setattr(
        (
            "dz_fastapi.services.supplier_order_responses."
            "_fetch_supplier_response_messages"
        ),
        fake_fetch_messages,
    )
    monkeypatch.setattr(
        "dz_fastapi.services.supplier_order_responses.SUPPLIER_RESPONSE_DIR",
        str(tmp_path),
    )

    result = await process_supplier_response_messages(test_session)
    await test_session.refresh(order)

    unmapped_rows = (
        await test_session.execute(
            select(ExternalStatusUnmapped).where(
                ExternalStatusUnmapped.source_key
                == EXTERNAL_STATUS_SOURCE_SUPPLIER_EMAIL,
                ExternalStatusUnmapped.provider_id == provider.id,
            )
        )
    ).scalars().all()
    message_row = (
        await test_session.execute(
            select(SupplierOrderMessage).where(
                SupplierOrderMessage.source_message_id
                == "supplier-msg-text-off"
            )
        )
    ).scalar_one()

    assert result["processed_messages"] == 1
    assert result["unmapped_statuses"] == 0
    assert order.response_status_raw is None
    assert message_row.raw_status is None
    assert message_row.message_type == "UNKNOWN"
    assert len(unmapped_rows) == 0


@pytest.mark.asyncio
async def test_skip_response_files_when_disabled(
    monkeypatch,
    test_session,
    created_providers,
    created_autopart,
    tmp_path,
):
    provider = created_providers[0]
    provider.email_contact = "supplier@example.com"
    provider.supplier_response_allow_response_files = False
    provider.supplier_response_allow_text_status = False
    order = SupplierOrder(
        provider_id=provider.id,
        status=SUPPLIER_ORDER_STATUS.SENT,
    )
    test_session.add(order)
    await test_session.flush()
    item = SupplierOrderItem(
        supplier_order_id=order.id,
        autopart_id=created_autopart.id,
        oem_number=created_autopart.oem_number,
        brand_name=created_autopart.brand.name,
        autopart_name=created_autopart.name,
        quantity=5,
        price=150.0,
    )
    test_session.add(item)
    await test_session.commit()

    response_frame = pd.DataFrame(
        [
            {
                "OEM": created_autopart.oem_number,
                "Brand": created_autopart.brand.name,
                "Qty": 2,
                "Price": 77.5,
            }
        ]
    )
    buffer = BytesIO()
    response_frame.to_excel(buffer, index=False)
    payload = buffer.getvalue()

    async def fake_fetch_messages(session, *, date_from, date_to=None):
        return [
            (
                SimpleNamespace(
                    from_="supplier@example.com",
                    subject=f"Заказ поставщику #{order.id}",
                    text="",
                    html=None,
                    attachments=[
                        SimpleNamespace(
                            filename=f"supplier_order_{order.id}.xlsx",
                            payload=payload,
                        )
                    ],
                    received_at=None,
                    date=None,
                    external_id="supplier-msg-files-off",
                    uid=None,
                    folder_name="INBOX",
                ),
                None,
            )
        ]

    monkeypatch.setattr(
        (
            "dz_fastapi.services.supplier_order_responses."
            "_fetch_supplier_response_messages"
        ),
        fake_fetch_messages,
    )
    monkeypatch.setattr(
        "dz_fastapi.services.supplier_order_responses.SUPPLIER_RESPONSE_DIR",
        str(tmp_path),
    )

    result = await process_supplier_response_messages(test_session)
    await test_session.refresh(item)

    assert result["processed_messages"] == 1
    assert result["parsed_response_files"] == 0
    assert item.confirmed_quantity is None
    assert item.response_price is None


@pytest.mark.asyncio
async def test_skip_shipping_docs_when_disabled(
    monkeypatch,
    test_session,
    created_providers,
    tmp_path,
):
    provider = created_providers[0]
    provider.email_contact = "supplier@example.com"
    provider.supplier_response_allow_shipping_docs = False
    provider.supplier_response_allow_text_status = False
    await test_session.commit()

    async def fake_fetch_messages(session, *, date_from, date_to=None):
        return [
            (
                SimpleNamespace(
                    from_="supplier@example.com",
                    subject="Документы по поставке",
                    text="Во вложении УПД",
                    html=None,
                    attachments=[
                        SimpleNamespace(
                            filename="upd_001.pdf",
                            payload=b"%PDF-test%",
                        )
                    ],
                    received_at=None,
                    date=None,
                    external_id="supplier-msg-upd-off",
                    uid=None,
                    folder_name="INBOX",
                ),
                None,
            )
        ]

    monkeypatch.setattr(
        (
            "dz_fastapi.services.supplier_order_responses."
            "_fetch_supplier_response_messages"
        ),
        fake_fetch_messages,
    )
    monkeypatch.setattr(
        "dz_fastapi.services.supplier_order_responses.SUPPLIER_RESPONSE_DIR",
        str(tmp_path),
    )

    result = await process_supplier_response_messages(test_session)
    message_row = (
        await test_session.execute(
            select(SupplierOrderMessage).where(
                SupplierOrderMessage.source_message_id
                == "supplier-msg-upd-off"
            )
        )
    ).scalar_one()
    attachment_row = (
        await test_session.execute(
            select(SupplierOrderAttachment).where(
                SupplierOrderAttachment.message_id == message_row.id
            )
        )
    ).scalar_one()

    assert result["processed_messages"] == 1
    assert message_row.message_type == "UNKNOWN"
    assert attachment_row.parsed_kind is None


@pytest.mark.asyncio
async def test_apply_provider_column_layout_for_response_file(
    monkeypatch,
    test_session,
    created_providers,
    created_autopart,
    tmp_path,
):
    provider = created_providers[0]
    provider.email_contact = "supplier@example.com"
    provider.supplier_response_start_row = 2
    provider.supplier_response_oem_col = 1
    provider.supplier_response_brand_col = 2
    provider.supplier_response_qty_col = 3
    provider.supplier_response_price_col = 4
    provider.supplier_response_comment_col = 5
    provider.supplier_response_status_col = 6
    order = SupplierOrder(
        provider_id=provider.id,
        status=SUPPLIER_ORDER_STATUS.SENT,
    )
    test_session.add(order)
    await test_session.flush()
    item = SupplierOrderItem(
        supplier_order_id=order.id,
        autopart_id=created_autopart.id,
        oem_number=created_autopart.oem_number,
        brand_name=created_autopart.brand.name,
        autopart_name=created_autopart.name,
        quantity=5,
        price=150.0,
    )
    test_session.add(item)
    await test_session.commit()

    response_frame = pd.DataFrame(
        [
            ["meta", "meta", "meta", "meta", "meta", "meta"],
            [
                created_autopart.oem_number,
                created_autopart.brand.name,
                3,
                88.5,
                "Подтверждено файлом",
                "готово",
            ],
        ]
    )
    buffer = BytesIO()
    response_frame.to_excel(buffer, index=False, header=False)
    payload = buffer.getvalue()

    async def fake_fetch_messages(session, *, date_from, date_to=None):
        return [
            (
                SimpleNamespace(
                    from_="supplier@example.com",
                    subject=f"Заказ поставщику #{order.id}",
                    text="",
                    html=None,
                    attachments=[
                        SimpleNamespace(
                            filename=f"supplier_order_{order.id}.xlsx",
                            payload=payload,
                        )
                    ],
                    received_at=None,
                    date=None,
                    external_id="supplier-msg-layout",
                    uid=None,
                    folder_name="INBOX",
                ),
                None,
            )
        ]

    monkeypatch.setattr(
        (
            "dz_fastapi.services.supplier_order_responses."
            "_fetch_supplier_response_messages"
        ),
        fake_fetch_messages,
    )
    monkeypatch.setattr(
        "dz_fastapi.services.supplier_order_responses.SUPPLIER_RESPONSE_DIR",
        str(tmp_path),
    )

    result = await process_supplier_response_messages(test_session)
    await test_session.refresh(item)

    assert result["processed_messages"] == 1
    assert result["parsed_response_files"] == 1
    assert item.confirmed_quantity == 3
    assert float(item.response_price) == 88.5
    assert item.response_comment == "Подтверждено файлом"
    assert item.response_status_raw == "готово"


@pytest.mark.asyncio
async def test_use_response_filename_pattern_for_spreadsheets(
    monkeypatch,
    test_session,
    created_providers,
    created_autopart,
    tmp_path,
):
    provider = created_providers[0]
    provider.email_contact = "supplier@example.com"
    provider.supplier_response_allow_text_status = False
    provider.supplier_response_filename_pattern = r"^answer_\d+\.xlsx$"
    order = SupplierOrder(
        provider_id=provider.id,
        status=SUPPLIER_ORDER_STATUS.SENT,
    )
    test_session.add(order)
    await test_session.flush()
    item = SupplierOrderItem(
        supplier_order_id=order.id,
        autopart_id=created_autopart.id,
        oem_number=created_autopart.oem_number,
        brand_name=created_autopart.brand.name,
        autopart_name=created_autopart.name,
        quantity=5,
        price=150.0,
    )
    test_session.add(item)
    await test_session.commit()

    response_frame = pd.DataFrame(
        [
            {
                "OEM": created_autopart.oem_number,
                "Brand": created_autopart.brand.name,
                "Qty": 4,
            }
        ]
    )
    buffer = BytesIO()
    response_frame.to_excel(buffer, index=False)
    payload = buffer.getvalue()

    async def fake_fetch_messages(session, *, date_from, date_to=None):
        return [
            (
                SimpleNamespace(
                    from_="supplier@example.com",
                    subject=f"Заказ поставщику #{order.id}",
                    text="",
                    html=None,
                    attachments=[
                        SimpleNamespace(
                            filename="supplier_order_response.xlsx",
                            payload=payload,
                        )
                    ],
                    received_at=None,
                    date=None,
                    external_id="supplier-msg-pattern",
                    uid=None,
                    folder_name="INBOX",
                ),
                None,
            )
        ]

    monkeypatch.setattr(
        (
            "dz_fastapi.services.supplier_order_responses."
            "_fetch_supplier_response_messages"
        ),
        fake_fetch_messages,
    )
    monkeypatch.setattr(
        "dz_fastapi.services.supplier_order_responses.SUPPLIER_RESPONSE_DIR",
        str(tmp_path),
    )

    result = await process_supplier_response_messages(test_session)
    await test_session.refresh(item)

    assert result["processed_messages"] == 1
    assert result["parsed_response_files"] == 0
    assert item.confirmed_quantity is None


@pytest.mark.asyncio
async def test_use_response_filename_pattern_for_mime_encoded_filename(
    monkeypatch,
    test_session,
    created_providers,
    created_autopart,
    tmp_path,
):
    provider = created_providers[0]
    provider.email_contact = "otvet@aruda.ru"
    provider.supplier_response_allow_text_status = False
    provider.supplier_response_filename_pattern = r"Ответ"
    order = SupplierOrder(
        provider_id=provider.id,
        status=SUPPLIER_ORDER_STATUS.SENT,
    )
    test_session.add(order)
    await test_session.flush()
    item = SupplierOrderItem(
        supplier_order_id=order.id,
        autopart_id=created_autopart.id,
        oem_number=created_autopart.oem_number,
        brand_name=created_autopart.brand.name,
        autopart_name=created_autopart.name,
        quantity=5,
        price=150.0,
    )
    test_session.add(item)
    await test_session.commit()

    response_frame = pd.DataFrame(
        [
            {
                "OEM": created_autopart.oem_number,
                "Brand": created_autopart.brand.name,
                "Qty": 2,
            }
        ]
    )
    buffer = BytesIO()
    response_frame.to_excel(buffer, index=False)
    payload = buffer.getvalue()

    async def fake_fetch_messages(session, *, date_from, date_to=None):
        return [
            (
                SimpleNamespace(
                    from_="otvet@aruda.ru",
                    subject=f"Ответ по заказу {order.id}",
                    text="",
                    html=None,
                    attachments=[
                        SimpleNamespace(
                            filename="=?UTF-8?B?0J7RgtCy0LXRgi5YTFM=?=",
                            payload=payload,
                        )
                    ],
                    received_at=None,
                    date=None,
                    external_id="supplier-msg-mime-pattern",
                    uid=None,
                    folder_name="INBOX",
                ),
                None,
            )
        ]

    monkeypatch.setattr(
        (
            "dz_fastapi.services.supplier_order_responses."
            "_fetch_supplier_response_messages"
        ),
        fake_fetch_messages,
    )
    monkeypatch.setattr(
        "dz_fastapi.services.supplier_order_responses.SUPPLIER_RESPONSE_DIR",
        str(tmp_path),
    )

    result = await process_supplier_response_messages(test_session)
    await test_session.refresh(item)

    assert result["processed_messages"] == 1
    assert result["parsed_response_files"] == 1
    assert item.confirmed_quantity == 2


@pytest.mark.asyncio
async def test_use_shipping_doc_filename_pattern(
    monkeypatch,
    test_session,
    created_providers,
    tmp_path,
):
    provider = created_providers[0]
    provider.email_contact = "supplier@example.com"
    provider.supplier_response_allow_text_status = False
    provider.supplier_shipping_doc_filename_pattern = r"^doc_\d+\.pdf$"
    await test_session.commit()

    async def fake_fetch_messages(session, *, date_from, date_to=None):
        return [
            (
                SimpleNamespace(
                    from_="supplier@example.com",
                    subject="Документы по поставке",
                    text="",
                    html=None,
                    attachments=[
                        SimpleNamespace(
                            filename="doc_123.pdf",
                            payload=b"%PDF-custom%",
                        )
                    ],
                    received_at=None,
                    date=None,
                    external_id="supplier-msg-doc-pattern",
                    uid=None,
                    folder_name="INBOX",
                ),
                None,
            )
        ]

    monkeypatch.setattr(
        (
            "dz_fastapi.services.supplier_order_responses."
            "_fetch_supplier_response_messages"
        ),
        fake_fetch_messages,
    )
    monkeypatch.setattr(
        "dz_fastapi.services.supplier_order_responses.SUPPLIER_RESPONSE_DIR",
        str(tmp_path),
    )

    result = await process_supplier_response_messages(test_session)
    message_row = (
        await test_session.execute(
            select(SupplierOrderMessage).where(
                SupplierOrderMessage.source_message_id
                == "supplier-msg-doc-pattern"
            )
        )
    ).scalar_one()
    attachment_row = (
        await test_session.execute(
            select(SupplierOrderAttachment).where(
                SupplierOrderAttachment.message_id == message_row.id
            )
        )
    ).scalar_one()

    assert result["processed_messages"] == 1
    assert message_row.message_type == "SHIPPING_DOC"
    assert attachment_row.parsed_kind == "SHIPPING_DOC"


@pytest.mark.asyncio
async def test_parse_text_response_using_supplier_response_config(
    monkeypatch,
    test_session,
    created_providers,
    created_autopart,
    tmp_path,
):
    provider = created_providers[0]
    provider.email_contact = None
    provider.email_incoming_price = None
    order = SupplierOrder(
        provider_id=provider.id,
        status=SUPPLIER_ORDER_STATUS.SENT,
    )
    test_session.add(order)
    await test_session.flush()
    item = SupplierOrderItem(
        supplier_order_id=order.id,
        autopart_id=created_autopart.id,
        oem_number=created_autopart.oem_number,
        brand_name=created_autopart.brand.name,
        autopart_name=created_autopart.name,
        quantity=5,
        price=150.0,
    )
    test_session.add(item)
    test_session.add(
        SupplierResponseConfig(
            provider_id=provider.id,
            name="Text parser config",
            sender_emails=["supplier@example.com"],
            response_type="text",
            confirm_keywords=["да", "есть"],
            reject_keywords=["нет", "0"],
            value_after_article_type="both",
            is_active=True,
        )
    )
    await test_session.commit()

    text_payload = (
        f"{created_autopart.oem_number} 3 "
        "UNKNOWN999 0 "
        f"{created_autopart.oem_number} да"
    )

    async def fake_fetch_messages(
            session, *, date_from, date_to=None, **kwargs
    ):
        return [
            (
                SimpleNamespace(
                    from_="supplier@example.com",
                    subject=f"Заказ поставщику #{order.id}",
                    text=text_payload,
                    html=None,
                    attachments=[],
                    received_at=None,
                    date=None,
                    external_id="supplier-msg-text-config",
                    uid=None,
                    folder_name="INBOX",
                ),
                None,
            )
        ]

    monkeypatch.setattr(
        (
            "dz_fastapi.services.supplier_order_responses."
            "_fetch_supplier_response_messages"
        ),
        fake_fetch_messages,
    )
    monkeypatch.setattr(
        "dz_fastapi.services.supplier_order_responses.SUPPLIER_RESPONSE_DIR",
        str(tmp_path),
    )

    result = await process_supplier_response_messages(test_session)
    await test_session.refresh(item)
    message_row = (
        await test_session.execute(
            select(SupplierOrderMessage).where(
                SupplierOrderMessage.source_message_id
                == "supplier-msg-text-config"
            )
        )
    ).scalar_one()

    assert result["processed_messages"] == 1
    assert result["parsed_text_positions"] == 3
    assert result["recognized_positions"] >= 2
    assert result["unresolved_positions"] >= 1
    assert item.confirmed_quantity == 5
    assert message_row.message_type == "TEXT_RESPONSE"


@pytest.mark.asyncio
async def test_auto_confirm_timeout_creates_draft_receipt(
    monkeypatch,
    test_session,
    created_providers,
    created_autopart,
):
    provider = created_providers[0]
    provider.email_contact = None
    provider.email_incoming_price = None
    order = SupplierOrder(
        provider_id=provider.id,
        status=SUPPLIER_ORDER_STATUS.SENT,
        sent_at=now_moscow() - timedelta(minutes=45),
    )
    test_session.add(order)
    await test_session.flush()
    item = SupplierOrderItem(
        supplier_order_id=order.id,
        autopart_id=created_autopart.id,
        oem_number=created_autopart.oem_number,
        brand_name=created_autopart.brand.name,
        autopart_name=created_autopart.name,
        quantity=3,
        price=150.0,
    )
    test_session.add(item)
    test_session.add(
        SupplierResponseConfig(
            provider_id=provider.id,
            name="Timeout config",
            sender_emails=["timeout@example.com"],
            response_type="text",
            auto_confirm_after_minutes=40,
            is_active=True,
        )
    )
    await test_session.commit()

    async def fake_fetch_messages(
        session, *, date_from, date_to=None, **kwargs
    ):
        return []

    monkeypatch.setattr(
        (
            "dz_fastapi.services.supplier_order_responses."
            "_fetch_supplier_response_messages"
        ),
        fake_fetch_messages,
    )

    result = await process_supplier_response_messages(test_session)
    await test_session.refresh(item)

    receipt = (
        await test_session.execute(
            select(SupplierReceipt).where(
                SupplierReceipt.provider_id == provider.id
            )
        )
    ).scalar_one()
    receipt_items = (
        await test_session.execute(
            select(SupplierReceiptItem).where(
                SupplierReceiptItem.receipt_id == receipt.id
            )
        )
    ).scalars().all()

    assert result["timeout_auto_confirmed_orders"] == 1
    assert result["created_receipts"] == 1
    assert result["draft_receipts"] == 1
    assert item.confirmed_quantity == 3
    assert receipt.posted_at is None
    assert len(receipt_items) == 1
    assert receipt_items[0].received_quantity == 3


@pytest.mark.asyncio
async def test_auto_confirm_timeout_ignores_import_error_message(
    monkeypatch,
    test_session,
    created_providers,
    created_autopart,
):
    provider = created_providers[0]
    provider.email_contact = None
    provider.email_incoming_price = None
    order = SupplierOrder(
        provider_id=provider.id,
        status=SUPPLIER_ORDER_STATUS.SENT,
        sent_at=now_moscow() - timedelta(minutes=50),
    )
    test_session.add(order)
    await test_session.flush()
    item = SupplierOrderItem(
        supplier_order_id=order.id,
        autopart_id=created_autopart.id,
        oem_number=created_autopart.oem_number,
        brand_name=created_autopart.brand.name,
        autopart_name=created_autopart.name,
        quantity=4,
        price=220.0,
    )
    test_session.add(item)
    test_session.add(
        SupplierResponseConfig(
            provider_id=provider.id,
            name="Timeout config",
            sender_emails=["timeout@example.com"],
            response_type="text",
            auto_confirm_after_minutes=40,
            is_active=True,
        )
    )
    test_session.add(
        SupplierOrderMessage(
            supplier_order_id=order.id,
            provider_id=provider.id,
            message_type="IMPORT_ERROR",
            sender_email="timeout@example.com",
            received_at=now_moscow() - timedelta(minutes=20),
            source_uid="1:INBOX:111",
        )
    )
    await test_session.commit()

    async def fake_fetch_messages(
        session, *, date_from, date_to=None, **kwargs
    ):
        return []

    monkeypatch.setattr(
        (
            "dz_fastapi.services.supplier_order_responses."
            "_fetch_supplier_response_messages"
        ),
        fake_fetch_messages,
    )

    result = await process_supplier_response_messages(test_session)
    await test_session.refresh(item)

    assert result["timeout_auto_confirmed_orders"] == 1
    assert item.confirmed_quantity == 4


@pytest.mark.asyncio
async def test_document_payload_includes_unmatched_rows_and_rejects_missing(
    monkeypatch,
    test_session,
    created_providers,
    created_autopart,
    tmp_path,
):
    provider = created_providers[0]
    provider.email_contact = "supplier@example.com"
    order = SupplierOrder(
        provider_id=provider.id,
        status=SUPPLIER_ORDER_STATUS.SENT,
    )
    test_session.add(order)
    await test_session.flush()
    matched_item = SupplierOrderItem(
        supplier_order_id=order.id,
        autopart_id=created_autopart.id,
        oem_number=created_autopart.oem_number,
        brand_name=created_autopart.brand.name,
        autopart_name=created_autopart.name,
        quantity=5,
        price=120.0,
    )
    missing_item = SupplierOrderItem(
        supplier_order_id=order.id,
        autopart_id=created_autopart.id,
        oem_number="MISS-002",
        brand_name=created_autopart.brand.name,
        autopart_name="Missing row item",
        quantity=4,
        price=90.0,
    )
    test_session.add_all([matched_item, missing_item])
    test_session.add(
        SupplierResponseConfig(
            provider_id=provider.id,
            name="Doc parser with extras",
            sender_emails=["supplier@example.com"],
            response_type="file",
            file_payload_type="document",
            file_format="excel",
            start_row=1,
            oem_col=1,
            brand_col=2,
            qty_col=3,
            price_col=4,
            process_shipping_docs=True,
            is_active=True,
        )
    )
    await test_session.commit()

    response_frame = pd.DataFrame(
        [
            [
                created_autopart.oem_number,
                created_autopart.brand.name,
                2,
                101.0,
            ],
            ["EXTRA-999", "EXTRA-BRAND", 7, 333.0],
        ]
    )
    buffer = BytesIO()
    response_frame.to_excel(buffer, index=False, header=False)
    payload = buffer.getvalue()

    async def fake_fetch_messages(
        session, *, date_from, date_to=None, **kwargs
    ):
        return [
            (
                SimpleNamespace(
                    from_="supplier@example.com",
                    subject=f"Заказ поставщику #{order.id}: документ поставки",
                    text="",
                    html=None,
                    attachments=[
                        SimpleNamespace(
                            filename="doc_with_extra.xlsx",
                            payload=payload,
                        )
                    ],
                    received_at=None,
                    date=None,
                    external_id="supplier-msg-doc-extra",
                    uid=None,
                    folder_name="INBOX",
                ),
                None,
            )
        ]

    monkeypatch.setattr(
        (
            "dz_fastapi.services.supplier_order_responses."
            "_fetch_supplier_response_messages"
        ),
        fake_fetch_messages,
    )
    monkeypatch.setattr(
        "dz_fastapi.services.supplier_order_responses.SUPPLIER_RESPONSE_DIR",
        str(tmp_path),
    )

    result = await process_supplier_response_messages(test_session)
    await test_session.refresh(matched_item)
    await test_session.refresh(missing_item)
    receipt = (
        await test_session.execute(
            select(SupplierReceipt).where(
                SupplierReceipt.provider_id == provider.id
            )
        )
    ).scalar_one()
    receipt_items = (
        await test_session.execute(
            select(SupplierReceiptItem).where(
                SupplierReceiptItem.receipt_id == receipt.id
            )
        )
    ).scalars().all()
    linked_rows = [
        row
        for row in receipt_items
        if row.supplier_order_item_id == matched_item.id
    ]
    unlinked_rows = [
        row for row in receipt_items if row.supplier_order_item_id is None
    ]

    assert result["processed_messages"] == 1
    assert result["posted_receipts"] == 1
    assert len(linked_rows) == 1
    assert linked_rows[0].received_quantity == 2
    assert len(unlinked_rows) == 1
    assert unlinked_rows[0].oem_number == "EXTRA999"
    assert unlinked_rows[0].received_quantity == 7
    assert missing_item.confirmed_quantity == 0
    assert missing_item.response_status_raw == "автоотказ по документу"


@pytest.mark.asyncio
async def test_text_response_global_keyword_confirm_creates_draft_receipt(
    monkeypatch,
    test_session,
    created_providers,
    created_autopart,
    tmp_path,
):
    provider = created_providers[0]
    provider.email_contact = None
    provider.email_incoming_price = None
    order = SupplierOrder(
        provider_id=provider.id,
        status=SUPPLIER_ORDER_STATUS.SENT,
    )
    test_session.add(order)
    await test_session.flush()
    item = SupplierOrderItem(
        supplier_order_id=order.id,
        autopart_id=created_autopart.id,
        oem_number=created_autopart.oem_number,
        brand_name=created_autopart.brand.name,
        autopart_name=created_autopart.name,
        quantity=3,
        price=115.0,
    )
    test_session.add(item)
    test_session.add(
        SupplierResponseConfig(
            provider_id=provider.id,
            name="Text status config",
            sender_emails=["supplier@example.com"],
            response_type="text",
            confirm_keywords=["подтверждаем", "в работе"],
            reject_keywords=["нет", "отказ"],
            is_active=True,
        )
    )
    await test_session.commit()

    async def fake_fetch_messages(
        session, *, date_from, date_to=None, **kwargs
    ):
        return [
            (
                SimpleNamespace(
                    from_="supplier@example.com",
                    subject=f"Ответ по заказу #{order.id}",
                    text="Подтверждаем заказ, готовим к отгрузке.",
                    html=None,
                    attachments=[],
                    received_at=None,
                    date=None,
                    external_id="supplier-msg-text-global",
                    uid=None,
                    folder_name="INBOX",
                ),
                None,
            )
        ]

    monkeypatch.setattr(
        (
            "dz_fastapi.services.supplier_order_responses."
            "_fetch_supplier_response_messages"
        ),
        fake_fetch_messages,
    )
    monkeypatch.setattr(
        "dz_fastapi.services.supplier_order_responses.SUPPLIER_RESPONSE_DIR",
        str(tmp_path),
    )

    result = await process_supplier_response_messages(test_session)
    await test_session.refresh(item)
    message_row = (
        await test_session.execute(
            select(SupplierOrderMessage).where(
                SupplierOrderMessage.source_message_id
                == "supplier-msg-text-global"
            )
        )
    ).scalar_one()
    receipt = (
        await test_session.execute(
            select(SupplierReceipt).where(
                SupplierReceipt.provider_id == provider.id
            )
        )
    ).scalar_one()
    receipt_item = (
        await test_session.execute(
            select(SupplierReceiptItem).where(
                SupplierReceiptItem.receipt_id == receipt.id
            )
        )
    ).scalar_one()

    assert result["processed_messages"] == 1
    assert result["draft_receipts"] == 1
    assert message_row.message_type == "TEXT_RESPONSE"
    assert message_row.import_error_details is None
    assert item.confirmed_quantity == 3
    assert receipt.posted_at is None
    assert receipt_item.received_quantity == 3


@pytest.mark.asyncio
async def test_text_response_parses_numeric_hyphen_oem_with_reserve_keyword(
    monkeypatch,
    test_session,
    created_providers,
    created_autopart,
    tmp_path,
):
    provider = created_providers[0]
    provider.email_contact = None
    provider.email_incoming_price = None
    order = SupplierOrder(
        provider_id=provider.id,
        status=SUPPLIER_ORDER_STATUS.SENT,
    )
    test_session.add(order)
    await test_session.flush()
    item = SupplierOrderItem(
        supplier_order_id=order.id,
        autopart_id=created_autopart.id,
        oem_number="90178-11001",
        brand_name=created_autopart.brand.name,
        autopart_name=created_autopart.name,
        quantity=9,
        price=70.0,
    )
    test_session.add(item)
    test_session.add(
        SupplierResponseConfig(
            provider_id=provider.id,
            name="Text reserve config",
            sender_emails=["supplier@example.com"],
            response_type="text",
            confirm_keywords=["в резерве"],
            reject_keywords=["нет", "0"],
            value_after_article_type="text",
            is_active=True,
        )
    )
    await test_session.commit()

    async def fake_fetch_messages(
        session, *, date_from, date_to=None, **kwargs
    ):
        return [
            (
                SimpleNamespace(
                    from_="supplier@example.com",
                    subject=f"Заказ поставщику #{order.id}",
                    text="90178-11001 В РЕЗЕРВЕ",
                    html=None,
                    attachments=[],
                    received_at=None,
                    date=None,
                    external_id="supplier-msg-text-reserve",
                    uid=None,
                    folder_name="INBOX",
                ),
                None,
            )
        ]

    monkeypatch.setattr(
        (
            "dz_fastapi.services.supplier_order_responses."
            "_fetch_supplier_response_messages"
        ),
        fake_fetch_messages,
    )
    monkeypatch.setattr(
        "dz_fastapi.services.supplier_order_responses.SUPPLIER_RESPONSE_DIR",
        str(tmp_path),
    )

    result = await process_supplier_response_messages(test_session)
    await test_session.refresh(item)
    message_row = (
        await test_session.execute(
            select(SupplierOrderMessage).where(
                SupplierOrderMessage.source_message_id
                == "supplier-msg-text-reserve"
            )
        )
    ).scalar_one()

    assert result["processed_messages"] == 1
    assert result["recognized_positions"] >= 1
    assert message_row.message_type == "TEXT_RESPONSE"
    assert message_row.import_error_details is None
    assert item.confirmed_quantity == 9


@pytest.mark.asyncio
async def test_subject_pattern_filters_supplier_response_messages(
    monkeypatch,
    test_session,
    created_providers,
    tmp_path,
):
    provider = created_providers[0]
    provider.email_contact = None
    provider.email_incoming_price = None
    test_session.add(
        SupplierResponseConfig(
            provider_id=provider.id,
            name="Subject filter config",
            sender_emails=["supplier@example.com"],
            response_type="text",
            subject_pattern=r"^Ответ по заказу\\s+#?\\d+$",
            is_active=True,
        )
    )
    await test_session.commit()

    async def fake_fetch_messages(
        session, *, date_from, date_to=None, **kwargs
    ):
        return [
            (
                SimpleNamespace(
                    from_="supplier@example.com",
                    subject="Свободная тема без номера",
                    text="Подтверждаем.",
                    html=None,
                    attachments=[],
                    received_at=None,
                    date=None,
                    external_id="supplier-msg-subject-filter",
                    uid=None,
                    folder_name="INBOX",
                ),
                None,
            )
        ]

    monkeypatch.setattr(
        (
            "dz_fastapi.services.supplier_order_responses."
            "_fetch_supplier_response_messages"
        ),
        fake_fetch_messages,
    )
    monkeypatch.setattr(
        "dz_fastapi.services.supplier_order_responses.SUPPLIER_RESPONSE_DIR",
        str(tmp_path),
    )

    result = await process_supplier_response_messages(test_session)
    message_rows = (
        await test_session.execute(
            select(SupplierOrderMessage).where(
                SupplierOrderMessage.source_message_id
                == "supplier-msg-subject-filter"
            )
        )
    ).scalars().all()

    assert result["processed_messages"] == 0
    assert result["skipped_messages"] == 1
    assert message_rows == []
