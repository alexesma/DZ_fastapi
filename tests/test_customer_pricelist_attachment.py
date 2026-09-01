from io import BytesIO
from types import SimpleNamespace

import pandas as pd
from openpyxl import load_workbook

from dz_fastapi.services.process import (
    CUSTOMER_PRICELIST_COLUMN_WIDTHS,
    _build_customer_pricelist_attachment_bytes,
)


def test_customer_pricelist_xlsx_uses_readable_column_widths():
    frame = pd.DataFrame(
        [
            {
                "Производитель": "DRAGONZAP",
                "Наименование": "Длинное наименование автомобильной детали",
                "Артикул": "DZ12345678901234567890",
                "Количество": 10,
                "Кратность": 2,
                "Цена": 1500,
                "ТН ВЭД": "8708999709",
                "ОКПД 2": "29.32.30.390",
                "Честный знак": "Не маркируется",
                "Номер сертификата ЕАС": "Не требует сертификации",
                "Ссылка ФГИС": "https://pub.fsa.gov.ru/example",
            }
        ]
    )
    config = SimpleNamespace(export_file_format="xlsx")

    content = _build_customer_pricelist_attachment_bytes(frame, config)
    workbook = load_workbook(BytesIO(content))
    sheet = workbook["Прайс"]

    for index, title in enumerate(frame.columns, start=1):
        letter = sheet.cell(row=2, column=index).column_letter
        assert sheet.column_dimensions[letter].width == (
            CUSTOMER_PRICELIST_COLUMN_WIDTHS[title]
        )

    assert sheet["B3"].alignment.wrap_text is True
    assert sheet["C3"].alignment.wrap_text is True
    assert sheet["J3"].value == "Не требует сертификации"
    assert sheet["J3"].alignment.wrap_text is True
    assert sheet.freeze_panes == "A3"
